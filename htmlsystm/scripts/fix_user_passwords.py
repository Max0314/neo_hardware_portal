#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
修复MySQL中用户密码的脚本
- 检查所有用户的密码是否已哈希
- 如果密码是明文，使用默认密码或从Excel读取的密码进行哈希处理
- 确保MySQL和Excel中的密码同步
"""

import os
import sys
import logging

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from server.mysql_connection_pool import get_mysql_connection_pool
from server.security import PasswordHasher
from server.config import DATA_DIR
import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def is_password_hashed(password):
    """检查密码是否已哈希"""
    if not password:
        return False
    # bcrypt哈希以$2b$开头
    if password.startswith('$2b$'):
        return True
    # SHA256哈希格式：salt:hash
    if ':' in password and len(password.split(':')) == 2:
        return True
    return False

def get_password_from_excel(user_id, username):
    """从Excel文件中获取用户的密码"""
    users_file = os.path.join(DATA_DIR, 'users.xlsx')
    if not os.path.exists(users_file):
        logger.warning(f"Excel文件不存在: {users_file}")
        return None
    
    try:
        wb = openpyxl.load_workbook(users_file)
        
        # 先查找默认工作表
        if wb.active:
            ws = wb.active
            for row in range(2, ws.max_row + 1):
                cell_id = ws.cell(row=row, column=1).value
                cell_username = ws.cell(row=row, column=2).value
                if (cell_id and int(cell_id) == user_id) or (cell_username and str(cell_username).strip() == username):
                    password = ws.cell(row=row, column=3).value
                    if password and str(password).strip() and str(password).strip() != '***已加密***':
                        return str(password).strip()
        
        # 再查找钉钉用户数据表
        if "钉钉用户数据" in wb.sheetnames:
            ws_dingtalk = wb["钉钉用户数据"]
            # 创建字段名到列索引的映射
            from server.user_manager import DINGTALK_USER_HEADERS
            field_to_col = {header: idx + 1 for idx, header in enumerate(DINGTALK_USER_HEADERS)}
            
            for row in range(2, ws_dingtalk.max_row + 1):
                userid = ws_dingtalk.cell(row=row, column=field_to_col.get('userid', 2)).value
                job_number = ws_dingtalk.cell(row=row, column=field_to_col.get('job_number', 1)).value
                password_col = field_to_col.get('密码', 24)
                password = ws_dingtalk.cell(row=row, column=password_col).value
                
                # 检查是否匹配
                if (userid and str(userid).strip() == username) or (job_number and str(job_number).strip() == username):
                    if password and str(password).strip() and str(password).strip() != '***已加密***':
                        return str(password).strip()
    except Exception as e:
        logger.error(f"从Excel读取密码失败: {e}", exc_info=True)
    
    return None

def fix_user_passwords():
    """修复所有用户的密码"""
    try:
        pool = get_mysql_connection_pool()
        fixed_count = 0
        skipped_count = 0
        error_count = 0
        
        with pool.get_cursor() as cursor:
            # 获取所有用户
            cursor.execute('SELECT id, username, password FROM users')
            users = cursor.fetchall()
            
            logger.info(f"找到 {len(users)} 个用户，开始检查密码...")
            
            for user in users:
                user_id = user.get('id') if isinstance(user, dict) else user[0]
                username = user.get('username') if isinstance(user, dict) else user[1]
                password = user.get('password') if isinstance(user, dict) else user[2]
                
                if not password:
                    logger.warning(f"用户 {username} (ID={user_id}) 的密码为空，跳过")
                    skipped_count += 1
                    continue
                
                # 检查密码是否已哈希
                if is_password_hashed(password):
                    logger.debug(f"用户 {username} (ID={user_id}) 的密码已哈希，跳过")
                    skipped_count += 1
                    continue
                
                # 密码是明文，需要哈希
                logger.info(f"发现明文密码: 用户 {username} (ID={user_id})")
                
                # 尝试从Excel获取原始密码（如果MySQL中的密码不是原始密码）
                excel_password = get_password_from_excel(user_id, username)
                if excel_password and not is_password_hashed(excel_password):
                    # 使用Excel中的密码
                    password_to_hash = excel_password
                    logger.info(f"使用Excel中的密码进行哈希: {username}")
                else:
                    # 使用MySQL中的密码（假设是明文）
                    password_to_hash = password
                    logger.info(f"使用MySQL中的密码进行哈希: {username}")
                
                # 对密码进行哈希
                hashed_password = PasswordHasher.hash_password(password_to_hash)
                
                # 更新MySQL中的密码
                try:
                    cursor.execute('UPDATE users SET password = %s WHERE id = %s', (hashed_password, user_id))
                    logger.info(f"✅ 已修复用户 {username} (ID={user_id}) 的密码")
                    fixed_count += 1
                except Exception as e:
                    logger.error(f"❌ 更新用户 {username} (ID={user_id}) 的密码失败: {e}", exc_info=True)
                    error_count += 1
        
        logger.info("=" * 60)
        logger.info(f"密码修复完成:")
        logger.info(f"  - 已修复: {fixed_count} 个用户")
        logger.info(f"  - 已跳过: {skipped_count} 个用户（已哈希或为空）")
        logger.info(f"  - 错误: {error_count} 个用户")
        logger.info("=" * 60)
        
        return fixed_count, skipped_count, error_count
        
    except Exception as e:
        logger.error(f"修复密码失败: {e}", exc_info=True)
        return 0, 0, 1

if __name__ == '__main__':
    print("=" * 60)
    print("MySQL用户密码修复脚本")
    print("=" * 60)
    print()
    print("此脚本将:")
    print("  1. 检查所有用户的密码是否已哈希")
    print("  2. 如果密码是明文，将其哈希后更新到MySQL")
    print("  3. 尝试从Excel文件中获取原始密码（如果存在）")
    print()
    
    response = input("是否继续？(y/n): ")
    if response.lower() != 'y':
        print("已取消")
        sys.exit(0)
    
    print()
    fixed, skipped, errors = fix_user_passwords()
    
    if fixed > 0:
        print()
        print("✅ 密码修复完成！请重启服务器以使更改生效。")
    elif errors > 0:
        print()
        print("❌ 修复过程中出现错误，请检查日志。")
    else:
        print()
        print("ℹ️  所有密码都已正确哈希，无需修复。")

