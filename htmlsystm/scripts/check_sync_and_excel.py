#!/usr/bin/env python3
"""
检查同步钉钉用户功能是否正常工作
- 检查MySQL表结构
- 检查Excel文件
- 检查数据一致性
"""

import os
import sys
import json

# 添加项目根目录到路径
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

# 加载MySQL环境变量
mysql_env_file = os.path.join(project_root, '.mysql.env')
if os.path.exists(mysql_env_file):
    try:
        from dotenv import load_dotenv
        load_dotenv(dotenv_path=mysql_env_file, override=True)
    except ImportError:
        with open(mysql_env_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    if line.startswith('export '):
                        line = line[7:]
                    if '=' in line:
                        key, value = line.split('=', 1)
                        key = key.strip()
                        value = value.strip().strip("'\"")
                        os.environ[key] = value

from server.config import DATA_DIR, USE_MYSQL, MYSQL_CONFIG
from server.logger import logger
from server.user_manager import UserManager
from server.mysql_connection_pool import get_mysql_connection_pool
import openpyxl

def check_mysql_table_structure():
    """检查MySQL表结构"""
    print("\n" + "=" * 60)
    print("检查MySQL表结构")
    print("=" * 60)
    
    if not USE_MYSQL:
        print("⚠️  MySQL未启用")
        return False
    
    try:
        pool = get_mysql_connection_pool()
        with pool.get_cursor() as cursor:
            # 检查表是否存在
            cursor.execute("""
                SELECT TABLE_NAME 
                FROM INFORMATION_SCHEMA.TABLES 
                WHERE TABLE_SCHEMA = %s 
                AND TABLE_NAME = 'users'
            """, (MYSQL_CONFIG['database'],))
            
            if not cursor.fetchone():
                print("❌ users表不存在")
                return False
            
            print("✅ users表存在")
            
            # 检查dingtalk_data字段
            cursor.execute("""
                SELECT COLUMN_NAME, DATA_TYPE, IS_NULLABLE
                FROM INFORMATION_SCHEMA.COLUMNS 
                WHERE TABLE_SCHEMA = %s 
                AND TABLE_NAME = 'users' 
                AND COLUMN_NAME = 'dingtalk_data'
            """, (MYSQL_CONFIG['database'],))
            
            row = cursor.fetchone()
            if row:
                # 处理不同的返回格式（dict或tuple）
                if isinstance(row, dict):
                    data_type = row.get('DATA_TYPE', 'unknown')
                    is_nullable = row.get('IS_NULLABLE', 'unknown')
                else:
                    data_type = row[1] if len(row) > 1 else 'unknown'
                    is_nullable = row[2] if len(row) > 2 else 'unknown'
                print(f"✅ dingtalk_data字段存在 (类型: {data_type}, 可空: {is_nullable})")
                return True
            else:
                print("❌ dingtalk_data字段不存在")
                print("   建议运行: python3 scripts/add_dingtalk_fields_to_mysql.py")
                return False
                
    except Exception as e:
        logger.error(f"检查MySQL表结构失败: {e}", exc_info=True)
        print(f"❌ 检查失败: {e}")
        return False

def check_excel_file():
    """检查Excel文件"""
    print("\n" + "=" * 60)
    print("检查Excel文件")
    print("=" * 60)
    
    try:
        user_manager = UserManager()
        users_file = user_manager.users_file
        
        if not os.path.exists(users_file):
            print(f"❌ Excel文件不存在: {users_file}")
            return False
        
        print(f"✅ Excel文件存在: {users_file}")
        
        # 检查文件是否损坏
        try:
            wb = openpyxl.load_workbook(users_file)
            print("✅ Excel文件可以正常打开")
        except Exception as e:
            print(f"❌ Excel文件损坏: {e}")
            return False
        
        # 检查工作表
        if "钉钉用户数据" in wb.sheetnames:
            ws = wb["钉钉用户数据"]
            row_count = ws.max_row - 1  # 减去表头
            print(f"✅ '钉钉用户数据'工作表存在，有 {row_count} 行数据（不含表头）")
            
            # 检查是否有数据
            if row_count > 0:
                print("✅ Excel中有钉钉用户数据")
                return True
            else:
                print("⚠️  Excel中'钉钉用户数据'工作表为空")
                return False
        else:
            print("❌ Excel中不存在'钉钉用户数据'工作表")
            return False
            
    except Exception as e:
        logger.error(f"检查Excel文件失败: {e}", exc_info=True)
        print(f"❌ 检查失败: {e}")
        return False

def check_data_consistency():
    """检查数据一致性"""
    print("\n" + "=" * 60)
    print("检查数据一致性")
    print("=" * 60)
    
    try:
        user_manager = UserManager()
        
        # 从Excel获取用户
        excel_users = user_manager.get_all_users()
        excel_dingtalk_users = [u for u in excel_users if u.get('source') == 'dingtalk']
        print(f"✅ Excel中有 {len(excel_dingtalk_users)} 个钉钉用户")
        
        if not USE_MYSQL:
            print("⚠️  MySQL未启用，跳过MySQL一致性检查")
            return
        
        # 从MySQL获取用户
        pool = get_mysql_connection_pool()
        with pool.get_cursor() as cursor:
            # 检查dingtalk_data字段是否存在
            cursor.execute("""
                SELECT COLUMN_NAME 
                FROM INFORMATION_SCHEMA.COLUMNS 
                WHERE TABLE_SCHEMA = %s 
                AND TABLE_NAME = 'users' 
                AND COLUMN_NAME = 'dingtalk_data'
            """, (MYSQL_CONFIG['database'],))
            
            has_dingtalk_data = cursor.fetchone() is not None
            
            if not has_dingtalk_data:
                print("⚠️  MySQL中dingtalk_data字段不存在，无法进行一致性检查")
                print("   建议运行: python3 scripts/add_dingtalk_fields_to_mysql.py")
                return
            
            # 统计MySQL中有dingtalk_data的用户
            cursor.execute("""
                SELECT COUNT(*) FROM users 
                WHERE dingtalk_data IS NOT NULL 
                AND dingtalk_data != ''
            """)
            mysql_count = cursor.fetchone()[0]
            print(f"✅ MySQL中有 {mysql_count} 个用户有dingtalk_data")
            
            # 检查一致性
            excel_userids = {str(u.get('userid', '')) for u in excel_dingtalk_users if u.get('userid')}
            mysql_userids = set()
            
            cursor.execute("""
                SELECT JSON_UNQUOTE(JSON_EXTRACT(dingtalk_data, '$.userid')) as userid
                FROM users 
                WHERE dingtalk_data IS NOT NULL 
                AND dingtalk_data != ''
            """)
            rows = cursor.fetchall()
            for row in rows:
                userid = row[0] if isinstance(row, dict) else row[0]
                if userid:
                    mysql_userids.add(str(userid))
            
            print(f"✅ Excel中有 {len(excel_userids)} 个唯一的userid")
            print(f"✅ MySQL中有 {len(mysql_userids)} 个唯一的userid")
            
            # 找出不一致的用户
            excel_only = excel_userids - mysql_userids
            mysql_only = mysql_userids - excel_userids
            common = excel_userids & mysql_userids
            
            print(f"\n一致性统计:")
            print(f"  - 共同用户: {len(common)} 个")
            if excel_only:
                print(f"  - 仅在Excel中: {len(excel_only)} 个")
                if len(excel_only) <= 10:
                    print(f"    userid: {', '.join(list(excel_only)[:10])}")
            if mysql_only:
                print(f"  - 仅在MySQL中: {len(mysql_only)} 个")
                if len(mysql_only) <= 10:
                    print(f"    userid: {', '.join(list(mysql_only)[:10])}")
            
            if not excel_only and not mysql_only:
                print("✅ Excel和MySQL数据完全一致")
            else:
                print("⚠️  Excel和MySQL数据存在不一致")
                
    except Exception as e:
        logger.error(f"检查数据一致性失败: {e}", exc_info=True)
        print(f"❌ 检查失败: {e}")

def check_sync_functionality():
    """检查同步功能"""
    print("\n" + "=" * 60)
    print("检查同步功能")
    print("=" * 60)
    
    try:
        from server.config import DINGTALK_CONFIG
        
        # 检查配置
        if not DINGTALK_CONFIG.get('client_id') or not DINGTALK_CONFIG.get('client_secret'):
            print("❌ 钉钉配置不完整")
            print("   请检查 server/config.py 中的 DINGTALK_CONFIG")
            return False
        
        print("✅ 钉钉配置完整")
        
        # 检查save_dingtalk_users方法
        user_manager = UserManager()
        if hasattr(user_manager, 'save_dingtalk_users'):
            print("✅ save_dingtalk_users方法存在")
        else:
            print("❌ save_dingtalk_users方法不存在")
            return False
        
        # 检查_save_dingtalk_users_to_mysql方法
        if hasattr(user_manager, '_save_dingtalk_users_to_mysql'):
            print("✅ _save_dingtalk_users_to_mysql方法存在")
        else:
            print("❌ _save_dingtalk_users_to_mysql方法不存在")
            return False
        
        print("✅ 同步功能代码完整")
        return True
        
    except Exception as e:
        logger.error(f"检查同步功能失败: {e}", exc_info=True)
        print(f"❌ 检查失败: {e}")
        return False

def main():
    """主函数"""
    print("=" * 60)
    print("检查同步钉钉用户功能和数据一致性")
    print("=" * 60)
    
    # 检查MySQL表结构
    mysql_ok = check_mysql_table_structure()
    
    # 检查Excel文件
    excel_ok = check_excel_file()
    
    # 检查同步功能
    sync_ok = check_sync_functionality()
    
    # 检查数据一致性
    if mysql_ok and excel_ok:
        check_data_consistency()
    
    print("\n" + "=" * 60)
    print("检查完成")
    print("=" * 60)
    print()
    
    if not mysql_ok:
        print("⚠️  建议操作:")
        print("   1. 运行: python3 scripts/add_dingtalk_fields_to_mysql.py")
        print("   2. 然后重新运行同步脚本")
    elif not excel_ok:
        print("⚠️  建议操作:")
        print("   1. 在用户管理页面点击'同步钉钉用户'")
        print("   2. 或运行: python3 scripts/fix_mysql_users_from_dingtalk.py")

if __name__ == '__main__':
    main()

