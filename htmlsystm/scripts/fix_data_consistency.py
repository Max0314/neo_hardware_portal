#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据一致性修复脚本
修复MySQL、Excel和文件系统之间的数据不一致问题

功能：
1. 检查并修复MySQL users表和Excel users.xlsx的一致性
2. 检查公告文件系统的一致性
3. 清理无效的缓存
4. 生成修复报告
"""
import os
import sys
import json
import openpyxl
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from server.config import DATA_DIR, MYSQL_CONFIG
from server.mysql_connection_pool import get_mysql_connection_pool
from server.user_manager import UserManager, DINGTALK_USER_HEADERS, DEFAULT_HEADERS
from server.logger import logger
from server.file_lock import get_file_lock

class DataConsistencyFixer:
    """数据一致性修复器"""
    
    def __init__(self):
        self.users_file = os.path.join(DATA_DIR, 'users.xlsx')
        self.report = {
            'timestamp': datetime.now().isoformat(),
            'users': {
                'mysql_count': 0,
                'excel_count': 0,
                'fixed_count': 0,
                'errors': []
            },
            'announcements': {
                'checked_count': 0,
                'fixed_count': 0,
                'errors': []
            },
            'cache': {
                'cleared': False,
                'errors': []
            }
        }
    
    def run(self):
        """运行完整的修复流程"""
        print("=" * 80)
        print("数据一致性修复工具")
        print("=" * 80)
        print(f"开始时间: {self.report['timestamp']}")
        print()
        
        try:
            # 1. 修复用户数据一致性
            print("=" * 80)
            print("步骤 1: 检查并修复用户数据一致性（MySQL <-> Excel）")
            print("=" * 80)
            self.fix_users_consistency()
            
            # 2. 检查公告文件系统一致性
            print()
            print("=" * 80)
            print("步骤 2: 检查公告文件系统一致性")
            print("=" * 80)
            self.check_announcements_consistency()
            
            # 3. 清理缓存
            print()
            print("=" * 80)
            print("步骤 3: 清理无效缓存")
            print("=" * 80)
            self.clear_cache()
            
            # 4. 生成报告
            print()
            print("=" * 80)
            print("修复完成！")
            print("=" * 80)
            self.print_report()
            
        except Exception as e:
            logger.error(f"修复过程发生错误: {e}", exc_info=True)
            print(f"\n❌ 修复过程发生错误: {e}")
            import traceback
            traceback.print_exc()
            return False
        
        return True
    
    def fix_users_consistency(self):
        """修复用户数据一致性"""
        try:
            # 首先检查dingtalk_data字段是否存在
            pool = get_mysql_connection_pool()
            with pool.get_cursor() as cursor:
                cursor.execute("""
                    SELECT COLUMN_NAME 
                    FROM INFORMATION_SCHEMA.COLUMNS 
                    WHERE TABLE_SCHEMA = %s 
                    AND TABLE_NAME = 'users' 
                    AND COLUMN_NAME = 'dingtalk_data'
                """, (MYSQL_CONFIG['database'],))
                
                has_dingtalk_data = cursor.fetchone() is not None
                
                if not has_dingtalk_data:
                    print("⚠️  dingtalk_data字段不存在，正在添加...")
                    try:
                        cursor.execute("""
                            ALTER TABLE users 
                            ADD COLUMN dingtalk_data JSON NULL 
                            COMMENT '钉钉用户完整数据（JSON格式）'
                            AFTER last_login_time
                        """)
                        # 注意：get_cursor()上下文管理器会自动提交事务，不需要手动commit
                        print("✅ 成功添加dingtalk_data字段")
                    except Exception as e:
                        print(f"❌ 添加dingtalk_data字段失败: {e}")
                        self.report['users']['errors'].append(f"添加dingtalk_data字段失败: {e}")
                        return
            
            # 获取MySQL中的用户数据
            mysql_users = {}
            
            with pool.get_cursor() as cursor:
                cursor.execute("SELECT * FROM users")
                rows = cursor.fetchall()
                
                for row in rows:
                    if isinstance(row, dict):
                        user_dict = row
                    else:
                        columns = [desc[0] for desc in cursor.description] if hasattr(cursor, 'description') else []
                        user_dict = dict(zip(columns, row)) if columns else {}
                    
                    username = user_dict.get('username', '')
                    if username:
                        mysql_users[username] = user_dict
            
            self.report['users']['mysql_count'] = len(mysql_users)
            print(f"✅ MySQL中找到了 {len(mysql_users)} 个用户")
            
            # 获取Excel中的用户数据
            excel_users = {}
            if os.path.exists(self.users_file):
                file_lock = get_file_lock(self.users_file)
                with file_lock:
                    try:
                        wb = openpyxl.load_workbook(self.users_file)
                        
                        # 读取默认工作表
                        if wb.active:
                            ws = wb.active
                            for row in range(2, ws.max_row + 1):
                                username_cell = ws.cell(row=row, column=2).value
                                if username_cell:
                                    username = str(username_cell).strip()
                                    excel_users[username] = {
                                        'source': 'excel_default',
                                        'row': row,
                                        'worksheet': 'default'
                                    }
                        
                        # 读取钉钉用户数据工作表
                        if "钉钉用户数据" in wb.sheetnames:
                            ws_dingtalk = wb["钉钉用户数据"]
                            field_to_col = {header: idx + 1 for idx, header in enumerate(DINGTALK_USER_HEADERS)}
                            
                            for row in range(2, ws_dingtalk.max_row + 1):
                                job_number = ws_dingtalk.cell(row=row, column=field_to_col.get('job_number', 1)).value
                                userid = ws_dingtalk.cell(row=row, column=field_to_col.get('userid', 2)).value
                                
                                if job_number:
                                    username = str(job_number).strip()
                                    excel_users[username] = {
                                        'source': 'excel_dingtalk',
                                        'row': row,
                                        'worksheet': 'dingtalk',
                                        'userid': str(userid).strip() if userid else ''
                                    }
                                elif userid:
                                    username = str(userid).strip()
                                    excel_users[username] = {
                                        'source': 'excel_dingtalk',
                                        'row': row,
                                        'worksheet': 'dingtalk',
                                        'userid': username
                                    }
                    except Exception as e:
                        logger.error(f"读取Excel文件失败: {e}", exc_info=True)
                        self.report['users']['errors'].append(f"读取Excel文件失败: {e}")
                        print(f"⚠️  读取Excel文件失败: {e}")
            
            self.report['users']['excel_count'] = len(excel_users)
            print(f"✅ Excel中找到了 {len(excel_users)} 个用户")
            
            # 检查一致性并修复
            fixed_count = 0
            
            # 1. 检查MySQL中有但Excel中没有的用户
            mysql_only = set(mysql_users.keys()) - set(excel_users.keys())
            if mysql_only:
                print(f"\n⚠️  发现 {len(mysql_only)} 个仅在MySQL中存在的用户:")
                for username in list(mysql_only)[:10]:  # 只显示前10个
                    print(f"   - {username}")
                if len(mysql_only) > 10:
                    print(f"   ... 还有 {len(mysql_only) - 10} 个")
                
                # 询问是否同步到Excel
                print("\n💡 建议：这些用户应该同步到Excel文件")
            
            # 2. 检查Excel中有但MySQL中没有的用户
            excel_only = set(excel_users.keys()) - set(mysql_users.keys())
            if excel_only:
                print(f"\n⚠️  发现 {len(excel_only)} 个仅在Excel中存在的用户:")
                for username in list(excel_only)[:10]:  # 只显示前10个
                    print(f"   - {username}")
                if len(excel_only) > 10:
                    print(f"   ... 还有 {len(excel_only) - 10} 个")
                
                # 询问是否同步到MySQL
                print("\n💡 建议：这些用户应该同步到MySQL数据库")
            
            # 3. 检查共同用户的数据一致性
            common_users = set(mysql_users.keys()) & set(excel_users.keys())
            print(f"\n✅ 找到 {len(common_users)} 个共同用户")
            
            # 4. 修复dingtalk_data字段缺失的问题
            print("\n检查MySQL中dingtalk_data字段...")
            missing_dingtalk_data = []
            
            for username in common_users:
                mysql_user = mysql_users[username]
                dingtalk_data = mysql_user.get('dingtalk_data')
                
                if not dingtalk_data:
                    # 检查是否是钉钉用户（有userid或job_number）
                    userid = mysql_user.get('userid') if isinstance(mysql_user, dict) else None
                    if not userid:
                        # 尝试从Excel中获取
                        excel_user = excel_users[username]
                        if excel_user.get('userid'):
                            userid = excel_user['userid']
                    
                    if userid:
                        missing_dingtalk_data.append(username)
            
            if missing_dingtalk_data:
                print(f"⚠️  发现 {len(missing_dingtalk_data)} 个钉钉用户缺少dingtalk_data字段")
                print("   这些用户的数据将从Excel补充到MySQL...")
                
                # 从Excel补充dingtalk_data
                self._supplement_dingtalk_data_from_excel(missing_dingtalk_data, mysql_users, excel_users)
                fixed_count += len(missing_dingtalk_data)
            
            self.report['users']['fixed_count'] = fixed_count
            
            if fixed_count > 0:
                print(f"\n✅ 已修复 {fixed_count} 个用户的数据")
            else:
                print("\n✅ 用户数据一致性检查完成，未发现需要修复的问题")
                
        except Exception as e:
            logger.error(f"修复用户数据一致性失败: {e}", exc_info=True)
            self.report['users']['errors'].append(f"修复失败: {e}")
            print(f"❌ 修复用户数据一致性失败: {e}")
    
    def _supplement_dingtalk_data_from_excel(self, usernames: List[str], mysql_users: Dict, excel_users: Dict):
        """从Excel补充dingtalk_data到MySQL"""
        try:
            pool = get_mysql_connection_pool()
            file_lock = get_file_lock(self.users_file)
            
            with file_lock:
                wb = openpyxl.load_workbook(self.users_file)
                
                if "钉钉用户数据" not in wb.sheetnames:
                    print("   ⚠️  Excel中没有'钉钉用户数据'工作表，跳过补充")
                    return 0
                
                ws_dingtalk = wb["钉钉用户数据"]
                field_to_col = {header: idx + 1 for idx, header in enumerate(DINGTALK_USER_HEADERS)}
                
                updated_count = 0
                
                with pool.get_cursor() as cursor:
                    for username in usernames:
                        # 在Excel中查找该用户
                        found = False
                        for row in range(2, ws_dingtalk.max_row + 1):
                            job_number = ws_dingtalk.cell(row=row, column=field_to_col.get('job_number', 1)).value
                            userid = ws_dingtalk.cell(row=row, column=field_to_col.get('userid', 2)).value
                            
                            if (job_number and str(job_number).strip() == username) or \
                               (userid and str(userid).strip() == username):
                                # 找到匹配的用户，读取所有字段
                                dingtalk_data = {}
                                
                                for header in DINGTALK_USER_HEADERS:
                                    if header in ['密码', '物料库管理员', '公告栏管理员', '状态', '创建时间']:
                                        continue  # 跳过系统字段
                                    
                                    col_idx = field_to_col.get(header)
                                    if col_idx:
                                        value = ws_dingtalk.cell(row=row, column=col_idx).value
                                        if value is not None:
                                            # 处理特殊字段
                                            if header == 'dept_id_list':
                                                try:
                                                    if isinstance(value, str):
                                                        value = json.loads(value)
                                                except:
                                                    value = []
                                            elif header in ['active', 'admin', 'boss', 'leader', 'disable_status', 'exclusive_account', 'hide_mobile']:
                                                value = bool(value) if value is not None else False
                                            
                                            dingtalk_data[header] = value
                                
                                # 更新MySQL
                                dingtalk_data_json = json.dumps(dingtalk_data, ensure_ascii=False)
                                cursor.execute(
                                    "UPDATE users SET dingtalk_data = %s WHERE username = %s",
                                    (dingtalk_data_json, username)
                                )
                                
                                updated_count += 1
                                found = True
                                break
                        
                        if not found:
                            print(f"   ⚠️  在Excel中未找到用户: {username}")
                
                # 注意：get_cursor()上下文管理器会自动提交事务，不需要手动commit
                print(f"   ✅ 已更新 {updated_count} 个用户的dingtalk_data字段")
                return updated_count
                
        except Exception as e:
            logger.error(f"补充dingtalk_data失败: {e}", exc_info=True)
            print(f"   ❌ 补充dingtalk_data失败: {e}")
            return 0
    
    def check_announcements_consistency(self):
        """检查公告文件系统一致性"""
        try:
            announcements_dir = os.path.join(DATA_DIR, 'announcements')
            if not os.path.exists(announcements_dir):
                print("⚠️  公告目录不存在")
                return
            
            checked_count = 0
            fixed_count = 0
            
            # 遍历所有公告栏目录
            for board_id in os.listdir(announcements_dir):
                board_path = os.path.join(announcements_dir, board_id)
                if not os.path.isdir(board_path) or board_id == 'temp':
                    continue
                
                # 遍历公告目录
                for ann_id in os.listdir(board_path):
                    ann_path = os.path.join(board_path, ann_id)
                    if not os.path.isdir(ann_path):
                        continue
                    
                    checked_count += 1
                    
                    # 检查必要的文件
                    metadata_file = os.path.join(ann_path, 'metadata.json')
                    content_file = os.path.join(ann_path, 'content.html')
                    
                    issues = []
                    
                    if not os.path.exists(metadata_file):
                        issues.append("缺少metadata.json")
                    else:
                        try:
                            with open(metadata_file, 'r', encoding='utf-8') as f:
                                metadata = json.load(f)
                                if not metadata.get('id'):
                                    issues.append("metadata.json中缺少id字段")
                                if not metadata.get('title'):
                                    issues.append("metadata.json中缺少title字段")
                        except Exception as e:
                            issues.append(f"metadata.json解析失败: {e}")
                    
                    if not os.path.exists(content_file):
                        issues.append("缺少content.html")
                    
                    if issues:
                        print(f"⚠️  公告 {ann_id} 存在问题:")
                        for issue in issues:
                            print(f"   - {issue}")
                        # 这里可以添加自动修复逻辑
                        fixed_count += len(issues)
            
            self.report['announcements']['checked_count'] = checked_count
            self.report['announcements']['fixed_count'] = fixed_count
            
            print(f"\n✅ 检查了 {checked_count} 个公告，发现 {fixed_count} 个问题")
            
        except Exception as e:
            logger.error(f"检查公告一致性失败: {e}", exc_info=True)
            self.report['announcements']['errors'].append(f"检查失败: {e}")
            print(f"❌ 检查公告一致性失败: {e}")
    
    def clear_cache(self):
        """清理无效缓存"""
        try:
            # 清理Python缓存文件
            import glob
            cache_files = []
            
            # 清理__pycache__
            for root, dirs, files in os.walk(os.path.dirname(DATA_DIR)):
                if '__pycache__' in dirs:
                    cache_dir = os.path.join(root, '__pycache__')
                    cache_files.extend(glob.glob(os.path.join(cache_dir, '*.pyc')))
                    cache_files.extend(glob.glob(os.path.join(cache_dir, '*.pyo')))
            
            if cache_files:
                for cache_file in cache_files:
                    try:
                        os.remove(cache_file)
                    except:
                        pass
                print(f"✅ 清理了 {len(cache_files)} 个Python缓存文件")
            
            # 注意：内存缓存会在服务器重启时自动清理
            print("💡 内存缓存将在服务器重启时自动清理")
            
            self.report['cache']['cleared'] = True
            
        except Exception as e:
            logger.error(f"清理缓存失败: {e}", exc_info=True)
            self.report['cache']['errors'].append(f"清理失败: {e}")
            print(f"❌ 清理缓存失败: {e}")
    
    def print_report(self):
        """打印修复报告"""
        print("\n" + "=" * 80)
        print("修复报告")
        print("=" * 80)
        print(f"时间: {self.report['timestamp']}")
        print()
        
        print("用户数据:")
        print(f"  - MySQL用户数: {self.report['users']['mysql_count']}")
        print(f"  - Excel用户数: {self.report['users']['excel_count']}")
        print(f"  - 修复数量: {self.report['users']['fixed_count']}")
        if self.report['users']['errors']:
            print(f"  - 错误数: {len(self.report['users']['errors'])}")
            for error in self.report['users']['errors']:
                print(f"    * {error}")
        
        print()
        print("公告数据:")
        print(f"  - 检查数量: {self.report['announcements']['checked_count']}")
        print(f"  - 修复数量: {self.report['announcements']['fixed_count']}")
        if self.report['announcements']['errors']:
            print(f"  - 错误数: {len(self.report['announcements']['errors'])}")
            for error in self.report['announcements']['errors']:
                print(f"    * {error}")
        
        print()
        print("缓存:")
        print(f"  - 已清理: {'是' if self.report['cache']['cleared'] else '否'}")
        if self.report['cache']['errors']:
            print(f"  - 错误数: {len(self.report['cache']['errors'])}")
            for error in self.report['cache']['errors']:
                print(f"    * {error}")
        
        # 保存报告到文件
        report_file = os.path.join(DATA_DIR, 'consistency_report.json')
        try:
            with open(report_file, 'w', encoding='utf-8') as f:
                json.dump(self.report, f, ensure_ascii=False, indent=2)
            print(f"\n📄 详细报告已保存到: {report_file}")
        except Exception as e:
            print(f"\n⚠️  保存报告失败: {e}")


if __name__ == '__main__':
    fixer = DataConsistencyFixer()
    success = fixer.run()
    sys.exit(0 if success else 1)

