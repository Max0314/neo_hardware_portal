#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MySQL数据查询脚本
用于诊断MySQL数据库中的数据情况，为修复脚本提供依据

功能：
1. 查询users表的数据统计
2. 检查dingtalk_data字段的情况
3. 对比MySQL和Excel的数据差异
4. 检查公告相关的数据
"""
import os
import sys
import json
import openpyxl
from collections import defaultdict
from typing import Dict, List, Any, Optional

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from server.config import DATA_DIR, MYSQL_CONFIG
from server.mysql_connection_pool import get_mysql_connection_pool
from server.user_manager import DINGTALK_USER_HEADERS
from server.logger import logger

class MySQLDataQuery:
    """MySQL数据查询器"""
    
    def __init__(self):
        self.users_file = os.path.join(DATA_DIR, 'users.xlsx')
        self.results = {
            'users': {},
            'announcements': {},
            'comparison': {}
        }
    
    def run(self):
        """运行完整的查询流程"""
        print("=" * 80)
        print("MySQL数据查询工具")
        print("=" * 80)
        print()
        
        try:
            # 1. 查询users表统计
            print("=" * 80)
            print("1. MySQL users表数据统计")
            print("=" * 80)
            self.query_users_statistics()
            
            # 2. 检查dingtalk_data字段
            print()
            print("=" * 80)
            print("2. dingtalk_data字段检查")
            print("=" * 80)
            self.check_dingtalk_data()
            
            # 3. 对比MySQL和Excel
            print()
            print("=" * 80)
            print("3. MySQL vs Excel 数据对比")
            print("=" * 80)
            self.compare_mysql_excel()
            
            # 4. 查询公告相关数据
            print()
            print("=" * 80)
            print("4. 公告相关数据查询")
            print("=" * 80)
            self.query_announcements_data()
            
            # 5. 生成报告
            print()
            print("=" * 80)
            print("查询完成！")
            print("=" * 80)
            self.save_report()
            
        except Exception as e:
            logger.error(f"查询过程发生错误: {e}", exc_info=True)
            print(f"\n❌ 查询过程发生错误: {e}")
            import traceback
            traceback.print_exc()
            return False
        
        return True
    
    def query_users_statistics(self):
        """查询users表统计信息"""
        try:
            pool = get_mysql_connection_pool()
            
            with pool.get_cursor() as cursor:
                # 总用户数
                cursor.execute("SELECT COUNT(*) as total FROM users")
                row = cursor.fetchone()
                if row:
                    total = row['total'] if isinstance(row, dict) else row[0]
                else:
                    total = 0
                print(f"✅ 总用户数: {total}")
                
                # 按状态统计
                cursor.execute("""
                    SELECT status, COUNT(*) as count 
                    FROM users 
                    GROUP BY status
                """)
                status_stats = {}
                for row in cursor.fetchall():
                    if isinstance(row, dict):
                        status = row['status']
                        count = row['count']
                    else:
                        status = row[0]
                        count = row[1]
                    status_stats[status] = count
                
                print("\n按状态统计:")
                for status, count in sorted(status_stats.items()):
                    print(f"  - {status}: {count}")
                
                # 按部门统计
                cursor.execute("""
                    SELECT department, COUNT(*) as count 
                    FROM users 
                    GROUP BY department
                    ORDER BY count DESC
                """)
                dept_stats = {}
                for row in cursor.fetchall():
                    if isinstance(row, dict):
                        dept = row['department']
                        count = row['count']
                    else:
                        dept = row[0]
                        count = row[1]
                    dept_stats[dept] = count
                
                print("\n按部门统计:")
                for dept, count in list(dept_stats.items())[:10]:
                    print(f"  - {dept}: {count}")
                if len(dept_stats) > 10:
                    print(f"  ... 还有 {len(dept_stats) - 10} 个部门")
                
                # 有job_position的用户数
                cursor.execute("""
                    SELECT COUNT(*) as count 
                    FROM users 
                    WHERE job_position IS NOT NULL AND job_position != ''
                """)
                row = cursor.fetchone()
                if row:
                    job_pos_count = row['count'] if isinstance(row, dict) else row[0]
                else:
                    job_pos_count = 0
                print(f"\n有岗位信息的用户数: {job_pos_count}")
                
                # 有roles的用户数
                cursor.execute("""
                    SELECT COUNT(*) as count 
                    FROM users 
                    WHERE roles IS NOT NULL AND roles != ''
                """)
                row = cursor.fetchone()
                if row:
                    roles_count = row['count'] if isinstance(row, dict) else row[0]
                else:
                    roles_count = 0
                print(f"有角色信息的用户数: {roles_count}")
                
                self.results['users'] = {
                    'total': total,
                    'status_stats': status_stats,
                    'dept_stats': dept_stats,
                    'job_pos_count': job_pos_count,
                    'roles_count': roles_count
                }
                
        except Exception as e:
            logger.error(f"查询users统计失败: {e}", exc_info=True)
            print(f"❌ 查询users统计失败: {e}")
    
    def check_dingtalk_data(self):
        """检查dingtalk_data字段"""
        try:
            pool = get_mysql_connection_pool()
            
            with pool.get_cursor() as cursor:
                # 检查dingtalk_data字段是否存在
                cursor.execute("""
                    SELECT COLUMN_NAME 
                    FROM INFORMATION_SCHEMA.COLUMNS 
                    WHERE TABLE_SCHEMA = %s 
                    AND TABLE_NAME = 'users' 
                    AND COLUMN_NAME = 'dingtalk_data'
                """, (MYSQL_CONFIG['database'],))
                
                has_dingtalk_data = cursor.fetchone() is not None
                
                if not has_dingtalk_data:
                    print("⚠️  dingtalk_data字段不存在！")
                    print("   建议运行: python3 scripts/add_dingtalk_fields_to_mysql.py")
                    return
                
                print("✅ dingtalk_data字段存在")
                
                # 统计有dingtalk_data的用户数
                cursor.execute("""
                    SELECT COUNT(*) as count 
                    FROM users 
                    WHERE dingtalk_data IS NOT NULL AND dingtalk_data != ''
                """)
                row = cursor.fetchone()
                if row:
                    has_data_count = row['count'] if isinstance(row, dict) else row[0]
                else:
                    has_data_count = 0
                
                # 总用户数
                cursor.execute("SELECT COUNT(*) as total FROM users")
                row = cursor.fetchone()
                if row:
                    total = row['total'] if isinstance(row, dict) else row[0]
                else:
                    total = 0
                
                missing_count = total - has_data_count
                
                print(f"\n有dingtalk_data的用户数: {has_data_count}")
                print(f"缺少dingtalk_data的用户数: {missing_count}")
                
                if missing_count > 0:
                    print(f"\n⚠️  发现 {missing_count} 个用户缺少dingtalk_data字段")
                    
                    # 检查这些用户是否有钉钉相关字段（通过job_position判断）
                    cursor.execute("""
                        SELECT username, name, job_position, department
                        FROM users 
                        WHERE (dingtalk_data IS NULL OR dingtalk_data = '')
                        AND job_position IS NOT NULL AND job_position != ''
                        LIMIT 10
                    """)
                    
                    missing_users = []
                    for row in cursor.fetchall():
                        if isinstance(row, dict):
                            missing_users.append({
                                'username': row['username'],
                                'name': row['name'],
                                'job_position': row['job_position'],
                                'department': row['department']
                            })
                        else:
                            missing_users.append({
                                'username': row[0],
                                'name': row[1],
                                'job_position': row[2],
                                'department': row[3]
                            })
                    
                    if missing_users:
                        print("\n缺少dingtalk_data的用户示例（前10个）:")
                        for user in missing_users:
                            print(f"  - {user['username']} ({user['name']}) - {user['job_position']}")
                
                # 检查dingtalk_data的完整性
                if has_data_count > 0:
                    print("\n检查dingtalk_data字段完整性...")
                    cursor.execute("""
                        SELECT username, name, dingtalk_data
                        FROM users 
                        WHERE dingtalk_data IS NOT NULL AND dingtalk_data != ''
                        LIMIT 5
                    """)
                    
                    sample_count = 0
                    complete_count = 0
                    incomplete_fields = defaultdict(int)
                    
                    for row in cursor.fetchall():
                        if isinstance(row, dict):
                            username = row['username']
                            name = row['name']
                            dingtalk_data_str = row['dingtalk_data']
                        else:
                            username = row[0]
                            name = row[1]
                            dingtalk_data_str = row[2]
                        
                        try:
                            dingtalk_data = json.loads(dingtalk_data_str) if isinstance(dingtalk_data_str, str) else dingtalk_data_str
                            
                            sample_count += 1
                            
                            # 检查关键字段
                            required_fields = ['userid', 'unionid', 'job_number', 'name', 'title', 'dept_id']
                            missing_fields = [f for f in required_fields if f not in dingtalk_data or not dingtalk_data[f]]
                            
                            if not missing_fields:
                                complete_count += 1
                            else:
                                for field in missing_fields:
                                    incomplete_fields[field] += 1
                                
                                if sample_count <= 3:
                                    print(f"  ⚠️  {username} ({name}) 缺少字段: {', '.join(missing_fields)}")
                        
                        except Exception as e:
                            print(f"  ❌ {username} ({name}) dingtalk_data解析失败: {e}")
                    
                    if sample_count > 0:
                        print(f"\n完整性统计（基于前5个样本）:")
                        print(f"  - 完整: {complete_count}/{sample_count}")
                        print(f"  - 不完整: {sample_count - complete_count}/{sample_count}")
                        
                        if incomplete_fields:
                            print(f"\n缺失字段统计:")
                            for field, count in sorted(incomplete_fields.items(), key=lambda x: x[1], reverse=True):
                                print(f"  - {field}: {count} 个用户缺失")
                
        except Exception as e:
            logger.error(f"检查dingtalk_data失败: {e}", exc_info=True)
            print(f"❌ 检查dingtalk_data失败: {e}")
    
    def compare_mysql_excel(self):
        """对比MySQL和Excel数据"""
        try:
            # 获取MySQL用户列表
            pool = get_mysql_connection_pool()
            mysql_usernames = set()
            
            with pool.get_cursor() as cursor:
                cursor.execute("SELECT username FROM users")
                for row in cursor.fetchall():
                    if isinstance(row, dict):
                        mysql_usernames.add(row['username'])
                    else:
                        mysql_usernames.add(row[0])
            
            print(f"✅ MySQL中有 {len(mysql_usernames)} 个用户")
            
            # 获取Excel用户列表
            excel_usernames = set()
            if os.path.exists(self.users_file):
                try:
                    wb = openpyxl.load_workbook(self.users_file)
                    
                    # 读取默认工作表
                    if wb.active:
                        ws = wb.active
                        for row in range(2, ws.max_row + 1):
                            username_cell = ws.cell(row=row, column=2).value
                            if username_cell:
                                excel_usernames.add(str(username_cell).strip())
                    
                    # 读取钉钉用户数据工作表
                    if "钉钉用户数据" in wb.sheetnames:
                        ws_dingtalk = wb["钉钉用户数据"]
                        field_to_col = {header: idx + 1 for idx, header in enumerate(DINGTALK_USER_HEADERS)}
                        
                        for row in range(2, ws_dingtalk.max_row + 1):
                            job_number = ws_dingtalk.cell(row=row, column=field_to_col.get('job_number', 1)).value
                            userid = ws_dingtalk.cell(row=row, column=field_to_col.get('userid', 2)).value
                            
                            if job_number:
                                excel_usernames.add(str(job_number).strip())
                            if userid:
                                excel_usernames.add(str(userid).strip())
                
                except Exception as e:
                    logger.error(f"读取Excel文件失败: {e}", exc_info=True)
                    print(f"⚠️  读取Excel文件失败: {e}")
                    return
            
            print(f"✅ Excel中有 {len(excel_usernames)} 个用户")
            
            # 对比
            mysql_only = mysql_usernames - excel_usernames
            excel_only = excel_usernames - mysql_usernames
            common = mysql_usernames & excel_usernames
            
            print(f"\n对比结果:")
            print(f"  - 共同用户: {len(common)}")
            print(f"  - 仅在MySQL中: {len(mysql_only)}")
            print(f"  - 仅在Excel中: {len(excel_only)}")
            
            if mysql_only:
                print(f"\n⚠️  仅在MySQL中的用户（前10个）:")
                for username in list(mysql_only)[:10]:
                    print(f"  - {username}")
                if len(mysql_only) > 10:
                    print(f"  ... 还有 {len(mysql_only) - 10} 个")
            
            if excel_only:
                print(f"\n⚠️  仅在Excel中的用户（前10个）:")
                for username in list(excel_only)[:10]:
                    print(f"  - {username}")
                if len(excel_only) > 10:
                    print(f"  ... 还有 {len(excel_only) - 10} 个")
            
            self.results['comparison'] = {
                'mysql_count': len(mysql_usernames),
                'excel_count': len(excel_usernames),
                'common_count': len(common),
                'mysql_only_count': len(mysql_only),
                'excel_only_count': len(excel_only),
                'mysql_only': list(mysql_only)[:20],
                'excel_only': list(excel_only)[:20]
            }
            
        except Exception as e:
            logger.error(f"对比MySQL和Excel失败: {e}", exc_info=True)
            print(f"❌ 对比MySQL和Excel失败: {e}")
    
    def query_announcements_data(self):
        """查询公告相关数据"""
        try:
            pool = get_mysql_connection_pool()
            
            with pool.get_cursor() as cursor:
                # 检查是否有announcements表
                cursor.execute("""
                    SELECT TABLE_NAME 
                    FROM INFORMATION_SCHEMA.TABLES 
                    WHERE TABLE_SCHEMA = %s 
                    AND TABLE_NAME = 'announcements'
                """, (MYSQL_CONFIG['database'],))
                
                has_announcements_table = cursor.fetchone() is not None
                
                if has_announcements_table:
                    cursor.execute("SELECT COUNT(*) as total FROM announcements")
                    row = cursor.fetchone()
                    if row:
                        total = row['total'] if isinstance(row, dict) else row[0]
                    else:
                        total = 0
                    print(f"✅ announcements表存在，有 {total} 条记录")
                else:
                    print("ℹ️  announcements表不存在（公告数据存储在文件系统中）")
                
                # 检查audit_logs表
                cursor.execute("""
                    SELECT TABLE_NAME 
                    FROM INFORMATION_SCHEMA.TABLES 
                    WHERE TABLE_SCHEMA = %s 
                    AND TABLE_NAME = 'audit_logs'
                """, (MYSQL_CONFIG['database'],))
                
                has_audit_logs = cursor.fetchone() is not None
                
                if has_audit_logs:
                    cursor.execute("SELECT COUNT(*) as total FROM audit_logs")
                    row = cursor.fetchone()
                    if row:
                        total = row['total'] if isinstance(row, dict) else row[0]
                    else:
                        total = 0
                    print(f"✅ audit_logs表存在，有 {total} 条记录")
                else:
                    print("ℹ️  audit_logs表不存在")
                
        except Exception as e:
            logger.error(f"查询公告数据失败: {e}", exc_info=True)
            print(f"❌ 查询公告数据失败: {e}")
    
    def save_report(self):
        """保存查询报告"""
        try:
            report_file = os.path.join(DATA_DIR, 'mysql_query_report.json')
            with open(report_file, 'w', encoding='utf-8') as f:
                json.dump(self.results, f, ensure_ascii=False, indent=2, default=str)
            print(f"\n📄 详细报告已保存到: {report_file}")
        except Exception as e:
            print(f"\n⚠️  保存报告失败: {e}")


if __name__ == '__main__':
    query = MySQLDataQuery()
    success = query.run()
    sys.exit(0 if success else 1)

