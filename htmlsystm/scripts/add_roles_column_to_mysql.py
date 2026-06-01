#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
为MySQL users表添加roles字段的迁移脚本
同时从Excel文件同步现有用户的roles信息
"""
import os
import sys

# 添加项目根目录到Python路径
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(script_dir)
sys.path.insert(0, project_root)

from server.db_adapter import get_connection_pool
from server.config import DATA_DIR
from server.logger import logger
import openpyxl

# 钉钉用户表头（与server/user_manager.py保持一致）
DINGTALK_USER_HEADERS = [
    "job_number", "userid", "unionid", "login_id", "name", "nickname", "title",
    "dept_id", "dept_id_list", "parent_id", "dept_order",
    "active", "admin", "boss", "leader", "disable_status",
    "exclusive_account", "exclusive_account_type", "exclusive_account_corp_id", "exclusive_account_corp_name",
    "avatar", "hide_mobile", "hired_date",
    "密码", "物料库管理员", "公告栏管理员", "状态", "创建时间"
]

def _parse_roles(roles_str):
    """解析角色字符串为列表"""
    if not roles_str:
        return []
    if isinstance(roles_str, list):
        return roles_str
    if isinstance(roles_str, str):
        # 支持逗号分隔和空格分隔
        roles = [r.strip() for r in roles_str.replace(',', ' ').split() if r.strip()]
        return roles
    return []

def _roles_to_string(roles):
    """将角色列表转换为字符串"""
    if isinstance(roles, list):
        return ' '.join(roles)
    return str(roles) if roles else ''

def sync_roles_from_excel(cursor):
    """从Excel文件同步roles到MySQL"""
    users_file = os.path.join(DATA_DIR, 'users.xlsx')
    
    if not os.path.exists(users_file):
        print("⚠️  Excel文件不存在，跳过roles同步")
        return
    
    print("\n正在从Excel文件同步roles信息...")
    
    try:
        wb = openpyxl.load_workbook(users_file)
        updated_count = 0
        
        # 1. 从默认工作表读取roles
        if wb.active:
            ws = wb.active
            for row in range(2, ws.max_row + 1):
                try:
                    user_id = ws.cell(row=row, column=1).value
                    if not user_id:
                        continue
                    
                    # 读取roles（第5列）
                    roles_str = ws.cell(row=row, column=5).value or ''
                    roles = _parse_roles(roles_str)
                    
                    if roles:
                        roles_db_str = _roles_to_string(roles)
                        cursor.execute('UPDATE users SET roles = %s WHERE id = %s', (roles_db_str, int(user_id)))
                        updated_count += 1
                        print(f"  ✅ 用户ID={user_id}: 同步roles={roles_db_str}")
                except Exception as e:
                    print(f"  ⚠️  处理用户行{row}失败: {e}")
                    continue
        
        # 2. 从钉钉用户数据工作表读取roles（公告栏管理员列）
        if "钉钉用户数据" in wb.sheetnames:
            ws_dingtalk = wb["钉钉用户数据"]
            field_to_col = {header: idx + 1 for idx, header in enumerate(DINGTALK_USER_HEADERS)}
            announcement_manager_col = field_to_col.get('公告栏管理员', 26)
            
            # 获取所有MySQL用户，用于匹配
            cursor.execute('SELECT id, username FROM users')
            rows = cursor.fetchall()
            mysql_users = {}
            for row in rows:
                if isinstance(row, dict):
                    mysql_users[row['username']] = row['id']
                else:
                    # 如果是tuple，假设(id, username)顺序
                    if len(row) >= 2:
                        mysql_users[row[1]] = row[0]
            
            for row in range(2, ws_dingtalk.max_row + 1):
                try:
                    job_number = ws_dingtalk.cell(row=row, column=field_to_col.get('job_number', 1)).value
                    userid = ws_dingtalk.cell(row=row, column=field_to_col.get('userid', 2)).value
                    announcement_manager = ws_dingtalk.cell(row=row, column=announcement_manager_col).value or ''
                    
                    # 尝试匹配MySQL用户（通过username）
                    user_id = None
                    if job_number and str(job_number).strip() in mysql_users:
                        user_id = mysql_users[str(job_number).strip()]
                    elif userid and str(userid).strip() in mysql_users:
                        user_id = mysql_users[str(userid).strip()]
                    
                    if user_id and announcement_manager:
                        roles = _parse_roles(announcement_manager)
                        if roles:
                            roles_db_str = _roles_to_string(roles)
                            cursor.execute('UPDATE users SET roles = %s WHERE id = %s', (roles_db_str, int(user_id)))
                            updated_count += 1
                            print(f"  ✅ 钉钉用户ID={user_id}: 同步roles={roles_db_str}")
                except Exception as e:
                    print(f"  ⚠️  处理钉钉用户行{row}失败: {e}")
                    continue
        
        print(f"✅ 成功同步 {updated_count} 个用户的roles信息")
        
    except Exception as e:
        print(f"⚠️  从Excel同步roles失败: {e}")
        import traceback
        traceback.print_exc()

def add_roles_column():
    """为users表添加roles字段，并从Excel同步现有数据"""
    print("=" * 60)
    print("为MySQL users表添加roles字段")
    print("=" * 60)
    
    try:
        pool = get_connection_pool()
        with pool.get_cursor() as cursor:
            # 检查roles字段是否已存在
            cursor.execute("""
                SELECT COLUMN_NAME 
                FROM INFORMATION_SCHEMA.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                AND TABLE_NAME = 'users' 
                AND COLUMN_NAME = 'roles'
            """)
            exists = cursor.fetchone()
            
            if exists:
                print("✅ roles字段已存在")
                # 即使字段已存在，也尝试从Excel同步（可能之前同步失败）
                sync_roles_from_excel(cursor)
                return True
            
            # 添加roles字段
            print("正在添加roles字段...")
            cursor.execute("""
                ALTER TABLE users 
                ADD COLUMN roles TEXT DEFAULT NULL 
                AFTER job_position
            """)
            print("✅ 成功添加roles字段")
            
            # 为现有用户设置默认roles（空字符串）
            print("正在为现有用户初始化roles字段...")
            cursor.execute("UPDATE users SET roles = '' WHERE roles IS NULL")
            print("✅ 已初始化现有用户的roles字段")
            
            # 从Excel同步roles信息
            sync_roles_from_excel(cursor)
            
            print("\n" + "=" * 60)
            print("✅ 迁移完成！")
            print("=" * 60)
            print("\n说明：")
            print("1. MySQL users表已添加roles字段")
            print("2. 钉钉用户的详细信息（job_number, userid, unionid等）仍然存储在Excel文件中")
            print("3. 系统会从Excel的'钉钉用户数据'工作表读取这些详细信息")
            print("4. roles字段已从Excel同步到MySQL")
            
            return True
            
    except Exception as e:
        print(f"❌ 添加roles字段失败: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    success = add_roles_column()
    sys.exit(0 if success else 1)

