#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
待办数据验证脚本
验证本地的待办创建、完成更新以及刷新后数据是否保存成功
"""

import os
import sys
import json
import openpyxl
from datetime import datetime
from pathlib import Path

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from server.config import DATA_DIR
from server.mysql_connection_pool import get_mysql_connection_pool
from server.user_manager import UserManager
from server.todo_manager import TodoManager
from server.announcement_manager import AnnouncementManager

def print_section(title):
    """打印分节标题"""
    print("\n" + "=" * 80)
    print(f"  {title}")
    print("=" * 80)

def verify_todo_excel_files():
    """验证待办Excel文件"""
    print_section("1. 验证待办Excel文件")
    
    todos_dir = os.path.join(DATA_DIR, 'todos')
    if not os.path.exists(todos_dir):
        print(f"❌ 待办目录不存在: {todos_dir}")
        return []
    
    todo_files = [f for f in os.listdir(todos_dir) if f.endswith('_todos.xlsx')]
    print(f"📁 找到 {len(todo_files)} 个待办Excel文件")
    
    valid_files = []
    invalid_files = []
    
    for filename in todo_files:
        file_path = os.path.join(todos_dir, filename)
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # 提取公告ID
            announcement_id = filename[13:-10] if filename.startswith('announcement_') else ''
            
            # 读取数据
            todos = []
            max_row = ws.max_row if ws.max_row > 1 else 1
            for row in range(2, max_row + 1):
                userid = ws.cell(row=row, column=4).value
                if userid:
                    todos.append({
                        'announcement_id': announcement_id,
                        'userid': str(userid).strip(),
                        'username': str(ws.cell(row=row, column=7).value or '').strip(),
                        'name': str(ws.cell(row=row, column=8).value or '').strip(),
                        'status': str(ws.cell(row=row, column=9).value or '未完成').strip(),
                        'done': str(ws.cell(row=row, column=9).value or '').strip() in ['已完成', 'done'],
                        'complete_time': str(ws.cell(row=row, column=10).value or '').strip(),
                    })
            
            if todos:
                valid_files.append({
                    'filename': filename,
                    'announcement_id': announcement_id,
                    'todos': todos,
                    'total': len(todos),
                    'done': sum(1 for t in todos if t['done']),
                    'pending': sum(1 for t in todos if not t['done'])
                })
                print(f"  ✅ {filename}: {len(todos)} 条待办 (已完成: {sum(1 for t in todos if t['done'])}, 待完成: {sum(1 for t in todos if not t['done'])})")
            else:
                invalid_files.append(filename)
                print(f"  ⚠️  {filename}: 文件存在但无数据")
        except Exception as e:
            invalid_files.append(filename)
            print(f"  ❌ {filename}: 读取失败 - {e}")
    
    if invalid_files:
        print(f"\n⚠️  有 {len(invalid_files)} 个文件存在问题:")
        for f in invalid_files:
            print(f"    - {f}")
    
    return valid_files

def verify_todo_manager():
    """验证TodoManager读取"""
    print_section("2. 验证TodoManager读取")
    
    todo_mgr = TodoManager()
    todos_dir = os.path.join(DATA_DIR, 'todos')
    
    if not os.path.exists(todos_dir):
        print(f"❌ 待办目录不存在: {todos_dir}")
        return {}
    
    todo_files = [f for f in os.listdir(todos_dir) if f.endswith('_todos.xlsx')]
    print(f"📁 找到 {len(todo_files)} 个待办Excel文件")
    
    results = {}
    
    for filename in todo_files:
        # 提取公告ID：announcement_<id>_todos.xlsx
        if filename.startswith('announcement_') and filename.endswith('_todos.xlsx'):
            announcement_id = filename[13:-10]  # 去掉前缀和后缀
        else:
            continue
        
        if not announcement_id:
            continue
        
        try:
            todos = todo_mgr.get_all_todos(announcement_id)
            if todos:
                results[announcement_id] = {
                    'total': len(todos),
                    'done': sum(1 for t in todos if t.get('done', False)),
                    'pending': sum(1 for t in todos if not t.get('done', False)),
                    'todos': todos
                }
                print(f"  ✅ {announcement_id[:8]}...: {len(todos)} 条待办 (已完成: {sum(1 for t in todos if t.get('done', False))}, 待完成: {sum(1 for t in todos if not t.get('done', False))})")
            else:
                print(f"  ⚠️  {announcement_id[:8]}...: TodoManager返回0条待办")
        except Exception as e:
            print(f"  ❌ {announcement_id[:8]}...: 读取失败 - {e}")
    
    return results

def verify_user_info():
    """验证用户信息获取"""
    print_section("3. 验证用户信息获取")
    
    user_mgr = UserManager()
    todos_dir = os.path.join(DATA_DIR, 'todos')
    
    if not os.path.exists(todos_dir):
        print(f"❌ 待办目录不存在: {todos_dir}")
        return {}
    
    todo_files = [f for f in os.listdir(todos_dir) if f.endswith('_todos.xlsx')]
    
    userids_found = set()
    for filename in todo_files:
        file_path = os.path.join(todos_dir, filename)
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            max_row = ws.max_row if ws.max_row > 1 else 1
            for row in range(2, max_row + 1):
                userid = ws.cell(row=row, column=4).value
                if userid:
                    userids_found.add(str(userid).strip())
        except:
            continue
    
    print(f"📋 从待办文件中找到 {len(userids_found)} 个不同的userid")
    
    success_count = 0
    fail_count = 0
    user_info_map = {}
    
    for userid in userids_found:
        try:
            user_info = user_mgr.get_user_by_userid(userid)
            if user_info:
                name = user_info.get('name', '') or user_info.get('username', '') or userid
                user_info_map[userid] = {
                    'name': name,
                    'username': user_info.get('username', ''),
                    'found': True
                }
                success_count += 1
                print(f"  ✅ {userid}: {name}")
            else:
                user_info_map[userid] = {
                    'name': userid,
                    'username': userid,
                    'found': False
                }
                fail_count += 1
                print(f"  ❌ {userid}: 未找到用户信息")
        except Exception as e:
            user_info_map[userid] = {
                'name': userid,
                'username': userid,
                'found': False,
                'error': str(e)
            }
            fail_count += 1
            print(f"  ❌ {userid}: 获取失败 - {e}")
    
    print(f"\n📊 统计: 成功 {success_count}, 失败 {fail_count}")
    return user_info_map

def verify_announcement_todo_stats():
    """验证公告待办统计"""
    print_section("4. 验证公告待办统计")
    
    announcement_mgr = AnnouncementManager()
    todo_mgr = TodoManager()
    
    # 获取所有公告
    announcements = announcement_mgr.get_announcements(status=None, include_temp=False)
    print(f"📋 找到 {len(announcements)} 条公告")
    
    stats_results = {}
    
    for ann in announcements:
        ann_id = ann.get('id', '')
        if not ann_id:
            continue
        
        try:
            todos = todo_mgr.get_all_todos(ann_id)
            total = len(todos)
            done = sum(1 for t in todos if t.get('done', False))
            pending = total - done
            
            stats_results[ann_id] = {
                'title': ann.get('title', ''),
                'total': total,
                'done': done,
                'pending': pending
            }
            
            if total > 0:
                print(f"  ✅ {ann_id[:8]}... ({ann.get('title', '')[:30]}): 总计 {total}, 已完成 {done}, 待完成 {pending}")
            else:
                print(f"  ⚠️  {ann_id[:8]}... ({ann.get('title', '')[:30]}): 无待办数据")
        except Exception as e:
            print(f"  ❌ {ann_id[:8]}...: 读取失败 - {e}")
    
    return stats_results

def verify_cache_consistency():
    """验证缓存一致性"""
    print_section("5. 验证缓存一致性")
    
    from server.data_preloader import get_data_preloader
    
    preloader = get_data_preloader()
    if not preloader:
        print("❌ 无法获取DataPreloader实例")
        return
    
    todos_dir = os.path.join(DATA_DIR, 'todos')
    if not os.path.exists(todos_dir):
        print(f"❌ 待办目录不存在: {todos_dir}")
        return
    
    todo_files = [f for f in os.listdir(todos_dir) if f.endswith('_todos.xlsx')]
    print(f"📁 找到 {len(todo_files)} 个待办Excel文件")
    
    cache_count = len(preloader.todos_cache) if hasattr(preloader, 'todos_cache') else 0
    print(f"💾 内存缓存中有 {cache_count} 个公告的待办缓存")
    
    mismatch_count = 0
    
    for filename in todo_files:
        announcement_id = filename[13:-10] if filename.startswith('announcement_') else ''
        if not announcement_id:
            continue
        
        # 从文件读取
        try:
            todo_mgr = TodoManager()
            file_todos = todo_mgr.get_all_todos(announcement_id)
            file_count = len(file_todos)
        except:
            file_count = 0
        
        # 从缓存读取
        try:
            cache_todos = preloader.get_todos_for_announcement_cached(announcement_id)
            cache_count = len(cache_todos) if cache_todos else 0
        except:
            cache_count = 0
        
        if file_count != cache_count:
            mismatch_count += 1
            print(f"  ⚠️  {announcement_id[:8]}...: 文件 {file_count} 条, 缓存 {cache_count} 条 (不一致)")
        elif file_count > 0:
            print(f"  ✅ {announcement_id[:8]}...: 文件 {file_count} 条, 缓存 {cache_count} 条 (一致)")
    
    if mismatch_count == 0:
        print("\n✅ 所有缓存数据一致")
    else:
        print(f"\n⚠️  有 {mismatch_count} 个公告的缓存数据不一致")

def main():
    """主函数"""
    print("=" * 80)
    print("  待办数据验证脚本")
    print("=" * 80)
    print(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    try:
        # 1. 验证待办Excel文件
        excel_results = verify_todo_excel_files()
        
        # 2. 验证TodoManager读取
        manager_results = verify_todo_manager()
        
        # 3. 验证用户信息获取
        user_info_map = verify_user_info()
        
        # 4. 验证公告待办统计
        stats_results = verify_announcement_todo_stats()
        
        # 5. 验证缓存一致性
        verify_cache_consistency()
        
        # 生成报告
        print_section("6. 验证报告")
        
        report = {
            'timestamp': datetime.now().isoformat(),
            'excel_files': {
                'total': len(excel_results),
                'valid': len(excel_results),
                'details': excel_results
            },
            'todo_manager': {
                'total_announcements': len(manager_results),
                'details': manager_results
            },
            'user_info': {
                'total_userids': len(user_info_map),
                'success': sum(1 for u in user_info_map.values() if u.get('found', False)),
                'failed': sum(1 for u in user_info_map.values() if not u.get('found', False)),
                'details': user_info_map
            },
            'announcement_stats': stats_results
        }
        
        report_file = os.path.join(DATA_DIR, 'todo_verification_report.json')
        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        
        print(f"📄 详细报告已保存到: {report_file}")
        
        print("\n" + "=" * 80)
        print("  验证完成")
        print("=" * 80)
        
    except Exception as e:
        import traceback
        print(f"\n❌ 验证过程发生错误: {e}")
        print(traceback.format_exc())
        sys.exit(1)

if __name__ == '__main__':
    main()

