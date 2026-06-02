import json
import os
import secrets
import shutil
import time
import zipfile
from datetime import datetime
from typing import Any, Dict, List, Optional
from zipfile import BadZipFile

import openpyxl

from server.config import (
    DATA_DIR,
    LIBRARY_CONFIG,
    JOB_POSITION_OPTIONS,
    LEGACY_JOB_POSITION_MAP,
    DEPARTMENT_OPTIONS,
    DEFAULT_LIBRARY_HANDOFF_USER_ID,
)
from server.logger import logger
from server.security import PasswordHasher, SUPER_ADMIN_USERNAME
from server.file_lock import get_file_lock

# 超级管理员凭据已移至 server.security 模块

# 岗位选项从配置文件导入，确保统一管理
JOB_ROLE_OPTIONS = JOB_POSITION_OPTIONS

DEFAULT_HEADERS = ["ID", "用户名", "密码", "姓名", "角色", "部门", "物料库管理员", "状态", "创建时间"]

# 钉钉用户表头（用于从钉钉导入的用户数据）
# 包含所有钉钉API返回的字段，按顺序排列
DINGTALK_USER_HEADERS = [
    # 基础信息
    "job_number", "userid", "unionid", "login_id", "name", "nickname", "title",
    # 部门信息
    "dept_id", "dept_id_list", "parent_id", "dept_order",
    # 状态信息
    "active", "admin", "boss", "leader", "disable_status",
    # 账号信息
    "exclusive_account", "exclusive_account_type", "exclusive_account_corp_id", "exclusive_account_corp_name",
    # 其他信息
    "avatar", "hide_mobile", "hired_date", "gender",
    # 系统字段
    "密码", "物料库管理员", "公告栏管理员", "状态", "创建时间"
]
# 注意：表头顺序必须与用户提供的格式一致

STATUS_ACTIVE = 'active'
STATUS_PENDING = 'pending'
STATUS_REJECTED = 'rejected'

# dingtalk_data 仅合并白名单字段，避免异常/超大 JSON 撑爆内存与 auth/check
DINGTALK_MERGE_KEYS = frozenset({
    'userid', 'unionid', 'job_number', 'name', 'nickname', 'title',
    'dept_id', 'dept_id_list', 'dept_order', 'parent_id', 'avatar',
    'login_id', 'active', 'admin', 'boss', 'leader', 'disable_status',
    'exclusive_account', 'exclusive_account_type', 'exclusive_account_corp_id',
    'exclusive_account_corp_name', 'hide_mobile', 'hired_date', 'gender',
})
MAX_DINGTALK_JSON_BYTES = 256 * 1024


class UserManager:
    def __init__(self):
        # 账号数据仅以 MySQL 为权威数据源（不再读写 users.xlsx）
        self._ensure_mysql_user_schema()
        self._ensure_super_admin()
        try:
            from server.auth.password_service import get_password_service
            get_password_service(self).audit_passwords()
        except Exception as e:
            logger.warning(f'启动密码审计跳过: {e}')

    def _ensure_mysql_user_schema(self):
        """确保 users 表含钉钉相关列（启动与同步前调用）。"""
        try:
            from server.mysql_connection_pool import get_mysql_connection_pool
            from server.mysql_schema import migrate_users_table

            pool = get_mysql_connection_pool()
            with pool.get_cursor() as cursor:
                migrate_users_table(cursor)
        except Exception as e:
            logger.warning(f"users 表结构检查/迁移失败: {e}", exc_info=True)

    def _legacy_users_file_path(self) -> Optional[str]:
        """遗留 Excel 路径（已停用，始终返回 None）。"""
        return None
    
    def _ensure_users_file(self):
        """已停用：账号数据仅存 MySQL。"""
        return

    def _ensure_users_file_legacy(self):
        """确保用户文件存在并具备所需列（线程安全，带文件锁和错误恢复）"""
        # 如果文件不存在，直接创建（不需要锁，因为文件不存在）
        if not os.path.exists(self.users_file):
            try:
                # 使用进程锁确保只有一个进程创建文件
                from server.file_lock import get_process_file_lock
                recovery_lock_file = f"{self.users_file}.recovering"
                # 对于 recovering 锁，使用较短的超时时间（30秒），因为清理逻辑会更积极地清理僵尸锁
                recovery_lock = get_process_file_lock(recovery_lock_file, timeout=30)
                
                try:
                    with recovery_lock:
                        # 再次检查文件是否存在（可能已被其他进程创建）
                        if os.path.exists(self.users_file):
                            logger.info("文件已被其他进程创建，跳过创建操作")
                            return
                        
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "用户数据"
                        for col, header in enumerate(DEFAULT_HEADERS, 1):
                            ws.cell(row=1, column=col, value=header)
                        
                        for row_idx, user in enumerate(self._default_users(), start=2):
                            self._write_user_row(ws, row_idx, user)
                        
                        # 创建"钉钉用户数据"工作表（重要：确保钉钉用户数据可以正常保存）
                        ws_dingtalk = wb.create_sheet("钉钉用户数据")
                        for col, header in enumerate(DINGTALK_USER_HEADERS, 1):
                            ws_dingtalk.cell(row=1, column=col, value=header)
                        
                        wb.save(self.users_file)
                        logger.info(f"创建新的用户文件（包含钉钉用户数据工作表）: {self.users_file}")
                        return
                except TimeoutError:
                    # 其他进程正在创建，等待完成（使用更长的等待时间）
                    import time
                    max_wait = 30  # 增加到30秒
                    waited = 0
                    check_interval = 0.5
                    while waited < max_wait:
                        if os.path.exists(self.users_file):
                            logger.info("文件已被其他进程创建，等待完成")
                            # 再等待一小段时间，确保文件写入完成
                            time.sleep(0.5)
                            return
                        time.sleep(check_interval)
                        waited += check_interval
                    # 如果等待超时但文件仍不存在，检查是否有僵尸锁
                    if not os.path.exists(self.users_file):
                        logger.warning("等待文件创建超时，尝试清理可能的僵尸锁并重试")
                        # 清理僵尸锁文件
                        lock_file = f"{recovery_lock_file}.lock"
                        if os.path.exists(lock_file):
                            try:
                                lock_mtime = os.path.getmtime(lock_file)
                                if time.time() - lock_mtime > 300:  # 5分钟
                                    os.remove(lock_file)
                                    logger.info("已清理僵尸锁文件，将重试创建文件")
                                    # 递归调用一次（但只允许一次）
                                    if not hasattr(self, '_retry_create_file'):
                                        self._retry_create_file = True
                                        return self._ensure_users_file()
                            except Exception as cleanup_error:
                                logger.warning(f"清理僵尸锁文件失败: {cleanup_error}")
                        logger.error("等待文件创建超时")
                        raise
            except Exception as e:
                logger.error(f"创建用户文件失败: {e}", exc_info=True)
                raise
        
        # 文件存在，使用文件锁保护读取和修改
        file_lock = get_file_lock(self.users_file)
        
        with file_lock:
            
            # 文件存在，尝试加载并检查完整性
            try:
                wb = openpyxl.load_workbook(self.users_file)
                ws = wb.active
                self._ensure_schema(ws)
                
                # 检查是否存在"钉钉用户数据"工作表，如果不存在则创建
                if "钉钉用户数据" not in wb.sheetnames:
                    logger.warning("用户文件中缺少'钉钉用户数据'工作表，正在创建...")
                    ws_dingtalk = wb.create_sheet("钉钉用户数据")
                    for col, header in enumerate(DINGTALK_USER_HEADERS, 1):
                        ws_dingtalk.cell(row=1, column=col, value=header)
                    logger.info("已创建'钉钉用户数据'工作表")
                
                wb.save(self.users_file)
            except BadZipFile as e:
                # 文件损坏，尝试恢复
                file_size = os.path.getsize(self.users_file) if os.path.exists(self.users_file) else 0
                logger.error(f"用户文件损坏 (BadZipFile): {self.users_file}, 错误: {e}, 文件大小: {file_size} 字节")
                
                # 如果文件太小（小于1KB），很可能是损坏的，直接删除并重建
                if file_size < 1024:
                    logger.warning(f"文件大小异常小（{file_size}字节），可能是损坏的文件，将直接重建")
                    try:
                        if os.path.exists(self.users_file):
                            corrupted_file = f"{self.users_file}.corrupted.{int(time.time())}"
                            try:
                                shutil.move(self.users_file, corrupted_file)
                                logger.info(f"已备份损坏的文件到: {corrupted_file}")
                            except Exception as move_error:
                                logger.warning(f"备份损坏文件失败: {move_error}，尝试直接删除")
                                try:
                                    os.remove(self.users_file)
                                except:
                                    pass
                    except Exception as cleanup_error:
                        logger.warning(f"清理损坏文件失败: {cleanup_error}")
                
                try:
                    logger.info("开始恢复损坏的用户文件...")
                    self._recover_corrupted_file()
                    logger.info("用户文件恢复完成，重新加载...")
                    
                    # 恢复后重试加载
                    if os.path.exists(self.users_file):
                        try:
                            wb = openpyxl.load_workbook(self.users_file)
                            ws = wb.active
                            self._ensure_schema(ws)
                            # 确保有钉钉用户数据工作表
                            if "钉钉用户数据" not in wb.sheetnames:
                                ws_dingtalk = wb.create_sheet("钉钉用户数据")
                                for col, header in enumerate(DINGTALK_USER_HEADERS, 1):
                                    ws_dingtalk.cell(row=1, column=col, value=header)
                                wb.save(self.users_file)
                            logger.info("用户文件恢复并验证成功")
                        except Exception as reload_error:
                            logger.error(f"恢复后重新加载失败: {reload_error}", exc_info=True)
                            # 如果重新加载失败，尝试从MySQL恢复
                            self._recover_from_mysql()
                    else:
                        logger.error("恢复后文件不存在，尝试从MySQL恢复")
                        self._recover_from_mysql()
                except Exception as recover_error:
                    logger.error(f"恢复用户文件失败: {recover_error}", exc_info=True)
                    # 如果恢复失败，尝试从MySQL恢复数据
                    try:
                        self._recover_from_mysql()
                    except Exception as mysql_error:
                        logger.error(f"从MySQL恢复也失败: {mysql_error}", exc_info=True)
                        # 最后尝试：创建最小可用文件
                        logger.warning("所有恢复方法都失败，创建最小可用文件")
                        self._create_minimal_users_file()
            except Exception as e:
                logger.error(f"加载用户文件失败: {e}", exc_info=True)
                # 尝试恢复
                try:
                    self._recover_corrupted_file()
                    # 恢复后重试加载
                    if os.path.exists(self.users_file):
                        wb = openpyxl.load_workbook(self.users_file)
                        ws = wb.active
                        self._ensure_schema(ws)
                        # 确保有钉钉用户数据工作表
                        if "钉钉用户数据" not in wb.sheetnames:
                            ws_dingtalk = wb.create_sheet("钉钉用户数据")
                            for col, header in enumerate(DINGTALK_USER_HEADERS, 1):
                                ws_dingtalk.cell(row=1, column=col, value=header)
                            wb.save(self.users_file)
                except Exception as recover_error:
                    logger.error(f"恢复用户文件失败: {recover_error}", exc_info=True)
                    # 如果恢复失败，尝试从MySQL恢复数据
                    self._recover_from_mysql()
    
    def _recover_corrupted_file(self):
        """恢复损坏的用户文件（使用非阻塞锁，避免多进程死锁）"""
        backup_file = f"{self.users_file}.backup"
        corrupted_file = f"{self.users_file}.corrupted"
        recovery_lock_file = f"{self.users_file}.recovering"
        
        logger.info(f"开始恢复损坏的用户文件: {self.users_file}")
        logger.info(f"备份文件: {backup_file} (存在: {os.path.exists(backup_file)})")
        logger.info(f"损坏文件备份位置: {corrupted_file}")
        
        # 使用非阻塞锁，如果其他进程正在恢复，等待它完成
        from server.file_lock import get_file_lock, get_process_file_lock
        recovery_lock = get_process_file_lock(recovery_lock_file, timeout=30)  # 增加超时时间到30秒，避免多进程竞争
        
        try:
            # 尝试获取恢复锁（非阻塞）
            try:
                with recovery_lock:
                    # 检查文件是否已经被其他进程恢复
                    if os.path.exists(self.users_file):
                        try:
                            test_wb = openpyxl.load_workbook(self.users_file)
                            test_wb.close()
                            logger.info("文件已被其他进程恢复，跳过恢复操作")
                            return
                        except:
                            pass  # 文件仍然损坏，继续恢复
                    
                    # 1. 尝试从备份恢复
                    if os.path.exists(backup_file):
                        logger.info(f"尝试从备份恢复: {backup_file}")
                        try:
                            # 验证备份文件是否有效
                            test_wb = openpyxl.load_workbook(backup_file)
                            test_wb.close()
                            # 备份文件有效，恢复
                            shutil.copy2(backup_file, self.users_file)
                            logger.info(f"成功从备份恢复用户文件")
                            return
                        except Exception as e:
                            logger.warning(f"备份文件也损坏: {e}")
                    
                    # 2. 备份损坏的文件
                    if os.path.exists(self.users_file):
                        try:
                            shutil.move(self.users_file, corrupted_file)
                            logger.info(f"已备份损坏的文件到: {corrupted_file}")
                        except Exception as e:
                            logger.warning(f"备份损坏文件失败: {e}")
                    
                    # 3. 创建新文件（包含默认工作表和钉钉用户数据工作表）
                    logger.warning("无法恢复文件，创建新的用户文件")
                    wb = openpyxl.Workbook()
                    
                    # 创建默认"用户数据"工作表
                    ws = wb.active
                    ws.title = "用户数据"
                    for col, header in enumerate(DEFAULT_HEADERS, 1):
                        ws.cell(row=1, column=col, value=header)
                    
                    for row_idx, user in enumerate(self._default_users(), start=2):
                        self._write_user_row(ws, row_idx, user)
                    
                    # 创建"钉钉用户数据"工作表（重要：确保钉钉用户数据可以正常保存）
                    ws_dingtalk = wb.create_sheet("钉钉用户数据")
                    for col, header in enumerate(DINGTALK_USER_HEADERS, 1):
                        ws_dingtalk.cell(row=1, column=col, value=header)
                    
                    # 使用文件锁保护保存操作
                    thread_lock = get_file_lock(self.users_file)
                    with thread_lock:
                        wb.save(self.users_file)
                    
                    logger.info(f"已创建新的用户文件（包含钉钉用户数据工作表）: {self.users_file}")
                    logger.warning("⚠️  注意：新文件只包含默认用户，钉钉用户数据已丢失。请立即执行钉钉同步以恢复用户数据！")
            except TimeoutError:
                # 其他进程正在恢复，等待文件恢复完成
                logger.info("其他进程正在恢复文件，等待恢复完成...")
                import time
                max_wait = 60  # 最多等待60秒
                wait_interval = 1
                waited = 0
                while waited < max_wait:
                    if os.path.exists(self.users_file):
                        try:
                            test_wb = openpyxl.load_workbook(self.users_file)
                            test_wb.close()
                            logger.info("文件已被其他进程恢复")
                            return
                        except:
                            pass
                    time.sleep(wait_interval)
                    waited += wait_interval
                logger.error("等待文件恢复超时")
                raise
        except Exception as e:
            logger.error(f"恢复用户文件失败: {e}", exc_info=True)
            raise
    
    def _create_minimal_users_file(self):
        """创建最小可用的用户文件（最后的恢复手段）"""
        try:
            logger.warning("创建最小可用的用户文件...")
            from server.file_lock import get_file_lock
            thread_lock = get_file_lock(self.users_file)
            
            with thread_lock:
                # 如果文件存在，先备份
                if os.path.exists(self.users_file):
                    corrupted_file = f"{self.users_file}.corrupted.{int(time.time())}"
                    try:
                        shutil.move(self.users_file, corrupted_file)
                        logger.info(f"已备份损坏的文件到: {corrupted_file}")
                    except:
                        try:
                            os.remove(self.users_file)
                        except:
                            pass
                
                # 创建新文件
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "用户数据"
                for col, header in enumerate(DEFAULT_HEADERS, 1):
                    ws.cell(row=1, column=col, value=header)
                
                for row_idx, user in enumerate(self._default_users(), start=2):
                    self._write_user_row(ws, row_idx, user)
                
                # 创建"钉钉用户数据"工作表
                ws_dingtalk = wb.create_sheet("钉钉用户数据")
                for col, header in enumerate(DINGTALK_USER_HEADERS, 1):
                    ws_dingtalk.cell(row=1, column=col, value=header)
                
                wb.save(self.users_file)
                logger.info(f"已创建最小可用的用户文件: {self.users_file}")
                logger.warning("⚠️  注意：新文件只包含默认用户，钉钉用户数据已丢失。请立即执行钉钉同步以恢复用户数据！")
        except Exception as e:
            logger.error(f"创建最小用户文件失败: {e}", exc_info=True)
            raise
    
    def _recover_from_mysql(self):
        """从MySQL恢复用户数据到Excel（如果MySQL中有数据）"""
        try:
            from server.config import USE_MYSQL
            if not USE_MYSQL:
                return
            
            from server.mysql_connection_pool import get_mysql_connection_pool
            pool = get_mysql_connection_pool()
            
            with pool.get_cursor() as cursor:
                # 检查MySQL中是否有用户数据
                cursor.execute("SELECT COUNT(*) FROM users")
                count = cursor.fetchone()[0]
                if count == 0:
                    logger.info("MySQL中没有用户数据，无法恢复")
                    return
                
                logger.info(f"从MySQL恢复 {count} 个用户到Excel...")
                
                # 创建新的Excel文件
                wb = openpyxl.Workbook()
                
                # 创建默认"用户数据"工作表
                ws = wb.active
                ws.title = "用户数据"
                for col, header in enumerate(DEFAULT_HEADERS, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # 添加默认用户
                for row_idx, user in enumerate(self._default_users(), start=2):
                    self._write_user_row(ws, row_idx, user)
                
                # 创建"钉钉用户数据"工作表
                ws_dingtalk = wb.create_sheet("钉钉用户数据")
                for col, header in enumerate(DINGTALK_USER_HEADERS, 1):
                    ws_dingtalk.cell(row=1, column=col, value=header)
                
                # 从MySQL读取钉钉用户数据
                cursor.execute("""
                    SELECT dingtalk_data FROM users 
                    WHERE dingtalk_data IS NOT NULL
                """)
                rows = cursor.fetchall()
                
                row_idx = 2
                for row in rows:
                    dingtalk_data = row[0]
                    if isinstance(dingtalk_data, str):
                        import json
                        dingtalk_data = json.loads(dingtalk_data)
                    
                    if isinstance(dingtalk_data, dict):
                        # 写入钉钉用户数据（按照DINGTALK_USER_HEADERS顺序）
                        for col, header in enumerate(DINGTALK_USER_HEADERS, 1):
                            # 获取字段值，如果不存在则使用空字符串
                            value = dingtalk_data.get(header, '')
                            # 处理特殊字段（如布尔值、列表等）
                            if isinstance(value, bool):
                                value = '是' if value else '否'
                            elif isinstance(value, list):
                                value = ','.join(str(v) for v in value) if value else ''
                            elif value is None:
                                value = ''
                            ws_dingtalk.cell(row=row_idx, column=col, value=value)
                        row_idx += 1
                
                # 保存文件
                from server.file_lock import get_file_lock
                thread_lock = get_file_lock(self.users_file)
                with thread_lock:
                    wb.save(self.users_file)
                
                logger.info(f"已从MySQL恢复 {row_idx - 2} 个钉钉用户到Excel")
        except Exception as e:
            logger.error(f"从MySQL恢复用户数据失败: {e}", exc_info=True)
    
    def _ensure_schema(self, ws):
        """保证旧表格也包含状态列等新字段"""
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        # 检查并添加缺失的列
        if "状态" not in headers:
            status_col = 7
            ws.insert_cols(status_col)
            ws.cell(row=1, column=status_col, value="状态")
            ws.cell(row=1, column=status_col + 1, value="创建时间")
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=status_col).value is None:
                    ws.cell(row=row, column=status_col, value=STATUS_ACTIVE)
        
        # 检查物料库管理员列
        if "物料库管理员" not in headers:
            # 在部门列后插入物料库管理员列
            library_col = 7
            ws.insert_cols(library_col)
            ws.cell(row=1, column=library_col, value="物料库管理员")
            # 调整后续列的位置
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=library_col, value="")
        
        for idx, header in enumerate(DEFAULT_HEADERS, 1):
            if ws.cell(row=1, column=idx).value != header:
                ws.cell(row=1, column=idx, value=header)
        
        status_col_idx = DEFAULT_HEADERS.index("状态") + 1
        library_col_idx = DEFAULT_HEADERS.index("物料库管理员") + 1
        for row in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row, column=status_col_idx)
            if not status_cell.value:
                status_cell.value = STATUS_ACTIVE
            # 确保物料库管理员列有值（即使是空）
            library_cell = ws.cell(row=row, column=library_col_idx)
            if library_cell.value is None:
                library_cell.value = ""
    
    def _default_users(self):
        """初始用户列表"""
        # 仅初始化最高管理员，其余账号通过同步或手动创建
        return [
            {
                'username': SUPER_ADMIN_USERNAME,
                'password': '',
                'name': '系统最高管理员',
                'roles': ['super_admin', 'admin', 'management'],
                'department': '管理组',
                'status': STATUS_ACTIVE,
                'create_time': datetime.now().strftime('%Y-%m-%d')
            }
        ]
    
    def _write_user_row(self, ws, row_idx, user):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=user['username'])
        ws.cell(row=row_idx, column=3, value=user['password'])
        ws.cell(row=row_idx, column=4, value=user['name'])
        ws.cell(row=row_idx, column=5, value=self._roles_to_string(user.get('roles', [])))
        ws.cell(row=row_idx, column=6, value=user.get('department', ''))
        ws.cell(row=row_idx, column=7, value=self._roles_to_string(user.get('library_roles', [])))
        ws.cell(row=row_idx, column=8, value=user.get('status', STATUS_ACTIVE))
        ws.cell(row=row_idx, column=9, value=user.get('create_time', datetime.now().strftime('%Y-%m-%d')))
    
    def _ensure_super_admin(self):
        """保证最高管理员存在（直接从数据库检查，不再使用Excel文件）"""
        try:
            from server.db_adapter import get_connection_pool
            from server.security import PasswordHasher
            from server.admin_credentials import (
                bootstrap_admin_password,
                load_credentials,
                print_admin_credentials_banner,
            )
            
            pool = get_connection_pool()
            password_hasher = PasswordHasher()
            
            with pool.get_cursor() as cursor:
                # 检查超级管理员是否存在
                cursor.execute('SELECT id, username, password, roles, status FROM users WHERE username = %s', (SUPER_ADMIN_USERNAME,))
                row = cursor.fetchone()
                
                if row:
                    if not load_credentials():
                        logger.warning(
                            '超级管理员 %s 已存在但缺少 data/admin_credentials.json；'
                            '请执行: docker exec stack-htmlsystm python scripts/view_and_reset_admin_password.py '
                            "--reset zzw --password '你的新密码'",
                        )
                    # 用户存在，检查是否需要更新
                    if isinstance(row, dict):
                        user_id = row.get('id')
                        current_password = row.get('password')
                        current_roles = row.get('roles', '')
                        current_status = row.get('status')
                    else:
                        user_id = row[0] if row else None
                        current_password = row[2] if len(row) > 2 else None
                        current_roles = row[3] if len(row) > 3 else ''
                        current_status = row[4] if len(row) > 4 else None
                    
                    # 检查角色与状态
                    roles_list = self._parse_roles(current_roles)
                    required_roles = ['super_admin', 'admin', 'management']
                    if not all(role in roles_list for role in required_roles):
                        roles_str = self._roles_to_string(list(set(roles_list + required_roles)))
                    else:
                        roles_str = current_roles
                    status_value = STATUS_ACTIVE if current_status != STATUS_ACTIVE else current_status
                    needs_update = (
                        roles_str != current_roles
                        or current_status != STATUS_ACTIVE
                    )

                    if needs_update:
                        set_parts = []
                        params = []
                        if roles_str != current_roles:
                            set_parts.append('roles = %s')
                            params.append(roles_str)
                        if status_value != current_status:
                            set_parts.append('status = %s')
                            params.append(status_value)
                        set_parts.append('updated_time = NOW()')
                        params.append(user_id)
                        update_sql = f"UPDATE users SET {', '.join(set_parts)} WHERE id = %s"
                        cursor.execute(update_sql, tuple(params))
                        logger.info(f"已更新超级管理员信息: {SUPER_ADMIN_USERNAME}")
                else:
                    admin_user, plain_password, is_new = bootstrap_admin_password(create_if_missing=True)
                    password_hash = password_hasher.hash_password(plain_password)
                    roles_str = self._roles_to_string(['super_admin', 'admin', 'management'])

                    cursor.execute('''
                        INSERT INTO users (username, password, name, department, job_position, roles, status, created_time, updated_time)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                    ''', (admin_user, password_hash, '系统最高管理员', '管理组', 'super_admin', roles_str, STATUS_ACTIVE))
                    logger.info(f"已创建超级管理员: {admin_user}")
                    if is_new:
                        print_admin_credentials_banner(plain_password, username=admin_user)
        except Exception as e:
            logger.error(f"_ensure_super_admin 失败: {e}", exc_info=True)
            # 不抛出异常，允许系统继续运行（数据库可能暂时不可用）
            logger.warning("超级管理员检查失败，但系统将继续运行")
    
    def _parse_dingtalk_data_safe(self, raw) -> Optional[Dict[str, Any]]:
        """解析 dingtalk_data，过大或非法时返回 None。"""
        if not raw:
            return None
        try:
            if isinstance(raw, (bytes, bytearray)):
                if len(raw) > MAX_DINGTALK_JSON_BYTES:
                    logger.warning(
                        "dingtalk_data 过大(%s bytes)，已跳过全量合并",
                        len(raw),
                    )
                    return None
                raw = raw.decode('utf-8', errors='ignore')
            if isinstance(raw, str):
                if len(raw.encode('utf-8', errors='ignore')) > MAX_DINGTALK_JSON_BYTES:
                    logger.warning(
                        "dingtalk_data 过大(%s chars)，已跳过全量合并",
                        len(raw),
                    )
                    return None
                data = json.loads(raw)
            elif isinstance(raw, dict):
                data = raw
            else:
                return None
            return data if isinstance(data, dict) else None
        except Exception as exc:
            logger.debug(f"解析 dingtalk_data 失败: {exc}")
            return None

    def _merge_dingtalk_fields(self, user: Dict[str, Any], dingtalk_data: Dict[str, Any]) -> None:
        """仅合并白名单钉钉字段，不写入 dingtalk_data 原始 blob。"""
        json_uid = str(dingtalk_data.get('userid') or '').strip()
        user['userid'] = user.get('userid') or json_uid
        user['unionid'] = user.get('unionid') or str(dingtalk_data.get('unionid') or '').strip()
        user['job_number'] = user.get('job_number') or str(dingtalk_data.get('job_number') or '').strip()
        if not user.get('name') and dingtalk_data.get('name'):
            user['name'] = dingtalk_data.get('name', '')
        for key in DINGTALK_MERGE_KEYS:
            if key in dingtalk_data and key not in ('id', 'password'):
                user[key] = dingtalk_data[key]
        user['source'] = 'dingtalk'

    def _strip_heavy_user_fields(self, user: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
        """移除不应进入会话/接口的大字段。"""
        if not user:
            return user
        user.pop('dingtalk_data', None)
        user.pop('password', None)
        return user

    def _parse_roles(self, roles_value) -> List[str]:
        if not roles_value:
            return []
        if isinstance(roles_value, list):
            return [role.strip() for role in roles_value if role]
        return [role.strip() for role in str(roles_value).split(',') if role.strip()]
    
    def _roles_to_string(self, roles: List[str]) -> str:
        return ','.join(sorted(set(role for role in roles if role)))
    
    def _primary_role(self, roles: List[str]) -> str:
        if not roles:
            return 'user'
        priority = ['super_admin', 'admin', 'management']
        for key in priority:
            if key in roles:
                return key
        return roles[0]
    
    def _row_to_dingtalk_user(self, ws, row: int, user_id: int) -> Optional[Dict[str, Any]]:
        """从Excel行读取钉钉用户数据（新表头格式）
        
        使用新的表头定义，通过字段名映射到列索引
        """
        try:
            # 创建字段名到列索引的映射
            field_to_col = {header: idx + 1 for idx, header in enumerate(DINGTALK_USER_HEADERS)}
            
            # 读取userid（第2列）
            userid = ws.cell(row=row, column=field_to_col.get('userid', 2)).value
            if not userid:
                return None
            
            # 读取各个字段
            job_number = ws.cell(row=row, column=field_to_col.get('job_number', 1)).value or ''
            password = ws.cell(row=row, column=field_to_col.get('密码', 0)).value or 'CHXW_HW_123456'  # 默认密码
            name = ws.cell(row=row, column=field_to_col.get('name', 5)).value or ''
            title = ws.cell(row=row, column=field_to_col.get('title', 7)).value or ''  # title在第7列
            unionid = ws.cell(row=row, column=field_to_col.get('unionid', 3)).value or ''  # unionid在第3列
            parent_id = ws.cell(row=row, column=field_to_col.get('parent_id', 0)).value
            dept_id = ws.cell(row=row, column=field_to_col.get('dept_id', 8)).value or ''
            library_manager = ws.cell(row=row, column=field_to_col.get('物料库管理员', 0)).value or ''
            announcement_manager = ws.cell(row=row, column=field_to_col.get('公告栏管理员', 0)).value or ''
            status_raw = ws.cell(row=row, column=field_to_col.get('状态', 0)).value
            # 处理状态值：转换为字符串，如果为空或None，使用 'inactive'（未激活）
            if status_raw is None or str(status_raw).strip() == '':
                status = 'inactive'  # 未设置状态的用户视为未激活
            else:
                status = str(status_raw).strip().lower()  # 转换为小写以便比较
                # 将 'active' 标准化
                if status == 'active' or status == '激活' or status == '1' or status == 'true':
                    status = STATUS_ACTIVE
                elif status not in [STATUS_ACTIVE, STATUS_PENDING, STATUS_REJECTED, 'inactive']:
                    status = 'inactive'  # 无效状态视为未激活
            create_time = ws.cell(row=row, column=field_to_col.get('创建时间', 0)).value or datetime.now().strftime('%Y-%m-%d')
            
            # 如果没有密码，使用默认密码
            if not password or str(password).strip() == '':
                password = 'CHXW_HW_123456'
            
            # 构建用户对象（转换为标准格式）
            # username优先使用job_number（如果存在），否则使用userid
            username = str(job_number).strip() if job_number and str(job_number).strip() else str(userid)
            
            # 验证状态值是否有效
            valid_statuses = [STATUS_ACTIVE, STATUS_PENDING, STATUS_REJECTED, 'inactive']
            if status not in valid_statuses:
                # 无效状态视为未激活，不默认设置为激活
                logger.warning(f"用户 {name} (userid={userid}) 的状态值无效: {status}，设置为 inactive")
                status = 'inactive'
            
            # 从"公告栏管理员"列读取角色信息
            roles = []
            if announcement_manager:
                # 解析角色字符串（可能是逗号分隔的多个角色）
                roles = self._parse_roles(announcement_manager)
            
            # 读取所有钉钉用户字段
            login_id = ws.cell(row=row, column=field_to_col.get('login_id', 4)).value or ''
            nickname = ws.cell(row=row, column=field_to_col.get('nickname', 6)).value or ''
            dept_id_list_raw = ws.cell(row=row, column=field_to_col.get('dept_id_list', 9)).value or '[]'
            dept_order = ws.cell(row=row, column=field_to_col.get('dept_order', 12)).value or ''
            active_raw = ws.cell(row=row, column=field_to_col.get('active', 13)).value
            admin_raw = ws.cell(row=row, column=field_to_col.get('admin', 14)).value
            boss_raw = ws.cell(row=row, column=field_to_col.get('boss', 15)).value
            leader_raw = ws.cell(row=row, column=field_to_col.get('leader', 16)).value
            disable_status_raw = ws.cell(row=row, column=field_to_col.get('disable_status', 17)).value
            exclusive_account_raw = ws.cell(row=row, column=field_to_col.get('exclusive_account', 18)).value
            exclusive_account_type = ws.cell(row=row, column=field_to_col.get('exclusive_account_type', 19)).value or ''
            exclusive_account_corp_id = ws.cell(row=row, column=field_to_col.get('exclusive_account_corp_id', 20)).value or ''
            exclusive_account_corp_name = ws.cell(row=row, column=field_to_col.get('exclusive_account_corp_name', 21)).value or ''
            avatar = ws.cell(row=row, column=field_to_col.get('avatar', 22)).value or ''
            hide_mobile_raw = ws.cell(row=row, column=field_to_col.get('hide_mobile', 23)).value
            hired_date = ws.cell(row=row, column=field_to_col.get('hired_date', 24)).value or ''
            gender = ws.cell(row=row, column=field_to_col.get('gender', 25)).value or ''
            
            # 处理布尔值字段
            active = bool(active_raw) if active_raw is not None else True
            admin = bool(admin_raw) if admin_raw is not None else False
            boss = bool(boss_raw) if boss_raw is not None else False
            leader = bool(leader_raw) if leader_raw is not None else False
            disable_status = bool(disable_status_raw) if disable_status_raw is not None else False
            exclusive_account = bool(exclusive_account_raw) if exclusive_account_raw is not None else False
            hide_mobile = bool(hide_mobile_raw) if hide_mobile_raw is not None else False
            
            # 处理dept_id_list（可能是JSON字符串）
            dept_id_list = []
            if dept_id_list_raw:
                try:
                    import json
                    if isinstance(dept_id_list_raw, str):
                        dept_id_list = json.loads(dept_id_list_raw)
                    else:
                        dept_id_list = dept_id_list_raw
                except:
                    dept_id_list = []
            
            user = {
                'id': user_id,
                'username': username,  # 优先使用job_number，否则使用userid
                'password': password,
                'name': name,
                'role': self._primary_role(roles),  # 从公告栏管理员列读取角色
                'roles': roles,  # 从公告栏管理员列读取角色
                'department': dept_id if dept_id else '',
                'library_roles': [library_manager] if library_manager else [],
                'status': status,
                'create_time': str(create_time),
                'source': 'dingtalk',  # 标记来源
                'job_number': job_number,
                'userid': str(userid),  # 确保userid字段存在
                'unionid': str(unionid) if unionid else '',  # 添加unionid字段
                'title': title,
                'parent_id': parent_id,
                'dept_id': dept_id,
                'library_manager': library_manager,
                'announcement_manager': announcement_manager,
                # 添加所有钉钉用户字段
                'login_id': str(login_id) if login_id else '',
                'nickname': str(nickname) if nickname else '',
                'dept_id_list': dept_id_list,
                'dept_order': str(dept_order) if dept_order else '',
                'active': active,
                'admin': admin,
                'boss': boss,
                'leader': leader,
                'disable_status': disable_status,
                'exclusive_account': exclusive_account,
                'exclusive_account_type': str(exclusive_account_type) if exclusive_account_type else '',
                'exclusive_account_corp_id': str(exclusive_account_corp_id) if exclusive_account_corp_id else '',
                'exclusive_account_corp_name': str(exclusive_account_corp_name) if exclusive_account_corp_name else '',
                'avatar': str(avatar) if avatar else '',
                'hide_mobile': hide_mobile,
                'hired_date': str(hired_date) if hired_date else '',
                'gender': str(gender) if gender else ''
            }
            
            return user
        except Exception as e:
            logger.warning(f"解析钉钉用户数据失败 (行 {row}): {e}")
            return None
    
    def _row_to_user(self, ws, row: int) -> Optional[Dict[str, Any]]:
        """从Excel行读取用户数据（兼容旧格式）"""
        try:
            user_id = ws.cell(row=row, column=1).value
            if not user_id:
                return None
            
            # 安全转换ID
            try:
                user_id = int(user_id)
            except (ValueError, TypeError):
                logger.warning(f"无效的用户ID (行 {row}): {user_id}")
                return None
            
            username = ws.cell(row=row, column=2).value
            if not username:
                return None
            
            # 解析角色（兼容空值）
            roles_value = ws.cell(row=row, column=5).value
            roles = self._parse_roles(roles_value) if roles_value else []
            
            # 检查是否有物料库管理员列（兼容旧数据）
            library_roles_col = 7
            status_col = 8
            create_time_col = 9
            
            # 检查列7的值，判断是否是旧格式（没有物料库管理员列）
            col7_value = ws.cell(row=row, column=7).value
            if col7_value in [STATUS_ACTIVE, STATUS_PENDING, STATUS_REJECTED]:
                # 旧格式：没有物料库管理员列
                library_roles_col = None
                status_col = 7
                create_time_col = 8
            elif col7_value is None and ws.max_column < 8:
                # 如果列数不足，也是旧格式
                library_roles_col = None
                status_col = 7
                create_time_col = 8
            
            # 读取物料库管理员角色
            library_roles = []
            if library_roles_col and library_roles_col <= ws.max_column:
                library_roles_value = ws.cell(row=row, column=library_roles_col).value
                if library_roles_value:
                    library_roles = self._parse_roles(library_roles_value)
            
            # 读取状态（兼容旧格式）
            if status_col <= ws.max_column:
                status = ws.cell(row=row, column=status_col).value or STATUS_ACTIVE
            else:
                status = STATUS_ACTIVE
            
            # 读取创建时间（可选）
            create_time = None
            if create_time_col <= ws.max_column:
                create_time = ws.cell(row=row, column=create_time_col).value
            
            # 读取密码并清理空格
            password_value = ws.cell(row=row, column=3).value
            if password_value is None:
                password_value = ''
            else:
                # 转换为字符串并去除前后空格
                password_value = str(password_value).strip()
            
            return {
                'id': user_id,
                'username': str(username).strip(),
                'password': password_value,
                'name': ws.cell(row=row, column=4).value or str(username).strip(),
                'roles': roles,
                'role': self._primary_role(roles),
                'department': ws.cell(row=row, column=6).value if ws.max_column >= 6 else '',
                'library_roles': library_roles,
                'status': status,
                'create_time': create_time
            }
        except Exception as e:
            logger.warning(f"解析用户数据失败 (行 {row}): {e}")
            return None
    
    def _build_user_from_db_row(self, user_dict: Dict[str, Any]) -> Dict[str, Any]:
        """将 MySQL users 行转为登录/会话用用户对象（不含 Excel 回退）。"""
        user = {
            'id': user_dict.get('id'),
            'username': user_dict.get('username', ''),
            'password': user_dict.get('password', ''),
            'name': user_dict.get('name', ''),
            'department': user_dict.get('department', ''),
            'job_position': user_dict.get('job_position'),
            'library_roles': user_dict.get('library_roles', ''),
            'roles': self._parse_roles(user_dict.get('roles', '')),
            'status': user_dict.get('status', 'active'),
            'created_time': user_dict.get('created_time'),
            'updated_time': user_dict.get('updated_time'),
            'last_login_time': user_dict.get('last_login_time'),
        }
        if user['library_roles'] and isinstance(user['library_roles'], str):
            user['library_roles'] = [r.strip() for r in user['library_roles'].split(',') if r.strip()]
        elif not user['library_roles']:
            user['library_roles'] = []
        if user.get('job_position'):
            user['title'] = user['job_position']
            user['source'] = 'dingtalk'
            if not user['department']:
                user['department'] = user['title']
        dingtalk_uid = str(user_dict.get('dingtalk_userid') or '').strip()
        if dingtalk_uid:
            user['userid'] = dingtalk_uid
            user['source'] = 'dingtalk'
        dingtalk_parsed = self._parse_dingtalk_data_safe(user_dict.get('dingtalk_data'))
        if dingtalk_parsed:
            self._merge_dingtalk_fields(user, dingtalk_parsed)
        return self._strip_heavy_user_fields(user)

    @staticmethod
    def _extract_password_from_row(row, cursor=None) -> str:
        """从 MySQL 行读取 password 列（勿经 _row_to_user_dict，其会移除 password）。"""
        if isinstance(row, dict):
            raw = row.get('password', '')
        elif cursor and getattr(cursor, 'description', None):
            columns = [desc[0] for desc in cursor.description]
            mapping = dict(zip(columns, row)) if columns else {}
            raw = mapping.get('password', '')
        else:
            raw = ''
        if raw is None:
            return ''
        return str(raw).strip()

    def _verify_password_against_stored(self, cleaned_password: str, stored_password: str) -> bool:
        stored_password = (stored_password or '').strip()
        if not cleaned_password or not stored_password:
            return False
        if PasswordHasher.verify_password(cleaned_password, stored_password):
            return True
        return False

    def _minimal_login_user(self, user_dict: Dict[str, Any], username: str) -> Dict[str, Any]:
        """构建最小登录用户对象（_build_user_from_db_row 失败时的兜底）。"""
        roles = self._parse_roles(user_dict.get('roles', ''))
        return {
            'id': user_dict.get('id'),
            'username': user_dict.get('username', username),
            'name': user_dict.get('name', ''),
            'roles': roles,
            'role': self._primary_role(roles),
            'department': user_dict.get('department', ''),
            'status': user_dict.get('status', 'active'),
        }

    def authenticate_user_for_login(self, username, password):
        """HTTP 登录专用：仅 MySQL 校验密码，不走 Excel 回退，避免附加逻辑导致误判失败。"""
        username = (username or '').strip()
        cleaned_password = (password or '').strip()
        if not username or not cleaned_password:
            return None
        try:
            from server.db_adapter import get_connection_pool
            pool = get_connection_pool()
            with pool.get_cursor() as cursor:
                cursor.execute('SELECT * FROM users WHERE username = %s', (username,))
                row = cursor.fetchone()
                if not row:
                    logger.warning(f"登录用户不存在: {username!r}")
                    return None
                # 密码须在 _row_to_user_dict 之前取出（该方法会 strip password 字段）
                stored_password = self._extract_password_from_row(row, cursor)
                user_dict = self._row_to_user_dict(row, cursor)
                if not user_dict:
                    return None
            if not self._verify_password_against_stored(cleaned_password, stored_password):
                logger.warning(f"登录密码错误: {username!r}")
                return None
            try:
                user = self._build_user_from_db_row(user_dict)
            except Exception as build_exc:
                logger.warning(f"登录用户对象构建失败，使用最小对象: {username!r}, {build_exc}")
                user = self._minimal_login_user(user_dict, username)
            logger.info(f"用户认证成功: {username} (登录专用 MySQL 路径)")
            return user
        except Exception as e:
            logger.warning(f"登录 MySQL 认证异常: {username!r}, {e}", exc_info=True)
            return None

    def authenticate_user(self, username, password):
        """验证用户登录（仅 MySQL，与 authenticate_user_for_login 一致）。"""
        return self.authenticate_user_for_login(username, password)

    def get_all_users(self, statuses: Optional[List[str]] = None, department: Optional[str] = None):
        """获取所有用户，可过滤状态和部门
        
        Args:
            statuses: 状态列表，如 ['active', 'pending']
            department: 部门代码，如 'hardware_rd'
        
        Returns:
            用户列表
        """
        try:
            return self._get_users_from_db(statuses=statuses, department=department)
        except Exception as e:
            logger.error(f"从 MySQL 读取用户失败: {e}", exc_info=True)
            return []
    
    def _get_users_from_db(self, statuses: Optional[List[str]] = None, department: Optional[str] = None):
        """从MySQL数据库获取用户列表
        
        Args:
            statuses: 状态列表
            department: 部门代码
        
        Returns:
            用户列表
        """
        try:
            from server.mysql_connection_pool import get_mysql_connection_pool
            
            pool = get_mysql_connection_pool()
            users = []
            
            with pool.get_cursor() as cursor:
                # 构建查询（MySQL使用%s占位符）
                query = "SELECT * FROM users WHERE 1=1"
                params = []  # 确保params是列表，不是字典
                
                if statuses:
                    # 确保statuses是列表，且每个元素都是字符串
                    if isinstance(statuses, dict):
                        statuses = list(statuses.values()) if statuses else []
                    elif not isinstance(statuses, list):
                        statuses = [statuses] if statuses else []
                    # 确保每个元素都是字符串（不是字典或其他类型）
                    statuses = [str(s) for s in statuses if s is not None]
                    if statuses:
                        placeholders = ','.join(['%s'] * len(statuses))
                        query += f" AND status IN ({placeholders})"
                        params.extend(statuses)
                
                if department:
                    # 确保department是字符串，不是字典
                    if isinstance(department, dict):
                        department = department.get('department', '') if department else ''
                    query += " AND department = %s"
                    params.append(str(department))
                
                query += " ORDER BY id"
                
                # 确保params是元组（pymysql要求）
                params_tuple = tuple(params) if params else ()
                
                cursor.execute(query, params_tuple)
                rows = cursor.fetchall()
                
                # 转换为字典格式（MySQL返回字典格式）
                for row in rows:
                    user = self._row_to_user_dict(row, cursor)
                    if user:
                        users.append(user)
            
            return users
        except Exception as e:
            logger.error(f"从MySQL获取用户列表失败: {e}", exc_info=True)
            return []
    
    def _row_to_user_dict(self, row, cursor=None):
        """将MySQL查询结果的一行转换为用户字典
        
        Args:
            row: MySQL查询结果的一行（可能是dict或tuple）
            cursor: 数据库游标（用于获取列名，如果row是tuple）
        
        Returns:
            用户字典，如果转换失败返回None
        """
        try:
            import json
            
            # 转换为字典格式
            if isinstance(row, dict):
                user_dict = row
            else:
                # 如果是tuple，需要获取列名
                if cursor and hasattr(cursor, 'description'):
                    columns = [desc[0] for desc in cursor.description]
                    user_dict = dict(zip(columns, row)) if columns else {}
                else:
                    # 如果没有cursor，尝试从_get_users_from_db的上下文获取
                    # 这种情况下，我们假设row已经是dict格式（pymysql默认返回dict）
                    user_dict = row if isinstance(row, dict) else {}
            
            if not user_dict:
                return None
                    
            # 转换为标准格式
            # 注意：password字段保留在内部使用（用于认证），但不应返回给前端
            user = {
                'id': user_dict.get('id'),
                'username': user_dict.get('username', ''),
                'password': user_dict.get('password', ''),  # 保留用于认证，但_public_user_payload会过滤
                'name': user_dict.get('name', ''),
                'department': user_dict.get('department', ''),
                'job_position': user_dict.get('job_position'),
                'roles': user_dict.get('roles', ''),
                'library_roles': user_dict.get('library_roles', ''),
                'status': user_dict.get('status', 'active'),
                'created_time': user_dict.get('created_time'),
                'updated_time': user_dict.get('updated_time'),
                'last_login_time': user_dict.get('last_login_time'),
                'userid': str(user_dict.get('dingtalk_userid') or '').strip(),
                'unionid': str(user_dict.get('dingtalk_unionid') or '').strip(),
                'job_number': str(user_dict.get('job_number') or '').strip(),
                'source': user_dict.get('user_source') or 'local',
            }
            if user['status'] == 'inactive':
                user['status'] = 'disabled'

            dingtalk_parsed = self._parse_dingtalk_data_safe(user_dict.get('dingtalk_data'))
            if dingtalk_parsed:
                self._merge_dingtalk_fields(user, dingtalk_parsed)
            
            # 确保unionid字段存在（即使为空）
            if 'unionid' not in user:
                user['unionid'] = ''
            
            # 解析roles（如果是字符串）
            if user['roles'] and isinstance(user['roles'], str):
                user['roles'] = self._parse_roles(user['roles'])
            elif not user['roles']:
                user['roles'] = []
            
            # 解析library_roles（如果是字符串）
            if user['library_roles'] and isinstance(user['library_roles'], str):
                user['library_roles'] = [r.strip() for r in user['library_roles'].split(',') if r.strip()]
            elif not user['library_roles']:
                user['library_roles'] = []
            
            # 如果job_position字段有值，可能是钉钉用户的title
            if user.get('job_position'):
                user['title'] = user['job_position']
                if user.get('source') != 'dingtalk' and user.get('userid'):
                    user['source'] = 'dingtalk'
                if not user['department']:
                    user['department'] = user['title']

            if not user.get('userid'):
                user['userid'] = ''
            if not user.get('unionid'):
                user['unionid'] = ''

            return self._strip_heavy_user_fields(user)
        except Exception as e:
            logger.error(f"转换MySQL行数据到用户字典失败: {e}", exc_info=True)
            return None
    
    def _get_users_from_excel_fallback(self, statuses: Optional[List[str]] = None, department: Optional[str] = None):
        """已停用 Excel 回退，账号数据仅以 MySQL 为准。"""
        return []
    
    def get_user_by_userid(self, userid: str) -> Optional[Dict[str, Any]]:
        """根据钉钉userid获取用户（优先从MySQL，然后从Excel）
        
        Args:
            userid: 钉钉userid（字符串）
        
        Returns:
            用户字典，如果未找到返回None
        """
        userid_str = str(userid).strip()
        if not userid_str:
            return None
        
        try:
            from server.mysql_connection_pool import get_mysql_connection_pool
            pool = get_mysql_connection_pool()
            with pool.get_cursor() as cursor:
                queries = [
                    ("SELECT * FROM users WHERE dingtalk_userid = %s LIMIT 1", (userid_str,)),
                    (
                        "SELECT * FROM users WHERE JSON_UNQUOTE(JSON_EXTRACT(dingtalk_data, '$.userid')) = %s "
                        "AND dingtalk_data IS NOT NULL LIMIT 1",
                        (userid_str,),
                    ),
                    ("SELECT * FROM users WHERE username = %s LIMIT 1", (userid_str,)),
                ]
                for sql, params in queries:
                    try:
                        cursor.execute(sql, params)
                        row = cursor.fetchone()
                        if row:
                            user = self._row_to_user_dict(row, cursor)
                            if user:
                                return user
                    except Exception as qe:
                        logger.debug(f"按 userid 查询失败 ({sql[:40]}...): {qe}")
            logger.warning(f"MySQL 中未找到用户 (userid={userid_str})")
            return None
        except Exception as e:
            logger.error(f"根据userid获取用户失败: {e}", exc_info=True)
            return None

    def get_user_by_dingtalk_identity(
        self,
        userid: str = '',
        unionid: str = '',
        job_number: str = '',
    ) -> Optional[Dict[str, Any]]:
        """按钉钉身份绑定现有用户：userid -> unionid -> 工号/用户名。"""
        userid = str(userid or '').strip()
        unionid = str(unionid or '').strip()
        job_number = str(job_number or '').strip()
        try:
            from server.db_adapter import get_connection_pool

            pool = get_connection_pool()
            queries = []
            if userid:
                queries.extend([
                    ("SELECT * FROM users WHERE dingtalk_userid = %s LIMIT 1", (userid,)),
                    (
                        "SELECT * FROM users WHERE JSON_UNQUOTE(JSON_EXTRACT(dingtalk_data, '$.userid')) = %s "
                        "AND dingtalk_data IS NOT NULL LIMIT 1",
                        (userid,),
                    ),
                    ("SELECT * FROM users WHERE username = %s LIMIT 1", (userid,)),
                ])
            if unionid:
                queries.extend([
                    ("SELECT * FROM users WHERE dingtalk_unionid = %s LIMIT 1", (unionid,)),
                    (
                        "SELECT * FROM users WHERE JSON_UNQUOTE(JSON_EXTRACT(dingtalk_data, '$.unionid')) = %s "
                        "AND dingtalk_data IS NOT NULL LIMIT 1",
                        (unionid,),
                    ),
                ])
            if job_number:
                queries.extend([
                    ("SELECT * FROM users WHERE job_number = %s LIMIT 1", (job_number,)),
                    ("SELECT * FROM users WHERE username = %s LIMIT 1", (job_number,)),
                ])

            with pool.get_cursor() as cursor:
                for sql, params in queries:
                    try:
                        cursor.execute(sql, params)
                        row = cursor.fetchone()
                        if row:
                            user = self._row_to_user_dict(row, cursor)
                            if user:
                                return user
                    except Exception as qe:
                        logger.debug(f"按钉钉身份查询失败 ({sql[:48]}...): {qe}")
            return None
        except Exception as e:
            logger.error(f"按钉钉身份获取用户失败: {e}", exc_info=True)
            return None

    def upsert_dingtalk_login_user(self, profile: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """钉钉登录后绑定或自动创建 hardware 用户。

        只更新身份资料与 last_login_time，不覆盖 roles/library_roles/status。
        """
        profile = profile or {}
        userid = str(profile.get('userid') or profile.get('userId') or '').strip()
        unionid = str(profile.get('unionid') or profile.get('unionId') or '').strip()
        job_number = str(profile.get('job_number') or profile.get('jobNumber') or '').strip()
        if not userid:
            logger.warning("钉钉登录用户缺少 userid，拒绝创建/绑定")
            return None

        name = (
            str(profile.get('name') or profile.get('nick') or profile.get('nickname') or userid)
            .strip()
            or userid
        )
        title = str(profile.get('title') or profile.get('job_position') or '').strip()
        department = str(profile.get('department') or profile.get('dept_name') or '').strip()
        if not department and title:
            department = title

        dingtalk_data = {}
        for key in DINGTALK_MERGE_KEYS:
            if key in profile:
                dingtalk_data[key] = profile.get(key)
        dingtalk_data['userid'] = userid
        if unionid:
            dingtalk_data['unionid'] = unionid
        if job_number:
            dingtalk_data['job_number'] = job_number
        if name:
            dingtalk_data['name'] = name
        if title:
            dingtalk_data['title'] = title

        try:
            from server.db_adapter import get_connection_pool

            self._ensure_mysql_user_schema()
            pool = get_connection_pool()
            existing = self.get_user_by_dingtalk_identity(userid, unionid, job_number)

            with pool.get_cursor() as cursor:
                if existing:
                    cursor.execute(
                        '''
                        UPDATE users
                        SET name = %s,
                            department = COALESCE(NULLIF(%s, ''), department),
                            job_position = COALESCE(NULLIF(%s, ''), job_position),
                            dingtalk_userid = %s,
                            dingtalk_unionid = COALESCE(NULLIF(%s, ''), dingtalk_unionid),
                            job_number = COALESCE(NULLIF(%s, ''), job_number),
                            user_source = 'dingtalk',
                            dingtalk_data = %s,
                            last_login_time = NOW(),
                            updated_time = NOW()
                        WHERE id = %s
                        ''',
                        (
                            name,
                            department,
                            title,
                            userid,
                            unionid,
                            job_number,
                            json.dumps(dingtalk_data, ensure_ascii=False, default=str),
                            int(existing['id']),
                        ),
                    )
                    user_id = int(existing['id'])
                    logger.info("钉钉用户已绑定现有账号: userid=%s, user_id=%s", userid, user_id)
                else:
                    username = job_number or userid
                    cursor.execute('SELECT id FROM users WHERE username = %s LIMIT 1', (username,))
                    if cursor.fetchone():
                        username = userid
                    unusable_password = PasswordHasher.hash_password(
                        'dingtalk-disabled-' + secrets.token_urlsafe(48)
                    )
                    cursor.execute(
                        '''
                        INSERT INTO users (
                            username, password, name, department, job_position,
                            roles, library_roles, status, created_time, updated_time,
                            last_login_time, dingtalk_userid, dingtalk_unionid,
                            job_number, user_source, dingtalk_data
                        )
                        VALUES (%s, %s, %s, %s, %s, 'user', '', %s, NOW(), NOW(),
                            NOW(), %s, %s, %s, 'dingtalk', %s)
                        ''',
                        (
                            username,
                            unusable_password,
                            name,
                            department,
                            title or None,
                            STATUS_ACTIVE,
                            userid,
                            unionid or None,
                            job_number or None,
                            json.dumps(dingtalk_data, ensure_ascii=False, default=str),
                        ),
                    )
                    user_id = int(cursor.lastrowid)
                    logger.info("钉钉首次登录自动创建普通用户: userid=%s, user_id=%s", userid, user_id)

            return self.get_user_by_id(user_id)
        except Exception as e:
            logger.error(f"钉钉登录用户绑定/创建失败: {e}", exc_info=True)
            return None
    
    def get_user_by_username(self, username: str) -> Optional[Dict[str, Any]]:
        """根据 username 获取用户（仅 MySQL）。"""
        username = (username or '').strip()
        if not username:
            return None
        try:
            from server.db_adapter import get_connection_pool
            pool = get_connection_pool()
            with pool.get_cursor() as cursor:
                cursor.execute('SELECT * FROM users WHERE username = %s LIMIT 1', (username,))
                row = cursor.fetchone()
                if row:
                    return self._row_to_user_dict(row, cursor)
            return None
        except Exception as e:
            logger.error(f"根据username获取用户失败: {e}", exc_info=True)
            return None

    def get_user_by_id(self, user_id: int) -> Optional[Dict[str, Any]]:
        """根据 ID 获取用户（仅 MySQL）。"""
        try:
            uid = int(user_id)
        except (TypeError, ValueError):
            return None
        try:
            from server.db_adapter import get_connection_pool
            pool = get_connection_pool()
            with pool.get_cursor() as cursor:
                cursor.execute('SELECT * FROM users WHERE id = %s LIMIT 1', (uid,))
                row = cursor.fetchone()
                if row:
                    user = self._row_to_user_dict(row, cursor)
                    if user and user.get('id') is not None:
                        try:
                            user['id'] = int(user['id'])
                        except (TypeError, ValueError):
                            pass
                    return user
            return None
        except Exception as e:
            logger.error(f"根据ID获取用户失败: {e}", exc_info=True)
            return None
    
    def get_user_count(self):
        """获取激活用户数量"""
        return len(self.get_all_users(statuses=[STATUS_ACTIVE]))
    
    def add_user(self, user_data):
        """添加用户（写入 MySQL）。"""
        username = (user_data.get('username') or '').strip()
        if not username:
            return False, "用户名不能为空"
        roles = user_data.get('roles', ['user'])
        if isinstance(roles, str):
            roles = self._parse_roles(roles)
        elif not isinstance(roles, list):
            roles = ['user']
        roles = self.sanitize_job_roles(roles)
        library_roles = user_data.get('library_roles', [])
        if isinstance(library_roles, str):
            library_roles = self._parse_roles(library_roles)
        elif not isinstance(library_roles, list):
            library_roles = []
        library_roles = self.sanitize_library_roles_list(library_roles)
        password_raw = user_data.get('password')
        if not password_raw:
            from server.auth.password_service import PasswordService
            password_raw = PasswordService.generate_temporary_password()
        password = password_raw
        if not (str(password).startswith('$2b$') or str(password).startswith('$2a$') or ':' in str(password)):
            password = PasswordHasher.hash_password(str(password))
        name = (user_data.get('name') or username).strip() or username
        department = (user_data.get('department') or '').strip()
        status = user_data.get('status', STATUS_ACTIVE)
        if status == 'inactive':
            status = 'disabled'
        roles_str = self._roles_to_string(roles)
        lib_str = ','.join(library_roles) if library_roles else ''
        system_roles = {'admin', 'super_admin', 'management'}
        job_roles = [r for r in roles if r not in system_roles and r in JOB_ROLE_OPTIONS.keys()]
        job_position = (user_data.get('job_position') or '').strip()
        if not job_position and job_roles:
            job_position = job_roles[0]
        try:
            from server.db_adapter import get_connection_pool
            pool = get_connection_pool()
            with pool.get_cursor() as cursor:
                cursor.execute('SELECT id FROM users WHERE username = %s', (username,))
                if cursor.fetchone():
                    return False, "用户名已存在"
                cursor.execute(
                    '''
                    INSERT INTO users (username, password, name, department, job_position,
                        roles, library_roles, status, created_time, updated_time)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                    ''',
                    (username, password, name, department, job_position or None, roles_str, lib_str, status),
                )
            return True, "用户添加成功"
        except Exception as e:
            logger.error(f"add_user MySQL 失败: {e}", exc_info=True)
            return False, f"添加用户失败: {e}"
    
    def register_user(self, registration_data):
        """注册新用户，状态为待审批"""
        required_fields = ['username', 'password', 'name', 'roles']
        for field in required_fields:
            if not registration_data.get(field):
                return False, f"缺少必要字段: {field}"
        
        chosen_roles = self.sanitize_job_roles(registration_data.get('roles', []))
        if not chosen_roles:
            return False, "请选择至少一个岗位"
        
        user_payload = {
            'username': registration_data['username'],
            'password': registration_data['password'],
            'name': registration_data['name'],
            'roles': chosen_roles,
            'department': registration_data.get('department', ''),
            'status': STATUS_PENDING,
            'create_time': datetime.now().isoformat()
        }
        return self.add_user(user_payload)
    
    def sanitize_job_roles(self, roles):
        role_list = self._parse_roles(roles)
        out: List[str] = []
        for role in role_list:
            if role in JOB_ROLE_OPTIONS:
                if role not in out:
                    out.append(role)
            elif role in LEGACY_JOB_POSITION_MAP:
                mapped = LEGACY_JOB_POSITION_MAP[role]
                if mapped not in out:
                    out.append(mapped)
        return out if out else ['user']

    def normalize_job_roles_for_form(self, roles) -> List[str]:
        """编辑表单回显：将历史岗位键映射为 user / manager。"""
        parsed = self._parse_roles(roles)
        out: List[str] = []
        job_keys = set(JOB_ROLE_OPTIONS.keys())
        for role in parsed:
            if role in job_keys and role not in out:
                out.append(role)
            elif role in LEGACY_JOB_POSITION_MAP:
                mapped = LEGACY_JOB_POSITION_MAP[role]
                if mapped not in out:
                    out.append(mapped)
        return out[:1] if out else ['user']
    
    def merge_roles(self, existing_roles: List[str], new_job_roles: List[str]) -> List[str]:
        """合并角色：保留系统角色（admin, super_admin, management），更新岗位角色"""
        # 系统角色列表（不会被岗位角色覆盖）
        system_roles = {'admin', 'super_admin', 'management'}
        job_role_keys = set(JOB_ROLE_OPTIONS.keys())
        
        # 解析现有角色
        existing_parsed = self._parse_roles(existing_roles)
        
        # 解析新角色
        new_parsed = self._parse_roles(new_job_roles)
        
        # 保留系统角色（从现有角色中）
        preserved_system = [role for role in existing_parsed if role in system_roles]
        
        # 保留非岗位角色的系统角色（从新角色中）
        new_system_roles = [role for role in new_parsed if role in system_roles]
        
        # 合并系统角色（去重）
        all_system_roles = list(set(preserved_system + new_system_roles))
        
        # 获取岗位角色（从新角色中，过滤掉系统角色）
        job_roles = [role for role in new_parsed if role in job_role_keys or role == 'management']
        
        # 如果新角色中包含系统角色但没有岗位角色，保留所有系统角色
        if new_system_roles and not job_roles:
            # 直接使用新角色（包含系统角色）
            return new_parsed
        
        # 合并：系统角色 + 岗位角色
        merged = all_system_roles + job_roles
        
        # 去重并返回
        return list(set(merged))
    
    def apply_role_update(self, existing_roles, selected_roles: List[str]) -> List[str]:
        """账号管理提交的完整岗位+系统角色：可增删 admin/management；保留 super_admin 与旧版 roles（如 *_manager）。"""
        existing_parsed = self._parse_roles(existing_roles)
        selected_parsed = self._parse_roles(selected_roles)
        system_set = {'admin', 'super_admin', 'management'}
        job_keys = set(JOB_ROLE_OPTIONS.keys())
        legacy = [r for r in existing_parsed if r not in system_set and r not in job_keys]
        selected_set = set(selected_parsed)
        if 'super_admin' in existing_parsed and 'super_admin' not in selected_set:
            selected_parsed = list(selected_set) + ['super_admin']
            selected_set = set(selected_parsed)
        system_kept = [r for r in selected_parsed if r in system_set]
        job_from_selected = [r for r in selected_parsed if r in job_keys]
        # 表单未勾选岗位时保留原有岗位角色，避免误提交 roles:[] 清空 circuit 等
        if not job_from_selected:
            job_from_selected = [r for r in existing_parsed if r in job_keys]
        core = system_kept + job_from_selected
        return list(dict.fromkeys(core + legacy))
    
    def resolve_handoff_user_id(self, raw) -> Optional[int]:
        """将接替人参数解析为 users.id（支持数字 ID 或用户名）。"""
        if raw is None:
            return None
        s = str(raw).strip()
        if not s:
            return None
        if s.isdigit():
            uid = int(s)
            u = self.get_user_by_id(uid)
            return uid if u else None
        u = self.get_user_by_username(s)
        if u and u.get('id') is not None:
            return int(u['id'])
        return None
    
    def transfer_library_roles(self, from_user_id: int, to_user_id: int) -> tuple:
        """把 from 用户的 library_roles 合并到 to 用户，并清空 from。"""
        if from_user_id == to_user_id:
            return False, "接替人不能与当前用户相同"
        from_u = self.get_user_by_id(int(from_user_id))
        to_u = self.get_user_by_id(int(to_user_id))
        if not from_u or not to_u:
            return False, "用户不存在"
        from_libs = self._parse_roles(from_u.get('library_roles', []))
        if not from_libs:
            return True, "无物料库权限需转移"
        to_libs = self._parse_roles(to_u.get('library_roles', []))
        merged = list(dict.fromkeys(to_libs + from_libs))
        ok, msg = self.update_user(int(to_user_id), {'library_roles': merged})
        if not ok:
            return False, msg or "交接失败"
        ok2, msg2 = self.update_user(int(from_user_id), {'library_roles': []})
        if not ok2:
            return False, msg2 or "清空原用户库权限失败"
        return True, "库权限已交接"
    
    def remove_library_id_from_all_users(self, lib_id: str) -> int:
        """从所有用户 library_roles 中移除指定物料库 id（逗号分隔列表）。"""
        if not lib_id or not str(lib_id).strip():
            return 0
        lib_id = str(lib_id).strip()
        changed = 0
        try:
            from server.db_adapter import get_connection_pool
            pool = get_connection_pool()
            with pool.get_cursor() as cursor:
                cursor.execute(
                    "SELECT id, library_roles FROM users WHERE library_roles IS NOT NULL "
                    "AND library_roles != '' AND FIND_IN_SET(%s, library_roles) > 0",
                    (lib_id,),
                )
                rows = cursor.fetchall() or []
                for row in rows:
                    if isinstance(row, dict):
                        uid = row.get('id')
                        lr = row.get('library_roles') or ''
                    else:
                        uid = row[0]
                        lr = row[1] or ''
                    parts = [p.strip() for p in self._parse_roles(lr) if p.strip() and p.strip() != lib_id]
                    new_str = ','.join(parts)
                    cursor.execute(
                        'UPDATE users SET library_roles = %s, updated_time = NOW() WHERE id = %s',
                        (new_str, uid),
                    )
                    changed += 1
            return changed
        except Exception as e:
            logger.error(f"remove_library_id_from_all_users 失败: {e}", exc_info=True)
            return changed
    
    def sanitize_library_roles_list(self, library_roles) -> List[str]:
        """只保留当前 material-db 中存在的库 id。"""
        try:
            from server import material_db_manager as mdb
            libs = mdb.list_libraries()
            valid = {str(l.get('id')) for l in libs if l and l.get('id')}
        except Exception as e:
            logger.warning(f"sanitize_library_roles_list: {e}")
            return self._parse_roles(library_roles)
        parsed = self._parse_roles(library_roles)
        return [r for r in parsed if r in valid]
    
    def update_job_roles(self, user_id: int, job_roles: List[str]):
        """仅更新岗位角色"""
        user = self.get_user_by_id(user_id)
        if not user:
            return False, "用户不存在"
        merged = self.merge_roles(user.get('roles', []), job_roles)
        return self.update_user(user_id, {'roles': merged})
    
    def update_user(self, user_id, user_data, skip_session_invalidation=False):
        """更新用户信息（只更新MySQL数据库，不再使用Excel文件）"""
        user_data = dict(user_data)
        handoff_raw = user_data.pop('library_handoff_user_id', None)
        if 'password' in user_data and not skip_session_invalidation:
            user_data.pop('password', None)
        try:
            # 首先尝试更新MySQL数据库
            try:
                from server.db_adapter import get_connection_pool
                
                pool = get_connection_pool()
                pending_password_session_clear = False
                pending_security_session_clear = False
                pending_update_ok = False
                pending_no_field_update = False
                with pool.get_cursor() as cursor:
                    # 检查用户是否存在，并获取用户信息
                    cursor.execute('SELECT * FROM users WHERE id = %s', (user_id,))
                    existing_user_row = cursor.fetchone()
                    
                    if existing_user_row:
                        # 转换为字典格式
                        if isinstance(existing_user_row, dict):
                            existing_user_dict = existing_user_row
                        else:
                            columns = [desc[0] for desc in cursor.description] if hasattr(cursor, 'description') else []
                            existing_user_dict = dict(zip(columns, existing_user_row)) if columns else {}
                        
                        # 禁用：仍持有物料库权限时需先交接（除非同请求显式清空 library_roles）
                        if 'status' in user_data:
                            st = user_data['status']
                            if st == 'inactive':
                                user_data['status'] = 'disabled'
                                st = 'disabled'
                            if st == 'disabled':
                                prev_st = str(existing_user_dict.get('status', '')).lower()
                                existing_libs = self._parse_roles(existing_user_dict.get('library_roles') or '')
                                explicit_clear = (
                                    'library_roles' in user_data
                                    and not self._parse_roles(user_data.get('library_roles'))
                                )
                                if existing_libs and prev_st != 'disabled' and not explicit_clear:
                                    hid = self.resolve_handoff_user_id(handoff_raw)
                                    if not hid and DEFAULT_LIBRARY_HANDOFF_USER_ID:
                                        hid = self.resolve_handoff_user_id(DEFAULT_LIBRARY_HANDOFF_USER_ID)
                                    if not hid or hid == int(user_id):
                                        return False, (
                                            "该用户仍持有物料库管理权限，禁用前请在请求体中加入 "
                                            "library_handoff_user_id（接替人用户数字 ID），或配置环境变量 "
                                            "DEFAULT_LIBRARY_HANDOFF_USER_ID。"
                                        )
                                    ok_t, msg_t = self.transfer_library_roles(int(user_id), hid)
                                    if not ok_t:
                                        return False, msg_t or "库权限交接失败"
                                    user_data.pop('library_roles', None)
                                    cursor.execute('SELECT * FROM users WHERE id = %s', (user_id,))
                                    existing_user_row = cursor.fetchone()
                                    if isinstance(existing_user_row, dict):
                                        existing_user_dict = existing_user_row
                                    else:
                                        columns = [desc[0] for desc in cursor.description] if hasattr(cursor, 'description') else []
                                        existing_user_dict = dict(zip(columns, existing_user_row)) if columns else {}
                        
                        # 构建更新语句
                        updates = []
                        params = []
                        password_changed = False
                        
                        if 'name' in user_data:
                            updates.append('name = %s')
                            params.append(user_data['name'])
                        
                        if 'password' in user_data and user_data['password']:
                            password_changed = True
                            # 对密码进行哈希处理（如果还不是哈希格式）
                            password_value = user_data['password']
                            # 检查是否已经是哈希格式（bcrypt哈希以$2b$开头，SHA256哈希包含冒号）
                            if not (password_value.startswith('$2b$') or ':' in password_value):
                                # 明文密码，需要哈希
                                password_value = PasswordHasher.hash_password(password_value)
                                logger.info(f"密码已哈希处理 (用户ID={user_id})")
                            updates.append('password = %s')
                            params.append(password_value)
                        
                        if 'department' in user_data:
                            updates.append('department = %s')
                            params.append(user_data['department'])
                        
                        # 处理岗位字段：优先从roles中提取岗位角色，然后是job_position，title，最后是department
                        job_position_value = None
                        
                        # 1. 优先从roles中提取岗位角色（最准确）
                        if 'roles' in user_data:
                            roles = user_data['roles']
                            if isinstance(roles, str):
                                roles = self._parse_roles(roles)
                            elif not isinstance(roles, list):
                                roles = []
                            
                            # 过滤出岗位角色（排除系统角色）
                            system_roles = {'admin', 'super_admin', 'management'}
                            job_roles = [r for r in roles if r not in system_roles and r in JOB_ROLE_OPTIONS.keys()]
                            
                            if job_roles:
                                job_position_value = job_roles[0]  # 取第一个岗位角色
                                logger.info(f"从roles中提取岗位角色: {job_position_value} (ID={user_id})")
                        
                        # 2. 如果roles中没有岗位角色，使用job_position
                        if not job_position_value and 'job_position' in user_data:
                            job_position_value = user_data['job_position']
                            logger.info(f"使用job_position字段: {job_position_value} (ID={user_id})")
                        
                        # 3. 如果还没有，使用title
                        if not job_position_value and 'title' in user_data:
                            job_position_value = user_data['title']
                            logger.info(f"使用title字段: {job_position_value} (ID={user_id})")
                        
                        # 4. 最后，如果都没有，且是钉钉用户，使用department（对于钉钉用户，department可能是岗位）
                        if not job_position_value and 'department' in user_data:
                            # 检查是否是钉钉用户（通过检查username是否匹配job_number或userid格式）
                            # 或者通过检查是否有title字段（钉钉用户通常有title）
                            username = existing_user_dict.get('username', '') if existing_user_dict else ''
                            is_dingtalk_user = (
                                username and (
                                    len(username) > 10 or  # 钉钉userid通常较长
                                    existing_user_dict.get('job_position')  # 如果有job_position，可能是钉钉用户
                                )
                            )
                            if is_dingtalk_user:
                                job_position_value = user_data['department']
                                logger.info(f"钉钉用户使用department作为岗位: {job_position_value} (ID={user_id})")
                        
                        if job_position_value:
                            updates.append('job_position = %s')
                            params.append(job_position_value)
                        
                        if 'library_roles' in user_data:
                            library_roles = self.sanitize_library_roles_list(user_data['library_roles'])
                            if isinstance(library_roles, list):
                                library_roles_str = ','.join(library_roles) if library_roles else ''
                            else:
                                library_roles_str = str(library_roles) if library_roles else ''
                            updates.append('library_roles = %s')
                            params.append(library_roles_str)
                        
                        if 'roles' in user_data:
                            roles = user_data['roles']
                            if isinstance(roles, list):
                                roles_str = self._roles_to_string(roles)
                            elif isinstance(roles, str):
                                roles_str = roles
                            else:
                                roles_str = ''
                            updates.append('roles = %s')
                            params.append(roles_str)
                            logger.info(f"更新roles字段: {roles_str} (ID={user_id})")
                        
                        if 'status' in user_data:
                            # 确保status值符合MySQL约束：'active', 'disabled', 'pending', 'rejected'
                            status_value = user_data['status']
                            # 将'inactive'映射为'disabled'（向后兼容）
                            if status_value == 'inactive':
                                status_value = 'disabled'
                                logger.info(f"将status值从'inactive'映射为'disabled' (ID={user_id})")
                            # 验证status值
                            valid_statuses = ['active', 'disabled', 'pending', 'rejected']
                            if status_value not in valid_statuses:
                                logger.warning(f"无效的status值: {status_value}，使用默认值'active' (ID={user_id})")
                                status_value = 'active'
                            updates.append('status = %s')
                            params.append(status_value)
                        
                        if updates:
                            params.append(user_id)
                            query = f"UPDATE users SET {', '.join(updates)}, updated_time = NOW() WHERE id = %s"
                            cursor.execute(query, params)
                            logger.info(f"成功更新MySQL用户 (ID={user_id})")
                            # 改密清会话在 with 外执行（事务已提交），见下方
                            pending_password_session_clear = password_changed
                            if 'roles' in user_data or 'status' in user_data or 'library_roles' in user_data:
                                pending_security_session_clear = True
                            pending_update_ok = True
                        else:
                            logger.warning(f"没有需要更新的字段 (ID={user_id})")
                            pending_password_session_clear = False
                            pending_update_ok = False
                            pending_no_field_update = True
                    else:
                        logger.warning(f"MySQL数据库中未找到用户 (ID={user_id})")
                        return False, f"用户不存在 (ID={user_id})"

                if not skip_session_invalidation and (pending_password_session_clear or pending_security_session_clear):
                    try:
                        from server.auth.password_service import get_password_service
                        removed = get_password_service(self).invalidate_user_sessions(int(user_id))
                        logger.info(
                            f"用户更新已清除会话 (ID={user_id}, sessions_removed={removed})"
                        )
                    except Exception as sess_err:
                        logger.warning(
                            f"更新后会话清理失败 (ID={user_id}): {sess_err}"
                        )

                if pending_update_ok:
                    return True, "用户更新成功"
                if pending_no_field_update:
                    return True, "没有需要更新的字段"
            except Exception as e:
                logger.error(f"更新MySQL用户失败: {e}", exc_info=True)
                return False, f"更新用户失败: {e}"
        
        except Exception as e:
            logger.error(f"更新用户异常: {e}", exc_info=True)
            return False, f"更新用户失败: {e}"
    
    def delete_user(self, user_id, library_handoff_user_id=None):
        """从 MySQL 删除用户；若仍持有物料库权限须先交接。"""
        try:
            uid = int(user_id)
        except (TypeError, ValueError):
            return False, "无效的用户 ID"
        user = self.get_user_by_id(uid)
        if not user:
            return False, "用户不存在"
        if self.is_super_admin(user):
            return False, "不能删除最高管理员"
        libs = self._parse_roles(user.get('library_roles', []))
        if libs:
            hid = self.resolve_handoff_user_id(library_handoff_user_id)
            if not hid and DEFAULT_LIBRARY_HANDOFF_USER_ID:
                hid = self.resolve_handoff_user_id(DEFAULT_LIBRARY_HANDOFF_USER_ID)
            if not hid or hid == uid:
                return False, (
                    "该用户仍持有物料库管理权限，删除前请传入 query 参数 library_handoff_user_id（接替人用户 ID），"
                    "或配置环境变量 DEFAULT_LIBRARY_HANDOFF_USER_ID。"
                )
            ok_t, msg_t = self.transfer_library_roles(uid, hid)
            if not ok_t:
                return False, msg_t or "库权限交接失败"
        try:
            from server.db_adapter import get_connection_pool
            pool = get_connection_pool()
            with pool.get_cursor() as cursor:
                cursor.execute('DELETE FROM users WHERE id = %s', (uid,))
                if cursor.rowcount == 0:
                    return False, "用户不存在"
            logger.info(f"成功从 MySQL 删除用户 (ID={uid})")
            return True, "用户删除成功"
        except Exception as e:
            logger.error(f"删除用户失败: {e}", exc_info=True)
            return False, f"删除用户失败: {str(e)}"
    
    def get_pending_users(self):
        return self.get_all_users(statuses=[STATUS_PENDING])
    
    def approve_registration(self, user_id, approve=True):
        """审批用户注册（MySQL）"""
        try:
            uid = int(user_id)
        except (TypeError, ValueError):
            return False, "无效的用户 ID"
        try:
            from server.db_adapter import get_connection_pool
            pool = get_connection_pool()
            with pool.get_cursor() as cursor:
                cursor.execute('SELECT id, status, roles FROM users WHERE id = %s', (uid,))
                row = cursor.fetchone()
                if not row:
                    return False, "用户不存在"
                if isinstance(row, dict):
                    current_status = row.get('status')
                    roles_val = row.get('roles') or ''
                else:
                    current_status = row[1] if len(row) > 1 else None
                    roles_val = row[2] if len(row) > 2 else ''
                if str(current_status).lower() != STATUS_PENDING:
                    return False, "该用户不是待审批状态"
                if approve:
                    new_status = STATUS_ACTIVE
                    roles_parsed = self._parse_roles(roles_val)
                    if not roles_parsed:
                        roles_str = 'user'
                    else:
                        roles_str = self._roles_to_string(roles_parsed)
                    cursor.execute(
                        'UPDATE users SET status = %s, roles = %s, updated_time = NOW() WHERE id = %s',
                        (new_status, roles_str, uid),
                    )
                else:
                    cursor.execute(
                        'UPDATE users SET status = %s, updated_time = NOW() WHERE id = %s',
                        (STATUS_REJECTED, uid),
                    )
            return True, "审批完成"
        except Exception as e:
            logger.error(f"approve_registration 失败: {e}", exc_info=True)
            return False, f"审批失败: {e}"
    
    def user_has_role(self, user, role):
        if isinstance(user, dict):
            roles = self._parse_roles(user.get('roles', []))
        elif isinstance(user, list):
            roles = user
        elif isinstance(user, str):
            roles = self._parse_roles(user)
        else:
            roles = []
        return role in roles
    
    def is_super_admin(self, user):
        from server.auth.permissions import is_super_admin as _is_super_admin
        return _is_super_admin(user)
    
    def can_manage_library(self, user, library_name):
        """检查用户是否有权限管理指定的物料库
        
        权限规则：
        1. 管理组（management部门或admin/super_admin角色）
        2. 对应物料库管理员（library_roles中包含该物料库ID）
        
        Args:
            user: 用户对象（dict）或用户角色列表/字符串
            library_name: 物料库名称
        """
        # 统一处理参数：支持user对象或roles
        if isinstance(user, dict):
            roles = user.get('roles', [])
            library_roles = user.get('library_roles', [])
            department = user.get('department', '')
            # 检查是否是超级管理员
            if self.is_super_admin(user):
                return True
        else:
            roles = self._parse_roles(user if not isinstance(user, dict) else user.get('roles', []))
            library_roles = []
            department = ''
        
        # 解析角色列表
        if isinstance(roles, str):
            roles = self._parse_roles(roles)
        if isinstance(library_roles, str):
            library_roles = self._parse_roles(library_roles)
        
        # 1. 检查是否是管理组（部门为 management / 管理组 文案，或角色含 management / admin / super_admin）
        mgmt_label = DEPARTMENT_OPTIONS.get('management', '')
        is_management = (
            department == 'management'
            or (mgmt_label and str(department).strip() == str(mgmt_label).strip())
            or 'management' in roles
            or 'admin' in roles
            or 'super_admin' in roles
        )
        
        # 2. 检查是否是对应物料库管理员（library_roles中包含该物料库ID）
        is_library_manager = library_name in library_roles
        
        # 3. 检查是否有manager_role（兼容旧的角色系统）
        library_config = LIBRARY_CONFIG.get(library_name, {})
        manager_role = library_config.get('manager_role')
        has_manager_role = manager_role in roles
        
        # 权限：管理组 或 对应物料库管理员 或 有manager_role
        return is_management or is_library_manager or has_manager_role
    
    def get_manageable_libraries(self, user_roles):
        """获取用户可以管理的物料库（静态 LIBRARY_CONFIG + material-db 动态库，按 id 去重）。"""
        if isinstance(user_roles, dict):
            roles = self._parse_roles(user_roles.get('roles', []))
            library_roles = self._parse_roles(user_roles.get('library_roles', []))
        else:
            roles = self._parse_roles(user_roles)
            library_roles = []
        
        manageable_libraries = []
        seen = set()
        
        def add_lib(lid: str, name: str):
            if not lid or lid in seen:
                return
            seen.add(lid)
            manageable_libraries.append({'id': lid, 'name': name or lid})
        
        for lib_id, lib_config in LIBRARY_CONFIG.items():
            if 'admin' in roles or 'super_admin' in roles:
                add_lib(lib_id, lib_config['name'])
            elif lib_config['manager_role'] in roles:
                add_lib(lib_id, lib_config['name'])
            elif lib_id in library_roles:
                add_lib(lib_id, lib_config['name'])
        
        try:
            from server import material_db_manager as mdb
            for lib in mdb.list_libraries():
                lid = str((lib or {}).get('id') or '').strip()
                if not lid:
                    continue
                nm = (lib.get('name') or lid) if isinstance(lib, dict) else lid
                if 'admin' in roles or 'super_admin' in roles:
                    add_lib(lid, nm)
                elif lid in library_roles:
                    add_lib(lid, nm)
        except Exception as e:
            logger.warning(f"get_manageable_libraries 读取 material-db 失败: {e}")
        
        return manageable_libraries
    
    def _is_hardware_rd_department(self, dept_id, dept_name=''):
        """检查部门是否是硬件研发部或其子部门
        
        Args:
            dept_id: 部门ID
            dept_name: 部门名称（可选，用于名称匹配）
        
        Returns:
            bool: 是否是硬件研发部或其子部门
        """
        try:
            from server.department_manager import DepartmentManager
            dept_mgr = DepartmentManager()
            
            # 使用新的方法获取硬件研发部及其所有子部门的ID列表
            hardware_rd_dept_ids = dept_mgr.get_hardware_rd_department_ids()
            
            if not hardware_rd_dept_ids:
                # 如果没有找到硬件研发部，尝试通过名称匹配
                if dept_name and ('硬件研发' in str(dept_name) or '研发部' in str(dept_name)):
                    return True
                return False
            
            # 检查dept_id是否在硬件研发部及其子部门的ID列表中
            # 确保类型一致（都转换为整数进行比较）
            try:
                dept_id_int = int(dept_id) if dept_id else None
                if dept_id_int is None:
                    return False
                return dept_id_int in hardware_rd_dept_ids
            except (ValueError, TypeError):
                # 如果转换失败，尝试直接比较（可能是字符串）
                return dept_id in hardware_rd_dept_ids or str(dept_id) in [str(d) for d in hardware_rd_dept_ids]
        except Exception as e:
            logger.warning(f"检查部门是否属于硬件研发部失败: {e}")
            # 如果检查失败，默认返回False（禁用）
            return False
    
    def _get_field_value(self, user: Dict[str, Any], field_name: str, default=None):
        """获取用户字段值，并进行必要的转换
        
        Args:
            user: 用户数据字典
            field_name: 字段名
            default: 默认值
        
        Returns:
            转换后的字段值
        """
        # 字段名映射（钉钉API可能使用驼峰命名，需要转换为下划线命名）
        field_name_mapping = {
            'unionId': 'unionid',
            'userId': 'userid',
            'jobNumber': 'job_number',
            'loginId': 'login_id',
            'deptId': 'dept_id',
            'deptIdList': 'dept_id_list',
            'deptOrder': 'dept_order',
            'parentId': 'parent_id',
            'disableStatus': 'disable_status',
            'exclusiveAccount': 'exclusive_account',
            'exclusiveAccountType': 'exclusive_account_type',
            'exclusiveAccountCorpId': 'exclusive_account_corp_id',
            'exclusiveAccountCorpName': 'exclusive_account_corp_name',
            'hideMobile': 'hide_mobile',
            'hiredDate': 'hired_date'
        }
        
        # 尝试直接获取字段值
        value = user.get(field_name, default)
        
        # 如果直接获取失败，尝试使用映射的字段名
        if value is None or value == default:
            mapped_field_name = field_name_mapping.get(field_name)
            if mapped_field_name:
                value = user.get(mapped_field_name, default)
        
        # 如果还是失败，尝试反向映射（从下划线到驼峰）
        if value is None or value == default:
            reverse_mapping = {v: k for k, v in field_name_mapping.items()}
            reverse_field_name = reverse_mapping.get(field_name)
            if reverse_field_name:
                value = user.get(reverse_field_name, default)
        
        # 处理数组字段（序列化为JSON字符串）
        if field_name == 'dept_id_list':
            if isinstance(value, list):
                return json.dumps(value, ensure_ascii=False)
            elif isinstance(value, str):
                # 如果已经是JSON字符串，直接返回
                return value
            else:
                return json.dumps([], ensure_ascii=False)
        
        # 处理时间戳字段（转换为日期字符串）
        if field_name == 'hired_date':
            if value:
                try:
                    if isinstance(value, str):
                        timestamp = int(value) / 1000
                    else:
                        timestamp = value / 1000
                    return datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
                except:
                    return str(value) if value else ''
            return ''
        
        # 处理布尔值字段（转换为字符串）
        if isinstance(value, bool):
            return 'true' if value else 'false'
        
        # 处理数字字段
        if isinstance(value, (int, float)):
            return value
        
        # 处理None值
        if value is None:
            return default if default is not None else ''
        
        # 其他类型直接返回字符串
        return str(value) if value else ''
    
    def _write_user_row(self, ws, row: int, user_data: Dict[str, Any], existing_data: Dict[str, Any] = None):
        """将用户数据写入Excel行
        
        Args:
            ws: Excel工作表
            row: 行号
            user_data: 从钉钉API获取的用户数据
            existing_data: 现有用户数据（用于保留密码和管理员信息）
        """
        # 创建字段名到列索引的映射
        field_to_col = {header: idx + 1 for idx, header in enumerate(DINGTALK_USER_HEADERS)}
        
        # 处理密码和管理员字段（保留现有值）
        password = 'CHXW_HW_123456'
        library_manager = ''
        announcement_manager = ''
        status = 'active'
        create_time = datetime.now().strftime('%Y-%m-%d')
        
        if existing_data:
            password = existing_data.get('密码') or existing_data.get('password') or password
            library_manager = existing_data.get('物料库管理员') or existing_data.get('library_manager') or ''
            announcement_manager = existing_data.get('公告栏管理员') or existing_data.get('announcement_manager') or ''
            status = existing_data.get('状态') or existing_data.get('status') or status
            create_time = existing_data.get('创建时间') or existing_data.get('create_time') or create_time
        
        # 检查是否是硬件研发部部长（leader或boss字段为True），设置为管理组
        is_leader = user_data.get('leader', False) or user_data.get('boss', False)
        dept_id = user_data.get('dept_id')
        title = user_data.get('title', '')
        
        # 判断是否是硬件研发部部长
        is_hardware_rd_leader = False
        if is_leader or is_leader is True:
            # 检查部门是否是硬件研发部
            if dept_id:
                try:
                    dept_id_int = int(dept_id) if dept_id else None
                    if dept_id_int is not None:
                        is_hardware_rd_leader = self._is_hardware_rd_department(dept_id_int, '')
                except (ValueError, TypeError):
                    pass
            
            # 也可以通过title判断
            if not is_hardware_rd_leader and title:
                if '硬件研发部' in str(title) and '部长' in str(title):
                    is_hardware_rd_leader = True
        
        # 如果是硬件研发部部长，设置为管理组（公告栏管理员）
        if is_hardware_rd_leader:
            announcement_manager = 'management'
            logger.info(f"用户 {user_data.get('name', '')} (userid={user_data.get('userid', '')}) 是硬件研发部部长，设置为管理组")
        
        # 处理创建时间（从hired_date转换）
        hired_date = user_data.get('hired_date', '')
        if hired_date:
            try:
                if isinstance(hired_date, str):
                    timestamp = int(hired_date) / 1000
                else:
                    timestamp = hired_date / 1000
                create_time = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d')
            except:
                pass
        
        # 写入所有字段
        for field_name in DINGTALK_USER_HEADERS:
            col = field_to_col.get(field_name)
            if col is None:
                continue
            
            if field_name == '密码':
                ws.cell(row=row, column=col, value=password)
            elif field_name == '物料库管理员':
                ws.cell(row=row, column=col, value=library_manager)
            elif field_name == '公告栏管理员':
                ws.cell(row=row, column=col, value=announcement_manager)
            elif field_name == '状态':
                ws.cell(row=row, column=col, value=status)
            elif field_name == '创建时间':
                ws.cell(row=row, column=col, value=create_time)
            else:
                # 优先直接从user_data获取（保持原始值）
                if field_name in user_data:
                    value = user_data[field_name]
                else:
                    value = self._get_field_value(user_data, field_name, '')
                
                # 处理特殊字段类型
                if field_name == 'dept_id_list' and isinstance(value, list):
                    value = json.dumps(value, ensure_ascii=False)
                elif field_name in ['active', 'admin', 'boss', 'leader', 'disable_status', 'exclusive_account', 'hide_mobile']:
                    # 布尔值字段：转换为字符串
                    if isinstance(value, bool):
                        value = 'true' if value else 'false'
                    elif value is None:
                        value = 'false'
                elif field_name == 'hired_date' and value:
                    # 时间戳字段：转换为日期字符串
                    try:
                        if isinstance(value, (int, float)):
                            timestamp = value / 1000 if value > 1e10 else value
                            value = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(value, str) and value.isdigit():
                            timestamp = int(value) / 1000
                            value = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        pass
                
                ws.cell(row=row, column=col, value=value)
    
    def _read_user_row(self, ws, row: int) -> Dict[str, Any]:
        """从Excel行读取用户数据
        
        Args:
            ws: Excel工作表
            row: 行号
        
        Returns:
            用户数据字典
        """
        user_data = {}
        field_to_col = {header: idx + 1 for idx, header in enumerate(DINGTALK_USER_HEADERS)}
        
        for field_name in DINGTALK_USER_HEADERS:
            col = field_to_col.get(field_name)
            if col is None:
                continue
            
            value = ws.cell(row=row, column=col).value
            
            # 处理数组字段（从JSON字符串反序列化）
            if field_name == 'dept_id_list':
                if isinstance(value, str) and value:
                    try:
                        value = json.loads(value)
                    except:
                        value = []
                elif value is None:
                    value = []
            
            user_data[field_name] = value
        
        return user_data
    
    def save_dingtalk_users(self, users: List[Dict[str, Any]], append: bool = False, source_dept_ids: List[int] = None):
        """保存从钉钉获取的用户数据到数据库（不再使用Excel文件）
        
        策略：
        1. 从数据库读取现有用户数据
        2. 对于新获取的用户，如果已存在（通过userid匹配），则更新信息（但保留管理员相关字段）
        3. 如果不存在，则新增
        4. 对于现有用户中不在新获取列表中的，将其状态设置为禁用（inactive），但如果是管理员则不清除
        5. 默认只激活硬件研发部及其子部门的用户，其他部门默认禁用
        
        Args:
            users: 用户列表，每个用户包含钉钉API返回的字段
            append: 是否追加模式（暂未使用）
            source_dept_ids: 当前拉取的部门ID列表（用于判断是否应该激活用户）
        
        Returns:
            bool: 是否保存成功
        """
        try:
            import json
            from server.db_adapter import get_connection_pool

            self._ensure_mysql_user_schema()
            pool = get_connection_pool()
            existing_user_data = {}  # {userid: {所有字段数据}}
            existing_users = {}  # {userid: user_id}
            
            with pool.get_cursor() as cursor:
                # 查询所有用户（钉钉用户的userid存储在dingtalk_data JSON字段中）
                cursor.execute('SELECT * FROM users')
                rows = cursor.fetchall()
                
                for row in rows:
                    if isinstance(row, dict):
                        user_dict = row
                        user_id = row.get('id')
                    else:
                        columns = [desc[0] for desc in cursor.description] if hasattr(cursor, 'description') else []
                        user_dict = dict(zip(columns, row)) if columns else {}
                        user_id = user_dict.get('id')
                    
                    userid = str(user_dict.get('dingtalk_userid') or '').strip()
                    dingtalk_data = user_dict.get('dingtalk_data')
                    if not userid and dingtalk_data:
                        try:
                            if isinstance(dingtalk_data, str):
                                dingtalk_data = json.loads(dingtalk_data)
                            if isinstance(dingtalk_data, dict):
                                userid = str(dingtalk_data.get('userid') or '').strip()
                        except Exception as e:
                            logger.debug(f"解析dingtalk_data失败: {e}")
                    if not userid:
                        username = user_dict.get('username', '')
                        if username and len(str(username)) > 10:
                            userid = str(username).strip()
                    
                    if userid:
                        existing_user_data[str(userid)] = user_dict
                        existing_users[str(userid)] = user_id
            
            # 获取新用户列表的userid集合
            new_userids = set()
            for user in users:
                userid = user.get('userid', '')
                if userid:
                    new_userids.add(str(userid))
            
            logger.info(f"准备处理 {len(users)} 个用户，现有用户 {len(existing_users)} 个")
            
            # 直接保存到MySQL数据库（不再使用Excel文件）
            try:
                self._save_dingtalk_users_to_mysql(users, existing_user_data, new_userids, source_dept_ids)
                logger.info(f"成功保存钉钉用户数据到MySQL")
                return True
            except Exception as e:
                logger.error(f"保存钉钉用户数据到MySQL失败: {e}", exc_info=True)
                return False
        except Exception as e:
            logger.error(f"保存钉钉用户数据失败: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    def _save_dingtalk_users_to_mysql(self, users: List[Dict[str, Any]], existing_user_data: Dict, new_userids: set, source_dept_ids: List[int] = None):
        """保存钉钉用户到MySQL数据库
        
        Args:
            users: 钉钉用户列表
            existing_user_data: 现有用户数据字典 {userid: {所有字段数据}}
            new_userids: 新用户ID集合
            source_dept_ids: 部门ID列表（用于判断是否激活用户）
        """
        try:
            from server.db_adapter import get_connection_pool
            from datetime import datetime
            
            pool = get_connection_pool()
            added_count = 0
            updated_count = 0
            
            with pool.get_cursor() as cursor:
                for user in users:
                    userid = str(user.get('userid', ''))
                    if not userid:
                        continue
                    
                    # 映射钉钉数据到数据库字段
                    job_number = user.get('job_number', '')
                    name = user.get('name', '')
                    title = user.get('title', '')
                    unionid = user.get('unionid', '')
                    
                    # 处理部门ID
                    dept_id_list = user.get('dept_id_list', [])
                    if isinstance(dept_id_list, str):
                        try:
                            dept_id_list = json.loads(dept_id_list)
                        except:
                            dept_id_list = []
                    
                    dept_id = None
                    if isinstance(dept_id_list, list) and len(dept_id_list) > 0:
                        if source_dept_ids:
                            for dept_id_item in dept_id_list:
                                try:
                                    dept_id_int = int(dept_id_item) if dept_id_item else None
                                    if dept_id_int is not None and dept_id_int in source_dept_ids:
                                        dept_id = dept_id_int
                                        break
                                except (ValueError, TypeError):
                                    continue
                        if dept_id is None:
                            try:
                                dept_id = int(dept_id_list[0]) if dept_id_list[0] else None
                            except (ValueError, TypeError):
                                dept_id = None
                    
                    # 确定用户名（优先使用job_number）
                    username = str(job_number).strip() if job_number and str(job_number).strip() else str(userid)
                    
                    # 确定状态（硬件研发部用户激活）
                    status = 'active'  # 默认激活
                    
                    # 确定部门代码（从部门ID映射到部门代码）
                    department = ''
                    if dept_id:
                        # 这里需要从部门管理器获取部门代码，暂时使用空字符串
                        # 后续可以通过部门ID查询部门代码
                        department = ''
                    
                    # 处理library_roles（从现有数据中获取，或使用空）
                    library_roles = ''
                    if userid in existing_user_data:
                        existing = existing_user_data[userid]
                        library_roles = existing.get('library_roles') or ''
                    
                    existing_user = None
                    cursor.execute(
                        'SELECT id FROM users WHERE dingtalk_userid = %s LIMIT 1',
                        (userid,),
                    )
                    existing_user = cursor.fetchone()
                    if not existing_user:
                        try:
                            cursor.execute(
                                """
                                SELECT id FROM users
                                WHERE JSON_UNQUOTE(JSON_EXTRACT(dingtalk_data, '$.userid')) = %s
                                  AND dingtalk_data IS NOT NULL
                                LIMIT 1
                                """,
                                (userid,),
                            )
                            existing_user = cursor.fetchone()
                        except Exception as e:
                            logger.debug(f"JSON 查找用户失败: {e}")
                    if not existing_user:
                        cursor.execute(
                            'SELECT id FROM users WHERE username = %s LIMIT 1',
                            (username,),
                        )
                        existing_user = cursor.fetchone()
                    
                    # 构建钉钉用户完整数据（JSON格式）
                    dingtalk_data = {
                        'userid': userid,
                        'unionid': unionid,
                        'job_number': job_number,
                        'login_id': user.get('login_id', ''),
                        'nickname': user.get('nickname', ''),
                        'title': title,
                        'dept_id': dept_id,
                        'dept_id_list': dept_id_list,
                        'parent_id': user.get('parent_id'),
                        'dept_order': user.get('dept_order'),
                        'active': user.get('active', True),
                        'admin': user.get('admin', False),
                        'boss': user.get('boss', False),
                        'leader': user.get('leader', False),
                        'disable_status': user.get('disable_status', False),
                        'exclusive_account': user.get('exclusive_account', False),
                        'exclusive_account_type': user.get('exclusive_account_type', ''),
                        'exclusive_account_corp_id': user.get('exclusive_account_corp_id', ''),
                        'exclusive_account_corp_name': user.get('exclusive_account_corp_name', ''),
                        'avatar': user.get('avatar', ''),
                        'hide_mobile': user.get('hide_mobile', False),
                        'hired_date': user.get('hired_date'),
                        'gender': user.get('gender', '')
                    }
                    dingtalk_data_json = json.dumps(dingtalk_data, ensure_ascii=False)
                    
                    if existing_user:
                        user_id = existing_user['id'] if isinstance(existing_user, dict) else existing_user[0]
                        cursor.execute(
                            '''
                            UPDATE users
                            SET name = %s, department = %s, job_position = %s,
                                library_roles = %s, status = %s,
                                dingtalk_data = %s, dingtalk_userid = %s, dingtalk_unionid = %s,
                                job_number = %s, user_source = 'dingtalk', updated_time = NOW()
                            WHERE id = %s
                            ''',
                            (
                                name, department, title, library_roles, status,
                                dingtalk_data_json, userid, unionid or None,
                                job_number or None, user_id,
                            ),
                        )
                        updated_count += 1
                    else:
                        from server.auth.password_service import PasswordService
                        default_password = PasswordService.default_password_hash()
                        default_roles = 'user'
                        cursor.execute(
                            '''
                            INSERT INTO users (
                                username, password, name, department, job_position,
                                roles, library_roles, status, dingtalk_data, dingtalk_userid,
                                dingtalk_unionid, job_number, user_source, created_time, updated_time
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'dingtalk', NOW(), NOW())
                            ''',
                            (
                                username, default_password, name, department, title,
                                default_roles, library_roles, status, dingtalk_data_json, userid,
                                unionid or None, job_number or None,
                            ),
                        )
                        added_count += 1
                
                # 处理未搜索到的用户（禁用非硬件研发部用户）
                if source_dept_ids:
                    # 获取所有用户
                    cursor.execute('SELECT id, username, department FROM users')
                    all_db_users = cursor.fetchall()
                    
                    for db_user in all_db_users:
                        user_id = db_user['id'] if isinstance(db_user, dict) else db_user[0]
                        username = db_user['username'] if isinstance(db_user, dict) else db_user[1]
                        dept = db_user['department'] if isinstance(db_user, dict) else db_user[2]
                        
                        # 如果用户不在新用户列表中，且不是硬件研发部用户，则禁用
                        # 这里简化处理：如果用户的username不在new_userids中（通过userid匹配），则可能需要禁用
                        # 但由于MySQL中没有userid字段，这里暂时跳过
                        pass
            
            logger.info(f"MySQL保存完成: 新增 {added_count} 个，更新 {updated_count} 个")
            
        except Exception as e:
            logger.error(f"保存钉钉用户到MySQL失败: {e}", exc_info=True)
            raise
