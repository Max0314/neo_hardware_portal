import http.server
import socketserver
import socket
import json
import urllib.parse
import os
import uuid
import sys
import base64
import threading
import traceback
import gzip
import hashlib
import time
import mimetypes
from io import BytesIO
from typing import List, Dict, Any, Optional
try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

# 添加项目根目录到Python路径，确保可以导入server模块
# 动态获取项目根目录，不依赖文件夹名称
_current_file_dir = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(_current_file_dir)
sys.path.insert(0, BASE_DIR)

from server.config import (
    PORT, HOST, STATIC_DIR, TEMPLATE_DIR,
    ENABLE_GZIP, ENABLE_CACHE, CACHE_MAX_AGE, MAX_WORKERS, REQUEST_TIMEOUT,
    JOB_POSITION_OPTIONS, DINGTALK_CONFIG, check_dingtalk_config,
    PUBLIC_BASE_URL, get_dingtalk_agent_id_numeric,
    DINGTALK_WORKSPACE_ID, DINGTALK_DOC_PARENT_NODE_ID, DINGTALK_DOC_OPERATOR_UNIONID,
    MAX_CONNECTIONS, CONNECTION_QUEUE_SIZE, CACHE_SIZE,
    GZIP_COMPRESSION_LEVEL, CHUNK_SIZE, SEND_BUFFER_SIZE, RECV_BUFFER_SIZE,
    PRELOAD_USERS, PRELOAD_ANNOUNCEMENTS, PRELOAD_DEPARTMENTS, PRELOAD_TODOS,
    MEMORY_CACHE_TTL, FILE_READ_BUFFER_SIZE,
    ANNOUNCEMENT_PAGE_SIZE, ANNOUNCEMENT_MAX_PAGE_SIZE,
    MAX_ATTACHMENT_SIZE, MAX_ATTACHMENTS_PER_ANNOUNCEMENT,
    ENABLE_RECYCLE_BIN, HOT_CACHE_TTL, NORMAL_CACHE_TTL,
    ANNOUNCEMENT_APPROVERS
)
from server.user_manager import (
    UserManager,
    STATUS_ACTIVE,
    STATUS_PENDING,
    STATUS_REJECTED,
)
from server.announcement_manager import AnnouncementManager
from server.announcement_config import ANNOUNCEMENT_BOARDS, ANNOUNCEMENT_MANAGE_ROLES
from server.quick_link_manager import QuickLinkManager
from server.department_manager import DepartmentManager
from server.todo_manager import TodoManager
from server.logger import logger
import logging
from server.security import InputValidator
from server.data_preloader import get_data_preloader
from server.session_manager import get_session_manager
from server import csrf as csrf_mod
from server.security_http import (
    apply_security_headers,
    apply_cors,
    safe_error_payload,
    send_safe_http_error,
    log_server_error,
)

# 钉钉待办SDK导入
try:
    from alibabacloud_dingtalk.todo_1_0.client import Client as TodoClient
    from alibabacloud_dingtalk.todo_1_0 import models as todo_models
    from alibabacloud_tea_openapi import models as open_api_models
    from alibabacloud_tea_util import models as util_models
    HAS_TODO_SDK = True
except ImportError:
    HAS_TODO_SDK = False
    logger.warning("钉钉待办SDK未安装，待办功能将不可用")

# 简单的会话管理（线程安全）
sessions = {}
sessions_lock = threading.Lock()  # 保护 sessions 字典的锁

try:
    from server.auth.session_sync import start_session_invalidation_listener
    start_session_invalidation_listener(sessions, sessions_lock)
except Exception as _sess_listener_err:
    logger.debug(f'会话失效监听未启动: {_sess_listener_err}')

# 高性能LRU缓存机制（线程安全，支持大小限制）
class LRUCache:
    """LRU缓存，用于提升性能，支持高并发，支持跨进程缓存同步"""
    def __init__(self, default_ttl=60, max_size=CACHE_SIZE):
        self.cache = {}  # {key: (value, timestamp, access_time)}
        self.default_ttl = default_ttl  # 默认缓存时间（秒）
        self.max_size = max_size  # 最大缓存条目数
        self.lock = threading.RLock()  # 使用可重入锁，提高性能
        self.access_order = []  # 访问顺序列表（用于LRU淘汰）
        self.cache_sync_manager = None  # 跨进程缓存同步管理器（延迟初始化）
        self.last_sync_check = {}  # {cache_type: timestamp} 记录最后检查时间，避免频繁检查
        self.sync_check_interval = 0.2  # 跨进程同步检查间隔（秒）
    
    def _get_cache_sync_manager(self):
        """获取缓存同步管理器（延迟初始化）"""
        if self.cache_sync_manager is None:
            try:
                from server.cache_sync import get_cache_sync_manager
                self.cache_sync_manager = get_cache_sync_manager()
            except Exception as e:
                # 如果导入失败，记录警告但不影响功能
                logger.warning(f"无法初始化缓存同步管理器: {e}")
        return self.cache_sync_manager
    
    def _get_cache_type_from_key(self, key):
        """根据缓存key判断缓存类型"""
        if not key:
            return None
        key_lower = key.lower()
        if 'announcement' in key_lower:
            return 'announcements'
        elif 'user' in key_lower:
            return 'users'
        elif 'department' in key_lower:
            return 'departments'
        elif 'todo' in key_lower:
            return 'todos'
        elif 'material' in key_lower:
            return 'materials'
        return None
    
    def _check_cross_process_invalidation(self, key):
        """检查跨进程缓存失效标记
        
        Returns:
            bool: True表示缓存需要失效，False表示缓存有效
        """
        try:
            cache_type = self._get_cache_type_from_key(key)
            if not cache_type:
                return False
            
            sync_manager = self._get_cache_sync_manager()
            if not sync_manager:
                return False
            
            # 避免过于频繁的检查（性能优化）
            current_time = time.time()
            last_check = self.last_sync_check.get(cache_type, 0)
            if current_time - last_check < self.sync_check_interval:
                return False
            
            # 检查缓存失效标记
            marker_time = sync_manager.check_cache_invalidation(cache_type)
            self.last_sync_check[cache_type] = current_time
            
            if marker_time:
                # 检测到失效标记，更新本地版本号
                sync_manager.update_local_version(cache_type, marker_time)
                logger.debug(f"检测到跨进程缓存失效: key={key}, cache_type={cache_type}")
                return True
            
            return False
        except Exception as e:
            # 检查失败不影响正常功能，只记录警告
            logger.debug(f"检查跨进程缓存失效标记失败: {e}")
            return False
    
    def get(self, key):
        """获取缓存值（O(1)时间复杂度），支持跨进程缓存同步检查"""
        with self.lock:
            # 先检查跨进程缓存失效标记
            if self._check_cross_process_invalidation(key):
                # 缓存已失效，删除并返回None
                self._remove_key(key)
                return None
            
            if key in self.cache:
                value, timestamp, _ = self.cache[key]
                # 检查是否过期
                current_time = time.time()
                if current_time - timestamp < self.default_ttl:
                    # 更新访问时间和顺序
                    self.cache[key] = (value, timestamp, current_time)
                    # 更新访问顺序（移到末尾）
                    if key in self.access_order:
                        self.access_order.remove(key)
                    self.access_order.append(key)
                    return value
                else:
                    # 过期，删除
                    self._remove_key(key)
            return None
    
    def set(self, key, value, ttl=None):
        """设置缓存值（O(1)时间复杂度）"""
        with self.lock:
            current_time = time.time()
            cache_ttl = ttl if ttl is not None else self.default_ttl
            
            # 如果缓存已满，删除最久未使用的项
            if len(self.cache) >= self.max_size and key not in self.cache:
                self._evict_lru()
            
            # 设置缓存
            self.cache[key] = (value, current_time, current_time)
            # 更新访问顺序
            if key in self.access_order:
                self.access_order.remove(key)
            self.access_order.append(key)
    
    def _remove_key(self, key):
        """删除缓存项"""
        if key in self.cache:
            del self.cache[key]
        if key in self.access_order:
            self.access_order.remove(key)
    
    def _evict_lru(self):
        """淘汰最久未使用的缓存项"""
        if self.access_order:
            lru_key = self.access_order.pop(0)
            # 确保键存在再删除，避免内存泄漏
            if lru_key in self.cache:
                del self.cache[lru_key]
            # 如果键不在cache中，access_order已经被清理，无需额外操作
    
    def invalidate(self, key):
        """使缓存失效"""
        with self.lock:
            self._remove_key(key)
    
    def clear(self):
        """清空所有缓存"""
        with self.lock:
            self.cache.clear()
            self.access_order.clear()
    
    def size(self):
        """获取当前缓存大小"""
        with self.lock:
            return len(self.cache)

# 全局缓存实例（高性能LRU缓存）
# 阶段1优化：使用普通缓存TTL（600秒），热点数据在API层面使用HOT_CACHE_TTL
api_cache = LRUCache(default_ttl=NORMAL_CACHE_TTL, max_size=CACHE_SIZE)

class HardwareRDBHandler(http.server.SimpleHTTPRequestHandler):
    # 类级别的管理器实例（共享，减少初始化开销，提高并发性能）
    _user_manager = None
    _announcement_mgr = None
    _quick_link_mgr = None
    _department_mgr = None
    _todo_mgr = None
    _lock = threading.Lock()
    
    def __init__(self, *args, **kwargs):
        # 延迟初始化管理器（线程安全，单例模式）
        if HardwareRDBHandler._user_manager is None:
            with HardwareRDBHandler._lock:
                if HardwareRDBHandler._user_manager is None:
                    HardwareRDBHandler._user_manager = UserManager()
                    HardwareRDBHandler._announcement_mgr = AnnouncementManager(base_dir=BASE_DIR)
                    HardwareRDBHandler._quick_link_mgr = QuickLinkManager()
                    HardwareRDBHandler._department_mgr = DepartmentManager()
                    HardwareRDBHandler._todo_mgr = TodoManager()
        
        self.user_manager = HardwareRDBHandler._user_manager
        self.announcement_mgr = HardwareRDBHandler._announcement_mgr
        self.quick_link_mgr = HardwareRDBHandler._quick_link_mgr
        self.department_mgr = HardwareRDBHandler._department_mgr
        self.todo_mgr = HardwareRDBHandler._todo_mgr
        # 设置静态文件目录为项目根目录下的static
        self.directory = STATIC_DIR
        # Python 3.6 不支持 directory 参数，需要移除
        # 我们将通过重写 translate_path 方法来处理目录
        import sys
        if sys.version_info >= (3, 7):
            # Python 3.7+ 支持 directory 参数
            super().__init__(*args, directory=self.directory, **kwargs)
        else:
            # Python 3.6 不支持 directory 参数，使用默认初始化
            super().__init__(*args, **kwargs)
    
    def _has_role(self, user, role):
        return self.user_manager.user_has_role(user, role) if user else False
    
    def _is_super_admin(self, user):
        return self.user_manager.is_super_admin(user) if user else False

    def _can_manage_quick_links(self, user) -> bool:
        """快捷链接增删：最高管理员、管理员、管理组"""
        if not user:
            return False
        if self._is_super_admin(user):
            return True
        return self._has_role(user, 'admin') or self._has_role(user, 'management')

    def _can_manage_boards(self, user) -> bool:
        """公告栏管理：最高管理员、管理员、管理组。"""
        if not user:
            return False
        if self._is_super_admin(user):
            return True
        return self._has_role(user, 'admin') or self._has_role(user, 'management')
    
    def _can_approve_announcement(self, user) -> bool:
        """检查用户是否具备公告审批权限（与 ANNOUNCEMENT_APPROVERS 配置一致）"""
        if not user:
            return False
        if self._is_super_admin(user):
            return True
        
        config = ANNOUNCEMENT_APPROVERS or {}
        roles = self.user_manager._parse_roles(user.get('roles', []))
        approver_roles = set(config.get('roles', ['management', 'admin']))
        approver_roles.add('super_admin')
        if any(role in approver_roles for role in roles):
            return True
        
        user_title = (user.get('job_position') or user.get('title') or '').strip()
        for title_pattern in config.get('titles', []):
            if title_pattern and title_pattern in user_title:
                return True
        
        user_identifiers = {
            str(user.get('userid', '')).strip(),
            str(user.get('username', '')).strip(),
            str(user.get('id', '')).strip(),
        }
        user_identifiers.discard('')
        for uid in config.get('userids', []):
            if uid and str(uid).strip() in user_identifiers:
                return True
        return False
    
    def _invalidate_announcement_cache(self, max_retries=3, retry_delay=0.1):
        """清除公告相关的缓存（带重试机制），支持跨进程缓存同步
        
        Args:
            max_retries: 最大重试次数
            retry_delay: 重试延迟（秒）
        """
        last_error = None
        for attempt in range(max_retries):
            try:
                # 使用CacheSyncManager通知所有worker进程缓存失效
                try:
                    from server.cache_sync import invalidate_cache
                    invalidate_cache('announcements')
                    logger.debug("已通知所有进程公告缓存失效")
                except Exception as e:
                    logger.warning(f"通知跨进程缓存失效失败: {e}")
                    # 继续执行本地缓存清除
                
                # 清除所有以 "announcement_" 开头的缓存
                # 注意：使用RLock避免死锁，invalidate内部也会获取锁
                with api_cache.lock:
                    keys_to_remove = [key for key in list(api_cache.cache.keys()) if key.startswith('announcement_')]
                    # 直接调用_remove_key避免重复获取锁（因为已经在锁内）
                    for key in keys_to_remove:
                        try:
                            api_cache._remove_key(key)
                        except Exception as e:
                            logger.warning(f"清除API缓存项失败: key={key}, error={e}")
                            last_error = e
                            continue
                
                # 清除内存预加载器中的公告缓存，并立即重新加载
                preloader = get_data_preloader()
                if preloader:
                    try:
                        preloader.invalidate_cache('announcements')
                        # 立即重新加载公告数据，确保缓存更新
                        preloader.reload_cache('announcements')
                        logger.debug("公告缓存已重新加载")
                    except Exception as e:
                        logger.warning(f"清除或重新加载公告缓存失败: {e}")
                        last_error = e
                        if attempt < max_retries - 1:
                            time.sleep(retry_delay)
                            continue
                
                # 如果执行到这里，说明清除成功
                if attempt > 0:
                    logger.info(f"公告缓存清除成功（第{attempt + 1}次尝试）")
                return True
                
            except Exception as e:
                last_error = e
                logger.warning(f"清除公告缓存失败（第{attempt + 1}次尝试）: {e}")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
        
        # 所有重试都失败
        logger.error(f"清除公告缓存失败（已重试{max_retries}次）: {last_error}")
        return False
    
    def _invalidate_user_cache(self, max_retries=3, retry_delay=0.1):
        """清除用户相关的缓存（带重试机制），支持跨进程缓存同步
        
        Args:
            max_retries: 最大重试次数
            retry_delay: 重试延迟（秒）
        """
        last_error = None
        for attempt in range(max_retries):
            try:
                # 使用CacheSyncManager通知所有worker进程缓存失效
                try:
                    from server.cache_sync import invalidate_cache
                    invalidate_cache('users')
                    logger.debug("已通知所有进程用户缓存失效")
                except Exception as e:
                    logger.warning(f"通知跨进程缓存失效失败: {e}")
                    # 继续执行本地缓存清除
                
                # 清除API缓存中的用户相关缓存
                # 注意：使用RLock避免死锁，invalidate内部也会获取锁
                with api_cache.lock:
                    keys_to_remove = [key for key in list(api_cache.cache.keys()) if 'user' in key.lower()]
                    # 直接调用_remove_key避免重复获取锁（因为已经在锁内）
                    for key in keys_to_remove:
                        try:
                            api_cache._remove_key(key)
                        except Exception as e:
                            logger.warning(f"清除API缓存项失败: key={key}, error={e}")
                            last_error = e
                            continue
                
                # 清除内存预加载器中的用户缓存
                preloader = get_data_preloader()
                if preloader:
                    try:
                        preloader.invalidate_cache('users')
                    except Exception as e:
                        logger.warning(f"清除用户缓存失败: {e}")
                        last_error = e
                        if attempt < max_retries - 1:
                            time.sleep(retry_delay)
                            continue
                
                # 如果执行到这里，说明清除成功
                if attempt > 0:
                    logger.info(f"用户缓存清除成功（第{attempt + 1}次尝试）")
                return True
                
            except Exception as e:
                last_error = e
                logger.warning(f"清除用户缓存失败（第{attempt + 1}次尝试）: {e}")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
        
        # 所有重试都失败
        logger.error(f"清除用户缓存失败（已重试{max_retries}次）: {last_error}")
        return False
    
    def _invalidate_todo_cache(self, announcement_id: str = None, max_retries=3, retry_delay=0.1):
        """清除待办相关的缓存（带重试机制），支持跨进程缓存同步
        
        Args:
            announcement_id: 公告ID（可选，如果提供则只清除该公告的待办缓存）
            max_retries: 最大重试次数
            retry_delay: 重试延迟（秒）
        """
        last_error = None
        for attempt in range(max_retries):
            try:
                # 使用CacheSyncManager通知所有worker进程缓存失效
                try:
                    from server.cache_sync import invalidate_cache
                    invalidate_cache('todos')
                    logger.debug("已通知所有进程待办缓存失效")
                except Exception as e:
                    logger.warning(f"通知跨进程缓存失效失败: {e}")
                    # 继续执行本地缓存清除

                with api_cache.lock:
                    keys_to_remove = [
                        key for key in list(api_cache.cache.keys())
                        if 'todo' in key.lower()
                    ]
                    for key in keys_to_remove:
                        try:
                            api_cache._remove_key(key)
                        except Exception as e:
                            logger.warning(f"清除待办 API 缓存项失败: key={key}, error={e}")
                            last_error = e
                
                # 清除内存预加载器中的待办缓存
                preloader = get_data_preloader()
                if preloader:
                    try:
                        if announcement_id:
                            preloader.invalidate_cache('todos', key=announcement_id)
                        else:
                            preloader.invalidate_cache('todos')
                    except Exception as e:
                        logger.warning(f"清除待办缓存失败: {e}")
                        last_error = e
                        if attempt < max_retries - 1:
                            time.sleep(retry_delay)
                            continue
                
                # 如果执行到这里，说明清除成功
                if attempt > 0:
                    logger.info(f"待办缓存清除成功（第{attempt + 1}次尝试）")
                return True
                
            except Exception as e:
                last_error = e
                logger.warning(f"清除待办缓存失败（第{attempt + 1}次尝试）: {e}")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
        
        # 所有重试都失败
        logger.error(f"清除待办缓存失败（已重试{max_retries}次）: {last_error}")
        return False
    
    def _allow_neo_internal_request(self) -> bool:
        """NEO 后端容器内拉取用户列表（免 Cookie，须携带共享密钥）。"""
        expected = (os.getenv('NEO_INTERNAL_SECRET') or '').strip()
        if not expected:
            return False
        return self.headers.get('X-Neo-Internal-Secret') == expected

    def _auth_check_wants_fast_path(self) -> bool:
        """NEO 内部校验会话：跳过 get_user_by_id 与会话写库补全。"""
        ua = (self.headers.get('User-Agent') or '').lower()
        if 'aiohttp' in ua:
            return True
        return self._allow_neo_internal_request()

    def _auth_check_should_refresh_user(self, user: Dict[str, Any]) -> bool:
        """浏览器路径始终从 DB 刷新 roles/status；NEO fast-path 跳过。"""
        if self._auth_check_wants_fast_path():
            return False
        return True

    def _user_status_allows_login(self, user: Optional[Dict[str, Any]]) -> bool:
        """会话存在但账号已禁用/待审时，auth/check 应视为未登录。"""
        if not user:
            return False
        st = str(user.get('status') or STATUS_ACTIVE).strip().lower()
        return st in ('active', STATUS_ACTIVE)

    def _neo_auth_user_payload(self, user: Dict[str, Any]) -> Dict[str, Any]:
        """NEO 内网校验用轻量用户结构（不查库补全 userid）。"""
        if not user:
            return {}
        roles = user.get('roles', [])
        if isinstance(roles, str):
            roles = self.user_manager._parse_roles(roles)
        elif not isinstance(roles, list):
            roles = []
        from server.neo_user_key import resolve_neo_user_key

        canonical_key = resolve_neo_user_key(user, None)
        uid = user.get('id')
        try:
            uid = int(uid) if uid is not None and str(uid).strip() != '' else uid
        except (TypeError, ValueError):
            pass
        return {
            'id': uid,
            'username': user.get('username'),
            'name': user.get('name'),
            'roles': roles,
            'role': user.get('role'),
            'department': user.get('department'),
            'job_position': user.get('job_position', ''),
            'status': user.get('status'),
            'userKey': canonical_key,
            'userid': canonical_key or user.get('userid') or user.get('username', ''),
        }

    def _handle_internal_neo_verify_session(self):
        """GET /api/internal/neo/verify-session — 仅 NEO 容器 + 共享密钥，极简校验。"""
        if not self._allow_neo_internal_request():
            self.send_json_response({'authenticated': False, 'error': '禁止访问'}, status=403)
            return
        user = self.get_current_user(skip_session_enrich=True)
        if user:
            from server.auth.capabilities import user_capabilities
            self.send_json_response({
                'authenticated': True,
                'user': self._neo_auth_user_payload(user),
                'capabilities': user_capabilities(user),
            })
        else:
            self.send_json_response({'authenticated': False})

    def _handle_health_check(self):
        """GET /api/health — 轻量存活探针（可选 ?db=1 校验 MySQL）。"""
        from urllib.parse import parse_qs, urlparse

        from server.health import health_payload

        parsed = urlparse(self.path)
        params = parse_qs(parsed.query, keep_blank_values=True)
        db_flag = (params.get('db') or [''])[0]
        check_db = str(db_flag).strip().lower() in ('1', 'true', 'yes')
        payload, status_code = health_payload(check_db=check_db)
        self.send_json_response(payload, status=status_code)

    def _handle_startup_status(self):
        """GET /api/startup/status — 启动进度（供登录页与运维查看）。"""
        from server.startup_gate import get_startup_status

        self.send_json_response(get_startup_status())

    def _handle_auth_config(self):
        """GET /api/auth/config — 钉钉鉴权配置（不返回密钥）。"""
        from server.auth.dingtalk_auth import build_auth_config_response

        status, body = build_auth_config_response(self)
        self.send_json_response(body, status=status)

    def _get_query_params(self, parsed_path):
        query = parsed_path.query or getattr(self, 'query_string', '') or ''
        if query:
            return urllib.parse.parse_qs(query, keep_blank_values=True)
        if hasattr(self, 'query_params') and isinstance(self.query_params, dict):
            return self.query_params
        return {}

    def _handle_dingtalk_web_start(self, parsed_path):
        """GET /api/auth/dingtalk/web/start — 生成钉钉 OAuth URL。"""
        from server.auth.dingtalk_auth import build_web_login_start_response

        params = self._get_query_params(parsed_path)
        return_url = (params.get('return_url') or [''])[0]
        status, body = build_web_login_start_response(self, return_url)
        self.send_json_response(body, status=status)

    def _send_html_response(self, html_text: str, status: int = 200, set_cookies: Optional[List[str]] = None):
        body = (html_text or '').encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
        self.send_header('Pragma', 'no-cache')
        self.send_header('Expires', '0')
        if set_cookies:
            for cookie in set_cookies:
                self.send_header('Set-Cookie', cookie)
        self.send_header('Content-Length', str(len(body)))
        apply_security_headers(self)
        self.end_headers()
        self.wfile.write(body)

    def _handle_dingtalk_web_callback(self, parsed_path):
        """GET /api/auth/dingtalk/callback — 钉钉网页登录回调。"""
        from server.auth.dingtalk_auth import handle_web_login_callback

        params = self._get_query_params(parsed_path)
        status, html_text, cookies = handle_web_login_callback(self, params)
        self._send_html_response(html_text, status=status, set_cookies=cookies)

    def _handle_dingtalk_inapp_login(self):
        """POST /api/auth/dingtalk/inapp-login — 钉钉内免登。"""
        from server.startup_gate import login_allowed
        from server.auth.dingtalk_auth import login_with_inapp_code

        ok_login, maint = login_allowed()
        if not ok_login and maint:
            self.send_json_response(
                {
                    'success': False,
                    'code': 503,
                    'error': maint.get('message') or '系统正在启动，请稍候',
                    'startup': maint,
                },
                status=503,
            )
            return

        try:
            content_length = int(self.headers.get('Content-Length', 0) or 0)
            if content_length == 0:
                self.send_json_response({'success': False, 'code': 400, 'error': '请求体不能为空'}, status=400)
                return
            raw = self.rfile.read(content_length).decode('utf-8')
            data = json.loads(raw)
        except json.JSONDecodeError:
            self.send_json_response({'success': False, 'code': 400, 'error': '无效的JSON数据'}, status=400)
            return

        auth_code = (
            data.get('code')
            or data.get('authCode')
            or data.get('auth_code')
            or ''
        ).strip()
        status, body, cookies = login_with_inapp_code(self, auth_code)
        self.send_json_response(body, status=status, set_cookies=cookies if cookies else None)

    def _handle_auth_check(self):
        """GET /api/auth/check — 兼容别名，同 session。"""
        self._handle_auth_session(legacy_check=True)

    def _handle_auth_session(self, legacy_check: bool = False):
        """GET /api/auth/session（及 check 别名）"""
        from server.startup_gate import login_allowed
        from server.auth.capabilities import user_capabilities

        ok_login, maint = login_allowed()
        if not ok_login and maint:
            self.send_json_response(
                {
                    'ok': False,
                    'authenticated': False,
                    'startup': maint,
                    'error': maint.get('message') or '系统正在启动，请稍候',
                },
                status=503,
            )
            return

        parsed = urllib.parse.urlparse(self.path)
        params = urllib.parse.parse_qs(parsed.query)
        lite = (params.get('lite') or [''])[0].strip().lower() in ('1', 'true', 'yes')

        fast = self._auth_check_wants_fast_path() or lite
        user = self.get_current_user(skip_session_enrich=fast)
        if user:
            uid = user.get('id')
            if uid is not None and self._auth_check_should_refresh_user(user):
                try:
                    fresh = self.user_manager.get_user_by_id(int(uid))
                    if fresh:
                        user = fresh
                except Exception as e:
                    logger.debug(f"auth/session 刷新用户失败，使用会话数据: {e}")
            if not self._user_status_allows_login(user):
                logger.info(
                    "auth/session 拒绝非 active 账号: username=%s, status=%s",
                    user.get('username'),
                    user.get('status'),
                )
                from server.auth.login_service import clear_auth_cookie_headers

                self.send_json_response(
                    {'ok': False, 'authenticated': False, 'error': '账号不可用'},
                    set_cookies=clear_auth_cookie_headers(secure=self._is_https_request()),
                )
                return
            caps = user_capabilities(user)
            if fast:
                user_payload = self._neo_auth_user_payload(user)
                self.send_json_response({
                    'ok': True,
                    'authenticated': True,
                    'user': user_payload,
                    'capabilities': caps,
                    'can_approve': False,
                })
                return
            user_payload = self._public_user_payload(user)
            cookie_header = self.headers.get('Cookie', '')
            cookies = {}
            for part in cookie_header.split(';'):
                part = part.strip()
                if '=' in part:
                    k, v = part.split('=', 1)
                    cookies[k.strip()] = v.strip()
            csrf_tok = cookies.get(csrf_mod.CSRF_COOKIE_NAME) or csrf_mod.new_token()
            set_cookies = [
                csrf_mod.cookie_header_value(csrf_tok, secure=self._is_https_request()),
            ]
            self.send_json_response(
                {
                    'ok': True,
                    'authenticated': True,
                    'user': user_payload,
                    'capabilities': caps,
                    'can_approve': self._can_approve_announcement(user),
                    'csrf_token': csrf_tok,
                },
                set_cookies=set_cookies,
            )
        else:
            from server.auth.login_service import clear_auth_cookie_headers

            self.send_json_response(
                {'ok': False, 'authenticated': False},
                set_cookies=clear_auth_cookie_headers(secure=self._is_https_request()),
            )

    def _list_neo_active_users(self) -> List[Dict[str, str]]:
        preloader = get_data_preloader()
        if preloader and PRELOAD_USERS:
            try:
                all_users = preloader.get_users(status='active')
                if not all_users:
                    all_users = self.user_manager.get_all_users(statuses=[STATUS_ACTIVE])
            except Exception:
                all_users = self.user_manager.get_all_users(statuses=[STATUS_ACTIVE])
        else:
            all_users = self.user_manager.get_all_users(statuses=[STATUS_ACTIVE])
        users_out: List[Dict[str, str]] = []
        for u in all_users:
            if str(u.get('status', '')).lower() not in ('active', STATUS_ACTIVE):
                continue
            from server.neo_user_key import resolve_neo_user_key

            user_key = resolve_neo_user_key(u, None)
            if not user_key:
                continue
            display_name = (u.get('name') or u.get('nickname') or u.get('username') or user_key)
            if display_name is not None:
                display_name = str(display_name).strip()
            else:
                display_name = user_key
            username = str(u.get('username') or '').strip()
            users_out.append({
                'userKey': user_key,
                'name': display_name or user_key,
                'username': username,
            })
        return users_out

    def _public_user_payload(self, user):
        if not user:
            return {}
        
        # 确保 roles 是数组格式
        roles = user.get('roles', [])
        if isinstance(roles, str):
            # 如果是字符串，尝试解析（可能是 JSON 字符串或逗号分隔的字符串）
            try:
                import json
                roles = json.loads(roles)
            except:
                # 如果不是 JSON，尝试按逗号分割
                roles = [r.strip() for r in roles.split(',') if r.strip()]
        elif not isinstance(roles, list):
            roles = []
        
        # 确保 library_roles 是数组格式
        library_roles = user.get('library_roles', [])
        if isinstance(library_roles, str):
            try:
                import json
                library_roles = json.loads(library_roles)
            except:
                library_roles = [r.strip() for r in library_roles.split(',') if r.strip()]
        elif not isinstance(library_roles, list):
            library_roles = []
        
        # 处理密码字段：不返回实际密码（安全考虑）
        password = user.get('password', '') or ''
        if password and not isinstance(password, str):
            if isinstance(password, (bytes, bytearray)):
                try:
                    password = password.decode('utf-8', 'ignore')
                except Exception:
                    password = ''
            else:
                password = str(password)
        if password:
            # 检查是否是哈希格式
            if password.startswith('$2b$') or ':' in password:
                # 是哈希值，不返回
                password = ''
            else:
                # 是明文，也不返回（安全考虑）
                password = ''
        
        _uid = user.get('id')
        try:
            _uid = int(_uid) if _uid is not None and str(_uid).strip() != '' else _uid
        except (TypeError, ValueError):
            pass
        ct = user.get('create_time') or user.get('created_time')
        if ct is not None and hasattr(ct, 'isoformat'):
            try:
                ct = ct.isoformat(sep=' ', timespec='seconds')
            except TypeError:
                ct = ct.isoformat()
        payload = {
            'id': _uid,
            'username': user.get('username'),
            'name': user.get('name'),
            'roles': roles,
            'role': user.get('role'),
            'department': user.get('department'),
            'job_position': user.get('job_position', ''),  # 添加岗位字段
            'library_roles': library_roles,
            'status': user.get('status'),
            'create_time': ct
        }
        
        # 始终添加所有可能的字段（无论是否为钉钉用户）
        # 这样前端可以显示所有可用信息，即使某些字段为空
        payload['job_number'] = user.get('job_number', '')
        from server.neo_user_key import resolve_neo_user_key

        canonical_key = resolve_neo_user_key(user, None)
        payload['userKey'] = canonical_key
        payload['userid'] = canonical_key or user.get('userid', user.get('username', ''))
        # 对于钉钉用户，title优先使用job_position，如果没有则使用title字段
        payload['title'] = user.get('job_position') or user.get('title', '')
        payload['source'] = user.get('source', '')
        
        # 添加所有钉钉用户字段（用于调试和完整显示）
        # 即使字段为空也返回，让前端显示 '-' 而不是完全不显示
        payload['unionid'] = user.get('unionid', '')
        payload['login_id'] = user.get('login_id', '')
        payload['nickname'] = user.get('nickname', '')
        payload['avatar'] = user.get('avatar', '')
        payload['dept_id'] = user.get('dept_id', '')
        
        # 处理dept_id_list（可能是字符串或列表）
        dept_id_list = user.get('dept_id_list', '')
        if isinstance(dept_id_list, str):
            try:
                import json
                dept_id_list = json.loads(dept_id_list) if dept_id_list else []
            except:
                dept_id_list = []
        payload['dept_id_list'] = dept_id_list if isinstance(dept_id_list, list) else []
        
        payload['dept_order'] = user.get('dept_order', '')
        payload['active'] = user.get('active', False)
        payload['admin'] = user.get('admin', False)
        payload['boss'] = user.get('boss', False)
        payload['leader'] = user.get('leader', False)
        payload['disable_status'] = user.get('disable_status', False)
        payload['exclusive_account'] = user.get('exclusive_account', False)
        payload['exclusive_account_type'] = user.get('exclusive_account_type', '')
        payload['exclusive_account_corp_id'] = user.get('exclusive_account_corp_id', '')
        payload['exclusive_account_corp_name'] = user.get('exclusive_account_corp_name', '')
        payload['hide_mobile'] = user.get('hide_mobile', False)
        payload['hired_date'] = user.get('hired_date', '')
        payload['gender'] = user.get('gender', '')
        payload['parent_id'] = user.get('parent_id', '')
        
        return payload
    
    def _should_compress(self, content_type):
        """判断是否应该压缩响应"""
        if not ENABLE_GZIP:
            return False
        
        compressible_types = [
            'text/html', 'text/css', 'text/javascript', 'application/javascript',
            'application/json', 'text/xml', 'application/xml', 'text/plain'
        ]
        return any(content_type.startswith(ct) for ct in compressible_types)
    
    def _compress_response(self, data):
        """压缩响应数据（优化压缩级别，充分利用带宽）"""
        if not data:
            return data
        
        buf = BytesIO()
        # 使用配置的压缩级别，平衡压缩率和性能
        with gzip.GzipFile(fileobj=buf, mode='wb', compresslevel=GZIP_COMPRESSION_LEVEL) as gz:
            # 使用更大的缓冲区写入，提高性能
            if isinstance(data, bytes):
                gz.write(data)
            else:
                gz.write(data.encode('utf-8'))
        return buf.getvalue()
    
    def _get_etag(self, data):
        """生成ETag"""
        return hashlib.md5(data).hexdigest()
    
    def _check_cache(self, etag):
        """检查客户端缓存"""
        if not ENABLE_CACHE:
            return False
        
        if_none_match = self.headers.get('If-None-Match')
        return if_none_match == etag

    def _is_https_request(self) -> bool:
        return (
            self.headers.get('X-Forwarded-Proto') == 'https'
            or self.headers.get('X-Forwarded-Ssl') == 'on'
            or self.headers.get('X-Forwarded-Protocol') == 'https'
        )

    def _session_cookie_flags(self) -> str:
        flags = 'Path=/; HttpOnly; SameSite=Lax; Max-Age=604800'
        if self._is_https_request():
            flags += '; Secure'
        return flags

    def _login_cookie_headers(self, session_id: str, csrf_token: Optional[str] = None) -> List[str]:
        csrf_token = csrf_token or csrf_mod.new_token()
        return [
            f'session_id={session_id}; {self._session_cookie_flags()}',
            csrf_mod.cookie_header_value(csrf_token, secure=self._is_https_request()),
        ]

    def _require_csrf(self) -> bool:
        if csrf_mod.validate(self):
            return True
        self.send_json_response({'success': False, 'error': 'CSRF 校验失败'}, status=403)
        return False

    def _handle_request_exception(self, method: str, exc: Exception) -> None:
        log_server_error(f'{method} {self.path}', exc)
        try:
            if self.path.startswith('/api/'):
                self.send_json_response(safe_error_payload(), status=500)
            else:
                send_safe_http_error(self, 500)
        except Exception:
            pass

    def do_OPTIONS(self):
        """CORS 预检"""
        try:
            self.send_response(204)
            apply_security_headers(self)
            apply_cors(self)
            self.end_headers()
        except Exception:
            pass
    
    def do_GET(self):
        """处理GET请求"""
        try:
            if self.path.startswith('/api/'):
                self.handle_api_get()
            else:
                self.handle_page_get()
        except BrokenPipeError:
            # 客户端断开连接，忽略错误
            pass
        except Exception as e:
            self._handle_request_exception('GET', e)
    
    def do_POST(self):
        """处理POST请求"""
        try:
            logger.debug(f"[POST] 收到POST请求: {self.path}")
            if self.path.startswith('/api/'):
                self.handle_api_post()
            else:
                logger.warning(f"[POST] 未知路径: {self.path}")
                self.send_error(404)
        except BrokenPipeError:
            # 客户端断开连接，忽略错误
            logger.debug(f"[POST] 客户端断开连接: {self.path}")
            pass
        except Exception as e:
            self._handle_request_exception('POST', e)
    
    def do_PUT(self):
        """处理PUT请求"""
        try:
            if self.path.startswith('/api/'):
                self.handle_api_put()
            else:
                self.send_error(404)
        except BrokenPipeError:
            # 客户端断开连接，忽略错误
            pass
        except Exception as e:
            self._handle_request_exception('PUT', e)
    
    def do_DELETE(self):
        """处理DELETE请求"""
        try:
            if self.path.startswith('/api/'):
                self.handle_api_delete()
            else:
                self.send_error(404)
        except BrokenPipeError:
            # 客户端断开连接，忽略错误
            pass
        except Exception as e:
            self._handle_request_exception('DELETE', e)
    
    def handle_api_delete(self):
        """处理API DELETE请求"""
        parsed_path = urllib.parse.urlparse(self.path)
        actual_path = parsed_path.path
        if not csrf_mod.is_exempt(actual_path, 'DELETE') and not csrf_mod.validate(self):
            self.send_json_response({'success': False, 'error': 'CSRF 校验失败'}, status=403)
            return
        if actual_path.startswith('/api/material-db/'):
            from server.material_db_api import MaterialDbApi
            MaterialDbApi(self).dispatch('DELETE', actual_path, parsed_path)
            return

        if self.path.startswith('/api/announcement/sub-boards/delete/'):
            if not self.check_auth():
                return
            
            user = self.get_current_user()
            if not self._can_manage_boards(user):
                self.send_json_response({'error': '仅管理员、管理组成员和最高管理员可以删除二级公告栏'}, status=403)
                return
            
            from server.sub_board_manager import SubBoardManager
            sub_board_mgr = SubBoardManager()
            
            # 删除二级公告栏
            # URL格式: /api/announcement/sub-boards/delete/{parent_board_id}/{sub_board_id}
            parts = self.path.split('/')
            logger.debug(f"[DELETE] 删除二级公告栏请求: path={self.path}, parts={parts}")
            
            if len(parts) >= 6:
                parent_board_id = urllib.parse.unquote(parts[-2])
                sub_board_id = urllib.parse.unquote(parts[-1])
                logger.debug(f"[DELETE] 解析参数: parent={parent_board_id}, sub={sub_board_id}")
                
                try:
                    success, message = sub_board_mgr.delete_sub_board(parent_board_id, sub_board_id)
                    logger.debug(f"[DELETE] delete_sub_board 返回: success={success}, message={message}")
                    if success:
                        logger.info(f"[DELETE] 删除二级公告栏成功: parent={parent_board_id}, sub={sub_board_id}")
                        self.send_json_response({'success': True, 'message': message})
                    else:
                        logger.error(f"[DELETE] 删除失败: {message}")
                        self.send_json_response({'success': False, 'error': message}, status=400)
                except Exception as e:
                    logger.error(f"[DELETE] 删除异常: {e}", exc_info=True)
                    import traceback
                    traceback.print_exc()
                    self.send_json_response({'success': False, 'error': f'删除异常: {str(e)}'}, status=500)
            else:
                logger.error(f"[DELETE] 参数错误: parts长度={len(parts)}, 需要>=6")
                self.send_json_response({'error': '参数错误'}, status=400)
        
        elif self.path.startswith('/api/auth/users/'):
            if not self.check_auth(require_super_admin=True):
                return
            
            parsed_del = urllib.parse.urlparse(self.path)
            path_only = parsed_del.path.rstrip('/')
            try:
                user_id = int(path_only.split('/')[-1])
            except ValueError:
                self.send_json_response({'success': False, 'message': '无效的用户 ID'}, status=400)
                return
            qdel = urllib.parse.parse_qs(parsed_del.query)
            handoff_del = (qdel.get('library_handoff_user_id') or [None])[0]
            handoff_id = int(handoff_del) if handoff_del and str(handoff_del).isdigit() else None
            success, message = self.user_manager.delete_user(user_id, library_handoff_user_id=handoff_id)
            if success:
                self._invalidate_user_cache()
                preloader = get_data_preloader()
                if preloader:
                    try:
                        preloader.invalidate_cache('users')
                    except Exception as e:
                        logger.warning(f"清除用户预加载缓存失败: {e}")
            self.send_json_response({'success': success, 'message': message})
        
        # 公告栏API
        elif self.path.startswith('/api/announcement/delete/'):
            if not self.check_auth():
                return
            
            parsed_delete = urllib.parse.urlparse(self.path)
            announcement_id = parsed_delete.path.split('/')[-1]
            if not announcement_id:
                self.send_json_response({'success': False, 'message': '无效的公告 ID'}, status=400)
                return
            
            # 检查是否有soft_delete参数
            query_params = urllib.parse.parse_qs(parsed_delete.query)
            
            soft_delete = query_params.get('soft_delete', ['true'])[0].lower() in ('true', '1', 'yes') if query_params.get('soft_delete') else None
            
            success, message = self.announcement_mgr.delete_announcement(announcement_id, soft_delete=soft_delete)
            if success:
                # 删除对应的待办Excel文件
                try:
                    todo_file_path = self.todo_mgr.get_todo_file_path(announcement_id)
                    if os.path.exists(todo_file_path):
                        os.remove(todo_file_path)
                        logger.debug(f"已删除公告 {announcement_id} 的待办Excel文件: {todo_file_path}")
                        
                        # 清除待办缓存
                        from server.data_preloader import get_data_preloader
                        preloader = get_data_preloader()
                        if preloader:
                            with preloader.cache_lock:
                                if announcement_id in preloader.todos_cache:
                                    del preloader.todos_cache[announcement_id]
                                    logger.debug(f"已清除公告 {announcement_id} 的待办缓存")
                except Exception as e:
                    logger.warning(f"删除待办Excel文件失败: {e}", exc_info=True)
                    # 不阻止删除公告的操作，只记录警告
                
                # 记录操作日志
                user = self.get_current_user()
                from server.audit_logger import AuditLogger
                AuditLogger.log_announcement_delete(
                    user_id=user.get('id') if user else None,
                    announcement_id=announcement_id,
                    title=message,
                    soft_delete=soft_delete if soft_delete is not None else ENABLE_RECYCLE_BIN,
                    ip_address=self.client_address[0],
                    user_agent=self.headers.get('User-Agent')
                )
                
                # 清除公告相关缓存
                self._invalidate_announcement_cache()
            self.send_json_response({'success': success, 'message': message})
        
        # 恢复公告（从回收站）
        elif self.path.startswith('/api/announcement/restore/'):
            if not self.check_auth():
                return
            
            # 路径格式：/api/announcement/restore/{announcement_id}?timestamp=YYYYMMDD_HHMMSS
            path_parts = self.path.split('/')
            announcement_id = path_parts[-1].split('?')[0] if '?' in path_parts[-1] else path_parts[-1]
            
            # 解析查询参数
            query_params = {}
            if '?' in self.path:
                query_string = self.path.split('?', 1)[1]
                query_params = urllib.parse.parse_qs(query_string)
            
            deleted_timestamp = query_params.get('timestamp', [None])[0]
            
            success, message = self.announcement_mgr.restore_announcement(announcement_id, deleted_timestamp)
            if success:
                # 清除公告相关缓存
                self._invalidate_announcement_cache()
            self.send_json_response({'success': success, 'message': message})
        
        # 清理回收站
        elif self.path == '/api/announcement/cleanup-recycle-bin':
            if not self.check_auth(require_admin=True):
                return
            
            cleaned_count, message = self.announcement_mgr.cleanup_recycle_bin()
            self.send_json_response({'success': True, 'cleaned_count': cleaned_count, 'message': message})
        
        # 快捷链接API - DELETE
        elif self.path.startswith('/api/quick-links/'):
            if not self.check_auth():
                return
            
            user = self.get_current_user()
            if not self._can_manage_quick_links(user):
                self.send_json_response({'success': False, 'error': '仅管理员、管理组成员和最高管理员可以管理快捷链接'}, status=403)
                return
            
            link_id = self.path.split('/')[-1]
            success, message = self.quick_link_mgr.delete_link(link_id)
            if success:
                self.send_json_response({'success': True, 'message': message})
            else:
                self.send_json_response({'success': False, 'error': message}, status=400)
        
        else:
            self.send_error(404)
    
    def handle_api_get(self):
        """处理API GET请求"""
        # 获取实际路径（去除查询参数）
        parsed_path = urllib.parse.urlparse(self.path)
        actual_path = parsed_path.path
        
        logger.debug(f"API GET请求: 原始路径={self.path}, 实际路径={actual_path}")
        
        # 验证码API（不需要认证，必须在最前面，避免被其他逻辑拦截）
        if actual_path == '/api/captcha' or self.path == '/api/captcha':
            logger.info(f"处理验证码API请求: actual_path={actual_path}, self.path={self.path}")
            try:
                from server.captcha import get_captcha_manager
                import base64
                
                captcha_mgr = get_captcha_manager()
                token, image_bytes = captcha_mgr.generate_captcha()
                
                # 将图片转换为base64
                image_base64 = base64.b64encode(image_bytes).decode('utf-8')
                
                # 判断是PNG还是SVG
                from server.captcha import HAS_PIL
                if HAS_PIL:
                    image_data = f'data:image/png;base64,{image_base64}'
                else:
                    # SVG格式
                    image_data = f'data:image/svg+xml;base64,{image_base64}'
                
                logger.info(f"验证码生成成功: token={token[:8]}..., image_size={len(image_bytes)}")
                
                # 直接发送响应，确保不会走到其他逻辑
                response_data = {
                    'success': True,
                    'token': token,
                    'image': image_data
                }
                self.send_json_response(response_data)
                logger.debug("验证码响应已发送")
                return
            except Exception as e:
                logger.error(f"生成验证码失败: {e}", exc_info=True)
                import traceback
                logger.error(traceback.format_exc())
                # 确保错误响应也是JSON格式
                try:
                    self.send_json_response({
                        'success': False,
                        'error': f'生成验证码失败: {str(e)}'
                    }, status=500)
                except:
                    # 如果发送JSON也失败，至少记录错误
                    logger.error("发送验证码错误响应也失败")
                return
        
        elif actual_path.startswith('/api/material-db/'):
            from server.material_db_api import MaterialDbApi
            MaterialDbApi(self).dispatch('GET', actual_path, parsed_path)
            return

        # NEO 内部：排行榜用户列表（Docker 内网 / 共享密钥，不走 Cookie）
        elif actual_path == '/api/internal/neo/active-users':
            if not self._allow_neo_internal_request():
                self.send_json_response({'success': False, 'error': '禁止访问'}, status=403)
                return
            try:
                users_out = self._list_neo_active_users()
                self.send_json_response({'success': True, 'users': users_out})
            except Exception as e:
                logger.error(f"内部 NEO 用户列表失败: {e}", exc_info=True)
                self.send_json_response({'success': False, 'error': str(e)}, status=500)
            return

        elif actual_path == '/api/internal/neo/verify-session':
            self._handle_internal_neo_verify_session()
            return

        elif actual_path == '/api/health':
            self._handle_health_check()
            return

        elif actual_path == '/api/startup/status':
            self._handle_startup_status()
            return

        elif actual_path == '/api/auth/config':
            self._handle_auth_config()
            return

        elif actual_path == '/api/auth/dingtalk/web/start':
            self._handle_dingtalk_web_start(parsed_path)
            return

        elif actual_path == '/api/auth/dingtalk/callback':
            self._handle_dingtalk_web_callback(parsed_path)
            return

        # 认证检查API
        elif actual_path == '/api/auth/check':
            self._handle_auth_check()
        elif actual_path == '/api/auth/session':
            self._handle_auth_session()

        # NEO 排行榜：已登录用户可获取激活用户名单（姓名 + 稳定 userKey）
        elif actual_path == '/api/neo/active-users':
            if not self.check_auth():
                return
            try:
                users_out = self._list_neo_active_users()
                self.send_json_response({'success': True, 'users': users_out})
            except Exception as e:
                logger.error(f"获取 NEO 激活用户列表失败: {e}", exc_info=True)
                self.send_json_response({'success': False, 'error': str(e)}, status=500)
            return

        # 退出登录API
        elif actual_path == '/api/auth/logout':
            from server.auth.login_service import perform_logout

            cookie_header = self.headers.get('Cookie', '')
            body, clear_cookies = perform_logout(
                cookie_header,
                secure=self._is_https_request(),
            )
            self.send_response(200)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            for c in clear_cookies:
                self.send_header('Set-Cookie', c)
            self.end_headers()
            self.wfile.write(json.dumps(body, ensure_ascii=False).encode('utf-8'))
        
        elif actual_path == '/api/auth/job-roles':
            # 从配置文件实时获取岗位选项，确保数据一致性
            role_options = [{'id': key, 'label': label} for key, label in JOB_POSITION_OPTIONS.items()]
            self.send_json_response(role_options)
        
        # 其他API都需要认证
        elif not self.check_auth():
            return
            
        elif actual_path == '/api/auth/users':
            logger.debug(f"[DEBUG] 用户列表API请求: path={self.path}")
            user = self.get_current_user()
            logger.debug(f"[DEBUG] 当前用户: {user.get('username') if user else 'None'}")
            if not user:
                logger.warning("[DEBUG] 获取用户列表失败: 用户未认证")
                self.send_json_response({'error': '未认证'}, status=401)
                return
            
            is_super = self._is_super_admin(user)
            logger.debug(f"[DEBUG] 用户 {user.get('username')} 请求用户列表, 是否超级管理员: {is_super}, 角色: {user.get('roles', [])}, 用户名: {user.get('username')}")
            
            if not is_super:
                logger.warning(f"用户 {user.get('username')} 无权限访问用户列表 (角色: {user.get('roles', [])})")
                self.send_json_response({'error': '仅超级管理员可访问'}, status=403)
                return
            
            try:
                # 支持按部门过滤和强制刷新（从查询参数获取）
                parsed_path = urllib.parse.urlparse(self.path)
                query_params = urllib.parse.parse_qs(parsed_path.query)
                department = query_params.get('department', [None])[0]  # 获取department参数
                force_refresh = query_params.get('_refresh', [None])[0] == 'true'  # 获取_refresh参数
                
                # 如果指定了强制刷新或部门过滤，直接从数据库读取（不使用缓存）
                # 这样可以确保在多进程环境下，所有worker都能看到最新的数据
                if force_refresh or department:
                    logger.debug(f"[用户列表API] 强制从数据库读取（force_refresh={force_refresh}, department={department}）")
                    users = self.user_manager.get_all_users(department=department)
                else:
                    # 优先从内存缓存获取（避免文件I/O），但需要验证缓存是否有效
                    preloader = get_data_preloader()
                    users = []
                    if preloader and PRELOAD_USERS:
                        try:
                            # 从预加载器获取（可能使用缓存）
                            users = preloader.get_users()
                            logger.debug(f"[用户列表API] 从预加载器获取到 {len(users)} 个用户")
                            # 如果预加载器返回空列表，回退到直接读取
                            if not users:
                                logger.warning(f"[用户列表API] 预加载器返回空列表，回退到直接读取")
                                users = self.user_manager.get_all_users(department=department)
                                logger.debug(f"[用户列表API] 从user_manager直接读取到 {len(users)} 个用户")
                        except Exception as e:
                            logger.warning(f"[用户列表API] 从内存缓存获取用户失败，使用数据库读取: {e}", exc_info=True)
                            users = self.user_manager.get_all_users(department=department)
                    else:
                        # 如果未启用预加载，直接从user_manager读取
                        logger.debug(f"[用户列表API] 预加载未启用，直接从user_manager读取")
                        users = self.user_manager.get_all_users(department=department)
                
                logger.debug(f"[用户列表API] 最终获取到 {len(users)} 个用户" + (f" (部门: {department})" if department else ""))
                sanitized = [self._public_user_payload(u) for u in users]
                logger.debug(f"[用户列表API] 获取用户列表成功: {len(sanitized)} 个用户, 准备发送响应")
                self.send_json_response(sanitized)
                logger.debug(f"[用户列表API] 用户列表响应已发送")
            except Exception as e:
                logger.error(f"获取用户列表失败: {e}", exc_info=True)
                self.send_json_response({'error': f'获取用户列表失败: {str(e)}'}, status=500)
        
        elif actual_path == '/api/auth/registrations':
            if not self.check_auth(required_roles=['management']):
                return
            pending = self.user_manager.get_pending_users()
            sanitized = [self._public_user_payload(u) for u in pending]
            self.send_json_response(sanitized)
        
        # 获取按title分组的用户列表（用于待办人员选择）
        elif self.path == '/api/users/by-title' or actual_path == '/api/auth/users/by-title':
            logger.info("收到获取按title分组的用户列表请求")
            
            if not self.check_auth():
                logger.warning("获取用户列表失败: 用户未认证")
                return
            
            user = self.get_current_user()
            logger.info(f"当前用户: {user.get('username', 'unknown')}, 角色: {user.get('roles', [])}")
            
            # 只有管理员、管理组和最高管理员可以查看用户列表
            if not (self._is_super_admin(user) or self._has_role(user, 'management') or self._has_role(user, 'admin')):
                logger.warning(f"用户 {user.get('username', 'unknown')} 无权限访问用户列表")
                self.send_json_response({'error': '仅管理员、管理组成员和最高管理员可以查看用户列表'}, status=403)
                return
            
            try:
                logger.info("开始获取所有激活状态的用户...")
                # 优先从内存缓存获取（避免文件I/O）
                preloader = get_data_preloader()
                if preloader and PRELOAD_USERS:
                    try:
                        all_users = preloader.get_users(status='active')
                        logger.info(f"从内存缓存获取到 {len(all_users)} 个激活用户")
                    except Exception as e:
                        logger.warning(f"从内存缓存获取用户失败，使用文件读取: {e}")
                        all_users = self.user_manager.get_all_users(statuses=[STATUS_ACTIVE])
                else:
                    # 如果未启用预加载，从文件读取
                    all_users = self.user_manager.get_all_users(statuses=[STATUS_ACTIVE])
                logger.info(f"获取到 {len(all_users)} 个激活用户（状态='active'）")
                
                # 按title分组
                users_by_title = {}
                title_counts = {}  # 用于调试
                for u in all_users:
                    if self._exclude_from_todo_user_selection(u):
                        continue
                    # 获取title，处理各种可能的情况
                    title = u.get('title', '')
                    if not title or str(title).strip() == '' or str(title).strip().lower() == 'none':
                        title = '未设置岗位'
                    else:
                        title = str(title).strip()
                    
                    if title not in users_by_title:
                        users_by_title[title] = []
                        title_counts[title] = 0
                    
                    users_by_title[title].append({
                        'userid': u.get('userid', ''),
                        'name': u.get('name', ''),
                        'title': title,
                        'job_number': u.get('job_number', '')
                    })
                    title_counts[title] += 1
                
                # 按title名称排序
                sorted_titles = sorted(users_by_title.keys())
                result = {
                    'titles': sorted_titles,
                    'users_by_title': {title: users_by_title[title] for title in sorted_titles}
                }
                
                # 输出调试信息
                logger.info(f"按title分组完成，共 {len(sorted_titles)} 个岗位组")
                for title, count in sorted(title_counts.items()):
                    logger.info(f"  - {title}: {count}人")
                self.send_json_response(result)
            except Exception as e:
                logger.error(f"获取按title分组的用户列表失败: {e}", exc_info=True)
                self.send_json_response({'error': f'获取用户列表失败: {str(e)}'}, status=500)
        
        # 快捷链接API
        elif self.path == '/api/quick-links':
            if not self.check_auth():
                return
            links = self.quick_link_mgr.get_links()
            self.send_json_response(links)
        
        # 公告栏API - 必须放在公告详情API之前，支持查询参数
        elif self.path.startswith('/api/announcement/boards'):
            # 检查是否是精确匹配（可能带查询参数）
            parsed = urllib.parse.urlparse(self.path)
            if parsed.path == '/api/announcement/boards':
                if not self.check_auth():
                    return
                # 从数据库读取一级公告栏
                from server.board_manager import BoardManager
                board_mgr = BoardManager()
                boards = board_mgr.get_all_boards()
                logger.debug(f"[API] 获取一级公告栏: 数量={len(boards)}")
                self.send_json_response(boards)
                return
            # 如果不是 /api/announcement/boards，继续检查其他路径
        
        elif self.path.startswith('/api/announcement/list'):
            # 减少日志输出（高并发下日志过多会影响性能）
            if not self.check_auth():
                return
            
            # 解析路径和查询参数
            # 优先使用 self.query_params（WSGI环境已设置），否则从 self.path 解析（兼容直接HTTP服务器）
            if hasattr(self, 'query_params') and isinstance(self.query_params, dict):
                query_params = self.query_params
            else:
                parsed = urllib.parse.urlparse(self.path)
                query_params = urllib.parse.parse_qs(parsed.query)
            
            # 获取查询参数，如果没有则使用默认值
            board_id = query_params.get('board', ['all'])[0]
            sub_board_id = query_params.get('sub_board', [None])[0]
            status = query_params.get('status', [None])[0]
            search = query_params.get('search', [None])[0]
            force_refresh = query_params.get('_refresh', [''])[0].lower() in ('1', 'true', 'yes')
            
            # 分页参数
            try:
                page = int(query_params.get('page', ['1'])[0])
                page_size = int(query_params.get('page_size', [str(ANNOUNCEMENT_PAGE_SIZE)])[0])
                # 限制最大页面大小，防止恶意请求
                page_size = min(page_size, ANNOUNCEMENT_MAX_PAGE_SIZE)
                page = max(1, page)  # 确保页码至少为1
            except (ValueError, TypeError):
                page = 1
                page_size = ANNOUNCEMENT_PAGE_SIZE
            
            # 构建缓存键（排除草稿和待审批，因为它们需要实时数据）
            cache_key = None
            if status not in ['draft', 'pending']:
                cache_key = f"announcement_list:{board_id}:{sub_board_id}:{status}:{search}:{page}:{page_size}"
                if not force_refresh:
                    cached_result = api_cache.get(cache_key)
                    if cached_result is not None:
                        self.send_json_response(cached_result)
                        return
            
            # 如果查询草稿，需要包含temp目录
            include_temp = (status == 'draft' or status == 'pending')
            
            # 跨进程/强制刷新：审批后其他 worker 或 _refresh 请求须从磁盘重载公告
            preloader = get_data_preloader()
            if force_refresh or (not include_temp and status in [None, 'approved']):
                try:
                    from server.cache_sync import check_cache_invalidation, update_local_version
                    marker_time = check_cache_invalidation('announcements')
                    if marker_time or force_refresh:
                        if preloader:
                            preloader.invalidate_cache('announcements')
                            preloader.reload_cache('announcements')
                        if marker_time:
                            update_local_version('announcements', marker_time)
                        if force_refresh:
                            with api_cache.lock:
                                keys_to_remove = [
                                    k for k in list(api_cache.cache.keys())
                                    if k.startswith('announcement_list:')
                                ]
                                for k in keys_to_remove:
                                    api_cache._remove_key(k)
                except Exception as e:
                    logger.debug(f"公告列表强制刷新缓存失败: {e}")
            
            # 对于已发布的公告，优先从内存缓存获取（避免文件I/O）
            announcements = None
            use_cache = False
            if not force_refresh and not include_temp and status in [None, 'approved'] and preloader and PRELOAD_ANNOUNCEMENTS:
                try:
                    # 从缓存获取公告（如果status=None，获取所有；如果status=approved，只获取approved）
                    cache_status = 'approved' if status == 'approved' else None
                    all_cached = preloader.get_announcements_cached(status=cache_status)
                    
                    # 如果缓存返回了数据（即使是空列表，也说明缓存已加载），使用缓存
                    # 重要：使用 is not None 而不是 if all_cached，因为空列表也是有效的响应
                    if all_cached is not None:
                        announcements = all_cached
                        use_cache = True
                        # 应用过滤条件
                        if board_id != 'all':
                            announcements = [a for a in announcements if a.get('board_id') == board_id]
                        if sub_board_id:
                            announcements = [a for a in announcements if a.get('sub_board_id') == sub_board_id]
                        
                        # 健康检查：缓存条数明显少于磁盘时从文件重载（审批后新增公告常见）
                        if self.announcement_mgr:
                            try:
                                file_announcements_check = self.announcement_mgr.get_announcements(
                                    status=status,
                                    board_id=board_id if board_id != 'all' else None,
                                    sub_board_id=sub_board_id,
                                    include_temp=include_temp
                                )
                                file_count = len(file_announcements_check or [])
                                cache_count = len(announcements)
                                if file_count > cache_count:
                                    logger.info(
                                        f"公告列表缓存落后于磁盘（缓存 {cache_count}，磁盘 {file_count}），从文件重载"
                                    )
                                    announcements = file_announcements_check
                                    use_cache = False
                                    if preloader:
                                        import threading
                                        threading.Thread(
                                            target=preloader._preload_announcements,
                                            daemon=True,
                                            name='announcement-cache-catchup',
                                        ).start()
                            except Exception as e:
                                logger.debug(f"公告列表健康检查失败（继续使用缓存）: {e}")
                        
                        # 兼容旧逻辑：缓存极少时做额外检查
                        if len(announcements) <= 2 and self.announcement_mgr:
                            try:
                                file_announcements_check = self.announcement_mgr.get_announcements(
                                    status=status, 
                                    board_id=board_id if board_id != 'all' else None,
                                    sub_board_id=sub_board_id, 
                                    include_temp=include_temp
                                )
                                # 如果文件中有更多公告，说明缓存异常，需要重新加载
                                if file_announcements_check and len(file_announcements_check) > len(announcements) + 2:
                                    logger.warning(f"检测到公告缓存异常：缓存中只有 {len(announcements)} 条，但文件中有 {len(file_announcements_check)} 条，将重新加载缓存")
                                    # 使用文件数据
                                    announcements = file_announcements_check
                                    use_cache = False
                                    # 在后台线程中重新加载缓存
                                    import threading
                                    def recover_cache():
                                        try:
                                            logger.info("开始恢复公告缓存...")
                                            preloader._preload_announcements()
                                            logger.info("公告缓存恢复完成")
                                        except Exception as e:
                                            logger.error(f"恢复公告缓存失败: {e}", exc_info=True)
                                    threading.Thread(target=recover_cache, daemon=True).start()
                            except Exception as e:
                                logger.debug(f"健康检查失败（继续使用缓存）: {e}")
                        
                        # 如果缓存为空，尝试从文件读取并更新缓存（但不阻塞当前请求）
                        if len(announcements) == 0 and self.announcement_mgr:
                            # 在后台线程中从文件读取并更新缓存
                            import threading
                            def update_cache_in_background():
                                try:
                                    file_anns = self.announcement_mgr.get_announcements(
                                        status=status, 
                                        sub_board_id=sub_board_id, 
                                        include_temp=include_temp
                                    )
                                    if file_anns and len(file_anns) > 0:
                                        # 更新缓存
                                        preloader._preload_announcements()
                                except Exception:
                                    pass
                            threading.Thread(target=update_cache_in_background, daemon=True).start()
                    else:
                        # 缓存返回None，说明缓存未初始化或加载失败，需要从文件读取
                        if logger.isEnabledFor(logging.DEBUG):
                            logger.debug("公告缓存未初始化，将从文件读取")
                except Exception as e:
                    # 只在真正出错时记录警告（减少日志）
                    if logger.isEnabledFor(logging.WARNING):
                        logger.warning(f"从内存缓存获取公告失败，使用文件读取: {e}")
                    announcements = None
            
            # 如果缓存获取失败或需要实时数据，从文件读取
            # 改进：如果缓存为空列表，也尝试从文件读取一次（确保数据最新）
            if not use_cache and announcements is None:
                try:
                    if board_id == 'all':
                        file_announcements = self.announcement_mgr.get_announcements(status=status, sub_board_id=sub_board_id, include_temp=include_temp)
                    else:
                        file_announcements = self.announcement_mgr.get_announcements(board_id=board_id, status=status, sub_board_id=sub_board_id, include_temp=include_temp)
                    
                    # 如果从文件读取到数据，使用文件数据并更新缓存
                    if file_announcements and len(file_announcements) > 0:
                        announcements = file_announcements
                        # 更新缓存（在后台进行，不阻塞请求）
                        if preloader:
                            try:
                                import threading
                                def update_cache():
                                    try:
                                        preloader._preload_announcements()
                                    except Exception:
                                        pass
                                threading.Thread(target=update_cache, daemon=True).start()
                            except Exception:
                                pass
                    else:
                        # 文件读取也返回空，返回空列表
                        announcements = []
                except Exception as e:
                    logger.error(f"从文件读取公告失败: {e}", exc_info=True)
                    # 即使读取失败，也返回空列表而不是抛出异常
                    announcements = []
            
            # 如果使用了缓存但缓存为空，也尝试从文件读取一次（确保数据最新）
            elif use_cache and announcements is not None and len(announcements) == 0 and not include_temp:
                try:
                    if board_id == 'all':
                        file_announcements = self.announcement_mgr.get_announcements(status=status, sub_board_id=sub_board_id, include_temp=include_temp)
                    else:
                        file_announcements = self.announcement_mgr.get_announcements(board_id=board_id, status=status, sub_board_id=sub_board_id, include_temp=include_temp)
                    
                    # 如果从文件读取到数据，使用文件数据并更新缓存
                    if file_announcements and len(file_announcements) > 0:
                        announcements = file_announcements
                        # 更新缓存（在后台进行，不阻塞请求）
                        if preloader:
                            try:
                                import threading
                                def update_cache():
                                    try:
                                        preloader._preload_announcements()
                                    except Exception:
                                        pass
                                threading.Thread(target=update_cache, daemon=True).start()
                            except Exception:
                                pass
                except Exception as e:
                    # 文件读取失败，继续使用空缓存（不记录错误，因为缓存已经是空的）
                    if logger.isEnabledFor(logging.DEBUG):
                        logger.debug(f"从文件读取公告失败（缓存为空时）: {e}")
            
            # 批量优化：一次性获取所有sub_board和待办统计（避免N+1查询）
            from server.sub_board_manager import SubBoardManager
            sub_board_mgr = SubBoardManager()
            preloader = get_data_preloader()
            
            # 确保preloader有正确的todo_mgr引用（如果未设置，使用当前handler的todo_mgr）
            if preloader and not preloader.todo_mgr and self.todo_mgr:
                preloader.todo_mgr = self.todo_mgr
            
            # 批量获取所有需要的sub_board信息（减少重复查询）
            sub_board_cache = {}
            for ann in announcements:
                board_id = ann.get('board_id')
                sub_board_id = ann.get('sub_board_id')
                if board_id and sub_board_id:
                    cache_key = f"{board_id}_{sub_board_id}"
                    if cache_key not in sub_board_cache:
                        sub_board = sub_board_mgr.get_sub_board(board_id, sub_board_id)
                        if sub_board:
                            sub_board_cache[cache_key] = sub_board['name']
                        else:
                            sub_board_cache[cache_key] = None
            
            # 批量获取所有待办统计（一次性从缓存获取，避免逐个查询）
            todo_stats_cache = {}
            
            # 优先使用预加载器（如果可用）
            if preloader and PRELOAD_TODOS and preloader.todo_mgr:
                # 一次性获取所有公告的待办数据
                for ann in announcements:
                    announcement_id = ann.get('id')
                    if announcement_id and announcement_id not in todo_stats_cache:
                        try:
                            todos = preloader.get_todos_for_announcement_cached(announcement_id)
                            # 确保todos是列表类型
                            if not isinstance(todos, list):
                                todos = []
                            
                            # 计算统计信息（即使todos为空也要计算）
                            total_count = len(todos)
                            completed_count = sum(1 for todo in todos if todo.get('done', False) or todo.get('status') == '已完成' or todo.get('status') == 'done')
                            pending_count = total_count - completed_count
                            
                            todo_stats_cache[announcement_id] = {
                                'total': total_count,
                                'completed': completed_count,
                                'pending': pending_count
                            }
                            
                            # 记录调试信息（仅当有数据时）
                            if total_count > 0:
                                logger.debug(f"公告 {announcement_id} 待办统计: 总计={total_count}, 已完成={completed_count}, 待完成={pending_count}")
                        except Exception as e:
                            logger.debug(f"从预加载器获取公告 {announcement_id} 的待办统计失败: {e}")
                            # 降级：尝试直接从TodoManager获取
                            try:
                                if self.todo_mgr:
                                    todos = self.todo_mgr.get_all_todos(announcement_id)
                                    if not isinstance(todos, list):
                                        todos = []
                                    total_count = len(todos)
                                    completed_count = sum(1 for todo in todos if todo.get('done', False) or todo.get('status') == '已完成' or todo.get('status') == 'done')
                                    pending_count = total_count - completed_count
                                    todo_stats_cache[announcement_id] = {
                                        'total': total_count,
                                        'completed': completed_count,
                                        'pending': pending_count
                                    }
                                else:
                                    todo_stats_cache[announcement_id] = {'total': 0, 'completed': 0, 'pending': 0}
                            except Exception as e2:
                                logger.debug(f"从TodoManager获取公告 {announcement_id} 的待办统计也失败: {e2}")
                                todo_stats_cache[announcement_id] = {'total': 0, 'completed': 0, 'pending': 0}
            else:
                # 如果没有预加载器，直接从TodoManager获取（降级方案）
                if self.todo_mgr:
                    for ann in announcements:
                        announcement_id = ann.get('id')
                        if announcement_id and announcement_id not in todo_stats_cache:
                            try:
                                todos = self.todo_mgr.get_all_todos(announcement_id)
                                if not isinstance(todos, list):
                                    todos = []
                                total_count = len(todos)
                                completed_count = sum(1 for todo in todos if todo.get('done', False) or todo.get('status') == '已完成' or todo.get('status') == 'done')
                                pending_count = total_count - completed_count
                                todo_stats_cache[announcement_id] = {
                                    'total': total_count,
                                    'completed': completed_count,
                                    'pending': pending_count
                                }
                            except Exception as e:
                                logger.debug(f"获取公告 {announcement_id} 的待办统计失败: {e}")
                                todo_stats_cache[announcement_id] = {'total': 0, 'completed': 0, 'pending': 0}
                else:
                    # 如果TodoManager也不可用，为所有公告创建空统计
                    for ann in announcements:
                        announcement_id = ann.get('id')
                        if announcement_id:
                            todo_stats_cache[announcement_id] = {'total': 0, 'completed': 0, 'pending': 0}
            
            # 快速填充公告数据（使用批量获取的缓存）
            for ann in announcements:
                # 添加sub_board名称
                board_id = ann.get('board_id')
                sub_board_id = ann.get('sub_board_id')
                if board_id and sub_board_id:
                    cache_key = f"{board_id}_{sub_board_id}"
                    if cache_key in sub_board_cache and sub_board_cache[cache_key]:
                        ann['sub_board_name'] = sub_board_cache[cache_key]
                
                # 添加待办统计
                announcement_id = ann.get('id')
                if announcement_id:
                    ann['todo_stats'] = todo_stats_cache.get(announcement_id, {'total': 0, 'completed': 0, 'pending': 0})
                else:
                    ann['todo_stats'] = {'total': 0, 'completed': 0, 'pending': 0}
            
            # 如果提供了搜索关键词，先进行搜索过滤（在所有公告上，使用完整content）
            # 注意：搜索必须在分页之前进行，这样才能搜索所有页的公告，而不是只搜索当前页
            if search:
                search_lower = search.lower()
                import re
                filtered_announcements = []
                for ann in announcements:
                    # 检查title
                    title_match = search_lower in (ann.get('title', '') or '').lower()
                    
                    # 检查author
                    author_match = search_lower in (ann.get('author', '') or '').lower()
                    
                    # 检查content（使用完整content，移除HTML标签后搜索）
                    content_match = False
                    if 'content' in ann:
                        content = ann.get('content', '')
                        if isinstance(content, str):
                            # 移除HTML标签后搜索完整内容
                            text_content = re.sub(r'<[^>]+>', '', content)
                            content_match = search_lower in text_content.lower()
                    
                    # 如果任一字段匹配，则包含此公告
                    if title_match or author_match or content_match:
                        filtered_announcements.append(ann)
                
                announcements = filtered_announcements
            
            # 移除完整的content字段，只保留摘要（列表API不需要完整内容）
            # 注意：这个操作在搜索之后进行，确保搜索使用的是完整内容
            # 如果需要完整内容，应该调用详情API
            for ann in announcements:
                if 'content' in ann:
                    content = ann['content']
                    # 如果是HTML，提取纯文本
                    if isinstance(content, str):
                        import re
                        # 移除HTML标签
                        text_content = re.sub(r'<[^>]+>', '', content)
                        # 限制摘要长度（最多200字符）
                        if len(text_content) > 200:
                            ann['content'] = text_content[:200] + '...'
                        else:
                            ann['content'] = text_content
                    # 如果content太大，直接移除（详情API会返回完整内容）
                    elif len(str(content)) > 500:
                        ann.pop('content', None)
            
            # 计算总数（分页前）
            total_count = len(announcements)
            
            # 分页处理
            # 注意：Python列表切片不会填充空白，如果只有21个公告，第二页只会返回1个，不会填充到20个
            start_idx = (page - 1) * page_size
            end_idx = start_idx + page_size
            paginated_announcements = announcements[start_idx:end_idx]
            
            # 确保不会超出列表范围（虽然Python切片会自动处理，但为了明确性）
            if start_idx >= total_count:
                # 如果起始索引超出范围，返回空列表
                paginated_announcements = []
            elif end_idx > total_count:
                # 如果结束索引超出范围，只返回到列表末尾
                paginated_announcements = announcements[start_idx:]
            
            # 构建响应（包含分页信息）
            response_data = {
                'announcements': paginated_announcements,
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total': total_count,
                    'total_pages': (total_count + page_size - 1) // page_size if total_count > 0 else 0
                }
            }
            
            # 缓存结果（仅缓存非草稿/待审批的查询）
            # 阶段1优化：根据公告数据判断是否为热点数据，使用不同的TTL
            if cache_key:
                # 判断是否为热点数据（最近30天的公告）
                cache_ttl = NORMAL_CACHE_TTL
                if announcements and len(announcements) > 0:
                    # 检查第一个公告的创建时间（假设列表已按时间排序）
                    first_ann = announcements[0]
                    created_time = first_ann.get('created_time') or first_ann.get('create_time') or first_ann.get('publish_time')
                    if created_time:
                        try:
                            from datetime import datetime
                            if isinstance(created_time, str):
                                created_dt = datetime.fromisoformat(created_time.replace('Z', '+00:00'))
                            else:
                                created_dt = created_time
                            days_old = (datetime.now() - created_dt.replace(tzinfo=None)).days
                            if days_old <= 30:  # 最近30天的公告使用热点缓存TTL
                                cache_ttl = HOT_CACHE_TTL
                        except Exception:
                            pass  # 解析失败，使用默认TTL
                api_cache.set(cache_key, response_data, ttl=cache_ttl)
            
            # 减少日志输出（高并发下日志过多会影响性能）
            # 只在需要调试时记录（通过环境变量控制）
            if os.getenv('ENABLE_DEBUG_LOGS', '').lower() in ('1', 'true', 'yes'):
                logger.debug(f"公告列表查询: status={status}, board={board_id}, total={total_count}, page={page}, page_size={page_size}, returned={len(paginated_announcements)}")
            
            self.send_json_response(response_data)
        
        # 二级公告栏管理API - 必须放在 startswith 之前，精确匹配优先
        elif self.path == '/api/announcement/sub-boards/all':
            if not self.check_auth():
                return
            
            # 所有登录用户都可以查看二级公告栏列表（用于显示）
            # 但只有最高管理员可以管理
            
            try:
                logger.debug("开始获取所有二级公告栏...")
                from server.sub_board_manager import SubBoardManager
                sub_board_mgr = SubBoardManager()
                logger.info("SubBoardManager 实例创建成功")
                
                all_sub_boards = sub_board_mgr.get_all_sub_boards()
                logger.info(f"get_all_sub_boards 返回类型: {type(all_sub_boards)}")
                logger.info(f"get_all_sub_boards 返回值: {all_sub_boards}")
                
                # 确保返回的是字典格式
                if not isinstance(all_sub_boards, dict):
                    logger.warning(f"get_all_sub_boards 返回了非字典类型: {type(all_sub_boards)}, 值: {all_sub_boards}")
                    all_sub_boards = {}
                
                logger.info(f"返回所有二级公告栏: {len(all_sub_boards)} 个父公告栏")
                total_sub_boards = 0
                for parent_id, sub_boards in all_sub_boards.items():
                    if isinstance(sub_boards, list):
                        total_sub_boards += len(sub_boards)
                        logger.info(f"  {parent_id}: {len(sub_boards)} 个二级公告栏")
                        for sub_board in sub_boards:
                            logger.debug(f"    - {sub_board.get('sub_board_id')}: {sub_board.get('name')}")
                    else:
                        logger.warning(f"  {parent_id}: 数据格式错误，期望列表但得到 {type(sub_boards)}")
                
                logger.info(f"总共 {total_sub_boards} 个二级公告栏，准备发送响应")
                self.send_json_response(all_sub_boards)
                logger.info("响应已发送")
            except Exception as e:
                logger.error(f"获取所有二级公告栏失败: {e}", exc_info=True)
                import traceback
                traceback.print_exc()
                self.send_json_response({'error': f'获取失败: {str(e)}'}, status=500)
        
        elif self.path.startswith('/api/announcement/sub-boards'):
            if not self.check_auth():
                return
            
            # 解析查询参数
            # 优先使用 self.query_params（WSGI环境已设置），否则从 self.path 解析（兼容直接HTTP服务器）
            if hasattr(self, 'query_params') and isinstance(self.query_params, dict):
                query_params = self.query_params
            else:
                parsed = urllib.parse.urlparse(self.path)
                query_params = urllib.parse.parse_qs(parsed.query)
            parent_board_id = query_params.get('parent', [None])[0]
            
            logger.debug(f"[API] /api/announcement/sub-boards 请求: path={self.path}, parent={parent_board_id}")
            
            if not parent_board_id:
                logger.warning(f"[API] 缺少 parent 参数，返回空数组")
                self.send_json_response([])
                return
            
            from server.sub_board_manager import SubBoardManager
            sub_board_mgr = SubBoardManager()
            sub_boards = sub_board_mgr.get_sub_boards(parent_board_id)
            logger.debug(f"[API] 获取二级公告栏: parent={parent_board_id}, 数量={len(sub_boards)}")
            if len(sub_boards) > 0:
                logger.debug(f"[API] 二级公告栏列表: {[s['sub_board_id'] + ':' + s['name'] for s in sub_boards]}")
            self.send_json_response(sub_boards)
        
        elif self.path == '/api/announcement/statistics':
            if not self.check_auth():
                return
            stats = self.announcement_mgr.get_statistics()
            self.send_json_response(stats)
        
        elif self.path == '/api/announcement/approvers':
            if not self.check_auth():
                return
            # 获取可选的审批人列表（管理组和管理员）
            try:
                approvers = self._get_approvers()
                # 返回格式化的审批人列表，供前端选择
                approver_list = []
                for approver in approvers:
                    # 优先使用userid，如果没有则使用username，最后使用id
                    identifier = approver.get('userid') or approver.get('username') or str(approver.get('id', ''))
                    approver_list.append({
                        'userid': approver.get('userid', ''),
                        'username': approver.get('username', ''),
                        'id': approver.get('id'),
                        'identifier': identifier,  # 统一标识符，用于前端选择
                        'name': approver.get('name', ''),
                        'unionid': approver.get('unionid', '')
                    })
                self.send_json_response({'success': True, 'approvers': approver_list})
            except Exception as e:
                logger.error(f"获取审批人列表失败: {e}", exc_info=True)
                self.send_json_response({'success': False, 'error': f'获取审批人列表失败: {str(e)}'})
        
        elif self.path == '/api/announcement/pending':
            # API请求，认证失败时返回JSON错误，而不是重定向
            user = self.get_current_user()
            if not user:
                self.send_json_response({'error': '未登录，请先登录'}, status=401)
                return
            try:
                logger.debug("获取待审批公告列表")
                pending_announcements = self.announcement_mgr.get_announcements(status='pending', include_temp=True)
                
                # 为每个公告添加二级公告栏名称
                try:
                    from server.sub_board_manager import SubBoardManager
                    sub_board_mgr = SubBoardManager()
                    for ann in pending_announcements:
                        if ann.get('sub_board_id'):
                            try:
                                sub_board = sub_board_mgr.get_sub_board(ann.get('board_id'), ann.get('sub_board_id'))
                                if sub_board:
                                    ann['sub_board_name'] = sub_board['name']
                            except Exception as e:
                                logger.warning(f"获取二级公告栏名称失败 (board_id={ann.get('board_id')}, sub_board_id={ann.get('sub_board_id')}): {e}")
                                continue
                except Exception as e:
                    logger.warning(f"初始化SubBoardManager失败: {e}")
                    # 继续执行，不添加二级公告栏名称
                
                logger.debug(f"找到 {len(pending_announcements)} 条待审批公告")
                self.send_json_response(pending_announcements)
            except Exception as e:
                logger.error(f"获取待审批公告失败: {e}", exc_info=True)
                self.send_json_response({
                    'error': f'加载待审批公告失败: {str(e)}'
                }, status=500)
        
        # 附件下载API
        elif self.path.startswith('/api/announcement/attachment/'):
            if not self.check_auth():
                return
            parts = self.path.split('/')
            print(f"附件下载请求: path={self.path}, parts={parts}, len={len(parts)}")
            if len(parts) >= 6:  # /api/announcement/attachment/{id}/{filename} 或 by-index/{n}
                announcement_id = parts[4]
                version_number = None
                filename = None

                if parts[5] == 'by-index' and len(parts) >= 7:
                    index_part = parts[6]
                    if '?' in index_part:
                        index_part, query_string = index_part.split('?', 1)
                        version_number = urllib.parse.parse_qs(query_string).get('version', [None])[0]
                    try:
                        att_index = int(index_part)
                    except ValueError:
                        self._send_attachment_api_error(400, '无效的附件索引')
                        return
                    ann = self.announcement_mgr.get_announcement_for_download(announcement_id)
                    if not ann:
                        ann = self.announcement_mgr.get_announcement(announcement_id)
                    attachments = (ann or {}).get('attachments') or []
                    if att_index < 0 or att_index >= len(attachments):
                        self._send_attachment_api_error(404, '附件索引不存在')
                        return
                    filename = attachments[att_index].get('name') or ''
                    if not filename:
                        self._send_attachment_api_error(404, '附件名为空')
                        return
                    logger.info(
                        f"附件下载(by-index): id={announcement_id}, index={att_index}, name={filename}"
                    )
                else:
                    filename_encoded = '/'.join(parts[5:])
                    if '?' in filename_encoded:
                        filename_encoded, query_string = filename_encoded.split('?', 1)
                        query_params = urllib.parse.parse_qs(query_string)
                        version_number = query_params.get('version', [None])[0]
                    try:
                        filename = urllib.parse.unquote(filename_encoded, encoding='utf-8')
                    except Exception as e:
                        logger.debug(f"解码文件名失败: {e}")
                        filename = filename_encoded

                if filename:
                    self.handle_download_attachment(announcement_id, filename, version_number)
            else:
                logger.warning(f"附件下载: 路径格式错误, path={self.path}")
                self._send_attachment_api_error(404, '无效的附件下载地址')
        
        # 获取所有历史版本API（历史公告文件夹）
        elif self.path == '/api/announcement/history' or self.path.startswith('/api/announcement/history?'):
            if not self.check_auth():
                return
            
            try:
                # 获取所有公告的历史版本列表
                all_versions = self.announcement_mgr.get_all_history_versions()
                self.send_json_response(all_versions)
            except Exception as e:
                logger.error(f"获取历史公告失败: {e}", exc_info=True)
                self.send_json_response({
                    'error': f'加载历史公告失败: {str(e)}'
                }, status=500)
        
        # 获取公告历史版本API
        elif self.path.startswith('/api/announcement/versions/'):
            if not self.check_auth():
                return
            
            # 解析路径：/api/announcement/versions/{announcement_id} 或 /api/announcement/versions/{announcement_id}/{version_number}
            path_parts = self.path.split('/')
            announcement_id = path_parts[-2] if len(path_parts) > 5 else path_parts[-1]
            version_number = path_parts[-1] if len(path_parts) > 5 else None
            
            if version_number and version_number != announcement_id:
                # 获取指定版本详情
                version = self.announcement_mgr.get_version(announcement_id, version_number)
                if version:
                    self.send_json_response(version)
                else:
                    self.send_json_response({'error': '历史版本不存在'}, status=404)
            else:
                # 获取所有版本列表
                versions = self.announcement_mgr.get_versions(announcement_id)
                self.send_json_response(versions)
        
        # 单个公告详情API - 必须放在最后
        elif self.path.startswith('/api/announcement/announcement/'):
            if not self.check_auth():
                return
            
            announcement_id = self.path.split('/')[-1]
            
            # 优先从文件读取（temp目录优先），确保获取最新内容
            # 因为编辑已发布公告时，会在temp目录创建待审批副本，需要优先读取
            announcement = self.announcement_mgr.get_announcement(announcement_id)
            
            if announcement:
                # 如果从temp目录读取到待审批版本，清除内存缓存（确保后续读取也是最新）
                if announcement.get('_is_pending_review') or announcement.get('status') in ['pending', 'draft']:
                    preloader = get_data_preloader()
                    if preloader:
                        preloader.invalidate_cache('announcements', announcement_id)
                    # 清除API缓存
                    cache_key = f"announcement_{announcement_id}"
                    api_cache.invalidate(cache_key)
                    logger.debug(f"从temp目录读取待审批公告，已清除缓存: {announcement_id}")
                else:
                    # 如果是已发布公告，可以尝试从缓存获取（但文件读取优先）
                    preloader = get_data_preloader()
                    cached_announcement = None
                    if preloader and PRELOAD_ANNOUNCEMENTS:
                        try:
                            cached_announcement = preloader.get_announcement(announcement_id)
                        except Exception:
                            pass
                    
                    # 如果缓存中的公告和文件中的公告状态一致，可以使用缓存（但文件读取已经完成，直接使用）
                    # 这里主要是为了性能优化，但为了确保数据一致性，优先使用文件读取的结果
                    logger.debug(f"从文件读取公告: {announcement_id}, status={announcement.get('status')}")
                
                # 添加二级公告栏名称
                if announcement.get('sub_board_id'):
                    from server.sub_board_manager import SubBoardManager
                    sub_board_mgr = SubBoardManager()
                    sub_board = sub_board_mgr.get_sub_board(announcement.get('board_id'), announcement.get('sub_board_id'))
                    if sub_board:
                        announcement['sub_board_name'] = sub_board['name']
                
                # 移除内部标记
                announcement.pop('_is_pending_review', None)
                
                self.send_json_response(announcement)
            else:
                self.send_json_response({'error': '公告不存在'}, status=404)
        
        # 查询用户待办任务（通过公告ID和sourceId）
        elif self.path.startswith('/api/todo/check/'):
            if not self.check_auth():
                return
            
            announcement_id = self.path.split('/')[-1]
            user = self.get_current_user()
            
            if not user:
                logger.warning(f"检查待办失败: 用户未登录")
                self.send_json_response({'success': False, 'error': '用户未登录，请先登录'}, status=401)
                return
            
            userid = user.get('userid', '')
            
            if not userid:
                # 尝试从用户数据中获取userid（如果session中没有）
                username = user.get('username', '')
                if username:
                    try:
                        db_user = self.user_manager.get_user_by_username(username)
                        if db_user and db_user.get('userid'):
                            userid = db_user.get('userid')
                            cookie_header = self.headers.get('Cookie', '')
                            cookies = {}
                            for cookie in cookie_header.split(';'):
                                cookie = cookie.strip()
                                if '=' in cookie:
                                    key, value = cookie.split('=', 1)
                                    cookies[key.strip()] = value.strip()
                            session_id = cookies.get('session_id')
                            if session_id:
                                from server.session_manager import sync_session_patch
                                sync_session_patch(
                                    session_id,
                                    {'userid': userid, 'unionid': db_user.get('unionid', '')},
                                    sessions,
                                    sessions_lock,
                                )
                                logger.info(f"已更新session中的userid: {userid}")
                    except Exception as e:
                        logger.warning(f"尝试从数据库获取userid失败: {e}")
                
                # 如果没有userid，仍然允许查看待办列表（但不能完成待办）
                if not userid:
                    logger.info(f"用户没有userid，但允许查看待办列表: username={user.get('username', 'N/A')}")
                    # 返回没有待办任务（因为没有userid无法匹配）
                    self.send_json_response({
                        'success': True,
                        'has_todo': False,
                        'done': False,
                        'status': '未完成',
                        'message': '您的账号未绑定钉钉userid，无法完成待办任务',
                        'announcement_id': announcement_id
                    })
                    return
            
            logger.info(f"检查用户待办: announcement_id={announcement_id}, userid={userid}")
            
            # 获取待办状态（todo_manager内部会优先从内存缓存读取）
            todo_status = self.todo_mgr.get_user_todo_status(announcement_id, userid)
            if todo_status:
                logger.info(f"找到用户待办记录: announcement_id={announcement_id}, userid={userid}, done={todo_status.get('done')}, status={todo_status.get('status')}")
                self.send_json_response({
                    'success': True,
                    'has_todo': True,
                    'done': todo_status.get('done', False),
                    'status': todo_status.get('status', '未完成'),
                    'task_id': todo_status.get('task_id', ''),
                    'source_id': todo_status.get('source_id', ''),
                    'announcement_id': announcement_id
                })
            else:
                logger.info(f"未找到用户待办记录: announcement_id={announcement_id}, userid={userid}")
                self.send_json_response({
                    'success': True,
                    'has_todo': False,
                    'announcement_id': announcement_id
                })
        
        # 获取公告的所有待办状态列表（所有人可见）
        elif self.path.startswith('/api/todo/list/'):
            if not self.check_auth():
                return
            
            # 路径格式：/api/todo/list/{announcement_id}
            path_parts = self.path.split('/')
            announcement_id = path_parts[-1] if len(path_parts) > 4 else None
            
            if not announcement_id:
                self.send_json_response({'success': False, 'error': '缺少announcement_id参数'}, status=400)
                return
            
            # 检查公告是否存在（所有人可见，无需权限检查）
            # 优先从内存缓存获取（避免文件I/O）
            preloader = get_data_preloader()
            announcement = None
            if preloader and PRELOAD_ANNOUNCEMENTS:
                try:
                    announcement = preloader.get_announcement(announcement_id)
                    if not announcement:
                        announcement = self.announcement_mgr.get_announcement(announcement_id)
                except Exception:
                    announcement = self.announcement_mgr.get_announcement(announcement_id)
            else:
                announcement = self.announcement_mgr.get_announcement(announcement_id)
            
            if not announcement:
                self.send_json_response({'success': False, 'error': '公告不存在'}, status=404)
                return
            
            # 优先从内存缓存获取待办记录（避免文件I/O）
            todos = None
            if preloader and PRELOAD_TODOS:
                try:
                    todos = preloader.get_todos_for_announcement_cached(announcement_id)
                    if todos is None:
                        todos = self.todo_mgr.get_all_todos(announcement_id)
                except Exception:
                    todos = self.todo_mgr.get_all_todos(announcement_id)
            else:
                # 获取所有待办记录
                todos = self.todo_mgr.get_all_todos(announcement_id)
            
            for todo in todos:
                self._enrich_todo_user_display(todo)
            
            # 统计完成情况
            total_count = len(todos)
            done_count = sum(1 for todo in todos if todo.get('done', False))
            pending_count = total_count - done_count
            
            self.send_json_response({
                'success': True,
                'todos': todos,
                'statistics': {
                    'total': total_count,
                    'done': done_count,
                    'pending': pending_count
                }
            })
        
        elif self.path == '/api/system/config/scheduled-notifications':
            if not self.check_auth(require_super_admin=True):
                return
            self.handle_get_scheduled_notifications_config()
        
        else:
            self.send_error(404)
    
    def handle_api_post(self):
        """处理API POST请求"""
        parsed_path = urllib.parse.urlparse(self.path)
        actual_path = parsed_path.path
        if not csrf_mod.is_exempt(actual_path, 'POST') and not csrf_mod.validate(self):
            self.send_json_response({'success': False, 'error': 'CSRF 校验失败'}, status=403)
            return
        if actual_path.startswith('/api/material-db/'):
            from server.material_db_api import MaterialDbApi
            MaterialDbApi(self).dispatch('POST', actual_path, parsed_path)
            return

        if actual_path.rstrip('/') == '/api/auth/dingtalk/inapp-login':
            self._handle_dingtalk_inapp_login()

        elif actual_path.rstrip('/') == '/api/auth/login':
            self.handle_login()
        
        # 通过userid登录（用于免登录）
        elif actual_path.rstrip('/') == '/api/auth/login-by-userid':
            self.handle_login_by_userid()
        elif actual_path.rstrip('/') == '/api/auth/register':
            self.handle_register()
        elif actual_path.rstrip('/') == '/api/auth/change-password':
            self.handle_auth_password(legacy_path=True)
        elif actual_path.rstrip('/') == '/api/auth/password':
            self.handle_auth_password(legacy_path=False)
        elif '/password' in actual_path and actual_path.startswith('/api/auth/users/'):
            if not self.check_auth(require_super_admin=True):
                return
            try:
                parts = actual_path.rstrip('/').split('/')
                uid = int(parts[-2])
            except (ValueError, IndexError):
                self.send_json_response({'ok': False, 'success': False, 'error': '无效的用户 ID'}, status=400)
                return
            content_length = int(self.headers.get('Content-Length', 0) or 0)
            body = self.rfile.read(content_length).decode('utf-8') if content_length else '{}'
            data = json.loads(body or '{}')
            new_password = (data.get('newPassword') or data.get('password') or '').strip()
            from server.auth.password_service import get_password_service
            ok, msg, meta = get_password_service(self.user_manager).admin_reset_password(
                uid, new_password, memory_sessions=sessions, memory_lock=sessions_lock,
            )
            if ok:
                self._invalidate_user_cache()
            self.send_json_response({
                'ok': ok, 'success': ok, 'message': msg, 'clearAutoLogin': meta.get('clearAutoLogin', True),
            } if ok else {'ok': False, 'success': False, 'error': msg})
        elif self.path == '/api/auth/users':
            if not self.check_auth(require_super_admin=True):
                return
            self.handle_add_user()
        elif self.path == '/api/auth/users/excel-debug':
            self.send_json_response({'ok': False, 'success': False, 'error': '该调试接口已移除'}, status=410)
        elif self.path == '/api/auth/sync-dingtalk-users':
            if not self.check_auth(require_super_admin=True):
                return
            self.handle_sync_dingtalk_users()
        elif self.path == '/api/auth/users/batch-status':
            if not self.check_auth(require_super_admin=True):
                return
            self.handle_batch_update_user_status()
        elif self.path == '/api/system/config/scheduled-notifications':
            if not self.check_auth(require_super_admin=True):
                return
            self.handle_set_scheduled_notifications_config()
        elif actual_path == '/api/announcement/notify-unread-readers':
            if not self.check_auth():
                return
            user = self.get_current_user()
            if not self._can_approve_announcement(user):
                self.send_json_response(
                    {'success': False, 'error': '仅公告审批人/管理组成员可发送催读通知'},
                    status=403,
                )
                return
            try:
                content_length = int(self.headers.get('Content-Length', 0))
                data = {}
                if content_length > 0:
                    data = json.loads(self.rfile.read(content_length).decode('utf-8'))
                announcement_id = (data.get('announcement_id') or '').strip() or None
                from server.announcement_read_notify import notify_pending_readers
                result = notify_pending_readers(
                    self.announcement_mgr,
                    self.todo_mgr,
                    announcement_id=announcement_id,
                    base_url=self._build_public_base_url(),
                    resolve_userids=self._resolve_dingtalk_userids,
                    filter_userids=self._filter_valid_todo_userids,
                )
                status = 200 if result.get('success') else 500
                self.send_json_response(result, status=status)
            except json.JSONDecodeError:
                self.send_json_response({'success': False, 'error': '无效的JSON数据'}, status=400)
            except Exception as e:
                log_server_error('催读通知', e)
                self.send_json_response(safe_error_payload(), status=500)
        # 公告栏API
        elif self.path == '/api/announcement/create':
            if not self.check_auth():
                return
            # 所有登录用户都可以创建公告（提交发布申请）
            
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length).decode('utf-8')
            data = json.loads(post_data)
            
            user = self.get_current_user()
            
            # 验证必要字段
            required_fields = ['title', 'content', 'board_id']
            for field in required_fields:
                if not data.get(field):
                    self.send_json_response({'success': False, 'error': f'缺少必要字段: {field}'})
                    return
            
            print(f"创建公告: 标题={data['title']}, 状态={data.get('status', 'draft')}, 用户={user['name']}, 用户ID={user.get('id')}")
            
            author_userid = str(user.get('userid') or '').strip()
            if not author_userid and user.get('username'):
                try:
                    full_user = self.user_manager.get_user_by_username(user['username'])
                    if full_user:
                        author_userid = str(full_user.get('userid') or '').strip()
                except Exception:
                    pass

            pending_approver_identifier = data.get('approver') or data.get('approver_userid')
            pending_approver_userid = None
            if pending_approver_identifier:
                pending_approver_userid = self._resolve_dingtalk_userid(pending_approver_identifier)

            announcement_id, message = self.announcement_mgr.create_announcement(
                board_id=data['board_id'],
                title=data['title'],
                content=data['content'],
                author=user['name'],
                priority=data.get('priority', 'normal'),
                status=data.get('status', 'draft'),
                attachments=data.get('attachments', []),
                sub_board_id=data.get('sub_board_id'),  # 支持二级公告栏
                user_id=user.get('id'),  # 传递用户ID，用于草稿时按用户组织
                author_userid=author_userid or None,
                pending_approver_identifier=pending_approver_identifier,
                pending_approver_userid=pending_approver_userid,
            )
            
            if announcement_id:
                # 清除公告列表缓存
                self._invalidate_announcement_cache()
                
                approval_notification_sent = None
                approval_notification_error = None
                # 如果公告状态为pending（待审批），发送工作通知给审批人
                if data.get('status') == 'pending':
                    try:
                        if pending_approver_identifier:
                            approval_notification_sent, approval_notification_error = self._notify_selected_approver(
                                announcement_id, data['title'], pending_approver_identifier
                            )
                        else:
                            approval_notification_sent, approval_notification_error = self._notify_department_heads_for_approval(
                                announcement_id, data['title']
                            )
                    except Exception as e:
                        logger.warning(f"发送审批通知失败: {e}", exc_info=True)
                        approval_notification_sent = False
                        approval_notification_error = str(e)
                
                resp = {'success': True, 'message': message, 'id': announcement_id}
                if approval_notification_sent is not None:
                    resp['approval_notification_sent'] = approval_notification_sent
                if approval_notification_error:
                    resp['approval_notification_error'] = approval_notification_error
                self.send_json_response(resp)
            else:
                self.send_json_response({'success': False, 'error': message})
        
        # 钉钉免登录API
        elif self.path == '/api/dingtalk/login':
            self.handle_dingtalk_login()
        
        # 钉钉获取AccessToken API（新API）
        elif self.path == '/api/dingtalk/get-access-token':
            self.handle_dingtalk_get_access_token()
        
        # 钉钉获取用户信息API（使用access_token和授权码）
        elif self.path == '/api/dingtalk/get-user-info':
            self.handle_dingtalk_get_user_info()
        
        # 钉钉获取部门列表API
        elif self.path == '/api/dingtalk/get-departments':
            self.handle_dingtalk_get_departments()
        
        # 钉钉获取部门用户列表API
        elif self.path == '/api/dingtalk/get-department-users':
            self.handle_dingtalk_get_department_users()
        
        elif actual_path == '/api/quick-links/refresh-icons':
            if not self.check_auth():
                return
            user = self.get_current_user()
            if not self._can_manage_quick_links(user):
                self.send_json_response(
                    {'success': False, 'error': '仅管理员、管理组成员和最高管理员可刷新快捷链接图标'},
                    status=403,
                )
                return
            try:
                self.quick_link_mgr.refresh_all_icons_on_startup()
                links = self.quick_link_mgr.get_links()
                with_icon = sum(1 for x in links if isinstance(x, dict) and x.get('icon_url'))
                self.send_json_response({
                    'success': True,
                    'message': f'图标刷新完成，{with_icon}/{len(links)} 个链接已缓存站点图标',
                    'links': links,
                })
            except Exception as e:
                logger.error('刷新快捷链接图标失败: %s', e, exc_info=True)
                self.send_json_response({'success': False, 'error': str(e)}, status=500)
            return

        # 快捷链接API - POST (添加)
        elif self.path == '/api/quick-links':
            if not self.check_auth():
                return
            
            user = self.get_current_user()
            if not self._can_manage_quick_links(user):
                self.send_json_response({'success': False, 'error': '仅管理员、管理组成员和最高管理员可以管理快捷链接'}, status=403)
                return
            
            try:
                content_length = int(self.headers.get('Content-Length', 0))
                if content_length == 0:
                    self.send_json_response({'success': False, 'error': '请求体不能为空'}, status=400)
                    return
                
                post_data = self.rfile.read(content_length).decode('utf-8')
                data = json.loads(post_data)
                
                name = data.get('name', '').strip()
                url = data.get('url', '').strip()
                icon = data.get('icon', '🔗').strip()
                description = data.get('description', '').strip()
                
                if not name or not url:
                    self.send_json_response({'success': False, 'error': '名称和URL不能为空'}, status=400)
                    return
                
                # 验证URL格式
                if not url.startswith('http://') and not url.startswith('https://'):
                    url = 'https://' + url
                
                success, message, icon_url = self.quick_link_mgr.add_link(name, url, icon, description)
                if success:
                    logger.info(f"添加快捷链接成功: {name} -> {url}")
                    payload = {'success': True, 'message': message}
                    if icon_url:
                        payload['icon_url'] = icon_url
                    self.send_json_response(payload)
                else:
                    logger.error(f"添加快捷链接失败: {message}")
                    self.send_json_response({'success': False, 'error': message}, status=400)
            except json.JSONDecodeError:
                self.send_json_response({'success': False, 'error': '无效的JSON数据'}, status=400)
            except Exception as e:
                logger.error(f"添加快捷链接异常: {e}", exc_info=True)
                self.send_json_response({'success': False, 'error': f'服务器错误: {str(e)}'}, status=500)
        
        # 一级公告栏管理API（POST - 创建/更新/删除）
        elif self.path == '/api/announcement/boards/manage':
            if not self.check_auth():
                return
            
            user = self.get_current_user()
            if not self._can_manage_boards(user):
                self.send_json_response({'error': '仅管理员、管理组成员和最高管理员可以管理公告栏'}, status=403)
                return
            
            try:
                content_length = int(self.headers.get('Content-Length', 0))
                if content_length == 0:
                    self.send_json_response({'success': False, 'error': '请求体不能为空'}, status=400)
                    return
                
                post_data = self.rfile.read(content_length).decode('utf-8')
                data = json.loads(post_data)
                
                from server.board_manager import BoardManager
                board_mgr = BoardManager()
                
                action = data.get('action')  # 'create', 'update', 'delete'
                board_id = data.get('board_id')
                name = data.get('name')
                description = data.get('description', '')
                display_order = data.get('display_order', 0)
                
                if action == 'create':
                    if not board_id or not name:
                        self.send_json_response({'success': False, 'error': '公告栏ID和名称不能为空'})
                        return
                    logger.debug(f"[API] 开始创建一级公告栏: {board_id}, name={name}")
                    try:
                        success, message = board_mgr.create_board(board_id, name, description, display_order)
                        if success:
                            logger.info(f"[API] 创建一级公告栏成功: board_id={board_id}, name={name}")
                        else:
                            logger.warning(f"[API] 创建一级公告栏失败: {message}")
                        
                        # 无论成功或失败，都返回当前的一级公告栏列表，方便前端刷新
                        try:
                            updated_boards = board_mgr.get_all_boards()
                            logger.debug(f"[API] 获取更新后的一级公告栏列表: {len(updated_boards)} 个")
                        except Exception as e:
                            logger.error(f"[API] 获取更新后的列表失败: {e}", exc_info=True)
                            updated_boards = []
                        
                        if success:
                            self.send_json_response({
                                'success': True, 
                                'message': message,
                                'updated_boards': updated_boards  # 返回更新后的列表
                            })
                        else:
                            self.send_json_response({
                                'success': False, 
                                'error': message,
                                'updated_boards': updated_boards  # 即使失败也返回当前列表
                            })
                        return  # 创建操作已发送响应，直接返回
                    except Exception as e:
                        logger.error(f"[API] 创建一级公告栏异常: {e}", exc_info=True)
                        import traceback
                        traceback.print_exc()
                        # 即使异常，也尝试返回当前列表
                        try:
                            updated_boards = board_mgr.get_all_boards()
                        except:
                            updated_boards = []
                        self.send_json_response({
                            'success': False, 
                            'error': f"创建失败: {str(e)}",
                            'updated_boards': updated_boards
                        })
                        return  # 异常处理已发送响应，直接返回
                elif action == 'update':
                    if not board_id:
                        self.send_json_response({'success': False, 'error': '公告栏ID不能为空'})
                        return
                    success, message = board_mgr.update_board(board_id, name, description, display_order)
                elif action == 'delete':
                    if not board_id:
                        self.send_json_response({'success': False, 'error': '公告栏ID不能为空'})
                        return
                    
                    logger.debug(f"[API] 开始删除一级公告栏: {board_id}")
                    try:
                        # 所有删除逻辑都在后端处理，前端只发送命令
                        success, message = board_mgr.delete_board(board_id)
                        if success:
                            logger.info(f"[API] 删除一级公告栏成功: board_id={board_id}")
                        else:
                            logger.warning(f"[API] 删除一级公告栏失败: {message}")
                        
                        # 无论成功或失败，都返回当前的一级公告栏列表，方便前端刷新
                        try:
                            updated_boards = board_mgr.get_all_boards()
                            logger.debug(f"[API] 获取更新后的一级公告栏列表: {len(updated_boards)} 个")
                        except Exception as e:
                            logger.error(f"[API] 获取更新后的列表失败: {e}", exc_info=True)
                            updated_boards = []
                        
                        if success:
                            self.send_json_response({
                                'success': True, 
                                'message': message,
                                'updated_boards': updated_boards  # 返回更新后的列表
                            })
                        else:
                            self.send_json_response({
                                'success': False, 
                                'error': message,
                                'updated_boards': updated_boards  # 即使失败也返回当前列表
                            })
                        return  # 删除操作已发送响应，直接返回
                    except Exception as e:
                        logger.error(f"[API] 删除一级公告栏异常: {e}", exc_info=True)
                        import traceback
                        traceback.print_exc()
                        # 即使异常，也尝试返回当前列表
                        try:
                            updated_boards = board_mgr.get_all_boards()
                        except:
                            updated_boards = []
                        self.send_json_response({
                            'success': False, 
                            'error': f"删除失败: {str(e)}",
                            'updated_boards': updated_boards
                        })
                        return  # 异常处理已发送响应，直接返回
                else:
                    self.send_json_response({'success': False, 'error': '无效的操作'})
                    return
                
                self.send_json_response({'success': success, 'message': message})
            except Exception as e:
                logger.error(f"管理一级公告栏失败: {e}", exc_info=True)
                import traceback
                traceback.print_exc()
                self.send_json_response({'success': False, 'error': f'操作失败: {str(e)}'}, status=500)
        
        # 二级公告栏管理API（POST - 创建/更新）
        elif self.path == '/api/announcement/sub-boards':
            if not self.check_auth():
                return
            
            user = self.get_current_user()
            if not self._can_manage_boards(user):
                self.send_json_response({'error': '仅管理员、管理组成员和最高管理员可以管理二级公告栏'}, status=403)
                return
            
            try:
                content_length = int(self.headers.get('Content-Length', 0))
                if content_length == 0:
                    self.send_json_response({'success': False, 'error': '请求体不能为空'}, status=400)
                    return
                
                post_data = self.rfile.read(content_length).decode('utf-8')
                print(f"收到二级公告栏创建请求: {post_data}")
                data = json.loads(post_data)
                
                from server.sub_board_manager import SubBoardManager
                sub_board_mgr = SubBoardManager()
                
                action = data.get('action')  # 'create' 或 'update'
                parent_board_id = data.get('parent_board_id')
                sub_board_id = data.get('sub_board_id')
                name = data.get('name')
                description = data.get('description', '')
                display_order = data.get('display_order', 0)
                
                print(f"处理二级公告栏请求: action={action}, parent={parent_board_id}, sub={sub_board_id}, name={name}")
                
                # 如果是创建且没有提供sub_board_id，自动生成
                if action == 'create' and not sub_board_id:
                    import re
                    # 基于名称生成ID：转换为小写，替换空格为下划线，移除特殊字符
                    sub_board_id = re.sub(r'[^a-z0-9_]', '', name.lower().replace(' ', '_').replace('-', '_'))
                    # 如果生成后为空，使用时间戳
                    if not sub_board_id:
                        import time
                        sub_board_id = f'sub_{int(time.time())}'
                    print(f"自动生成二级公告栏ID: {sub_board_id}")
                
                if action == 'create':
                    if not parent_board_id or not name:
                        self.send_json_response({'success': False, 'error': '缺少必要字段：父公告栏和名称'}, status=400)
                        return
                    success, message = sub_board_mgr.create_sub_board(
                        parent_board_id, sub_board_id, name, description, display_order
                    )
                    print(f"创建二级公告栏结果: success={success}, message={message}")
                elif action == 'update':
                    if not parent_board_id or not sub_board_id:
                        self.send_json_response({'success': False, 'error': '缺少必要字段'}, status=400)
                        return
                    success, message = sub_board_mgr.update_sub_board(
                        parent_board_id, sub_board_id, name, description, display_order
                    )
                    print(f"更新二级公告栏结果: success={success}, message={message}")
                else:
                    self.send_json_response({'success': False, 'error': '无效的操作'}, status=400)
                    return
                
                if success:
                    print(f"✅ 二级公告栏创建/更新成功，返回成功响应")
                    self.send_json_response({'success': True, 'message': message})
                else:
                    print(f"❌ 二级公告栏创建/更新失败: {message}")
                    self.send_json_response({'success': False, 'error': message}, status=400)
            except json.JSONDecodeError as e:
                print(f"JSON解析错误: {e}")
                traceback.print_exc()
                self.send_json_response({'success': False, 'error': '无效的JSON数据'}, status=400)
            except Exception as e:
                print(f"处理二级公告栏请求时出错: {e}")
                traceback.print_exc()
                self.send_json_response({'success': False, 'error': f'服务器错误: {str(e)}'}, status=500)
        
        elif self.path.startswith('/api/announcement/update/'):
            if not self.check_auth():
                return
            
            announcement_id = self.path.split('/')[-1]
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length).decode('utf-8')
            data = json.loads(post_data)
            
            # 获取当前用户作为编辑人
            user = self.get_current_user()
            editor = user.get('name') or user.get('username', '未知')
            
            selected_approver = data.get('approver') or data.get('approver_userid')
            if selected_approver:
                data['pending_approver_identifier'] = selected_approver
                resolved_uid = self._resolve_dingtalk_userid(selected_approver)
                if resolved_uid:
                    data['pending_approver_userid'] = resolved_uid

            # 传递editor参数
            success, message = self.announcement_mgr.update_announcement(announcement_id, editor=editor, **data)
            
            approval_notification_sent = None
            approval_notification_error = None
            # 更新成功后，清除相关缓存（确保后续读取获取最新数据）
            if success:
                # 检查是否创建了待审批副本（编辑已发布公告时）
                is_pending_review_created = "已创建待审批副本" in message
                
                # 获取更新后的公告信息（在清除缓存之前读取）
                updated_announcement = self.announcement_mgr.get_announcement(announcement_id)
                if updated_announcement:
                    title = updated_announcement.get('title', '')
                    new_status = updated_announcement.get('status', '')
                    
                    # 所有提交审核的情况（状态变为pending）都要发送审批通知给部长
                    # 包括：
                    # 1. 草稿编辑后提交审核（draft -> pending）
                    # 2. 编辑已发布公告后提交审核（approved -> pending，创建待审批副本）
                    # 3. 其他任何状态变为pending的情况
                    if new_status == 'pending' or is_pending_review_created:
                        if is_pending_review_created:
                            logger.info(f"检测到已创建待审批副本（编辑已发布公告），准备发送审批通知: {announcement_id}, 标题: {title}")
                        else:
                            logger.info(f"检测到公告状态变为pending（提交审核），准备发送审批通知: {announcement_id}, 标题: {title}")
                        
                        if selected_approver:
                            logger.info(f"使用用户选择的审批人: {selected_approver}")
                            approval_notification_sent, approval_notification_error = self._notify_selected_approver(
                                announcement_id, title, selected_approver
                            )
                        else:
                            logger.info(f"未指定审批人，使用默认审批人（管理组）")
                            approval_notification_sent, approval_notification_error = self._notify_department_heads_for_approval(
                                announcement_id, title
                            )
                else:
                    # 如果无法获取公告信息，但创建了待审批副本，仍然发送通知
                    if is_pending_review_created:
                        logger.info(f"检测到已创建待审批副本（编辑已发布公告），但无法获取公告信息，使用默认标题发送审批通知: {announcement_id}")
                        # 尝试从data中获取标题
                        title = data.get('title', '待审批公告')
                        if selected_approver:
                            logger.info(f"使用用户选择的审批人: {selected_approver}")
                            approval_notification_sent, approval_notification_error = self._notify_selected_approver(
                                announcement_id, title, selected_approver
                            )
                        else:
                            logger.info(f"未指定审批人，使用默认审批人（管理组）")
                            approval_notification_sent, approval_notification_error = self._notify_department_heads_for_approval(
                                announcement_id, title
                            )
                    else:
                        logger.warning(f"无法获取更新后的公告信息，跳过发送审批通知: {announcement_id}")
                
                # 清除该公告的内存缓存
                preloader = get_data_preloader()
                if preloader:
                    preloader.invalidate_cache('announcements', announcement_id)
                
                # 清除API缓存
                # 注意：使用RLock避免死锁，invalidate内部也会获取锁
                with api_cache.lock:
                    keys_to_remove = [key for key in list(api_cache.cache.keys()) if announcement_id in key]
                    # 直接调用_remove_key避免重复获取锁（因为已经在锁内）
                    for key in keys_to_remove:
                        api_cache._remove_key(key)
                
                # 清除公告列表缓存（因为可能有新的待审批公告）
                self._invalidate_announcement_cache()
            
            resp = {'success': success, 'message': message}
            if approval_notification_sent is not None:
                resp['approval_notification_sent'] = approval_notification_sent
            if approval_notification_error:
                resp['approval_notification_error'] = approval_notification_error
            self.send_json_response(resp)
        
        # 快捷链接API - POST (添加)
        elif self.path == '/api/quick-links':
            if not self.check_auth():
                return
            
            user = self.get_current_user()
            if not self._can_manage_quick_links(user):
                self.send_json_response({'success': False, 'error': '仅管理员、管理组成员和最高管理员可以管理快捷链接'}, status=403)
                return
            
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length).decode('utf-8')
            data = json.loads(post_data)
            
            name = data.get('name', '').strip()
            url = data.get('url', '').strip()
            icon = data.get('icon', '🔗').strip()
            description = data.get('description', '').strip()
            
            if not name or not url:
                self.send_json_response({'success': False, 'error': '名称和URL不能为空'}, status=400)
                return
            
            # 验证URL格式
            if not url.startswith('http://') and not url.startswith('https://'):
                url = 'https://' + url
            
            success, message, icon_url = self.quick_link_mgr.add_link(name, url, icon, description)
            if success:
                payload = {'success': True, 'message': message}
                if icon_url:
                    payload['icon_url'] = icon_url
                self.send_json_response(payload)
            else:
                self.send_json_response({'success': False, 'error': message}, status=400)
        
        else:
            self.send_error(404)
    
    def handle_api_put(self):
        """处理PUT请求"""
        parsed_path = urllib.parse.urlparse(self.path)
        actual_path = parsed_path.path
        if not csrf_mod.is_exempt(actual_path, 'PUT') and not csrf_mod.validate(self):
            self.send_json_response({'success': False, 'error': 'CSRF 校验失败'}, status=403)
            return
        if actual_path.startswith('/api/material-db/'):
            from server.material_db_api import MaterialDbApi
            MaterialDbApi(self).dispatch('PUT', actual_path, parsed_path)
            return

        # 快捷链接 API - PUT（更新名称 / URL / 备注等）
        if actual_path.startswith('/api/quick-links/'):
            link_id = actual_path.rstrip('/').split('/')[-1]
            if link_id in ('quick-links', 'refresh-icons'):
                self.send_error(404)
                return
            if not self.check_auth():
                return
            user = self.get_current_user()
            if not self._can_manage_quick_links(user):
                self.send_json_response(
                    {'success': False, 'error': '仅管理员、管理组成员和最高管理员可以管理快捷链接'},
                    status=403,
                )
                return
            try:
                content_length = int(self.headers.get('Content-Length', 0))
                if content_length == 0:
                    self.send_json_response({'success': False, 'error': '请求体不能为空'}, status=400)
                    return
                put_data = self.rfile.read(content_length).decode('utf-8')
                data = json.loads(put_data)
                name = data.get('name')
                url = data.get('url')
                icon = data.get('icon')
                description = data.get('description')
                if name is not None:
                    name = str(name).strip()
                if url is not None:
                    url = str(url).strip()
                    if url and not url.startswith('http://') and not url.startswith('https://'):
                        url = 'https://' + url
                if icon is not None:
                    icon = str(icon).strip()
                if description is not None:
                    description = str(description).strip()
                if name is not None and not name:
                    self.send_json_response({'success': False, 'error': '名称不能为空'}, status=400)
                    return
                if url is not None and not url:
                    self.send_json_response({'success': False, 'error': 'URL不能为空'}, status=400)
                    return
                success, message = self.quick_link_mgr.update_link(
                    link_id,
                    name=name if 'name' in data else None,
                    url=url if 'url' in data else None,
                    icon=icon if 'icon' in data else None,
                    description=description if 'description' in data else None,
                )
                if success:
                    self.send_json_response({'success': True, 'message': message})
                else:
                    self.send_json_response({'success': False, 'error': message}, status=400)
            except json.JSONDecodeError:
                self.send_json_response({'success': False, 'error': '无效的JSON数据'}, status=400)
            except Exception as e:
                logger.error('更新快捷链接异常: %s', e, exc_info=True)
                self.send_json_response({'success': False, 'error': f'服务器错误: {str(e)}'}, status=500)
            return

        if self.path.startswith('/api/auth/users/'):
            if not self.check_auth(require_super_admin=True):
                return
            
            path_only = actual_path.rstrip('/')
            try:
                user_id = int(path_only.split('/')[-1])
            except ValueError:
                self.send_json_response({'success': False, 'message': '无效的用户 ID'}, status=400)
                return
            content_length = int(self.headers.get('Content-Length', 0))
            put_data = self.rfile.read(content_length).decode('utf-8')
            data = json.loads(put_data)

            # 先获取用户信息，用于合并角色和判断用户类型
            existing_user = self.user_manager.get_user_by_id(user_id)
            if not existing_user:
                self.send_json_response({'success': False, 'message': '用户不存在'})
                return
            
            # 处理 roles：账号页提交完整岗位+系统角色，支持显式撤销 admin/management
            if 'roles' in data:
                roles = data['roles']
                if isinstance(roles, str):
                    roles = [roles]
                elif not isinstance(roles, list):
                    roles = []
                data['roles'] = self.user_manager.apply_role_update(existing_user.get('roles', []), roles)
            
            # 处理 library_roles（仅保留当前 material-db 中存在的库 id）
            if 'library_roles' in data:
                library_roles = data['library_roles']
                if isinstance(library_roles, str):
                    library_roles = [library_roles] if library_roles else []
                elif not isinstance(library_roles, list):
                    library_roles = []
                data['library_roles'] = self.user_manager.sanitize_library_roles_list(library_roles)
            
            # 对于钉钉用户，如果提供了department但没有title，将department同时映射到title和job_position
            if existing_user.get('source') == 'dingtalk' and 'department' in data and 'title' not in data:
                data['title'] = data['department']
                data['job_position'] = data['department']
            
            # 记录更新前的用户数据（用于调试，不记录明文密码）
            log_payload = dict(data)
            if log_payload.get('password'):
                log_payload['password'] = '***'
            logger.info(f"[用户更新] 开始更新用户 ID={user_id}, 数据={json.dumps(log_payload, ensure_ascii=False)}")

            pwd_msg = None
            if data.get('password'):
                from server.auth.password_service import get_password_service
                ok_pwd, pwd_msg, _ = get_password_service(self.user_manager).admin_reset_password(
                    user_id,
                    str(data.pop('password')),
                    memory_sessions=sessions,
                    memory_lock=sessions_lock,
                )
                if not ok_pwd:
                    self.send_json_response({'success': False, 'message': pwd_msg})
                    return

            success, message = self.user_manager.update_user(user_id, data)
            if pwd_msg and success:
                message = f"{message}；{pwd_msg}"
            
            if success:
                # 文件已更新，清除用户缓存（确保内存与文件一致）
                self._invalidate_user_cache()
                
                # 强制清除预加载器缓存，确保下次从数据库重新加载
                preloader = get_data_preloader()
                if preloader:
                    preloader.invalidate_cache('users')
                    logger.info(f"[用户更新] 已清除预加载器用户缓存（后台刷新）")
                    preloader.schedule_users_reload()
                
                # 验证更新后的数据（从MySQL直接读取）
                try:
                    from server.db_adapter import get_connection_pool
                    pool = get_connection_pool()
                    with pool.get_cursor() as cursor:
                        cursor.execute('SELECT id, name, department, status, job_position FROM users WHERE id = %s', (user_id,))
                        verify_row = cursor.fetchone()
                        if verify_row:
                            if isinstance(verify_row, dict):
                                verify_data = verify_row
                            else:
                                columns = [desc[0] for desc in cursor.description] if hasattr(cursor, 'description') else []
                                verify_data = dict(zip(columns, verify_row)) if columns else {}
                            logger.info(f"[用户更新] ✅ MySQL验证: ID={verify_data.get('id')}, 姓名={verify_data.get('name')}, 部门={verify_data.get('department')}, 状态={verify_data.get('status')}")
                except Exception as e:
                    logger.warning(f"[用户更新] MySQL验证失败: {e}", exc_info=True)
            else:
                logger.error(f"[用户更新] ❌ 更新失败: {message}")
            
            self.send_json_response({'success': success, 'message': message})
        
        elif self.path.startswith('/api/auth/registrations/'):
            if not self.check_auth(required_roles=['management']):
                return
            
            registration_id = int(self.path.split('/')[-1])
            content_length = int(self.headers.get('Content-Length', 0))
            put_data = self.rfile.read(content_length).decode('utf-8')
            data = json.loads(put_data)
            
            action = data.get('action')
            roles = data.get('roles', [])
            if isinstance(roles, str):
                roles = [roles]
            
            if action == 'approve':
                if roles:
                    role_success, role_message = self.user_manager.update_job_roles(registration_id, roles)
                    if not role_success:
                        self.send_json_response({'success': False, 'message': role_message})
                        return
                success, message = self.user_manager.approve_registration(registration_id, approve=True)
            elif action == 'reject':
                success, message = self.user_manager.approve_registration(registration_id, approve=False)
            else:
                success, message = False, "无效的操作"
            
            self.send_json_response({'success': success, 'message': message})
        
        # 获取按title分组的用户列表（用于待办人员选择）
        elif self.path == '/api/users/by-title' or actual_path == '/api/auth/users/by-title':
            logger.info("收到获取按title分组的用户列表请求")
            
            if not self.check_auth():
                logger.warning("获取用户列表失败: 用户未认证")
                return
            
            user = self.get_current_user()
            logger.info(f"当前用户: {user.get('username', 'unknown')}, 角色: {user.get('roles', [])}")
            
            # 只有管理员、管理组和最高管理员可以查看用户列表
            if not (self._is_super_admin(user) or self._has_role(user, 'management') or self._has_role(user, 'admin')):
                logger.warning(f"用户 {user.get('username', 'unknown')} 无权限访问用户列表")
                self.send_json_response({'error': '仅管理员、管理组成员和最高管理员可以查看用户列表'}, status=403)
                return
            
            try:
                logger.info("开始获取所有激活状态的用户...")
                # 优先从内存缓存获取（避免文件I/O）
                preloader = get_data_preloader()
                if preloader and PRELOAD_USERS:
                    try:
                        all_users = preloader.get_users(status='active')
                        logger.info(f"从内存缓存获取到 {len(all_users)} 个激活用户")
                    except Exception as e:
                        logger.warning(f"从内存缓存获取用户失败，使用文件读取: {e}")
                        all_users = self.user_manager.get_all_users(statuses=[STATUS_ACTIVE])
                else:
                    # 如果未启用预加载，从文件读取
                    all_users = self.user_manager.get_all_users(statuses=[STATUS_ACTIVE])
                logger.info(f"获取到 {len(all_users)} 个激活用户（状态='active'）")
                
                # 按title分组
                users_by_title = {}
                for u in all_users:
                    if self._exclude_from_todo_user_selection(u):
                        continue
                    title = u.get('title', '') or '未设置岗位'
                    if title not in users_by_title:
                        users_by_title[title] = []
                    users_by_title[title].append({
                        'userid': u.get('userid', ''),
                        'name': u.get('name', ''),
                        'title': title,
                        'job_number': u.get('job_number', '')
                    })
                
                # 按title名称排序
                sorted_titles = sorted(users_by_title.keys())
                result = {
                    'titles': sorted_titles,
                    'users_by_title': {title: users_by_title[title] for title in sorted_titles}
                }
                
                logger.info(f"按title分组完成，共 {len(sorted_titles)} 个岗位组")
                self.send_json_response(result)
            except Exception as e:
                logger.error(f"获取按title分组的用户列表失败: {e}", exc_info=True)
                self.send_json_response({'error': f'获取用户列表失败: {str(e)}'}, status=500)
        
        # 公告栏API
        elif self.path.startswith('/api/announcement/approve/'):
            if not self.check_auth():
                return
            
            user = self.get_current_user()
            if not self._can_approve_announcement(user):
                self.send_json_response({'error': '您没有公告审批权限'}, status=403)
                return
            
            announcement_id = self.path.split('/')[-1]
            content_length = int(self.headers.get('Content-Length', 0))
            put_data = self.rfile.read(content_length).decode('utf-8')
            data = json.loads(put_data)
            
            action = data.get('action')  # 'approve' or 'reject'
            comment = data.get('comment', '')
            todo_users = self._filter_valid_todo_userids(data.get('todo_users', []))
            create_todo = data.get('create_todo', False)  # 是否创建待办任务（默认False）
            knowledge_doc_error = None
            notification_warning = None
            read_notification_sent = None
            read_notification_warning = None
            
            # 获取当前用户作为审核人
            approver = user.get('name') or user.get('username', '未知')
            
            logger.info(f"审批公告: id={announcement_id}, action={action}, approver={approver}, todo_users={len(todo_users)}人, create_todo={create_todo}")
            
            success, message = self.announcement_mgr.approve_announcement(
                announcement_id, 
                approve=(action == 'approve'),
                comment=comment,
                approver=approver
            )
            
            # 如果审批通过，处理待办任务和通知
            todo_error = None
            source_id = None
            user_task_ids = {}
            user_source_ids = {}
            notification_sent = False
            access_token = None  # 初始化access_token变量
            
            if success and action == 'approve':
                # 审批通过后，立即清除缓存，确保后续读取获取最新数据
                preloader = get_data_preloader()
                if preloader:
                    preloader.invalidate_cache('announcements', announcement_id)
                    preloader.reload_cache('announcements')
                    logger.info(f"公告审批后已清除并重新加载缓存: {announcement_id}")
                
                # 清除API缓存
                with api_cache.lock:
                    keys_to_remove = [key for key in list(api_cache.cache.keys()) if announcement_id in key]
                    for key in keys_to_remove:
                        api_cache._remove_key(key)
                
                # 确保用户对象包含完整的userid信息（如果session中没有，从数据库查找）
                if not user.get('userid'):
                    username = user.get('username', '')
                    if username:
                        try:
                            full_user = self.user_manager.get_user_by_username(username)
                            if full_user and full_user.get('userid'):
                                user['userid'] = full_user.get('userid')
                                user['unionid'] = full_user.get('unionid', '')
                                logger.info(f"审批时补充用户userid: {username} -> {user['userid']}")
                        except Exception as e:
                            logger.warning(f"审批时从数据库获取userid失败: {e}")
                
                # 获取公告信息，用于发送通知（在清除缓存后重新读取）
                announcement = self.announcement_mgr.get_announcement(announcement_id)
                if not announcement:
                    logger.warning(f"无法获取公告信息: {announcement_id}")
                else:
                    title = announcement.get('title', '')
                    
                    # 待办人员：创建本地记录并发送阅读工作通知（与知识库、create_todo 解耦）
                    source_id = None
                    user_task_ids = {}
                    user_source_ids = {}
                    failed_userids = []
                    if todo_users:
                        if create_todo:
                            logger.info(f"✅ 开始为公告 {announcement_id} 创建待办任务，待办用户数: {len(todo_users)}")
                            try:
                                source_id, user_task_ids, user_source_ids, _, failed_userids = self._create_announcement_todos(
                                    announcement_id, todo_users, user, skip_notification=True
                                )
                                if source_id and user_task_ids:
                                    logger.info(f"✅ 成功为公告 {announcement_id} 创建待办任务，基础sourceId: {source_id}, taskIds: {len(user_task_ids)}个")
                                elif source_id:
                                    logger.info(
                                        f"✅ 已为公告 {announcement_id} 创建本地待办 Excel（sourceId: {source_id}），"
                                        f"当前不创建钉钉待办 taskId（共 {len(todo_users)} 人）"
                                    )
                                else:
                                    logger.warning(f"⚠️ 为公告 {announcement_id} 创建待办 Excel 失败，但不影响审批结果")
                            except Exception as e:
                                logger.error(f"❌ 创建待办任务时发生错误: {e}", exc_info=True)
                        else:
                            logger.info(f"ℹ️ 审批时未勾选创建待办，跳过待办 Excel 预创建（将发送阅读通知）")

                        try:
                            if not access_token:
                                access_token = self._get_dingtalk_access_token_simple()
                            read_notification_sent, read_err = self._notify_announcement_readers(
                                announcement_id, title, todo_users, access_token
                            )
                            if read_notification_sent:
                                logger.info(f"成功为公告 {announcement_id} 发送阅读通知给 {len(todo_users)} 人")
                            else:
                                read_notification_warning = read_err or '发送阅读通知失败'
                                notification_warning = read_notification_warning
                                logger.warning(f"为公告 {announcement_id} 发送阅读通知失败: {read_notification_warning}")
                        except Exception as e:
                            read_notification_warning = str(e)
                            notification_warning = read_notification_warning
                            logger.error(f"发送阅读通知时发生错误: {e}", exc_info=True)

                        # 在本地 Excel 中记录待办信息（未走 _create_announcement_todos 时）
                        if todo_users and not (create_todo and source_id):
                            try:
                                # 获取所有待办用户的unionid和用户信息（包括创建失败的用户）
                                unionids = []
                                valid_userids = []
                                user_names = []  # 存储用户姓名
                                user_usernames = []  # 存储用户名
                                for userid in todo_users:
                                    unionid = self._get_user_unionid_by_userid(userid)
                                    # 即使unionid获取失败，也记录到Excel（使用空unionid）
                                    unionids.append(unionid if unionid else '')
                                    valid_userids.append(userid)
                                    
                                    # 获取用户信息（姓名和用户名）
                                    user_info = self.user_manager.get_user_by_userid(userid)
                                    if user_info:
                                        user_name = user_info.get('name', '') or user_info.get('username', '') or str(userid)
                                        user_username = user_info.get('username', '') or str(userid)
                                    else:
                                        # 如果获取失败，使用userid作为默认值
                                        user_name = str(userid)
                                        user_username = str(userid)
                                        logger.warning(f"用户 {userid} 的信息获取失败，使用userid作为默认值")
                                    
                                    user_names.append(user_name)
                                    user_usernames.append(user_username)
                                    
                                    if not unionid:
                                        logger.warning(f"用户 {userid} 的unionid获取失败，但仍会创建Excel记录（unionid为空）")
                                
                                # 创建本地Excel文件存储待办状态
                                # 即使钉钉创建失败（user_task_ids为空），也要记录待办信息（task_id为空表示未创建）
                                # 即使unionid获取失败，也要记录待办信息（unionid为空）
                                if len(valid_userids) > 0 and len(unionids) == len(valid_userids):
                                    # 使用基础source_id（如果不存在则生成一个）
                                    if not source_id:
                                        import time
                                        source_id = f"announcement_{announcement_id}_{int(time.time() * 1000)}"
                                    
                                    # 创建Excel文件，即使task_id为空也要记录（用于显示统计）
                                    # 传入用户名和姓名列表，确保Excel文件中保存了用户信息
                                    todo_success = self.todo_mgr.create_todo_file(
                                        announcement_id, 
                                        source_id, 
                                        title, 
                                        valid_userids,  # 所有选择的待办用户
                                        unionids,       # 与valid_userids一一对应
                                        task_ids=user_task_ids if user_task_ids else {},  # 只有成功创建的才有task_id
                                        user_source_ids=user_source_ids if user_source_ids else {},  # 用户独立的sourceId映射
                                        user_names=user_names if user_names else None,  # 用户姓名列表
                                        user_usernames=user_usernames if user_usernames else None  # 用户名列表
                                    )
                                    if todo_success:
                                        logger.info(f"成功创建待办Excel文件: {announcement_id}, 共 {len(valid_userids)} 条记录（其中 {len(user_task_ids)} 条已成功创建钉钉待办）")
                                        # 清除待办缓存（文件已更新，内存缓存需要失效）
                                        self._invalidate_todo_cache(announcement_id)
                                    else:
                                        logger.warning(f"创建待办Excel文件失败: {announcement_id}")
                                else:
                                    logger.warning(f"用户列表和unionid列表长度不一致或为空，跳过创建Excel: announcement_id={announcement_id}, userids={len(valid_userids)}, unionids={len(unionids)}")
                            except Exception as e:
                                logger.warning(f"创建待办Excel文件失败: {e}", exc_info=True)
                    else:
                        read_notification_warning = '未选择阅读人，未发送阅读工作通知'
                        logger.info(f"公告 {announcement_id} 审批通过但未选择阅读人员，跳过阅读通知")

                    # 创建知识库文档（仅在首次审批通过时创建，避免重复创建）
                    try:
                        existing_doc_url = announcement.get('knowledge_doc_url')
                        if not access_token:
                            try:
                                access_token = self._get_dingtalk_access_token_simple()
                            except Exception as e:
                                logger.warning(f"获取access_token失败，跳过创建知识库文档: {e}")
                                access_token = None

                        if not access_token:
                            config_ok, config_err = check_dingtalk_config()
                            knowledge_doc_error = config_err if not config_ok else '无法获取 access_token，跳过创建知识库文档'
                        else:
                            author_name = announcement.get('author', '') or announcement.get('original_author', '')
                            if not author_name and not announcement.get('author_userid'):
                                knowledge_doc_error = '无法获取公告发起人信息，跳过创建知识库文档'
                            else:
                                if existing_doc_url:
                                    logger.info(f"公告 {announcement_id} 已有知识库文档URL: {existing_doc_url}")
                                    doc_url = existing_doc_url
                                else:
                                    doc_url, doc_err = self._create_knowledge_doc(announcement, access_token)
                                    if doc_err:
                                        knowledge_doc_error = doc_err
                                    if doc_url:
                                        self.announcement_mgr.update_announcement(
                                            announcement_id,
                                            knowledge_doc_url=doc_url
                                        )
                                        logger.info(f"成功创建知识库文档并保存URL: {announcement_id}, url={doc_url}")
                                    elif not doc_err:
                                        knowledge_doc_error = '创建知识库文档失败（未知原因）'

                                if doc_url and not knowledge_doc_error:
                                    _, notify_userid, _ = self._resolve_knowledge_doc_operator(announcement)
                                    if notify_userid:
                                        notification_text = (
                                            f"请复制添加公告内容到此文档内保存，方便AI学习。\n\n"
                                            f"您的公告《{title}》已通过审批并发布。"
                                        )
                                        if existing_doc_url:
                                            notification_text += "请更新知识库文档内容，点击查看文档。"
                                        else:
                                            notification_text += "知识库文档已创建，点击查看文档。"
                                        doc_notify_ok, doc_notify_err = self._send_announcement_notification(
                                            announcement_id,
                                            f"知识库文档{'更新' if existing_doc_url else '已创建'}：{title}",
                                            doc_url,
                                            [notify_userid],
                                            access_token,
                                            text=notification_text
                                        )
                                        if not doc_notify_ok:
                                            logger.warning(f"发送知识库文档通知失败: {doc_notify_err}")
                                    else:
                                        logger.warning(f"无法解析发起人 userid，跳过知识库文档通知: {author_name}")
                    except Exception as e:
                        knowledge_doc_error = str(e)
                        logger.warning(f"创建知识库文档时发生错误: {e}", exc_info=True)
            
            # 如果有待办/通知/知识库提示，在响应中说明（不影响审批成功与否）
            if todo_error and success:
                message = f"{message}（注意：{todo_error}）"
            if read_notification_warning and success and action == 'approve' and not todo_users:
                message = f"{message}（{read_notification_warning}）"
            if notification_warning and success:
                message = f"{message}（通知：{notification_warning}）"
            if knowledge_doc_error and success:
                message = f"{message}（知识库：{knowledge_doc_error}）"
            
            # 清除公告相关缓存（在审批完成后，确保所有worker进程都能看到最新数据）
            if success:
                try:
                    # 清除特定公告的缓存
                    preloader = get_data_preloader()
                    if preloader:
                        preloader.invalidate_cache('announcements', key=announcement_id)
                    
                    # 清除API缓存中与该公告相关的所有键
                    with api_cache.lock:
                        keys_to_remove = [key for key in list(api_cache.cache.keys()) if announcement_id in key]
                        for key in keys_to_remove:
                            api_cache._remove_key(key)
                    
                    # 清除公告列表缓存（因为状态从pending变为approved）
                    self._invalidate_announcement_cache()
                    logger.info(f"公告审批后已清除并重新加载缓存: {announcement_id}")
                except Exception as e:
                    logger.warning(f"清除公告缓存失败: {e}")
            
            resp = {'success': success, 'message': message}
            if notification_warning:
                resp['notification_warning'] = notification_warning
            if read_notification_sent is not None:
                resp['read_notification_sent'] = read_notification_sent
            if read_notification_warning:
                resp['read_notification_warning'] = read_notification_warning
            if knowledge_doc_error:
                resp['knowledge_doc_error'] = knowledge_doc_error
            self.send_json_response(resp)
        
        # 获取公告的所有待办状态列表（管理员或公告创建者可见）
        elif self.path.startswith('/api/todo/list/'):
            if not self.check_auth():
                return
            
            # 路径格式：/api/todo/list/{announcement_id}
            path_parts = self.path.split('/')
            announcement_id = path_parts[-1] if len(path_parts) > 4 else None
            
            if not announcement_id:
                self.send_json_response({'success': False, 'error': '缺少announcement_id参数'}, status=400)
                return
            
            # 检查权限：只有管理员或公告创建者可以查看待办列表
            user = self.get_current_user()
            roles = user.get('roles', [])
            is_admin = any(role in ['admin', 'super_admin', 'management'] for role in roles)
            
            # 获取公告信息，检查是否是创建者
            announcement = self.announcement_mgr.get_announcement(announcement_id)
            if not announcement:
                self.send_json_response({'success': False, 'error': '公告不存在'}, status=404)
                return
            
            author = announcement.get('author', '')
            user_name = user.get('name', '') or user.get('username', '')
            is_author = user_name == author
            
            if not is_admin and not is_author:
                self.send_json_response({'success': False, 'error': '无权查看待办列表'}, status=403)
                return
            
            # 获取所有待办记录
            todos = self.todo_mgr.get_all_todos(announcement_id)
            
            for todo in todos:
                self._enrich_todo_user_display(todo)
            
            # 统计完成情况
            total_count = len(todos)
            done_count = sum(1 for todo in todos if todo.get('done', False))
            pending_count = total_count - done_count
            
            self.send_json_response({
                'success': True,
                'todos': todos,
                'statistics': {
                    'total': total_count,
                    'done': done_count,
                    'pending': pending_count
                }
            })
        
        # 重试创建失败的待办任务
        elif self.path.startswith('/api/todo/retry/'):
            if not self.check_auth():
                return
            
            # 路径格式：/api/todo/retry/{announcement_id}
            path_parts = self.path.split('/')
            announcement_id = path_parts[-1] if len(path_parts) > 4 else None
            
            if not announcement_id:
                self.send_json_response({'success': False, 'error': '缺少announcement_id参数'}, status=400)
                return
            
            # 检查权限：只有管理员或公告创建者可以重试创建待办
            user = self.get_current_user()
            roles = user.get('roles', [])
            is_admin = any(role in ['admin', 'super_admin', 'management'] for role in roles)
            
            # 获取公告信息，检查是否是创建者
            announcement = self.announcement_mgr.get_announcement(announcement_id)
            if not announcement:
                self.send_json_response({'success': False, 'error': '公告不存在'}, status=404)
                return
            
            author = announcement.get('author', '')
            user_name = user.get('name', '') or user.get('username', '')
            is_author = user_name == author
            
            if not is_admin and not is_author:
                self.send_json_response({'success': False, 'error': '无权重试创建待办任务'}, status=403)
                return
            
            try:
                # 从Excel读取所有待办记录，找出失败的用户（没有task_id的用户）
                all_todos = self.todo_mgr.get_all_todos(announcement_id)
                failed_userids = []
                for todo in all_todos:
                    userid = todo.get('userid', '')
                    task_id = todo.get('task_id', '') or todo.get('taskId', '')
                    # 如果没有task_id，说明创建失败
                    if userid and not task_id:
                        failed_userids.append(userid)
                
                if not failed_userids:
                    self.send_json_response({
                        'success': True,
                        'message': '没有需要重试的待办任务（所有用户的待办都已成功创建）',
                        'retried_count': 0,
                        'failed_count': 0
                    })
                    return
                
                logger.info(f"开始重试为公告 {announcement_id} 创建待办任务，失败用户数: {len(failed_userids)}")
                
                # 使用重试模式创建待办（只重试失败的用户）
                from server.config import TODO_CREATE_RETRY_COUNT
                source_id, user_task_ids, user_source_ids, notification_sent, still_failed_userids = self._create_announcement_todos(
                    announcement_id, 
                    failed_userids, 
                    user, 
                    retry_failed_only=True,  # 启用重试模式，跳过已成功的用户
                    max_retries=TODO_CREATE_RETRY_COUNT
                )
                
                # 更新Excel文件（添加新成功创建的待办）
                if user_task_ids:
                    try:
                        # 获取成功创建待办的用户信息
                        successful_userids = list(user_task_ids.keys())
                        unionids = []
                        valid_userids = []
                        for userid in successful_userids:
                            unionid = self._get_user_unionid_by_userid(userid)
                            if unionid:
                                unionids.append(unionid)
                                valid_userids.append(userid)
                        
                        if len(valid_userids) > 0 and len(unionids) == len(valid_userids):
                            # 更新Excel文件，添加新成功创建的待办记录
                            title = announcement.get('title', '')
                            base_source_id = source_id or (list(user_source_ids.values())[0] if user_source_ids else '')
                            
                            # 读取现有待办记录
                            existing_todos = self.todo_mgr.get_all_todos(announcement_id)
                            existing_userids = {str(todo.get('userid', '')) for todo in existing_todos}
                            
                            # 只添加新成功创建的待办（不在现有记录中的）
                            new_userids = [uid for uid in valid_userids if str(uid) not in existing_userids]
                            new_unionids = [unionids[valid_userids.index(uid)] for uid in new_userids]
                            
                            if new_userids:
                                # 追加新记录到Excel文件
                                file_path = self.todo_mgr.get_todo_file_path(announcement_id)
                                if os.path.exists(file_path):
                                    import openpyxl
                                    from datetime import datetime
                                    wb = openpyxl.load_workbook(file_path)
                                    ws = wb.active
                                    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                    
                                    for idx, (new_userid, new_unionid) in enumerate(zip(new_userids, new_unionids)):
                                        row = ws.max_row + 1
                                        user_source_id = user_source_ids.get(new_userid, base_source_id)
                                        task_id = user_task_ids.get(new_userid, '')
                                        
                                        ws.cell(row=row, column=1, value=announcement_id)
                                        ws.cell(row=row, column=2, value=title)
                                        ws.cell(row=row, column=3, value=user_source_id)
                                        ws.cell(row=row, column=4, value=str(new_userid))
                                        ws.cell(row=row, column=5, value=str(new_unionid))
                                        ws.cell(row=row, column=6, value=task_id)
                                        ws.cell(row=row, column=7, value='')
                                        ws.cell(row=row, column=8, value='')
                                        ws.cell(row=row, column=9, value='未完成')
                                        ws.cell(row=row, column=10, value='')
                                        ws.cell(row=row, column=11, value=now)
                                    
                                    wb.save(file_path)
                                    logger.info(f"成功更新待办Excel文件，添加了 {len(new_userids)} 条新记录")
                                
                                # 清除待办缓存
                                self._invalidate_todo_cache(announcement_id)
                    except Exception as e:
                        logger.error(f"更新待办Excel文件失败: {e}", exc_info=True)
                
                retried_count = len(user_task_ids) if user_task_ids else 0
                still_failed_count = len(still_failed_userids) if still_failed_userids else 0
                
                if retried_count > 0:
                    message = f"成功重试创建 {retried_count} 个待办任务"
                    if still_failed_count > 0:
                        message += f"，仍有 {still_failed_count} 个待办创建失败"
                else:
                    message = f"重试失败，仍有 {still_failed_count} 个待办创建失败"
                
                self.send_json_response({
                    'success': retried_count > 0,
                    'message': message,
                    'retried_count': retried_count,
                    'failed_count': still_failed_count,
                    'failed_userids': still_failed_userids[:10]  # 只返回前10个失败的用户ID
                })
                
            except Exception as e:
                logger.error(f"重试创建待办任务失败: {e}", exc_info=True)
                self.send_json_response({'success': False, 'error': f'重试创建待办任务失败: {str(e)}'}, status=500)
        
        # 重试同步待办任务状态到钉钉
        elif self.path.startswith('/api/todo/sync-retry/'):
            if not self.check_auth():
                return
            
            # 路径格式：/api/todo/sync-retry/{announcement_id}/{userid}
            path_parts = self.path.split('/')
            if len(path_parts) < 5:
                self.send_json_response({'success': False, 'error': '参数错误'}, status=400)
                return
            
            announcement_id = path_parts[-2]
            userid = path_parts[-1]
            
            # 检查权限：只有管理员或公告创建者可以重试同步
            user = self.get_current_user()
            roles = user.get('roles', [])
            is_admin = any(role in ['admin', 'super_admin', 'management'] for role in roles)
            
            # 获取公告信息，检查是否是创建者
            announcement = self.announcement_mgr.get_announcement(announcement_id)
            if not announcement:
                self.send_json_response({'success': False, 'error': '公告不存在'}, status=404)
                return
            
            author = announcement.get('author', '')
            user_name = user.get('name', '') or user.get('username', '')
            is_author = user_name == author
            
            # 允许用户重试自己的待办同步
            is_own_todo = str(user.get('userid', '')) == str(userid)
            
            if not is_admin and not is_author and not is_own_todo:
                self.send_json_response({'success': False, 'error': '无权重试同步待办状态'}, status=403)
                return
            
            try:
                # 获取待办状态
                todo_status = self.todo_mgr.get_user_todo_status(announcement_id, userid)
                if not todo_status:
                    self.send_json_response({'success': False, 'error': '待办记录不存在'}, status=404)
                    return
                
                task_id = todo_status.get('task_id', '')
                if not task_id:
                    self.send_json_response({'success': False, 'error': '待办任务ID不存在，无法同步'}, status=400)
                    return
                
                # 获取用户的unionid
                unionid = self._get_user_unionid_by_userid(userid)
                if not unionid:
                    self.send_json_response({'success': False, 'error': '无法获取用户unionid'}, status=400)
                    return
                
                # 获取公告信息
                announcement_title = announcement.get('title', '')
                updated_subject = f"已阅读公告-{announcement_title}"
                description = announcement_title
                
                # 计算dueTime
                from datetime import datetime, timedelta
                priority = announcement.get('priority', 'normal')
                now = datetime.now()
                if priority in ['high', 'urgent']:
                    due_time = int((now + timedelta(days=1)).timestamp() * 1000)
                else:
                    due_time = int((now + timedelta(days=3)).timestamp() * 1000)
                
                # 获取当前完成状态
                done = todo_status.get('done', False)
                
                # 重试同步到钉钉
                from server.config import TODO_SYNC_RETRY_COUNT
                dingtalk_success = self._update_todo_status(
                    unionid=unionid,
                    task_id=task_id,
                    title=updated_subject,
                    description=description,
                    due_time=due_time,
                    executor_ids=[unionid],
                    participant_ids=[unionid],
                    done=done,
                    max_retries=TODO_SYNC_RETRY_COUNT
                )
                
                if dingtalk_success:
                    logger.info(f"重试同步待办状态成功: announcement_id={announcement_id}, userid={userid}")
                    self.send_json_response({
                        'success': True,
                        'message': '待办状态同步成功'
                    })
                else:
                    logger.error(f"重试同步待办状态失败: announcement_id={announcement_id}, userid={userid}")
                    self.send_json_response({
                        'success': False,
                        'error': '待办状态同步失败，请稍后重试或联系管理员'
                    }, status=500)
                    
            except Exception as e:
                logger.error(f"重试同步待办状态失败: {e}", exc_info=True)
                self.send_json_response({'success': False, 'error': f'重试同步失败: {str(e)}'}, status=500)
        
        # 更新待办任务状态
        elif self.path.startswith('/api/todo/update/'):
            if not self.check_auth():
                return
            
            # 路径格式：/api/todo/update/{announcement_id}
            path_parts = self.path.split('/')
            announcement_id = path_parts[-1] if len(path_parts) > 4 else None
            
            if not announcement_id:
                self.send_json_response({'success': False, 'error': '缺少announcement_id参数'}, status=400)
                return
            
            user = self.get_current_user()
            if not user:
                self.send_json_response({'success': False, 'error': '未登录或会话已过期'}, status=401)
                return
            
            userid = user.get('userid', '')
            
            # 如果session中没有userid，尝试从数据库查找
            if not userid:
                username = user.get('username', '')
                if username:
                    try:
                        logger.info(f"session中没有userid，尝试从数据库查找: username={username}")
                        db_user = self.user_manager.get_user_by_username(username)
                        if db_user and db_user.get('userid'):
                            userid = db_user.get('userid')
                            user['userid'] = userid
                            user['unionid'] = db_user.get('unionid', '')
                            # 更新session
                            cookie_header = self.headers.get('Cookie', '')
                            cookies = {}
                            for cookie in cookie_header.split(';'):
                                cookie = cookie.strip()
                                if '=' in cookie:
                                    key, value = cookie.split('=', 1)
                                    cookies[key.strip()] = value.strip()
                            session_id = cookies.get('session_id')
                            if session_id:
                                from server.session_manager import sync_session_patch
                                sync_session_patch(
                                    session_id,
                                    {'userid': userid, 'unionid': db_user.get('unionid', '')},
                                    sessions,
                                    sessions_lock,
                                )
                            logger.info(f"已从数据库获取并更新userid: {userid} (username={username})")
                        else:
                            logger.warning(f"从数据库未找到userid: username={username}")
                    except Exception as e:
                        logger.error(f"从数据库获取userid失败: {e}", exc_info=True)
            
            if not userid:
                logger.warning(f"无法获取userid: username={user.get('username', 'N/A')}, user_id={user.get('id', 'N/A')}")
                self.send_json_response({
                    'success': False, 
                    'error': '无法获取用户ID，请重新登录或联系管理员'
                }, status=400)
                return
            
            # 读取请求体
            content_length = int(self.headers.get('Content-Length', 0))
            done = True  # 默认为完成
            if content_length > 0:
                try:
                    put_data = self.rfile.read(content_length).decode('utf-8')
                    if put_data.strip():  # 如果请求体不为空
                        data = json.loads(put_data)
                        done = data.get('done', True)  # 默认为完成
                except json.JSONDecodeError as e:
                    logger.warning(f"解析待办更新请求体失败: {e}，使用默认值 done=True")
                    # 继续使用默认值
                except Exception as e:
                    logger.warning(f"读取待办更新请求体失败: {e}，使用默认值 done=True")
                    # 继续使用默认值
            
            # 获取用户的待办信息（包含taskId和unionid）
            todo_status = self.todo_mgr.get_user_todo_status(announcement_id, userid)
            if not todo_status:
                self.send_json_response({'success': False, 'error': '未找到待办任务'}, status=404)
                return
            
            task_id = todo_status.get('task_id', '')
            unionid = todo_status.get('unionid', '')
            
            # 先更新本地Excel状态（无论是否有task_id或unionid）
            # 这样可以确保用户点击"已阅读完成"后，本地状态立即更新
            local_success = self.todo_mgr.update_todo_status(
                announcement_id, 
                userid, 
                done, 
                update_cache=True, 
                immediate_save=True  # 立即保存，确保刷新后能看到
            )
            
            if not local_success:
                logger.error(f"本地Excel更新失败: announcement_id={announcement_id}, userid={userid}, done={done}")
                self.send_json_response({
                    'success': False, 
                    'error': '更新待办状态失败，请稍后重试'
                }, status=500)
                return
            
            # 更新本 worker 热缓存后广播失效标记，通知其他 worker 从文件重载
            preloader = get_data_preloader()
            if preloader and self.todo_mgr:
                try:
                    import time
                    todos = self.todo_mgr.get_all_todos(announcement_id)
                    if todos is not None:
                        with preloader.cache_lock:
                            preloader.todos_cache[announcement_id] = (todos, time.time())
                        logger.debug(f"已更新待办缓存: announcement_id={announcement_id}, count={len(todos)}")
                    else:
                        with preloader.cache_lock:
                            if announcement_id not in preloader.todos_cache:
                                preloader.todos_cache[announcement_id] = ([], time.time())
                                logger.warning(f"从文件加载待办数据失败，使用空缓存: announcement_id={announcement_id}")
                except Exception as e:
                    logger.warning(f"更新待办缓存失败: {e}")
            try:
                from server.cache_sync import broadcast_cache_invalidation, update_local_version
                broadcast_cache_invalidation('todos')
                update_local_version('todos', time.time())
                with api_cache.lock:
                    for key in [k for k in list(api_cache.cache.keys()) if 'todo' in k.lower()]:
                        api_cache._remove_key(key)
            except Exception as e:
                logger.warning(f"广播待办缓存失效标记失败: {e}")
            
            # 如果有task_id和unionid，尝试同步到钉钉
            dingtalk_success = False
            if task_id and unionid:
                # 获取公告信息（用于构建请求体）
                announcement = self.announcement_mgr.get_announcement(announcement_id)
                if announcement:
                    announcement_title = announcement.get('title', '')
                    # 使用"已阅读公告-标题"格式
                    updated_subject = f"已阅读公告-{announcement_title}"
                    description = announcement_title  # 描述使用原公告标题
                    
                    # 计算dueTime（根据优先级）
                    from datetime import datetime, timedelta
                    priority = announcement.get('priority', 'normal')
                    now = datetime.now()
                    if priority in ['high', 'urgent']:
                        due_time = int((now + timedelta(days=1)).timestamp() * 1000)
                    else:
                        due_time = int((now + timedelta(days=3)).timestamp() * 1000)
                    
                    # 调用钉钉API更新待办状态
                    dingtalk_success = self._update_todo_status(
                        unionid=unionid,
                        task_id=task_id,
                        title=updated_subject,
                        description=description,
                        due_time=due_time,
                        executor_ids=[unionid],  # 执行者只有当前用户
                        participant_ids=[unionid],  # 参与者只有当前用户
                        done=done
                    )
                    
                    if dingtalk_success:
                        logger.info(f"成功更新待办状态（钉钉API和本地Excel）: announcement_id={announcement_id}, userid={userid}, done={done}")
                    else:
                        logger.warning(f"钉钉API更新失败，但本地状态已更新: announcement_id={announcement_id}, userid={userid}, task_id={task_id}")
            else:
                logger.info(f"本地待办无task_id或unionid，仅更新本地状态: announcement_id={announcement_id}, userid={userid}")
            
            # 无论钉钉同步是否成功，本地状态已更新，返回成功
            message = '待办状态更新成功'
            if task_id and unionid:
                if dingtalk_success:
                    message += '（已同步到钉钉）'
                else:
                    message += '（仅本地更新，钉钉同步失败）'
            else:
                message += '（仅本地更新，无钉钉待办）'
            
            # 缓存已在上面更新（从文件重新加载），这里不需要再次验证
            # 因为update_todo_status已经更新了缓存，而且上面又从文件重新加载了最新数据
            
            self.send_json_response({
                'success': True, 
                'message': message
            })
        
        else:
            self.send_error(404)
    
    def _resolve_static_file_path(self):
        """解析静态资源路径，返回磁盘上的绝对路径或 None"""
        raw_path = self.path.split('?', 1)[0]
        if not raw_path or raw_path == '/':
            return None
        rel = raw_path.lstrip('/').replace('\\', '/')
        if '..' in rel.split('/'):
            return None
        # 快捷链接缓存图标（data/quick_link_icons，启动时抓取）
        icon_prefix = 'static/quick_link_icons/'
        if rel.startswith(icon_prefix):
            from server.quick_link_manager import QUICK_LINK_ICONS_DIR

            name = rel[len(icon_prefix):]
            if name and '/' not in name and '\\' not in name:
                file_path = os.path.normpath(os.path.join(QUICK_LINK_ICONS_DIR, name))
                icon_root = os.path.normpath(QUICK_LINK_ICONS_DIR)
                if file_path.startswith(icon_root + os.sep) and os.path.isfile(file_path):
                    return file_path
            return None
        # URL 为 /static/...，STATIC_DIR 本身已是 .../static，需去掉前缀避免 static/static/...
        if rel.startswith('static/'):
            rel = rel[len('static/'):]
        elif rel == 'static':
            return None
        file_path = os.path.normpath(os.path.join(STATIC_DIR, rel))
        static_root = os.path.normpath(STATIC_DIR)
        if not file_path.startswith(static_root + os.sep) and file_path != static_root:
            return None
        if os.path.isfile(file_path):
            return file_path
        return None

    def _serve_static_file(self, file_path):
        """直接读取并返回静态文件（兼容 WSGI，无需 super().do_GET()）"""
        try:
            with open(file_path, 'rb') as f:
                content = f.read()
            mime_type, _ = mimetypes.guess_type(file_path)
            if not mime_type:
                mime_type = 'application/octet-stream'
            self.send_response(200)
            self.send_header('Content-Type', mime_type)
            self.send_header('Content-Length', str(len(content)))
            if ENABLE_CACHE:
                self.send_header('Cache-Control', f'max-age={CACHE_MAX_AGE}')
            self.end_headers()
            self.wfile.write(content)
            return True
        except Exception as e:
            logger.error(f"静态文件读取失败: {file_path}, {e}", exc_info=True)
            self.send_error(500, '静态文件读取失败')
            return True

    def handle_page_get(self):
        """处理页面请求"""
        parsed_page = urllib.parse.urlparse(self.path)

        # 站点图标（浏览器默认请求 /favicon.ico）
        if parsed_page.path == '/favicon.ico':
            favicon_path = os.path.join(STATIC_DIR, 'neo-logo.svg')
            if os.path.isfile(favicon_path):
                self._serve_static_file(favicon_path)
            else:
                self.send_response(204)
                self.end_headers()
            return

        # 静态资源（logo/css/js 等）无需登录，优先处理
        static_file = self._resolve_static_file_path()
        if static_file:
            self._serve_static_file(static_file)
            return

        # NEO 子应用路由（htmlsystm 根路径误访问时重定向）
        if parsed_page.path == '/replacement-pairs':
            self.send_response(302)
            self.send_header('Location', '/neo/replacement-pairs')
            self.end_headers()
            return

        # 登录页面不需要认证
        if self.path == '/login':
            self.serve_template('login.html')
            return
        if self.path.startswith('/neo-bridge'):
            self.serve_template('neo_bridge.html')
            return
        if self.path == '/register':
            self.send_response(302)
            self.send_header('Location', '/login')
            self.end_headers()
            return
        
        # 检查认证
        if not self.check_auth():
            self.redirect_to_login()
            return
        
        # 用户管理页面需要管理员权限
        if self.path == '/user-management':
            user = self.get_current_user()
            if not (self._is_super_admin(user) or self._has_role(user, 'management')):
                self.redirect_to_home()
                return
        
        # 审批中心页面需要具备公告审批权限
        if self.path == '/review-center':
            if not self.check_auth():
                return
            user = self.get_current_user()
            if not self._can_approve_announcement(user):
                self.redirect_to_announcement('error=no_review_permission')
                return
        
        if self.path == '/':
            self.serve_template('index.html')
        elif self.path == '/bom-comparison':
            self.serve_template('bom-comparison.html')
        elif self.path == '/announcement':
            self.serve_template('announcement.html')
        elif self.path == '/dingtalk-token-test':
            if not self.check_auth(require_super_admin=True):
                return
            self.serve_template('dingtalk-token-test.html')
        elif self.path == '/dingtalk-department-test':
            if not self.check_auth(require_super_admin=True):
                return
            self.serve_template('dingtalk-department-test.html')
        elif self.path == '/dingtalk-user-test':
            if not self.check_auth(require_super_admin=True):
                return
            self.serve_template('dingtalk-user-test.html')
        elif self.path == '/announcement-editor' or self.path.startswith('/announcement-editor?'):
            self.serve_template('announcement-editor.html')
        elif self.path.startswith('/announcement-version/'):
            # 历史版本详情页面：/announcement-version/{announcement_id}/{version_number}
            self.serve_template('announcement-version.html')
        elif self.path.startswith('/announcement-detail/'):
            self.serve_template('announcement-detail.html')
        elif self.path == '/board-management' or self.path == '/sub-board-management' or self.path == '/board-management-unified':
            if not self.check_auth():
                return
            user = self.get_current_user()
            if not self._can_manage_boards(user):
                self.redirect_to_home()
                return
            # 使用合并后的公告栏管理页面
            self.serve_template('board-management-unified.html')
        elif self.path == '/user-management':
            self.serve_template('user_management.html')
        elif self.path == '/review-center':
            self.serve_template('review-center.html')
        else:
            self.send_error(404)
    
    def serve_template(self, template_name, substitutions=None):
        """服务模板文件（优化版：支持压缩和缓存）。substitutions 为 {占位符: 替换文本}。"""
        template_path = os.path.join(TEMPLATE_DIR, template_name)
        if os.path.exists(template_path):
            with open(template_path, 'r', encoding='utf-8') as f:
                content = f.read()
            if substitutions:
                for key, val in substitutions.items():
                    content = content.replace(key, str(val))
            if '</body>' in content and 'csrf-fetch.js' not in content:
                inject = '<script src="/js/csrf-fetch.js"></script>\n'
                if template_name not in ('login.html', 'register.html'):
                    inject += '<script src="/js/session-guard.js"></script>\n'
                    inject += '<script src="/js/idle-guard.js"></script>\n'
                content = content.replace('</body>', inject + '</body>', 1)

            content_bytes = content.encode('utf-8')
            etag = self._get_etag(content_bytes)
            
            # 检查缓存
            if self._check_cache(etag):
                self.send_response(304)
                self.send_header('ETag', etag)
                self.end_headers()
                return
            
            # 压缩响应
            compressed = False
            if self._should_compress('text/html'):
                compressed_data = self._compress_response(content_bytes)
                if len(compressed_data) < len(content_bytes):
                    content_bytes = compressed_data
                    compressed = True
            
            # 发送响应
            self.send_response(200)
            self.send_header('Content-type', 'text/html; charset=utf-8')
            self.send_header('Content-Length', str(len(content_bytes)))
            
            if compressed:
                self.send_header('Content-Encoding', 'gzip')
            
            if ENABLE_CACHE:
                self.send_header('ETag', etag)
                self.send_header('Cache-Control', f'max-age={CACHE_MAX_AGE}')

            apply_security_headers(self)
            self.end_headers()
            self.wfile.write(content_bytes)
        else:
            self.send_error(404)
    
    def send_json_response(self, data, status=200, set_cookies=None):
        """发送JSON响应（优化：使用GZIP压缩和分块传输，充分利用带宽，兼容钉钉浏览器）"""
        try:
            # 序列化JSON数据（确保使用UTF-8编码）
            try:
                response_data = json.dumps(data, ensure_ascii=False, default=str).encode('utf-8')
            except (TypeError, ValueError) as e:
                logger.error(f"JSON序列化失败: {e}, data={data}")
                # 如果序列化失败，返回错误信息
                error_data = {'success': False, 'error': '数据序列化失败'}
                response_data = json.dumps(error_data, ensure_ascii=False).encode('utf-8')
                status = 500
            
            # 检查是否应该压缩（充分利用带宽）
            # 钉钉浏览器可能对压缩有特殊要求，小数据不压缩以确保兼容性
            should_compress = self._should_compress('application/json')
            if should_compress and len(response_data) > 2048:  # 大于2KB才压缩，提高兼容性
                try:
                    compressed_data = self._compress_response(response_data)
                    # 如果压缩后更小，使用压缩版本
                    if len(compressed_data) < len(response_data):
                        response_data = compressed_data
                        content_encoding = 'gzip'
                    else:
                        content_encoding = None
                except Exception as e:
                    logger.warning(f"压缩失败，使用未压缩数据: {e}")
                    content_encoding = None
            else:
                content_encoding = None
            
            self.send_response(status)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            apply_cors(self)
            self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
            self.send_header('Pragma', 'no-cache')
            self.send_header('Expires', '0')
            if set_cookies:
                for cookie in set_cookies:
                    self.send_header('Set-Cookie', cookie)
            if content_encoding:
                self.send_header('Content-Encoding', content_encoding)
            self.send_header('Content-Length', str(len(response_data)))
            apply_security_headers(self)
            self.end_headers()
            
            # 使用分块传输，充分利用带宽（大文件时）
            # 钉钉浏览器兼容性：小文件直接发送，避免分块问题
            if len(response_data) > CHUNK_SIZE * 2:  # 大于2倍块大小才分块
                try:
                    # 分块发送
                    for i in range(0, len(response_data), CHUNK_SIZE):
                        chunk = response_data[i:i + CHUNK_SIZE]
                        self.wfile.write(chunk)
                        self.wfile.flush()  # 确保数据及时发送
                except (BrokenPipeError, ConnectionResetError):
                    # 客户端断开连接，忽略错误
                    pass
            else:
                # 小文件直接发送（钉钉浏览器兼容性更好）
                try:
                    self.wfile.write(response_data)
                    self.wfile.flush()  # 确保数据立即发送
                except (BrokenPipeError, ConnectionResetError):
                    # 客户端断开连接，忽略错误
                    pass
        except BrokenPipeError:
            # 客户端断开连接，忽略错误
            pass
        except Exception as e:
            logger.error(f"发送JSON响应失败: {e}", exc_info=True)
            # 尝试发送错误响应
            try:
                error_response = json.dumps({'success': False, 'error': '服务器内部错误'}, ensure_ascii=False).encode('utf-8')
                self.send_response(500)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.send_header('Content-Length', str(len(error_response)))
                self.end_headers()
                self.wfile.write(error_response)
            except:
                pass  # 如果发送错误响应也失败，忽略(f"发送JSON响应错误: {e}", exc_info=True)
            try:
                # 如果UTF-8编码失败，尝试使用ASCII安全的方式
                self.send_response(status)
                self.send_header('Content-type', 'application/json; charset=utf-8')
                self.end_headers()
                safe_data = self._make_json_safe(data)
                response_data = json.dumps(safe_data, ensure_ascii=True).encode('utf-8')
                self.wfile.write(response_data)
            except:
                pass  # 如果还是失败，忽略错误
    
    def _make_json_safe(self, data):
        """确保数据可以安全地序列化为JSON"""
        if isinstance(data, dict):
            return {k: self._make_json_safe(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [self._make_json_safe(item) for item in data]
        elif isinstance(data, str):
            # 移除或替换无法编码的字符
            return data.encode('utf-8', 'ignore').decode('utf-8')
        else:
            return data
    
    def send_error(self, code, message=None):
        """发送错误响应"""
        try:
            self.send_response(code)
            self.send_header('Content-type', 'text/html; charset=utf-8')
            self.end_headers()
            if message:
                # 确保消息可以安全编码
                safe_message = message.encode('utf-8', 'ignore').decode('utf-8')
                error_html = f"""
                <html>
                <head><title>Error {code}</title></head>
                <body>
                <h1>Error {code}</h1>
                <p>{safe_message}</p>
                </body>
                </html>
                """
                self.wfile.write(error_html.encode('utf-8'))
        except BrokenPipeError:
            # 客户端断开连接，忽略错误
            pass
        except Exception as e:
            logger.error(f"发送错误响应失败: {e}", exc_info=True)
    
    def check_auth(self, require_admin=False, required_roles=None, require_super_admin=False):
        """检查用户认证与权限"""
        from server.auth.permissions import check_access

        user = self.get_current_user()
        result = check_access(
            user,
            super_admin=require_super_admin,
            admin=require_admin,
            roles=required_roles,
        )
        if result.allowed:
            return True
        if self.path.startswith('/api/'):
            payload = {'success': False, 'error': result.error, 'authenticated': False}
            self.send_json_response(payload, status=result.status)
            return False
        if result.status == 401:
            self.redirect_to_login()
        else:
            self.send_error(result.status, result.error or '无访问权限')
        return False
    
    def get_current_user(self, skip_session_enrich: bool = False):
        """获取当前用户（签名 Cookie + jti 索引）。"""
        cookie_header = self.headers.get('Cookie', '')
        from server.auth.login_service import resolve_user_from_cookies

        return resolve_user_from_cookies(
            cookie_header,
            lite=skip_session_enrich,
            skip_session_enrich=skip_session_enrich,
        )
    
    def handle_login(self):
        """处理登录（签名 Cookie）。"""
        self.send_json_response(
            {
                'success': False,
                'code': 403,
                'error': '本地账号密码登录已关闭，请使用钉钉登录',
            },
            status=403,
        )
        return

        from server.auth.login_service import perform_login

        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length).decode('utf-8')
        data = urllib.parse.parse_qs(post_data)

        username = data.get('username', [''])[0]
        password = data.get('password', [''])[0]
        captcha_token = data.get('captcha_token', [''])[0]
        captcha_code = data.get('captcha_code', [''])[0]

        client_ip = self.client_address[0] if hasattr(self, 'client_address') else self.headers.get('X-Real-IP', '0.0.0.0')
        if not client_ip or client_ip == '0.0.0.0':
            forwarded_for = self.headers.get('X-Forwarded-For', '')
            if forwarded_for:
                client_ip = forwarded_for.split(',')[0].strip()
            else:
                client_ip = '0.0.0.0'

        ok, body, set_cookies = perform_login(
            username=username,
            password=password,
            captcha_token=captcha_token,
            captcha_code=captcha_code,
            client_ip=client_ip,
            secure=self._is_https_request(),
        )
        status = 200 if ok else 200
        self.send_json_response(body, status=status, set_cookies=set_cookies if ok else None)
    
    def handle_login_by_userid(self):
        """通过钉钉 authCode 验证身份后登录（企业内部应用网页免登）。"""
        self._handle_dingtalk_inapp_login()
        return

        from server.startup_gate import login_allowed

        ok_login, maint = login_allowed()
        if not ok_login and maint:
            self.send_json_response(
                {
                    'success': False,
                    'error': maint.get('message') or '系统正在启动，请稍候',
                    'startup': maint,
                },
                status=503,
            )
            return
        try:
            content_length = int(self.headers.get('Content-Length', 0) or 0)
            if content_length == 0:
                self.send_json_response({'success': False, 'error': '请求体不能为空'})
                return
            
            post_data = self.rfile.read(content_length).decode('utf-8')
            data = json.loads(post_data)

            auth_code = (
                data.get('authCode')
                or data.get('auth_code')
                or data.get('code')
                or ''
            ).strip()
            if not auth_code:
                self.send_json_response({'success': False, 'error': '缺少有效的钉钉授权码'})
                return

            config_ok, config_error = check_dingtalk_config()
            if not config_ok:
                logger.warning('钉钉免登配置不完整: %s', config_error)
                self.send_json_response({
                    'success': False,
                    'error': '钉钉配置不完整，请联系管理员',
                })
                return

            try:
                user_info = self._get_dingtalk_user_info(
                    auth_code,
                    '企业内部应用网页免登',
                )
            except Exception as api_err:
                logger.warning('钉钉 authCode 校验失败: %s', api_err)
                self.send_json_response({'success': False, 'error': '钉钉授权无效或已过期，请重试'})
                return

            userid = str(user_info.get('userid') or '').strip()
            if not userid:
                self.send_json_response({'success': False, 'error': '无法从钉钉获取用户身份'})
                return

            client_userid = str(data.get('userid') or '').strip()
            if client_userid and client_userid != userid:
                logger.warning(
                    '钉钉免登 userid 不匹配: client=%s, verified=%s',
                    client_userid,
                    userid,
                )
                self.send_json_response({'success': False, 'error': '授权码与用户不匹配'})
                return

            logger.info('钉钉 authCode 免登尝试: userid=%s', userid)

            # 通过 userid 查找用户
            user = self.user_manager.get_user_by_userid(userid)
            if user:
                status = user.get('status')
                if status != STATUS_ACTIVE:
                    if status == STATUS_PENDING:
                        error = '账号正在审批，审批通过后即可登录'
                    elif status == STATUS_REJECTED:
                        error = '账号申请已被拒绝，请联系管理组'
                    else:
                        error = '账号不可用'
                    self.send_json_response({'success': False, 'error': error})
                    return
                
                session_user = {
                    'id': user['id'],
                    'username': user['username'],
                    'name': user['name'],
                    'roles': user.get('roles', []),
                    'role': user.get('role'),
                    'department': user.get('department'),
                    'job_position': user.get('job_position', ''),
                    'status': user.get('status'),
                    'userid': user.get('userid'),
                    'unionid': user.get('unionid', ''),
                }

                from server.auth.login_service import issue_session_for_user

                try:
                    response, set_cookies = issue_session_for_user(
                        session_user,
                        secure=self._is_https_request(),
                    )
                except Exception as e:
                    logger.error(f"钉钉免登签发会话失败: {e}", exc_info=True)
                    self.send_json_response(
                        {'success': False, 'error': '会话服务繁忙，请稍后重试'},
                        status=503,
                    )
                    return

                logger.info(
                    f"通过userid登录成功: userid={userid}, 用户ID={user['id']}, 用户名={user['username']}"
                )
                self.send_json_response(response, set_cookies=set_cookies)
            else:
                logger.warning(f"通过userid登录失败: userid={userid} (用户不存在)")
                self.send_json_response({
                    'success': False,
                    'error': '用户不存在，请先注册或联系管理员'
                })
        except json.JSONDecodeError:
            logger.error("通过userid登录请求：无效的JSON数据")
            self.send_json_response({'success': False, 'error': '无效的JSON数据'})
        except Exception as e:
            log_server_error('通过userid登录', e)
            self.send_json_response(safe_error_payload(), status=500)
    
    def handle_auth_password(self, legacy_path: bool = True):
        """POST /api/auth/password 或兼容 /api/auth/change-password"""
        self.send_json_response(
            {
                'ok': False,
                'success': False,
                'code': 403,
                'error': '本地密码入口已关闭，请使用钉钉登录',
            },
            status=403,
        )
        return

        from server.auth.password_service import get_password_service

        try:
            content_length = int(self.headers.get('Content-Length', 0) or 0)
            if content_length == 0:
                self.send_json_response({'ok': False, 'success': False, 'error': '请求体不能为空'})
                return

            post_data = self.rfile.read(content_length).decode('utf-8')
            data = json.loads(post_data)

            action = (data.get('action') or '').strip().lower()
            new_password = (data.get('newPassword') or data.get('new_password') or '').strip()
            old_password = (data.get('oldPassword') or data.get('old_password') or '').strip()
            username = (data.get('username') or '').strip()

            current_user = self.get_current_user()
            is_logged_in = current_user is not None
            pwd_svc = get_password_service(self.user_manager)
            extra_cookies = []
            meta = {}

            if action == 'reset' or (not legacy_path and action == 'reset'):
                if not self.check_auth(require_super_admin=True):
                    return
                target_id = data.get('userId') or data.get('user_id')
                if not target_id and self.path.startswith('/api/auth/users/'):
                    try:
                        target_id = int(self.path.rstrip('/').split('/')[-2 if self.path.endswith('/password') else -1])
                    except ValueError:
                        target_id = None
                if not target_id:
                    self.send_json_response({'ok': False, 'success': False, 'error': '缺少 userId'})
                    return
                ok, msg, meta = pwd_svc.admin_reset_password(
                    int(target_id),
                    new_password,
                    memory_sessions=sessions,
                    memory_lock=sessions_lock,
                )
            elif is_logged_in:
                if not old_password:
                    self.send_json_response({'ok': False, 'success': False, 'error': '请输入当前密码'})
                    return
                ok, msg, meta = pwd_svc.change_own_password(
                    user_id=int(current_user.get('id')),
                    username=current_user.get('username') or username,
                    old_password=old_password,
                    new_password=new_password,
                    memory_sessions=sessions,
                    memory_lock=sessions_lock,
                )
                if ok:
                    extra_cookies.append('session_id=; Path=/; HttpOnly; Max-Age=0')
            else:
                if not username or not old_password:
                    self.send_json_response({'ok': False, 'success': False, 'error': '请先登录或提供用户名和当前密码'})
                    return
                user = self.user_manager.authenticate_user(username, old_password)
                if not user:
                    self.send_json_response({'ok': False, 'success': False, 'error': '当前密码错误'})
                    return
                ok, msg, meta = pwd_svc.change_own_password(
                    user_id=int(user.get('id')),
                    username=username,
                    old_password=old_password,
                    new_password=new_password,
                    memory_sessions=sessions,
                    memory_lock=sessions_lock,
                )

            if ok:
                preloader = get_data_preloader()
                if preloader:
                    preloader.invalidate_cache('users')
                resp = {
                    'ok': True,
                    'success': True,
                    'message': msg,
                    'clearAutoLogin': meta.get('clearAutoLogin', True),
                }
                self.send_json_response(resp, set_cookies=extra_cookies if extra_cookies else None)
            else:
                self.send_json_response({'ok': False, 'success': False, 'error': msg})
        except json.JSONDecodeError:
            self.send_json_response({'ok': False, 'success': False, 'error': '无效的JSON数据'})
        except Exception as e:
            logger.error(f"改密异常: {e}", exc_info=True)
            self.send_json_response({'ok': False, 'success': False, 'error': f'服务器错误: {str(e)}'})

    def handle_change_password(self):
        """兼容旧名"""
        self.handle_auth_password(legacy_path=True)
    
    def handle_dingtalk_login(self):
        """处理钉钉免登录"""
        try:
            content_length = int(self.headers.get('Content-Length', 0) or 0)
            if content_length == 0:
                self.send_json_response({'success': False, 'error': '请求体不能为空'})
                return
            
            post_data = self.rfile.read(content_length).decode('utf-8')
            data = json.loads(post_data)
            
            auth_code = data.get('authCode', '').strip()
            is_mock = data.get('isMock', False)
            
            if not auth_code:
                self.send_json_response({'success': False, 'error': '授权码不能为空'})
                return
            
            logger.info(f"钉钉免登录请求: authCode={auth_code[:10]}..., isMock={is_mock}")
            
            # 如果是模拟测试，返回模拟数据
            if is_mock or auth_code.startswith('MOCK_'):
                logger.info("使用模拟模式返回测试数据")
                mock_user_info = {
                    'nick': '测试用户',
                    'unionid': 'mock_unionid_' + str(uuid.uuid4()),
                    'openid': 'mock_openid_' + str(uuid.uuid4()),
                    'dingId': 'mock_dingid_' + str(uuid.uuid4())
                }
                self.send_json_response({
                    'success': True,
                    'message': '模拟测试成功（非真实钉钉环境）',
                    'userInfo': mock_user_info,
                    'isMock': True
                })
                return
            
            # 真实环境：调用钉钉API获取用户信息
            # 检查配置是否完整
            config_ok, config_error = check_dingtalk_config()
            if not config_ok:
                logger.warning(f"钉钉配置不完整: {config_error}")
                self.send_json_response({
                    'success': False,
                    'error': '钉钉配置不完整',
                    'message': config_error,
                    'hint': '请在 server/config.py 中配置 DINGTALK_CONFIG 的 client_secret，或设置环境变量 DINGTALK_CLIENT_SECRET'
                })
                return
            
            # 调用钉钉API获取用户信息
            try:
                # 判断免登方式（企业内部应用网页免登 vs 小程序免登）
                method = data.get('method', '')
                user_info = self._get_dingtalk_user_info(auth_code, method)
                if user_info:
                    logger.info(f"钉钉免登录成功: 方式={method}, userid={user_info.get('userid', user_info.get('unionid', 'N/A'))}")
                    self.send_json_response({
                        'success': True,
                        'message': '获取用户信息成功',
                        'userInfo': user_info,
                        'method': method
                    })
                else:
                    logger.error("钉钉API返回空用户信息")
                    self.send_json_response({
                        'success': False,
                        'error': '获取用户信息失败',
                        'message': '钉钉API返回空数据，请检查授权码是否有效'
                    })
            except Exception as api_error:
                logger.error(f"调用钉钉API失败: {api_error}", exc_info=True)
                self.send_json_response({
                    'success': False,
                    'error': '调用钉钉API失败',
                    'message': str(api_error)
                })
            
        except json.JSONDecodeError:
            logger.error("钉钉免登录请求：无效的JSON数据")
            self.send_json_response({'success': False, 'error': '无效的JSON数据'})
        except Exception as e:
            logger.error(f"钉钉免登录异常: {e}", exc_info=True)
            self.send_json_response({'success': False, 'error': f'服务器错误: {str(e)}'})
    
    def handle_dingtalk_get_access_token(self):
        """处理获取钉钉AccessToken的请求（使用官方API格式）
        
        参照官方API文档：
        POST /v1.0/oauth2/{corp_id}/token HTTP/1.1
        Host: api.dingtalk.com
        Content-Type: application/json
        
        {
            "client_id": "suite123",
            "client_secret": "********",
            "grant_type": "client_credentials"
        }
        
        返回格式：
        HTTP/1.1 200 OK
        Content-Type: application/json
        {
            "access_token": "2bf******9be361a5084f1e2b8",
            "expires_in": 7200
        }
        """
        try:
            logger.info("=" * 60)
            logger.info("收到获取钉钉AccessToken请求（官方API格式）")
            logger.info("=" * 60)
            
            # 读取请求体
            content_length = int(self.headers.get('Content-Length', 0) or 0)
            if content_length == 0:
                logger.warning("获取钉钉AccessToken请求：请求体为空")
                self.send_json_response({'success': False, 'error': '请求体不能为空'})
                return
            
            post_data = self.rfile.read(content_length).decode('utf-8')
            logger.info('收到获取钉钉 AccessToken 请求')
            try:
                json.loads(post_data)
            except json.JSONDecodeError:
                pass
            
            # 仅使用服务端配置，不接受客户端传入的 client_secret
            client_id = DINGTALK_CONFIG.get('client_id', '').strip()
            client_secret = DINGTALK_CONFIG.get('client_secret', '').strip()
            grant_type = 'client_credentials'
            corp_id = DINGTALK_CONFIG.get('corp_id', '').strip()

            config_ok, config_error = check_dingtalk_config()
            if not config_ok or not client_id or not corp_id:
                logger.error('钉钉 AccessToken 配置不完整: %s', config_error)
                self.send_json_response({
                    'success': False,
                    'error': '钉钉配置不完整',
                    'message': config_error or '请在环境变量 DINGTALK_CLIENT_SECRET 中设置密钥',
                })
                return
            
            # 构建官方API URL
            url = f"https://api.dingtalk.com/v1.0/oauth2/{corp_id}/token"
            
            headers = {
                'Content-Type': 'application/json',
                'Host': 'api.dingtalk.com',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            
            # 使用官方API的参数格式
            payload = {
                'client_id': client_id,
                'client_secret': client_secret,
                'grant_type': grant_type
            }
            
            logger.info(f"请求URL: {url}")
            logger.info(f"请求Headers: {headers}")
            logger.info(f"请求Payload: client_id={client_id[:10]}..., client_secret=***, grant_type={grant_type}")
            logger.info("正在使用urllib发送HTTP POST请求...")
            
            try:
                # 使用Python标准库urllib发送HTTP请求（不依赖requests库）
                import urllib.request
                import urllib.error
                import ssl
                
                # 准备请求数据
                request_data_bytes = json.dumps(payload).encode('utf-8')
                logger.info(f"请求体JSON: {json.dumps(payload, ensure_ascii=False)}")
                logger.info(f"请求体长度: {len(request_data_bytes)} 字节")
                
                # 创建请求对象
                req = urllib.request.Request(url, data=request_data_bytes, headers=headers, method='POST')
                
                # 创建SSL上下文（跳过证书验证）
                ssl_context = ssl.create_default_context()
                ssl_context.check_hostname = False
                ssl_context.verify_mode = ssl.CERT_NONE
                
                # 发送请求
                logger.info("发送HTTP POST请求到钉钉API...")
                with urllib.request.urlopen(req, timeout=30, context=ssl_context) as response:
                    response_text = response.read().decode('utf-8')
                    response_code = response.getcode()
                    
                    logger.info("=" * 60)
                    logger.info("✅ 请求完成！")
                    logger.info(f"状态码: {response_code}")
                    logger.info(f"响应长度: {len(response_text)} 字节")
                    logger.info(f"响应内容: {response_text[:200]}...")
                    logger.info("=" * 60)
                    
                    if response_code == 200:
                        # 检查响应是否是HTML（可能是重定向或错误页面）
                        if response_text.strip().startswith('<!DOCTYPE') or response_text.strip().startswith('<html'):
                            logger.warning("⚠️  API返回了HTML页面而不是JSON，可能是重定向或错误页面")
                            logger.warning(f"响应内容预览: {response_text[:500]}")
                            # 自动重试一次
                            logger.info("   自动重试获取AccessToken...")
                            time.sleep(1)  # 等待1秒后重试
                            
                            # 重试请求
                            retry_req = urllib.request.Request(url, data=request_data_bytes, headers=headers, method='POST')
                            with urllib.request.urlopen(retry_req, timeout=30, context=ssl_context) as retry_response:
                                retry_response_text = retry_response.read().decode('utf-8')
                                retry_response_code = retry_response.getcode()
                                
                                if retry_response_code == 200:
                                    if retry_response_text.strip().startswith('<!DOCTYPE') or retry_response_text.strip().startswith('<html'):
                                        logger.error("❌ 重试后仍然返回HTML页面")
                                        self.send_json_response({
                                            'success': False,
                                            'error': 'API返回了HTML页面而不是JSON',
                                            'status_code': retry_response_code,
                                            'response_preview': retry_response_text[:500] if len(retry_response_text) > 500 else retry_response_text
                                        })
                                        return
                                    else:
                                        response_text = retry_response_text
                                        response_code = retry_response_code
                                        logger.info("✅ 重试成功，获取到JSON响应")
                        
                        try:
                            result = json.loads(response_text)
                            logger.info(f"解析JSON成功: {json.dumps(result, ensure_ascii=False)}")
                            
                            # 官方API返回格式：access_token 和 expires_in
                            access_token = result.get('access_token')
                            expire_in = result.get('expires_in', 7200)
                            
                            if access_token:
                                logger.info("🎉 成功获取AccessToken！")
                                logger.info(f"AccessToken: {access_token[:20]}...")
                                logger.info(f"有效时间: {expire_in} 秒")
                                logger.info("=" * 60)
                                
                                # 返回结果，统一使用accessToken和expireIn字段名（前端兼容）
                                self.send_json_response({
                                    'success': True,
                                    'message': '成功获取AccessToken',
                                    'accessToken': access_token,  # 前端使用的字段名
                                    'access_token': access_token,  # 官方字段名
                                    'expireIn': expire_in,  # 前端使用的字段名
                                    'expires_in': expire_in,  # 官方字段名
                                    'full_response': result  # 包含完整的原始响应
                                })
                            else:
                                logger.error("❌ API返回成功但未包含access_token")
                                logger.error(f"响应内容: {result}")
                                logger.info("=" * 60)
                                self.send_json_response({
                                    'success': False,
                                    'error': 'API返回成功但未包含access_token',
                                    'response': result
                                })
                        except json.JSONDecodeError as json_err:
                            logger.error(f"解析JSON响应失败: {json_err}")
                            logger.error(f"响应内容预览: {response_text[:500]}")
                            self.send_json_response({
                                'success': False,
                                'error': 'API返回了无效的JSON格式',
                                'status_code': response_code,
                                'response_preview': response_text[:500] if len(response_text) > 500 else response_text
                            })
                    else:
                        error_text = response_text if response_text else '无响应内容'
                        logger.error(f"API返回错误，状态码: {response_code}, 响应: {error_text[:200]}")
                        try:
                            error_json = json.loads(error_text)
                            error_code = error_json.get('error', '')
                            error_msg = error_json.get('error_description') or error_json.get('error') or str(error_json)
                            
                            # 根据官方错误码提供详细说明
                            error_explanation = ""
                            if error_code == 'invalid.client':
                                error_explanation = "无效的ClientID或ClientSecret"
                            elif error_code == 'unsupported.grant.type':
                                error_explanation = "不支持此授权类型，请检查授权类型参数"
                            elif error_code == 'unauthorized.client':
                                error_explanation = "应用未被授权"
                            elif error_code == 'server.error':
                                error_explanation = "服务器意外错误"
                            
                            self.send_json_response({
                                'success': False,
                                'error': f'API返回错误，状态码: {response_code}',
                                'status_code': response_code,
                                'error_code': error_code,
                                'error_detail': error_msg,
                                'error_explanation': error_explanation,
                                'response': error_json
                            })
                        except (json.JSONDecodeError, ValueError):
                            self.send_json_response({
                                'success': False,
                                'error': f'API返回错误，状态码: {response_code}',
                                'status_code': response_code,
                                'response': error_text[:500]
                            })
                        
            except urllib.error.HTTPError as e:
                error_body = ''
                try:
                    if e.fp:
                        error_body = e.read().decode('utf-8')
                except:
                    pass
                
                logger.error("=" * 60)
                logger.error(f"❌ HTTP错误: {e.code}")
                logger.error(f"错误详情: {error_body[:200] if error_body else str(e)}")
                logger.error("=" * 60)
                
                try:
                    error_json = json.loads(error_body) if error_body else {}
                    error_code = error_json.get('error', '')
                    error_msg = error_json.get('error_description') or error_json.get('error') or str(error_json)
                    
                    self.send_json_response({
                        'success': False,
                        'error': f'HTTP错误 {e.code}',
                        'status_code': e.code,
                        'error_code': error_code,
                        'error_detail': error_msg,
                        'response': error_json if error_json else error_body[:500]
                    })
                except (json.JSONDecodeError, ValueError):
                    self.send_json_response({
                        'success': False,
                        'error': f'HTTP错误 {e.code}',
                        'status_code': e.code,
                        'response': error_body[:500] if error_body else str(e)
                    })
            except urllib.error.URLError as e:
                logger.error("=" * 60)
                logger.error("❌ 网络错误")
                logger.error(f"错误详情: {e}")
                logger.error("=" * 60)
                self.send_json_response({
                    'success': False,
                    'error': '网络错误',
                    'error_detail': str(e),
                    'request_url': url,
                    'message': '无法连接到钉钉API服务器，请检查网络连接'
                })
            except Exception as e:
                logger.error("=" * 60)
                logger.error("❌ 发生未知错误")
                logger.error(f"错误类型: {type(e).__name__}")
                logger.error(f"错误详情: {e}")
                import traceback
                logger.error(traceback.format_exc())
                logger.error("=" * 60)
                self.send_json_response({
                    'success': False,
                    'error': '发生未知错误',
                    'error_type': type(e).__name__,
                    'error_detail': str(e)
                })
                
        except json.JSONDecodeError as e:
            logger.error(f"获取钉钉AccessToken请求：无效的JSON数据: {e}")
            self.send_json_response({'success': False, 'error': '无效的JSON数据', 'message': f'后端未能解析前端发送的JSON请求体: {str(e)}'})
        except Exception as e:
            logger.error(f"处理获取AccessToken请求的顶层异常: {e}", exc_info=True)
            self.send_json_response({'success': False, 'error': f'服务器错误: {str(e)}', 'message': '服务器在处理请求时发生意外错误。'})
    
    def handle_dingtalk_get_user_info(self):
        """处理获取钉钉用户信息的请求（使用access_token和授权码）
        
        参照官方API文档：
        POST https://oapi.dingtalk.com/topapi/v2/user/getuserinfo?access_token=ACCESS_TOKEN
        请求正文：{"code":"a22e11cde7cd3a8b9976fa127a0c9353"}
        
        返回格式：
        {
            "errcode": 0,
            "result": {
                "unionid": "yRQptxVcfvdJQWRNcGPLhQiEiE",
                "device_id": "84ffd06697bc96c8c02f3a1f92d98f4f",
                "sys_level": 2,
                "name": "张志伟",
                "sys": true,
                "userid": "533918221524183112"
            },
            "errmsg": "ok",
            "request_id": "16kr0c8dbbxw4"
        }
        """
        try:
            logger.info("=" * 60)
            logger.info("收到获取钉钉用户信息请求")
            logger.info("=" * 60)
            
            content_length = int(self.headers.get('Content-Length', 0) or 0)
            if content_length == 0:
                logger.warning("获取用户信息请求：请求体为空")
                self.send_json_response({'success': False, 'error': '请求体不能为空'})
                return
            
            post_data = self.rfile.read(content_length).decode('utf-8')
            request_data = json.loads(post_data)
            
            code = request_data.get('code', '').strip()

            if not code:
                self.send_json_response({
                    'success': False,
                    'error': '参数不完整',
                    'message': '请提供 code（免登录授权码）'
                })
                return
            
            config_ok, config_error = check_dingtalk_config()
            if not config_ok:
                self.send_json_response({
                    'success': False,
                    'error': '钉钉配置不完整',
                    'message': config_error,
                })
                return

            client_id = DINGTALK_CONFIG.get('client_id', '').strip()
            client_secret = DINGTALK_CONFIG.get('client_secret', '').strip()
            try:
                access_token = self._get_dingtalk_access_token(client_id, client_secret)
            except Exception as token_err:
                logger.error('获取钉钉 access_token 失败: %s', token_err)
                self.send_json_response({
                    'success': False,
                    'error': '获取钉钉凭证失败',
                })
                return

            logger.info('使用服务端 access_token 和授权码获取用户信息: code=%s...', code[:10])

            try:
                user_info = self._get_userinfo_by_code(access_token, code)
                
                if user_info:
                    logger.info("=" * 60)
                    logger.info("✅ 成功获取用户信息！")
                    logger.info(f"用户ID: {user_info.get('userid')}")
                    logger.info(f"用户名: {user_info.get('name')}")
                    logger.info(f"UnionID: {user_info.get('unionid')}")
                    logger.info("=" * 60)
                    
                    self.send_json_response({
                        'success': True,
                        'message': '成功获取用户信息',
                        'userInfo': user_info
                    })
                else:
                    logger.error("❌ 获取用户信息返回空结果")
                    self.send_json_response({
                        'success': False,
                        'error': '获取用户信息返回空结果'
                    })
            except Exception as e:
                logger.error("=" * 60)
                logger.error(f"❌ 获取用户信息失败: {e}")
                logger.error("=" * 60)
                self.send_json_response({
                    'success': False,
                    'error': f'获取用户信息失败: {str(e)}'
                })
                
        except json.JSONDecodeError as e:
            logger.error(f"获取用户信息请求：无效的JSON数据: {e}")
            self.send_json_response({'success': False, 'error': '无效的JSON数据', 'message': f'后端未能解析前端发送的JSON请求体: {str(e)}'})
        except Exception as e:
            logger.error(f"处理获取用户信息请求的顶层异常: {e}", exc_info=True)
            self.send_json_response({'success': False, 'error': f'服务器错误: {str(e)}', 'message': '服务器在处理请求时发生意外错误。'})
    
    def handle_dingtalk_get_departments(self):
        """处理获取钉钉部门列表的请求
        
        参照官方API文档：
        POST https://oapi.dingtalk.com/topapi/v2/department/listsub?access_token=ACCESS_TOKEN
        请求正文：{"language":"zh_CN", "dept_id":1}
        
        返回格式：
        {
            "errcode": 0,
            "errmsg": "ok",
            "result": [
                {
                    "auto_add_user": true,
                    "create_dept_group": true,
                    "dept_id": 37xxxx95,
                    "name": "市场部",
                    "parent_id": 1
                }
            ],
            "request_id": "5um7ykyaalsj"
        }
        """
        try:
            logger.info("=" * 60)
            logger.info("收到获取钉钉部门列表请求")
            logger.info("=" * 60)
            
            content_length = int(self.headers.get('Content-Length', 0) or 0)
            if content_length == 0:
                logger.warning("获取部门列表请求：请求体为空")
                self.send_json_response({'success': False, 'error': '请求体不能为空'})
                return
            
            post_data = self.rfile.read(content_length).decode('utf-8')
            request_data = json.loads(post_data)
            
            access_token = request_data.get('access_token', '').strip()
            clear_existing = request_data.get('clear_existing', False)  # 是否清空现有数据
            
            logger.info(f"从请求体读取参数: access_token={bool(access_token)}, clear_existing={clear_existing}")
            
            if not access_token:
                self.send_json_response({
                    'success': False,
                    'error': '参数不完整',
                    'message': '请提供 access_token'
                })
                return
            
            # 如果选择清空现有数据
            if clear_existing:
                self.department_mgr.clear_departments()
                logger.info("已清空现有部门数据")
            
            # 递归获取所有部门
            all_departments = []
            visited_dept_ids = set()  # 防止重复获取
            
            def get_departments_recursive(dept_id, parent_id=None):
                """递归获取部门列表"""
                if dept_id in visited_dept_ids:
                    return
                visited_dept_ids.add(dept_id)
                
                try:
                    departments = self._get_department_list(access_token, dept_id)
                    logger.info(f"获取到 dept_id={dept_id} 下的 {len(departments)} 个部门")
                    
                    for dept in departments:
                        dept_info = {
                            'parent_id': dept.get('parent_id', parent_id),
                            'name': dept.get('name', ''),
                            'dept_id': dept.get('dept_id'),
                            'create_dept_group': dept.get('create_dept_group', False)
                        }
                        
                        if dept_info['dept_id']:
                            all_departments.append(dept_info)
                            
                            # 如果 create_dept_group 为 true，递归获取子部门
                            if dept_info['create_dept_group']:
                                logger.info(f"部门 {dept_info['name']} (dept_id={dept_info['dept_id']}) 有子部门，继续获取...")
                                get_departments_recursive(dept_info['dept_id'], dept_info['dept_id'])
                
                except Exception as e:
                    logger.error(f"获取 dept_id={dept_id} 的部门列表失败: {e}")
            
            # 从根部门开始（dept_id=1）
            logger.info("开始从根部门 (dept_id=1) 获取部门列表...")
            get_departments_recursive(1, 1)
            
            # 保存到Excel文件
            if all_departments:
                success = self.department_mgr.save_departments(all_departments, append=False)
                if success:
                    logger.info("=" * 60)
                    logger.info(f"✅ 成功获取并保存 {len(all_departments)} 个部门！")
                    logger.info("=" * 60)
                    
                    self.send_json_response({
                        'success': True,
                        'message': f'成功获取并保存 {len(all_departments)} 个部门',
                        'count': len(all_departments),
                        'departments': all_departments  # 返回所有部门
                    })
                else:
                    self.send_json_response({
                        'success': False,
                        'error': '获取部门成功，但保存到文件失败',
                        'count': len(all_departments)
                    })
            else:
                logger.warning("未获取到任何部门")
                self.send_json_response({
                    'success': False,
                    'error': '未获取到任何部门',
                    'message': '请检查 access_token 是否有效'
                })
                
        except json.JSONDecodeError as e:
            logger.error(f"获取部门列表请求：无效的JSON数据: {e}")
            self.send_json_response({'success': False, 'error': '无效的JSON数据', 'message': f'后端未能解析前端发送的JSON请求体: {str(e)}'})
        except Exception as e:
            logger.error(f"处理获取部门列表请求的顶层异常: {e}", exc_info=True)
            self.send_json_response({'success': False, 'error': f'服务器错误: {str(e)}', 'message': '服务器在处理请求时发生意外错误。'})
    
    def _get_department_list(self, access_token, dept_id):
        """调用钉钉API获取指定部门的子部门列表
        
        Args:
            access_token: 访问令牌
            dept_id: 部门ID（默认为1，根部门）
        
        Returns:
            部门列表
        """
        import urllib.request
        import urllib.error
        import ssl
        
        url = "https://oapi.dingtalk.com/topapi/v2/department/listsub"
        
        # 构建请求参数
        params = {
            'language': 'zh_CN',
            'dept_id': dept_id
        }
        
        # 构建完整URL（access_token作为查询参数）
        url_with_params = f"{url}?access_token={urllib.parse.quote(access_token)}"
        
        logger.info(f"调用获取部门列表API: {url}")
        logger.info(f"请求参数: dept_id={dept_id}")
        
        request_data = json.dumps(params).encode('utf-8')
        
        logger.info(f"使用urllib调用获取部门列表API: {url_with_params}")
        logger.info(f"请求体: {json.dumps(params, ensure_ascii=False)}")
        
        req = urllib.request.Request(
            url_with_params,
            data=request_data,
            headers={'Content-Type': 'application/json'},
            method='POST'
        )
        
        # 创建SSL上下文（跳过证书验证）
        ssl_context = ssl.create_default_context()
        ssl_context.check_hostname = False
        ssl_context.verify_mode = ssl.CERT_NONE
        
        try:
            with urllib.request.urlopen(req, timeout=30, context=ssl_context) as response:
                response_text = response.read().decode('utf-8')
                response_code = response.getcode()
                
                logger.info(f"获取部门列表响应状态码: {response_code}")
                
                if response_code == 200:
                    result = json.loads(response_text)
                    errcode = result.get('errcode', -1)
                    errmsg = result.get('errmsg', '')
                    
                    logger.info(f"获取部门列表响应: errcode={errcode}, errmsg={errmsg}")
                    
                    if errcode == 0:
                        departments = result.get('result', [])
                        logger.info(f"成功获取 {len(departments)} 个部门")
                        return departments
                    else:
                        logger.error(f"获取部门列表失败: {errmsg} (errcode: {errcode})")
                        raise Exception(f"获取部门列表失败: {errmsg} (errcode: {errcode})")
                else:
                    logger.error(f"获取部门列表HTTP错误: {response_code}, 响应: {response_text[:200]}")
                    raise Exception(f"HTTP错误: {response_code}")
        except urllib.error.HTTPError as e:
            error_body = ''
            try:
                if e.fp:
                    error_body = e.read().decode('utf-8')
            except:
                pass
            logger.error(f"获取部门列表HTTP错误 {e.code}: {error_body[:200] if error_body else str(e)}")
            raise Exception(f"HTTP错误 {e.code}: {error_body[:200] if error_body else str(e)}")
        except urllib.error.URLError as e:
            logger.error(f"获取部门列表网络错误: {e}")
            raise Exception(f"网络错误: {str(e)}")
    
    def handle_dingtalk_get_department_users(self):
        """处理获取钉钉部门用户列表的请求
        
        参照官方API文档：
        POST https://oapi.dingtalk.com/topapi/v2/user/list?access_token=ACCESS_TOKEN
        请求正文：{
            "cursor":"0",
            "contain_access_limit":"false",
            "size":"10",
            "order_field":"modify_desc",
            "language":"zh_CN",
            "dept_id":"10"
        }
        """
        try:
            logger.info("=" * 60)
            logger.info("收到获取钉钉部门用户列表请求")
            logger.info("=" * 60)
            
            content_length = int(self.headers.get('Content-Length', 0) or 0)
            if content_length == 0:
                logger.warning("获取部门用户列表请求：请求体为空")
                self.send_json_response({'success': False, 'error': '请求体不能为空'})
                return
            
            post_data = self.rfile.read(content_length).decode('utf-8')
            request_data = json.loads(post_data)
            
            access_token = request_data.get('access_token', '').strip()
            dept_id = request_data.get('dept_id', '1').strip()
            clear_existing = request_data.get('clear_existing', False)
            
            logger.info(f"从请求体读取参数: access_token={bool(access_token)}, dept_id={dept_id}, clear_existing={clear_existing}")
            
            if not access_token:
                self.send_json_response({
                    'success': False,
                    'error': '参数不完整',
                    'message': '请提供 access_token'
                })
                return
            
            # 获取所有用户（处理分页）
            all_users = []
            cursor = "0"
            has_more = True
            page_count = 0
            
            while has_more:
                try:
                    page_count += 1
                    logger.info(f"正在获取第 {page_count} 页用户数据（cursor={cursor}）...")
                    result = self._get_department_user_list(access_token, dept_id, cursor)
                    users = result.get('list', [])
                    if isinstance(users, dict):
                        # 如果返回的是单个用户对象，转换为列表
                        users = [users]
                    
                    logger.info(f"第 {page_count} 页获取到 {len(users)} 个用户")
                    # 记录第一个用户的字段，用于调试
                    if users and page_count == 1:
                        first_user = users[0]
                        logger.info(f"第一个用户的字段: {list(first_user.keys())}")
                        logger.info(f"第一个用户数据示例: {json.dumps(first_user, ensure_ascii=False, indent=2)[:500]}")
                    all_users.extend(users)
                    
                    has_more = result.get('has_more', False)
                    if has_more:
                        cursor = result.get('next_cursor', '0')
                        logger.info(f"还有更多用户，继续获取，当前已获取 {len(all_users)} 个用户...")
                    else:
                        logger.info(f"已获取所有用户，共 {len(all_users)} 个（共 {page_count} 页）")
                        break
                except Exception as e:
                    logger.error(f"获取部门用户列表失败（第 {page_count} 页）: {e}")
                    break
            
            # 保存到Excel文件（智能合并，不清除管理员信息）
            if all_users:
                # 不再支持清空现有数据，总是智能合并
                # 策略：新增/更新用户，未搜索到的用户设为禁用（管理员除外）
                success = self.user_manager.save_dingtalk_users(all_users, append=False)
                if success:
                    logger.info("=" * 60)
                    logger.info(f"✅ 成功获取并保存 {len(all_users)} 个用户！")
                    logger.info("=" * 60)
                    
                    # 文件已更新，清除用户缓存（确保内存与文件一致）
                    self._invalidate_user_cache()
                    
                    self.send_json_response({
                        'success': True,
                        'message': f'成功获取并保存 {len(all_users)} 个用户',
                        'count': len(all_users),
                        'users': all_users  # 返回所有用户
                    })
                else:
                    self.send_json_response({
                        'success': False,
                        'error': '获取用户成功，但保存到文件失败',
                        'count': len(all_users)
                    })
            else:
                logger.warning("未获取到任何用户")
                self.send_json_response({
                    'success': False,
                    'error': '未获取到任何用户',
                    'message': '请检查 access_token 和 dept_id 是否有效'
                })
                
        except json.JSONDecodeError as e:
            logger.error(f"获取部门用户列表请求：无效的JSON数据: {e}")
            self.send_json_response({'success': False, 'error': '无效的JSON数据', 'message': f'后端未能解析前端发送的JSON请求体: {str(e)}'})
        except Exception as e:
            logger.error(f"处理获取部门用户列表请求的顶层异常: {e}", exc_info=True)
            self.send_json_response({'success': False, 'error': f'服务器错误: {str(e)}', 'message': '服务器在处理请求时发生意外错误。'})
    
    def _get_department_user_list(self, access_token, dept_id, cursor="0", size=100):
        """调用钉钉API获取指定部门的用户列表
        
        Args:
            access_token: 访问令牌
            dept_id: 部门ID
            cursor: 游标，用于分页
            size: 每页大小
        
        Returns:
            包含用户列表和分页信息的字典
        """
        import urllib.request
        import urllib.error
        import ssl
        
        url = "https://oapi.dingtalk.com/topapi/v2/user/list"
        
        # 构建请求参数（按照官方API格式）
        params = {
            'cursor': int(cursor) if isinstance(cursor, str) and cursor.isdigit() else cursor,
            'contain_access_limit': False,  # 布尔值，不是字符串
            'size': int(size) if isinstance(size, (str, int)) else 100,  # 数字，不是字符串
            'order_field': 'modify_desc',
            'language': 'zh_CN',
            'dept_id': int(dept_id) if isinstance(dept_id, str) and dept_id.isdigit() else dept_id  # 数字，不是字符串
        }
        
        # 构建完整URL（access_token作为查询参数）
        url_with_params = f"{url}?access_token={urllib.parse.quote(access_token)}"
        
        logger.info(f"调用获取部门用户列表API: {url}")
        logger.info(f"请求参数: dept_id={dept_id}, cursor={cursor}, size={size}")
        
        request_data = json.dumps(params).encode('utf-8')
        
        logger.info(f"使用urllib调用获取部门用户列表API: {url_with_params}")
        logger.info(f"请求体: {json.dumps(params, ensure_ascii=False)}")
        
        req = urllib.request.Request(
            url_with_params,
            data=request_data,
            headers={'Content-Type': 'application/json'},
            method='POST'
        )
        
        # 创建SSL上下文（跳过证书验证）
        ssl_context = ssl.create_default_context()
        ssl_context.check_hostname = False
        ssl_context.verify_mode = ssl.CERT_NONE
        
        try:
            with urllib.request.urlopen(req, timeout=30, context=ssl_context) as response:
                response_text = response.read().decode('utf-8')
                response_code = response.getcode()
                
                logger.info(f"获取部门用户列表响应状态码: {response_code}")
                
                if response_code == 200:
                    result = json.loads(response_text)
                    errcode = result.get('errcode', -1)
                    errmsg = result.get('errmsg', '')
                    
                    logger.info(f"获取部门用户列表响应: errcode={errcode}, errmsg={errmsg}")
                    
                    if errcode == 0:
                        result_data = result.get('result', {})
                        users = result_data.get('list', [])
                        if isinstance(users, dict):
                            users = [users]
                        logger.info(f"成功获取 {len(users)} 个用户")
                        return {
                            'list': users,
                            'has_more': result_data.get('has_more', False),
                            'next_cursor': result_data.get('next_cursor', '0')
                        }
                    else:
                        logger.error(f"获取部门用户列表失败: {errmsg} (errcode: {errcode})")
                        raise Exception(f"获取部门用户列表失败: {errmsg} (errcode: {errcode})")
                else:
                    logger.error(f"获取部门用户列表HTTP错误: {response_code}, 响应: {response_text[:200]}")
                    raise Exception(f"HTTP错误: {response_code}")
        except urllib.error.HTTPError as e:
            error_body = ''
            try:
                if e.fp:
                    error_body = e.read().decode('utf-8')
            except:
                pass
            logger.error(f"获取部门用户列表HTTP错误 {e.code}: {error_body[:200] if error_body else str(e)}")
            raise Exception(f"HTTP错误 {e.code}: {error_body[:200] if error_body else str(e)}")
        except urllib.error.URLError as e:
            logger.error(f"获取部门用户列表网络错误: {e}")
            raise Exception(f"网络错误: {str(e)}")
    
    def _get_dingtalk_user_info(self, auth_code, method=''):
        """调用钉钉API获取用户信息
        
        Args:
            auth_code: 授权码
            method: 免登方式 ('企业内部应用网页免登' 或 '小程序免登')
        """
        client_id = DINGTALK_CONFIG.get('client_id')
        client_secret = DINGTALK_CONFIG.get('client_secret')
        
        if not client_id or not client_secret:
            raise ValueError('钉钉Client ID或Client Secret未配置')
        
        # 判断免登方式
        if '企业内部应用网页免登' in method or 'requestAuthCode' in method:
            # 企业内部应用网页免登流程
            return self._get_enterprise_user_info(auth_code, client_id, client_secret)
        else:
            # 小程序免登流程（原有逻辑）
            return self._get_miniprogram_user_info(auth_code, client_id, client_secret)
    
    def _get_enterprise_user_info(self, auth_code, client_id, client_secret):
        """企业内部应用网页免登：通过授权码获取用户信息"""
        # 步骤1: 获取access_token
        access_token = self._get_dingtalk_access_token(client_id, client_secret)
        if not access_token:
            raise Exception('获取access_token失败')
        
        # 步骤2: 通过access_token和授权码获取用户信息
        user_info = self._get_userinfo_by_code(access_token, auth_code)
        if not user_info:
            raise Exception('获取用户信息失败')
        
        # 步骤3: 获取用户详情（可选，获取更完整的用户信息）
        userid = user_info.get('userid')
        if userid:
            try:
                user_detail = self._get_user_detail(access_token, userid)
                if user_detail:
                    # 合并用户信息和详情
                    user_info.update(user_detail)
            except Exception as e:
                logger.warning(f"获取用户详情失败，使用基本信息: {e}")
        
        return user_info
    
    def _get_miniprogram_user_info(self, auth_code, client_id, client_secret):
        """小程序免登：通过授权码获取用户信息"""
        api_url = DINGTALK_CONFIG.get('sns_api_url', 'https://oapi.dingtalk.com/sns/getuserinfo_bycode')
        
        # 构建请求参数
        params = {
            'tmp_auth_code': auth_code
        }
        
        # 使用requests库（如果可用）或urllib
        if HAS_REQUESTS:
            response = requests.post(
                api_url,
                json=params,
                params={
                    'accessKey': client_id,
                    'timestamp': str(int(time.time() * 1000))
                },
                headers={
                    'Content-Type': 'application/json'
                },
                timeout=10
            )
            
            if response.status_code == 200:
                result = response.json()
                if result.get('errcode') == 0:
                    return result.get('user_info', {})
                else:
                    raise Exception(f"钉钉API错误: {result.get('errmsg', '未知错误')} (errcode: {result.get('errcode')})")
            else:
                raise Exception(f"HTTP错误: {response.status_code}")
        else:
            # 使用urllib作为备选方案
            import urllib.request
            import urllib.error
            
            timestamp = str(int(time.time() * 1000))
            request_data = json.dumps(params).encode('utf-8')
            url_with_params = f"{api_url}?accessKey={urllib.parse.quote(client_id)}&timestamp={timestamp}"
            
            req = urllib.request.Request(
                url_with_params,
                data=request_data,
                headers={
                    'Content-Type': 'application/json'
                }
            )
            
            try:
                with urllib.request.urlopen(req, timeout=10) as response:
                    result = json.loads(response.read().decode('utf-8'))
                    if result.get('errcode') == 0:
                        return result.get('user_info', {})
                    else:
                        raise Exception(f"钉钉API错误: {result.get('errmsg', '未知错误')} (errcode: {result.get('errcode')})")
            except urllib.error.HTTPError as e:
                error_body = e.read().decode('utf-8') if e.fp else ''
                raise Exception(f"HTTP错误 {e.code}: {error_body}")
            except urllib.error.URLError as e:
                raise Exception(f"网络错误: {str(e)}")
    
    def _get_dingtalk_access_token(self, client_id, client_secret):
        """获取钉钉access_token（使用新API格式）
        
        参照官方API文档（新版本）：
        POST /v1.0/oauth2/{corp_id}/token HTTP/1.1
        Host: api.dingtalk.com
        Content-Type: application/json
        
        {
            "client_id": "suite123",
            "client_secret": "********",
            "grant_type": "client_credentials"
        }
        
        返回格式：
        {
            "access_token": "2bf******9be361a5084f1e2b8",
            "expires_in": 7200
        }
        """
        # 获取corp_id（必需参数）
        corp_id = DINGTALK_CONFIG.get('corp_id', '')
        if not corp_id:
            logger.error("获取access_token失败: corp_id未配置")
            raise Exception("获取access_token失败: corp_id未配置，请在配置文件中设置corp_id")
        
        # 构建新API URL
        url = f"https://api.dingtalk.com/v1.0/oauth2/{corp_id}/token"
        
        headers = {
            'Content-Type': 'application/json',
            'Host': 'api.dingtalk.com',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        # 使用新API的参数格式
        payload = {
            'client_id': client_id,
            'client_secret': client_secret,
            'grant_type': 'client_credentials'
        }
        
        logger.info(f"调用获取access_token API (新格式): {url}")
        logger.info(f"请求参数: client_id={client_id[:10]}..., client_secret=***, grant_type=client_credentials")
        
        try:
            # 使用Python标准库urllib发送HTTP请求（与成功的接口保持一致）
            import urllib.request
            import urllib.error
            import ssl
            
            # 准备请求数据
            request_data_bytes = json.dumps(payload).encode('utf-8')
            
            # 创建请求对象
            req = urllib.request.Request(url, data=request_data_bytes, headers=headers, method='POST')
            
            # 创建SSL上下文（跳过证书验证）
            ssl_context = ssl.create_default_context()
            ssl_context.check_hostname = False
            ssl_context.verify_mode = ssl.CERT_NONE
            
            # 发送请求
            logger.info("发送HTTP POST请求到钉钉API...")
            with urllib.request.urlopen(req, timeout=30, context=ssl_context) as response:
                response_text = response.read().decode('utf-8')
                response_code = response.getcode()
                
                logger.info(f"获取access_token响应状态码: {response_code}")
                
                if response_code == 200:
                    # 检查响应是否是HTML（可能是重定向或错误页面）
                    if response_text.strip().startswith('<!DOCTYPE') or response_text.strip().startswith('<html'):
                        logger.warning("⚠️  API返回了HTML页面而不是JSON，可能是重定向或错误页面")
                        logger.warning(f"响应内容预览: {response_text[:500]}")
                        # 自动重试一次
                        logger.info("   自动重试获取AccessToken...")
                        import time
                        time.sleep(1)  # 等待1秒后重试
                        
                        # 重试请求
                        retry_req = urllib.request.Request(url, data=request_data_bytes, headers=headers, method='POST')
                        with urllib.request.urlopen(retry_req, timeout=30, context=ssl_context) as retry_response:
                            retry_response_text = retry_response.read().decode('utf-8')
                            retry_response_code = retry_response.getcode()
                            
                            if retry_response_code == 200:
                                if retry_response_text.strip().startswith('<!DOCTYPE') or retry_response_text.strip().startswith('<html'):
                                    logger.error("❌ 重试后仍然返回HTML页面")
                                    raise Exception(f"获取access_token失败: 钉钉API返回了HTML页面而不是JSON")
                                else:
                                    response_text = retry_response_text
                                    response_code = retry_response_code
                                    logger.info("✅ 重试成功，获取到JSON响应")
                            else:
                                # 重试失败，抛出异常
                                raise Exception(f"获取access_token失败: 重试请求返回状态码 {retry_response_code}")
                    # 如果原始响应不是HTML，直接使用（不需要else，因为response_text和response_code已经是正确的值）
                    
                    # 尝试解析JSON
                    try:
                        result = json.loads(response_text)
                        logger.info(f"获取access_token响应: {json.dumps(result, ensure_ascii=False)}")
                    
                        # 新API返回格式：access_token 和 expires_in
                        access_token = result.get('access_token')
                        expires_in = result.get('expires_in', 7200)
                        
                        if access_token:
                            logger.info(f"成功获取access_token: {access_token[:20]}... (有效期: {expires_in}秒)")
                            return access_token
                        else:
                            error_code = result.get('error', '')
                            error_msg = result.get('error_description') or result.get('error') or '未知错误'
                            logger.error(f"获取access_token失败: {error_msg} (error: {error_code})")
                            raise Exception(f"获取access_token失败: {error_msg} (error: {error_code})")
                    except json.JSONDecodeError as json_err:
                        logger.error(f"解析JSON响应失败: {json_err}")
                        logger.error(f"响应内容预览: {response_text[:500]}")
                        raise Exception(f"获取access_token失败: 钉钉API返回了无效的JSON格式")
                else:
                    error_text = response_text if response_text else '无响应内容'
                    logger.error(f"API返回错误，状态码: {response_code}, 响应: {error_text[:200]}")
                    raise Exception(f"HTTP错误: {response_code}")
        except urllib.error.HTTPError as e:
            error_body = e.read().decode('utf-8') if e.fp else ''
            logger.error(f"获取access_token HTTP错误 {e.code}: {error_body[:200] if error_body else str(e)}")
            raise Exception(f"HTTP错误 {e.code}: {error_body[:200] if error_body else str(e)}")
        except urllib.error.URLError as e:
            logger.error(f"获取access_token 网络错误: {str(e)}")
            raise Exception(f"网络错误: {str(e)}")
        except Exception as e:
            logger.error(f"获取access_token时发生错误: {e}", exc_info=True)
            raise
    
    def _get_userinfo_by_code(self, access_token, auth_code):
        """通过access_token和授权码获取用户信息
        
        参照官方API文档：
        POST https://oapi.dingtalk.com/topapi/v2/user/getuserinfo?access_token=ACCESS_TOKEN
        请求正文：{"code":"bab02f63c1e030fbbxxxx"}
        
        返回格式：
        {
            "errcode": 0,
            "result": {
                "associated_unionid": "N2o5U3axxxx",
                "unionid": "gliiW0piiii02zBUjUxxxx",
                "device_id": "12drtfxxxxx",
                "sys_level": 1,
                "name": "张xx",
                "sys": true,
                "userid": "userid123"
            },
            "errmsg": "ok"
        }
        """
        userinfo_url = DINGTALK_CONFIG.get('userinfo_url', 'https://oapi.dingtalk.com/topapi/v2/user/getuserinfo')
        
        # 构建请求参数（完全按照官方示例格式）
        params = {
            'code': auth_code
        }
        
        # 构建完整URL（access_token作为查询参数）
        url = f"{userinfo_url}?access_token={access_token}"
        
        logger.info(f"调用获取用户信息API: {userinfo_url}")
        logger.info(f"请求参数: code={auth_code[:10]}...")
        
        # 使用urllib（不依赖requests库，使用HTTP方式）
        import urllib.request
        import urllib.error
        import urllib.parse
        import ssl
        
        request_data = json.dumps(params).encode('utf-8')
        url_with_params = f"{userinfo_url}?access_token={urllib.parse.quote(access_token)}"
        
        logger.info(f"使用urllib调用获取用户信息API: {url_with_params}")
        logger.info(f"请求体: {json.dumps(params, ensure_ascii=False)}")
        
        req = urllib.request.Request(
            url_with_params,
            data=request_data,
            headers={'Content-Type': 'application/json'},
            method='POST'
        )
        
        # 创建SSL上下文（跳过证书验证）
        ssl_context = ssl.create_default_context()
        ssl_context.check_hostname = False
        ssl_context.verify_mode = ssl.CERT_NONE
        
        try:
            with urllib.request.urlopen(req, timeout=30, context=ssl_context) as response:
                response_text = response.read().decode('utf-8')
                response_code = response.getcode()
                
                logger.info(f"获取用户信息响应状态码: {response_code}")
                
                if response_code == 200:
                    result = json.loads(response_text)
                    logger.info(f"获取用户信息响应: errcode={result.get('errcode')}, errmsg={result.get('errmsg')}")
                    
                    if result.get('errcode') == 0:
                        user_info = result.get('result', {})
                        logger.info(f"成功获取用户信息: userid={user_info.get('userid')}, name={user_info.get('name')}")
                        return user_info
                    else:
                        errmsg = result.get('errmsg', '未知错误')
                        errcode = result.get('errcode')
                        logger.error(f"获取用户信息失败: {errmsg} (errcode: {errcode})")
                        raise Exception(f"获取用户信息失败: {errmsg} (errcode: {errcode})")
                else:
                    logger.error(f"获取用户信息HTTP错误: {response_code}, 响应: {response_text[:200]}")
                    raise Exception(f"HTTP错误: {response_code}")
        except urllib.error.HTTPError as e:
            error_body = ''
            try:
                if e.fp:
                    error_body = e.read().decode('utf-8')
            except:
                pass
            logger.error(f"获取用户信息HTTP错误 {e.code}: {error_body[:200] if error_body else str(e)}")
            raise Exception(f"HTTP错误 {e.code}: {error_body[:200] if error_body else str(e)}")
        except urllib.error.URLError as e:
            logger.error(f"获取用户信息网络错误: {e}")
            raise Exception(f"网络错误: {str(e)}")
    
    def _exclude_from_todo_user_selection(self, user: Dict[str, Any]) -> bool:
        """待办/阅读人选中排除系统最高管理员（不必选、也不应收到待办）。"""
        if not user:
            return True
        if self.user_manager.is_super_admin(user):
            return True
        name = (user.get('name') or '').strip()
        return name in ('系统最高管理员', '最高管理员')

    def _filter_valid_todo_userids(self, userids: List[str]) -> List[str]:
        """过滤空/重复及系统最高管理员的待办 userid。"""
        seen = set()
        valid = []
        for uid in userids or []:
            uid_str = str(uid).strip() if uid is not None else ''
            if not uid_str or uid_str in seen:
                continue
            try:
                u = self.user_manager.get_user_by_userid(uid_str)
                if u and self._exclude_from_todo_user_selection(u):
                    continue
            except Exception:
                pass
            seen.add(uid_str)
            valid.append(uid_str)
        return valid

    def _resolve_dingtalk_userid(self, identifier: str) -> Optional[str]:
        """将 userid / username / 数字 id 解析为钉钉 userid。"""
        ident = str(identifier or '').strip()
        if not ident:
            return None

        try:
            user = self.user_manager.get_user_by_userid(ident)
            if user and user.get('userid'):
                return str(user['userid']).strip()
        except Exception:
            pass

        try:
            user = self.user_manager.get_user_by_username(ident)
            if user and user.get('userid'):
                return str(user['userid']).strip()
        except Exception:
            pass

        try:
            user = self.user_manager.get_user_by_id(int(ident))
            if user and user.get('userid'):
                return str(user['userid']).strip()
        except (ValueError, TypeError):
            pass

        if ident.isdigit() and len(ident) >= 10:
            return ident

        for user in self.user_manager.get_all_users():
            if (
                str(user.get('userid', '')).strip() == ident
                or str(user.get('username', '')).strip() == ident
                or str(user.get('id', '')).strip() == ident
            ):
                uid = str(user.get('userid', '') or '').strip()
                if uid:
                    return uid
        return None

    def _resolve_dingtalk_userids(self, identifiers: List[str]) -> List[str]:
        """批量解析并去重钉钉 userid。"""
        seen = set()
        resolved = []
        for ident in identifiers or []:
            uid = self._resolve_dingtalk_userid(ident) or str(ident or '').strip()
            if uid and uid not in seen:
                seen.add(uid)
                resolved.append(uid)
        return resolved

    def _build_public_base_url(self) -> str:
        """构建对外访问根 URL（工作通知、钉钉跳转）。"""
        if PUBLIC_BASE_URL:
            return PUBLIC_BASE_URL.rstrip('/')
        host = (self.headers.get('Host') or '').strip()
        if host:
            proto = (self.headers.get('X-Forwarded-Proto') or 'http').split(',')[0].strip()
            return f"{proto}://{host}".rstrip('/')
        return ''

    def _build_announcement_detail_url(self, announcement_id: str) -> str:
        """构建公告详情页的钉钉 openapp 跳转链接。"""
        from server.dingtalk_url_util import build_announcement_detail_dingtalk_url
        return build_announcement_detail_dingtalk_url(announcement_id, self._build_public_base_url())

    def _build_review_center_url(self) -> str:
        """构建审核中心的钉钉 openapp 跳转链接。"""
        from server.dingtalk_url_util import build_review_center_dingtalk_url
        return build_review_center_dingtalk_url(self._build_public_base_url())

    def _enrich_todo_user_display(self, todo: Dict[str, Any]) -> None:
        """就地补充待办记录的姓名（优先保留 Excel 已有姓名）。"""
        userid = str(todo.get('userid', '') or '').strip()
        excel_name = str(todo.get('name', '') or '').strip()
        excel_username = str(todo.get('username', '') or '').strip()

        if not userid:
            if not excel_name and not excel_username:
                todo['name'] = '未知用户'
                todo['username'] = '未知用户'
            return

        user_info = self.user_manager.get_user_by_userid(userid)
        if user_info:
            todo['name'] = user_info.get('name', '') or user_info.get('username', '') or excel_name or str(userid)
            todo['username'] = user_info.get('username', '') or excel_username or str(userid)
            if user_info.get('unionid'):
                todo['unionid'] = user_info.get('unionid')
            return

        if excel_name and excel_name != userid:
            todo['name'] = excel_name
        else:
            todo['name'] = f"用户({userid})"
        if excel_username and excel_username != userid:
            todo['username'] = excel_username
        else:
            todo['username'] = str(userid)

    def _resolve_unionid_from_identifier(self, identifier: str) -> Optional[str]:
        """将 userid 或 unionid 配置项解析为钉钉 API 所需的 unionid。"""
        ident = str(identifier or '').strip()
        if not ident:
            return None

        user = self.user_manager.get_user_by_userid(ident)
        if user and user.get('unionid'):
            return str(user['unionid']).strip()

        # 纯数字且库中无 unionid：不能当作 unionid 传给钉钉（会导致 403/500）
        if ident.isdigit():
            logger.warning(
                f"标识 {ident} 形如 userid，但用户库中无 unionid，请先同步钉钉用户"
            )
            return None

        # 非纯数字：视为 unionid（钉钉 unionid 通常含字母）
        for u in self.user_manager.get_all_users():
            if str(u.get('unionid', '')).strip() == ident:
                return ident
        return ident

    def _sanitize_knowledge_doc_name(self, title: str) -> str:
        """知识库文档名称清洗（长度与非法控制字符）。"""
        import re
        name = (title or '未命名公告').strip()
        name = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', name)
        if len(name) > 200:
            name = name[:200]
        return name or '未命名公告'

    def _resolve_knowledge_doc_operator(self, announcement: Dict[str, Any]) -> tuple:
        """
        解析知识库文档 operatorId（unionid）。
        Returns: (operator_unionid, author_userid_for_notify, error_message)
        """
        author_userid = str(announcement.get('author_userid') or '').strip()
        author_name = announcement.get('author', '') or announcement.get('original_author', '')

        user = None
        if author_userid:
            user = self.user_manager.get_user_by_userid(author_userid)
        if not user and author_name:
            for u in self.user_manager.get_all_users():
                if u.get('name') == author_name or u.get('username') == author_name:
                    user = u
                    break

        unionid = (user.get('unionid', '') if user else '') or ''
        notify_userid = (user.get('userid', '') if user else '') or author_userid

        if not unionid and DINGTALK_DOC_OPERATOR_UNIONID:
            delegate_unionid = self._resolve_unionid_from_identifier(DINGTALK_DOC_OPERATOR_UNIONID)
            if delegate_unionid:
                logger.info(
                    f"发起人 {author_name!r} 无 unionid，使用代创建人 unionid "
                    f"({delegate_unionid[:8]}...)"
                )
                return delegate_unionid, notify_userid, None
            return None, notify_userid, (
                f"环境变量 DINGTALK_DOC_OPERATOR_UNIONID 当前值为 {DINGTALK_DOC_OPERATOR_UNIONID!r}，"
                "看起来像钉钉 userid 而非 unionid。请改为 unionid（用户管理/钉钉用户数据中的 unionid 列），"
                "或先执行「同步钉钉用户」后再填 userid。"
            )

        if not unionid:
            return None, notify_userid, (
                f"无法找到发起人 {author_name!r} 的 unionid；"
                f"请在钉钉知识库授权该用户，或配置环境变量 DINGTALK_DOC_OPERATOR_UNIONID（须为 unionid 或可解析的 userid）"
            )
        return unionid, notify_userid, None

    def _get_user_unionid_by_userid(self, userid: str):
        """根据userid获取用户的unionid
        
        Args:
            userid: 钉钉userid
        
        Returns:
            unionid字符串，如果未找到返回None
        """
        try:
            user = self.user_manager.get_user_by_userid(userid)
            if user:
                return user.get('unionid', '')
            return None
        except Exception as e:
            logger.error(f"获取用户unionid失败 (userid={userid}): {e}", exc_info=True)
            return None
    
    def _create_announcement_todos(self, announcement_id: str, todo_userids: List[str], approver_user: Dict[str, Any], retry_failed_only: bool = False, max_retries: int = None, skip_notification: bool = False) -> tuple:
        """为公告创建本地待办记录（不再调用钉钉API，仅创建本地Excel文件并发送工作通知）
        
        Args:
            announcement_id: 公告ID
            todo_userids: 待办人员userid列表
            approver_user: 审批管理员用户信息
            retry_failed_only: 是否只重试失败的用户（已废弃，保留兼容性）
            max_retries: 最大重试次数（已废弃，保留兼容性）
            skip_notification: 为 True 时不发送阅读工作通知（由审批流程统一发送）
        
        Returns:
            返回(source_id, user_task_ids, user_source_ids, notification_sent, failed_userids)
            source_id: 基础sourceId（用于本地记录）
            user_task_ids: {}（空字典，不再创建钉钉待办）
            user_source_ids: {}（空字典，不再创建钉钉待办）
            notification_sent: 是否已发送工作通知（True/False）
            failed_userids: 创建失败的用户ID列表（通常为空）
        """
        import urllib.parse
        from datetime import datetime
        import time
        
        try:
            todo_userids = self._filter_valid_todo_userids(todo_userids)
            # 获取公告信息（包括temp目录中的公告）
            announcement = self.announcement_mgr.get_announcement(announcement_id)
            if not announcement:
                logger.error(f"获取公告信息失败: {announcement_id}")
                return None, {}, {}, False, []
            
            title = announcement.get('title', '')
            
            # 获取待办人员的userid列表（去重，避免同一用户被多次选择）
            valid_userids = []  # 保存有效的userid
            unionids = []  # 保存对应的unionid
            user_names = []  # 保存用户姓名
            user_usernames = []  # 保存用户名
            seen_userids = set()  # 用于去重
            skipped_users = []  # 记录被跳过的用户及原因
            
            for userid in todo_userids:
                # 过滤空值
                if not userid or not str(userid).strip():
                    skipped_users.append((userid, "userid为空"))
                    logger.warning(f"跳过空的待办用户: userid={repr(userid)}")
                    continue
                
                # 去重
                if userid in seen_userids:
                    skipped_users.append((userid, "重复的用户"))
                    logger.warning(f"跳过重复的待办用户: {userid}")
                    continue
                seen_userids.add(userid)
                
                # 获取unionid和用户信息
                unionid = self._get_user_unionid_by_userid(userid)
                # 即使unionid获取失败，也记录到Excel（使用空unionid）
                unionids.append(unionid if unionid else '')
                valid_userids.append(userid)
                
                # 获取用户信息（姓名和用户名）
                user_info = self.user_manager.get_user_by_userid(userid)
                if user_info:
                    user_name = user_info.get('name', '') or user_info.get('username', '') or str(userid)
                    user_username = user_info.get('username', '') or str(userid)
                else:
                    # 如果获取失败，使用userid作为默认值
                    user_name = str(userid)
                    user_username = str(userid)
                    logger.warning(f"用户 {userid} 的信息获取失败，使用userid作为默认值")
                
                user_names.append(user_name)
                user_usernames.append(user_username)
                
                if not unionid:
                    logger.warning(f"用户 {userid} 的unionid获取失败，但仍会创建Excel记录（unionid为空）")
            
            if not valid_userids:
                logger.warning("没有有效的待办人员，跳过创建待办")
                return None, {}, {}, False, []
            
            # 记录统计信息
            total_input = len(todo_userids)
            total_valid = len(valid_userids)
            total_skipped = len(skipped_users)
            logger.info(f"准备为公告 {announcement_id} 创建本地待办记录，输入: {total_input}人, 有效: {total_valid}人, 跳过: {total_skipped}人")
            if skipped_users:
                logger.info(f"跳过的用户详情: {', '.join([f'{uid}({reason})' for uid, reason in skipped_users[:10]])}" + (f'... 等{len(skipped_users)}个' if len(skipped_users) > 10 else ''))
            
            detail_url_str = self._build_announcement_detail_url(announcement_id)
            
            # 构建sourceId（使用公告ID和当前时间戳）
            timestamp = int(time.time() * 1000)
            source_id = f"announcement_{announcement_id}_{timestamp}"
            
            access_token = self._get_dingtalk_access_token_simple()
            
            # 创建本地Excel文件存储待办状态（不再调用钉钉API）
            logger.info(f"创建本地待办Excel文件: {announcement_id}, 共 {len(valid_userids)} 条记录")
            todo_success = self.todo_mgr.create_todo_file(
                announcement_id, 
                source_id, 
                title, 
                valid_userids,  # 所有选择的待办用户
                unionids,       # 与valid_userids一一对应
                task_ids={},    # 空字典，不再创建钉钉待办
                user_source_ids={},  # 空字典，不再创建钉钉待办
                user_names=user_names,  # 用户姓名列表
                user_usernames=user_usernames  # 用户名列表
            )
            
            if not todo_success:
                logger.error(f"创建本地待办Excel文件失败: {announcement_id}")
                return None, {}, {}, False, valid_userids
            
            logger.info(f"✅ 成功创建本地待办Excel文件: {announcement_id}, 共 {len(valid_userids)} 条记录（不再调用钉钉API创建待办）")
            
            notification_sent = False
            if not skip_notification:
                if not access_token:
                    access_token = self._get_dingtalk_access_token_simple()
                if access_token and valid_userids:
                    try:
                        logger.info(f"发送工作通知: 公告={title}, 接收者={len(valid_userids)}人")
                        notification_sent, notify_err = self._send_announcement_notification(
                            announcement_id,
                            title,
                            detail_url_str,
                            valid_userids,
                            access_token,
                            text=f"请阅读公告：{title}",
                        )
                        if notification_sent:
                            logger.info(f"✅ 成功为公告 {announcement_id} 发送阅读通知给 {len(valid_userids)} 人")
                        else:
                            logger.warning(f"⚠️ 为公告 {announcement_id} 发送阅读通知失败: {notify_err}")
                    except Exception as e:
                        logger.error(f"发送工作通知时发生异常: {e}", exc_info=True)
                elif not access_token:
                    logger.warning("access_token为空，无法发送工作通知")
            
            return source_id, {}, {}, notification_sent, []
                
        except Exception as e:
            logger.error(f"创建本地待办记录时发生错误: {e}", exc_info=True)
            return None, {}, {}, False, []
    
    def _query_user_todo_by_source_id(self, unionid: str, source_id: str) -> Optional[Dict[str, Any]]:
        """通过sourceId查询用户的待办任务
        
        Args:
            unionid: 用户的unionid
            source_id: 待办的sourceId
        
        Returns:
            待办任务信息，如果未找到返回None
        """
        import urllib.request
        import urllib.error
        import urllib.parse
        import ssl
        
        try:
            # 获取access_token
            access_token = self._get_dingtalk_access_token(
                DINGTALK_CONFIG.get('client_id'),
                DINGTALK_CONFIG.get('client_secret')
            )
            if not access_token:
                logger.error("获取钉钉access_token失败，无法查询待办")
                return None
            
            # 构建查询API URL（通过sourceId查询）
            # 注意：钉钉API可能需要通过其他方式查询，这里先尝试通过sourceId查询
            api_url = f"https://api.dingtalk.com/v1.0/todo/users/{unionid}/tasks?sourceId={urllib.parse.quote(source_id)}"
            
            # 创建SSL上下文
            ssl_context = ssl.create_default_context()
            ssl_context.check_hostname = False
            ssl_context.verify_mode = ssl.CERT_NONE
            
            req = urllib.request.Request(
                api_url,
                headers={
                    'x-acs-dingtalk-access-token': access_token
                },
                method='GET'
            )
            
            with urllib.request.urlopen(req, timeout=30, context=ssl_context) as response:
                if response.status == 200:
                    response_data = json.loads(response.read().decode('utf-8'))
                    # 查找匹配sourceId的待办
                    if isinstance(response_data, dict):
                        tasks = response_data.get('result', {}).get('tasks', []) or response_data.get('tasks', [])
                        for task in tasks:
                            if task.get('sourceId') == source_id:
                                return task
                    elif isinstance(response_data, list):
                        for task in response_data:
                            if task.get('sourceId') == source_id:
                                return task
                else:
                    logger.warning(f"查询待办失败，状态码: {response.status}")
                    return None
        except urllib.error.HTTPError as e:
            error_body = e.read().decode('utf-8') if e.fp else ''
            logger.warning(f"查询待办HTTP错误 {e.code}: {error_body[:200]}")
            return None
        except Exception as e:
            logger.warning(f"查询待办时发生错误: {e}")
            return None
        
        return None
    
    def _send_announcement_notification(
        self,
        announcement_id: str,
        title: str,
        detail_url: str,
        userids: List[str],
        access_token: str,
        text: str = None,
        max_retries: int = None,
        retry_delay: float = None,
    ) -> tuple:
        """发送钉钉工作通知。Returns: (success, error_message)"""
        from server.dingtalk_notify_util import send_corpconversation_with_retry

        resolved = self._resolve_dingtalk_userids(userids)
        valid_userids = self._filter_valid_todo_userids(resolved)
        if not valid_userids:
            return False, '没有有效的待通知用户 userid'

        if text is None:
            text = f"请及时阅读公告：{title}"

        msg_content = {
            "msgtype": "link",
            "link": {
                "title": title,
                "text": text,
                "messageUrl": detail_url,
                "picUrl": "https://img.alicdn.com/imgextra/i1/O1CN01Kq8eYq1xWqJY5Y5Y5_!!6000000006441-2-tps-200-200.png",
            },
        }
        logger.info(f"发送工作通知: 公告={title}, 接收者={len(valid_userids)}人")
        return send_corpconversation_with_retry(
            access_token,
            valid_userids,
            msg_content,
            max_retries=max_retries,
            retry_base_sec=retry_delay,
            log_context=f" 公告={announcement_id}",
        )

    def _notify_announcement_readers(
        self,
        announcement_id: str,
        title: str,
        userids: List[str],
        access_token: Optional[str] = None,
    ) -> tuple:
        """向阅读人员发送工作通知。Returns: (sent, error_message)"""
        if not userids:
            return False, '未指定阅读人员'
        if not access_token:
            access_token = self._get_dingtalk_access_token_simple()
        if not access_token:
            config_ok, config_err = check_dingtalk_config()
            return False, config_err if not config_ok else '无法获取 access_token'
        detail_url = self._build_announcement_detail_url(announcement_id)
        return self._send_announcement_notification(
            announcement_id,
            title,
            detail_url,
            userids,
            access_token,
            text=f"请阅读公告：{title}",
        )

    def _send_pending_approval_notification(
        self,
        announcement_id: str,
        title: str,
        approver_identifier: Optional[str] = None,
    ) -> tuple:
        """发送待审批工作通知。Returns: (sent, error_message)"""
        try:
            access_token = self._get_dingtalk_access_token_simple()
            if not access_token:
                config_ok, config_err = check_dingtalk_config()
                return False, config_err if not config_ok else '无法获取 access_token'

            detail_url = self._build_review_center_url()
            notification_text = f"有新的公告《{title}》需要您审批，请及时处理。"

            if approver_identifier:
                approver_userid = self._resolve_dingtalk_userid(approver_identifier)
                if not approver_userid:
                    return False, f'无法解析审批人 userid: {approver_identifier}'
                success, err = self._send_announcement_notification(
                    announcement_id=announcement_id,
                    title=f"待审批公告：{title}",
                    detail_url=detail_url,
                    userids=[approver_userid],
                    access_token=access_token,
                    text=notification_text,
                )
                if success:
                    logger.info(f"成功发送审批通知给审批人: {approver_identifier} (userid={approver_userid})")
                else:
                    logger.warning(f"发送审批通知失败: {err}")
                return success, err

            approvers = self._get_approvers()
            if not approvers:
                return False, '未找到审批人'

            userids = []
            for approver in approvers:
                ident = approver.get('userid') or approver.get('username') or str(approver.get('id', ''))
                uid = self._resolve_dingtalk_userid(ident)
                if uid and uid not in userids:
                    userids.append(uid)

            if not userids:
                return False, '审批人没有有效的 userid，请先同步钉钉用户'

            success, err = self._send_announcement_notification(
                announcement_id=announcement_id,
                title=f"待审批公告：{title}",
                detail_url=detail_url,
                userids=userids,
                access_token=access_token,
                text=notification_text,
            )
            if success:
                logger.info(f"成功发送审批通知给 {len(userids)} 位审批人")
            else:
                logger.warning(f"发送审批通知失败: {err}")
            return success, err
        except Exception as e:
            logger.error(f"发送待审批通知时发生错误: {e}", exc_info=True)
            return False, str(e)
    
    def _get_dingtalk_access_token_simple(self) -> Optional[str]:
        """获取钉钉access_token（简化版本，从配置中读取client_id和client_secret）
        
        Returns:
            access_token字符串，如果获取失败返回None
        """
        try:
            from server.config import DINGTALK_CONFIG
            client_id = DINGTALK_CONFIG.get('client_id', '')
            client_secret = DINGTALK_CONFIG.get('client_secret', '')
            
            if not client_id or not client_secret:
                logger.warning("钉钉配置不完整，无法获取access_token")
                return None
            
            return self._get_dingtalk_access_token(client_id, client_secret)
        except Exception as e:
            logger.error(f"获取钉钉access_token失败: {e}", exc_info=True)
            return None
    
    def _get_approvers(self) -> List[Dict[str, Any]]:
        """获取审批人列表（仅从管理组中选取）
        
        只返回具有 management、admin 或 super_admin 角色的用户
        支持所有管理组用户，即使没有userid也返回（使用username或id作为标识）
        
        Returns:
            审批人用户列表（包含userid、username、id、name等字段）
        """
        try:
            # 优先从内存缓存获取所有激活状态的用户（避免文件I/O）
            preloader = get_data_preloader()
            all_users = None
            if preloader and PRELOAD_USERS:
                try:
                    all_users = preloader.get_users(status='active')
                    logger.debug(f"从内存缓存获取到 {len(all_users)} 个激活用户")
                except Exception as e:
                    logger.warning(f"从内存缓存获取用户失败，使用文件读取: {e}")
                    all_users = self.user_manager.get_all_users(statuses=[STATUS_ACTIVE])
            else:
                # 如果未启用预加载，从文件读取
                all_users = self.user_manager.get_all_users(statuses=[STATUS_ACTIVE])
            
            approvers = []
            approver_identifiers = set()  # 用于去重（使用userid、username或id）
            
            config = ANNOUNCEMENT_APPROVERS or {}
            approver_roles = set(config.get('roles', ['management', 'admin']))
            approver_roles.update(['super_admin'])
            title_patterns = config.get('titles', [])
            configured_userids = {str(uid).strip() for uid in config.get('userids', []) if uid}
            
            for user in all_users:
                user_roles = self.user_manager._parse_roles(user.get('roles', []))
                has_management_role = any(role in approver_roles for role in user_roles)
                
                user_title = (user.get('job_position') or user.get('title') or '').strip()
                has_title_match = any(
                    pattern and pattern in user_title for pattern in title_patterns
                )
                
                user_identifiers = {
                    str(user.get('userid', '')).strip(),
                    str(user.get('username', '')).strip(),
                    str(user.get('id', '')).strip(),
                }
                user_identifiers.discard('')
                has_userid_match = bool(user_identifiers & configured_userids)
                
                # 符合审批条件的用户加入列表（不要求必须有userid）
                if has_management_role or has_title_match or has_userid_match:
                    userid = user.get('userid', '')
                    username = user.get('username', '')
                    user_id = user.get('id')
                    
                    # 使用userid、username或id作为唯一标识
                    identifier = userid or username or str(user_id) if user_id else None
                    
                    if identifier and identifier not in approver_identifiers:
                        approvers.append({
                            'userid': str(userid) if userid else '',
                            'username': str(username) if username else '',
                            'id': user_id,
                            'name': user.get('name', ''),
                            'unionid': user.get('unionid', ''),
                            'source': 'management_group'
                        })
                        approver_identifiers.add(identifier)
            
            logger.info(f"找到 {len(approvers)} 位审批人（来源: 管理组）")
            return approvers
        except Exception as e:
            logger.error(f"获取审批人列表失败: {e}", exc_info=True)
            return []
    
    def _get_hardware_dept_heads(self) -> List[Dict[str, Any]]:
        """获取硬件研发部部长角色的用户（兼容旧方法）
        
        Returns:
            硬件研发部部长用户列表（包含userid和unionid）
        """
        # 使用新的配置化方法
        return self._get_approvers()
    
    def _notify_department_heads_for_approval(self, announcement_id: str, title: str) -> tuple:
        """通知默认审批人（管理组）。Returns: (sent, error_message)"""
        return self._send_pending_approval_notification(announcement_id, title, approver_identifier=None)

    def _notify_selected_approver(self, announcement_id: str, title: str, approver_identifier: str) -> tuple:
        """通知指定审批人。Returns: (sent, error_message)"""
        if not approver_identifier:
            logger.warning("审批人标识符为空，跳过发送审批通知")
            return False, '审批人标识符为空'
        return self._send_pending_approval_notification(
            announcement_id, title, approver_identifier=approver_identifier
        )
    
    def _post_knowledge_doc_request(
        self, api_url: str, access_token: str, request_body: Dict[str, Any], ssl_context
    ) -> tuple:
        """单次调用钉钉创建知识库文档。Returns: (doc_url, error_message, http_code)"""
        import urllib.request
        import urllib.error

        request_data = json.dumps(request_body, ensure_ascii=False).encode('utf-8')
        req = urllib.request.Request(
            api_url,
            data=request_data,
            headers={
                'Content-Type': 'application/json',
                'x-acs-dingtalk-access-token': access_token,
                'Host': 'api.dingtalk.com',
            },
            method='POST',
        )
        with urllib.request.urlopen(req, timeout=30, context=ssl_context) as response:
            response_data = json.loads(response.read().decode('utf-8'))
            if response.status == 200 and response_data.get('url'):
                return response_data.get('url', ''), None, response.status
            error_msg = response_data.get('message', response_data.get('errmsg', '未知错误'))
            return None, str(error_msg), response.status

    def _create_knowledge_doc(self, announcement: Dict[str, Any], access_token: str) -> tuple:
        """创建钉钉知识库文档。Returns: (doc_url or None, error_message or None)"""
        import urllib.request
        import urllib.error
        import ssl

        raw_title = announcement.get('title', '') or '未命名公告'
        doc_name = self._sanitize_knowledge_doc_name(raw_title)
        author_name = announcement.get('author', '') or announcement.get('original_author', '')

        operator_unionid, _, resolve_err = self._resolve_knowledge_doc_operator(announcement)
        if resolve_err:
            logger.warning(f"创建知识库文档跳过: {resolve_err}")
            return None, resolve_err

        workspace_id = (DINGTALK_WORKSPACE_ID or '').strip()
        if not workspace_id:
            return None, (
                '未配置钉钉知识库工作空间 ID（DINGTALK_WORKSPACE_ID）。'
                '请在 .env 中设置，或确认 config.py 默认值未被空环境变量覆盖。'
            )
        parent_node_id = (DINGTALK_DOC_PARENT_NODE_ID or '').strip()
        api_url = f"https://api.dingtalk.com/v1.0/doc/workspaces/{workspace_id}/docs"
        ssl_context = ssl.create_default_context()
        ssl_context.check_hostname = False
        ssl_context.verify_mode = ssl.CERT_NONE

        base_body = {
            "name": doc_name,
            "docType": "DOC",
            "operatorId": operator_unionid,
        }
        attempts = []
        if parent_node_id:
            body_with_parent = dict(base_body, parentNodeId=parent_node_id)
            attempts.append(("指定目录", body_with_parent))
        attempts.append(("知识库根目录", dict(base_body)))

        last_err = None
        last_code = None
        for label, body in attempts:
            try:
                logger.info(
                    f"创建知识库文档({label}): name={doc_name!r}, operatorId={operator_unionid[:12]}..., "
                    f"workspace={workspace_id}, parent={body.get('parentNodeId', '(根)')}"
                )
                doc_url, err_msg, code = self._post_knowledge_doc_request(
                    api_url, access_token, body, ssl_context
                )
                if doc_url:
                    logger.info(f"成功创建知识库文档({label}): url={doc_url}")
                    return doc_url, None
                last_err = err_msg
                last_code = code
                logger.warning(f"创建知识库文档({label})失败: {err_msg}")
            except urllib.error.HTTPError as e:
                error_body = ''
                try:
                    if e.fp:
                        error_body = e.read().decode('utf-8')
                except Exception:
                    pass
                last_code = e.code
                last_err = error_body[:500] if error_body else str(e)
                logger.error(f"创建知识库文档({label}) HTTP {e.code}: {last_err}")
                if e.code == 403:
                    return None, (
                        "用户无知识库操作权限(403)。请在钉钉知识库中授权操作人，"
                        "或将 DINGTALK_DOC_OPERATOR_UNIONID 设为有权限用户的 unionid。"
                    )
            except urllib.error.URLError as e:
                return None, f"网络错误: {e}"
            except Exception as e:
                logger.error(f"创建知识库文档({label})异常: {e}", exc_info=True)
                last_err = str(e)

        if last_code == 404 and last_err and 'InvalidAction.NotFound' in last_err:
            return None, (
                "钉钉知识库接口路径无效(404 InvalidAction.NotFound)。常见原因："
                "① DINGTALK_WORKSPACE_ID 为空或错误（Docker 部署时 .env 未配置且 compose 注入了空值）；"
                "② 知识库已被删除或 workspaceId 与当前企业不匹配。"
                f" 当前 workspace={workspace_id!r}。详情: {last_err}"
            )
        if last_code == 500:
            return None, (
                "钉钉知识库接口返回系统错误(500)。常见原因："
                "① DINGTALK_DOC_OPERATOR_UNIONID 填成了 userid（须为 unionid 或先同步用户后填 userid）；"
                "② 工作空间/父目录 ID 无效；"
                "③ 操作人在该知识库无写权限。"
                f" 详情: {last_err or 'internalError'}"
            )
        return None, f"创建失败(HTTP {last_code}): {last_err or '未知错误'}"
    
    def _update_todo_status(self, unionid: str, task_id: str, title: str, description: str, due_time: int, executor_ids: List[str], participant_ids: List[str], done: bool = True, max_retries: int = None) -> bool:
        """更新待办任务状态（调用钉钉API，支持重试机制）
        
        Args:
            unionid: 用户的unionid
            task_id: 待办任务ID
            title: 待办标题
            description: 待办描述
            due_time: 截止时间（毫秒时间戳）
            executor_ids: 执行者unionid列表
            participant_ids: 参与者unionid列表
            done: 是否完成（True=已完成，False=未完成）
            max_retries: 最大重试次数（None时使用配置中的默认值）
        
        Returns:
            是否更新成功
        """
        import urllib.request
        import urllib.error
        import urllib.parse
        import ssl
        
        # 如果max_retries为None，使用配置中的默认值
        if max_retries is None:
            from server.config import TODO_SYNC_RETRY_COUNT
            max_retries = TODO_SYNC_RETRY_COUNT
        
        try:
            # 获取access_token
            access_token = self._get_dingtalk_access_token(
                DINGTALK_CONFIG.get('client_id'),
                DINGTALK_CONFIG.get('client_secret')
            )
            if not access_token:
                logger.error("获取钉钉access_token失败，无法更新待办")
                return False
            
            # 构建更新API URL
            api_url = f"https://api.dingtalk.com/v1.0/todo/users/{unionid}/tasks/{task_id}?operatorId={unionid}"
            
            # 构建请求体（根据API文档）
            request_body = {
                "subject": title,
                "description": description,
                "dueTime": due_time,
                "done": done,
                "executorIds": executor_ids,
                "participantIds": participant_ids
            }
            
            # 创建SSL上下文
            ssl_context = ssl.create_default_context()
            ssl_context.check_hostname = False
            ssl_context.verify_mode = ssl.CERT_NONE
            
            request_data = json.dumps(request_body, ensure_ascii=False).encode('utf-8')
            
            # 重试逻辑
            last_error = None
            for retry_attempt in range(max_retries + 1):  # 0到max_retries，共max_retries+1次尝试
                try:
                    req = urllib.request.Request(
                        api_url,
                        data=request_data,
                        headers={
                            'Content-Type': 'application/json',
                            'x-acs-dingtalk-access-token': access_token
                        },
                        method='PUT'
                    )
                    
                    with urllib.request.urlopen(req, timeout=30, context=ssl_context) as response:
                        if response.status == 200:
                            response_data = json.loads(response.read().decode('utf-8'))
                            result = response_data.get('result', False)
                            if result:
                                logger.info(f"成功更新钉钉待办状态: taskId={task_id}, done={done}" + (f" (重试第{retry_attempt}次)" if retry_attempt > 0 else ""))
                                return True
                            else:
                                last_error = f"API返回失败: {response_data}"
                                logger.warning(f"更新钉钉待办状态失败（尝试 {retry_attempt + 1}/{max_retries + 1}）: {last_error}")
                        else:
                            error_body = response.read().decode('utf-8') if hasattr(response, 'read') else ''
                            last_error = f"状态码: {response.status}, 响应: {error_body[:200]}"
                            logger.error(f"更新钉钉待办状态失败（尝试 {retry_attempt + 1}/{max_retries + 1}）: {last_error}")
                except urllib.error.HTTPError as e:
                    error_body = ''
                    try:
                        if e.fp:
                            error_body = e.read().decode('utf-8')
                    except:
                        pass
                    last_error = f"HTTP错误 {e.code}: {error_body[:200] if error_body else str(e)}"
                    logger.warning(f"更新钉钉待办状态失败（尝试 {retry_attempt + 1}/{max_retries + 1}）: {last_error}")
                except urllib.error.URLError as e:
                    last_error = f"网络错误: {str(e)}"
                    logger.warning(f"更新钉钉待办状态失败（尝试 {retry_attempt + 1}/{max_retries + 1}）: {last_error}")
                except Exception as e:
                    last_error = f"未知错误: {str(e)}"
                    logger.error(f"更新钉钉待办状态时发生错误（尝试 {retry_attempt + 1}/{max_retries + 1}）: {last_error}", exc_info=True)
                
                # 如果不是最后一次尝试，等待后重试
                if retry_attempt < max_retries:
                    from server.config import TODO_SYNC_RETRY_INTERVAL
                    import time
                    wait_time = TODO_SYNC_RETRY_INTERVAL * (retry_attempt + 1)  # 递增等待时间
                    logger.info(f"等待 {wait_time} 秒后重试更新待办状态...")
                    time.sleep(wait_time)
            
            # 所有重试都失败
            logger.error(f"更新钉钉待办状态最终失败（已重试 {max_retries} 次）: taskId={task_id}, done={done}, 错误: {last_error}")
            return False
        except Exception as e:
            logger.error(f"更新钉钉待办状态时发生错误: {e}", exc_info=True)
            return False
    
    def _get_user_detail(self, access_token, userid):
        """获取用户详情"""
        user_detail_url = DINGTALK_CONFIG.get('user_detail_url', 'https://oapi.dingtalk.com/topapi/v2/user/get')
        
        params = {
            'userid': userid
        }
        
        if HAS_REQUESTS:
            response = requests.post(
                f"{user_detail_url}?access_token={access_token}",
                json=params,
                headers={'Content-Type': 'application/json'},
                timeout=10
            )
            if response.status_code == 200:
                result = response.json()
                if result.get('errcode') == 0:
                    return result.get('result', {})
                else:
                    logger.warning(f"获取用户详情失败: {result.get('errmsg', '未知错误')} (errcode: {result.get('errcode')})")
                    return None
            else:
                logger.warning(f"获取用户详情HTTP错误: {response.status_code}")
                return None
        else:
            import urllib.request
            import urllib.error
            
            request_data = json.dumps(params).encode('utf-8')
            url_with_params = f"{user_detail_url}?access_token={urllib.parse.quote(access_token)}"
            req = urllib.request.Request(
                url_with_params,
                data=request_data,
                headers={'Content-Type': 'application/json'}
            )
            
            try:
                with urllib.request.urlopen(req, timeout=10) as response:
                    result = json.loads(response.read().decode('utf-8'))
                    if result.get('errcode') == 0:
                        return result.get('result', {})
                    else:
                        logger.warning(f"获取用户详情失败: {result.get('errmsg', '未知错误')} (errcode: {result.get('errcode')})")
                        return None
            except Exception as e:
                logger.warning(f"获取用户详情异常: {e}")
                return None
    
    def handle_register(self):
        """处理注册申请"""
        self.send_json_response(
            {
                'success': False,
                'code': 403,
                'error': '本地注册入口已关闭，请使用钉钉登录',
            },
            status=403,
        )
        return

        content_length = int(self.headers.get('Content-Length', 0) or 0)
        if content_length == 0:
            self.send_json_response({'success': False, 'error': '请求体不能为空'})
            return
        post_data = self.rfile.read(content_length).decode('utf-8')
        try:
            data = json.loads(post_data)
        except json.JSONDecodeError:
            self.send_json_response({'success': False, 'error': '无效的JSON数据'})
            return
        
        # 验证验证码
        captcha_token = data.get('captcha_token', '')
        captcha_code = data.get('captcha_code', '')
        if captcha_token and captcha_code:
            try:
                from server.captcha import get_captcha_manager
                captcha_mgr = get_captcha_manager()
                if not captcha_mgr.verify_captcha(captcha_token, captcha_code):
                    self.send_json_response({'success': False, 'error': '验证码错误'})
                    return
            except Exception as e:
                logger.error(f"验证码验证失败: {e}", exc_info=True)
                self.send_json_response({'success': False, 'error': '验证码验证失败'})
                return
        else:
            # 注册时必须提供验证码
            self.send_json_response({'success': False, 'error': '请输入验证码'})
            return
        
        # 验证密码强度
        password = data.get('password', '')
        if password:
            password_valid, password_error = InputValidator.validate_password(password, check_strength=True)
            if not password_valid:
                self.send_json_response({'success': False, 'error': password_error or '密码不符合要求'})
                return
        
        roles = data.get('roles', [])
        if isinstance(roles, str):
            roles = [roles]
        data['roles'] = roles
        
        success, message = self.user_manager.register_user(data)
        if success:
            # 文件已更新，清除用户缓存（确保内存与文件一致）
            self._invalidate_user_cache()
            self.send_json_response({'success': True, 'message': '注册申请已提交，等待审批'})
        else:
            self.send_json_response({'success': False, 'error': message or '注册失败'})
    
    def handle_add_user(self):
        """处理添加用户"""
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length).decode('utf-8')
        data = json.loads(post_data)
        
        roles = data.get('roles', [])
        if isinstance(roles, str):
            roles = [roles]
        library_roles = data.get('library_roles', [])
        if isinstance(library_roles, str):
            library_roles = [library_roles]
        
        payload = {
            'username': data.get('username'),
            'password': data.get('password'),
            'name': data.get('name', data.get('username')),
            'roles': self.user_manager.apply_role_update([], roles),
            'library_roles': self.user_manager.sanitize_library_roles_list(library_roles),
            'department': data.get('department', ''),
            'status': data.get('status', STATUS_ACTIVE)
        }
        
        success, message = self.user_manager.add_user(payload)
        if success:
            # 文件已更新，清除用户缓存（确保内存与文件一致）
            self._invalidate_user_cache()
        self.send_json_response({'success': success, 'message': message})
    
    def handle_get_excel_debug(self):
        """获取Excel文件中的原始用户数据（用于调试）"""
        try:
            import openpyxl
            from server.config import DATA_DIR
            
            users_file = os.path.join(DATA_DIR, 'users.xlsx')
            if not os.path.exists(users_file):
                self.send_json_response({
                    'success': False,
                    'error': 'Excel文件不存在',
                    'file_path': users_file
                }, status=404)
                return
            
            wb = openpyxl.load_workbook(users_file)
            # 读取所有工作表的数据
            excel_data = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_data = []
                for row in ws.iter_rows(values_only=True):
                    sheet_data.append(row)
                excel_data[sheet_name] = sheet_data
            
            self.send_json_response({
                'success': True,
                'data': excel_data,
                'file_path': users_file
            })
        except Exception as e:
            logger.error(f"读取Excel文件失败: {e}", exc_info=True)
            self.send_json_response({
                'success': False,
                'error': f'读取Excel文件失败: {str(e)}'
            }, status=500)
    
    def handle_sync_dingtalk_users(self):
        """处理手动同步钉钉用户请求（同步执行，返回实际结果）。"""
        try:
            logger.info("收到手动同步钉钉用户请求")
            result = sync_dingtalk_data() or {}
            if result.get('ok'):
                self.send_json_response({
                    'success': True,
                    'message': result.get('message', '同步完成'),
                    'users_synced': result.get('users_synced', 0),
                    'users_with_dingtalk_id': result.get('users_with_dingtalk_id', 0),
                })
            else:
                self.send_json_response({
                    'success': False,
                    'error': result.get('error', '同步失败'),
                }, status=500)
        except Exception as e:
            logger.error(f"处理同步钉钉用户请求失败: {e}", exc_info=True)
            self.send_json_response({
                'success': False,
                'error': f'同步失败: {str(e)}'
            }, status=500)
    
    def handle_get_scheduled_notifications_config(self):
        """获取定时通知配置（超级管理员）"""
        try:
            from server.system_config_manager import (
                get_config_manager, 
                CONFIG_KEY_SCHEDULED_NOTIFICATIONS_ENABLED,
                CONFIG_KEY_SCHEDULED_NOTIFICATIONS_TIMES
            )
            
            config_mgr = get_config_manager()
            enabled = config_mgr.get_config_bool(CONFIG_KEY_SCHEDULED_NOTIFICATIONS_ENABLED, default_value=True)
            
            # 获取通知时间配置（默认值：8:00, 12:30, 17:30）
            times_json = config_mgr.get_config(CONFIG_KEY_SCHEDULED_NOTIFICATIONS_TIMES, default_value='[{"hour":8,"minute":0},{"hour":12,"minute":30},{"hour":17,"minute":30}]')
            try:
                notification_times = json.loads(times_json)
            except:
                notification_times = [{"hour":8,"minute":0},{"hour":12,"minute":30},{"hour":17,"minute":30}]
            
            self.send_json_response({
                'success': True,
                'enabled': enabled,
                'notification_times': notification_times
            })
        except Exception as e:
            logger.error(f"获取定时通知配置失败: {e}", exc_info=True)
            self.send_json_response({
                'success': False,
                'error': f'获取配置失败: {str(e)}'
            })
    
    def handle_set_scheduled_notifications_config(self):
        """设置定时通知配置（超级管理员）"""
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                self.send_json_response({'success': False, 'error': '请求体不能为空'})
                return
            
            post_data = self.rfile.read(content_length).decode('utf-8')
            data = json.loads(post_data)
            
            from server.system_config_manager import (
                get_config_manager, 
                CONFIG_KEY_SCHEDULED_NOTIFICATIONS_ENABLED,
                CONFIG_KEY_SCHEDULED_NOTIFICATIONS_TIMES
            )
            
            config_mgr = get_config_manager()
            success = True
            messages = []
            
            # 处理 enabled 参数
            if 'enabled' in data:
                enabled = data.get('enabled')
                if enabled is not None:
                    if not isinstance(enabled, bool):
                        # 尝试转换
                        if isinstance(enabled, str):
                            enabled = enabled.lower() in ('true', '1', 'yes', 'on', 'enabled')
                        else:
                            enabled = bool(enabled)
                    
                    if config_mgr.set_config_bool(
                        CONFIG_KEY_SCHEDULED_NOTIFICATIONS_ENABLED,
                        enabled,
                        description='定时通知功能开关'
                    ):
                        status_text = '已启用' if enabled else '已禁用'
                        messages.append(f'定时通知功能已{status_text}')
                        logger.info(f"超级管理员 {self.get_current_user().get('username', 'unknown')} {status_text}定时通知功能")
                    else:
                        success = False
                        messages.append('保存开关状态失败')
            
            # 处理 notification_times 参数
            if 'notification_times' in data:
                notification_times = data.get('notification_times')
                if isinstance(notification_times, list):
                    # 验证时间格式
                    valid_times = []
                    for time_item in notification_times:
                        if isinstance(time_item, dict):
                            hour = time_item.get('hour')
                            minute = time_item.get('minute')
                            if isinstance(hour, int) and isinstance(minute, int) and 0 <= hour <= 23 and 0 <= minute <= 59:
                                valid_times.append({'hour': hour, 'minute': minute})
                    
                    if valid_times:
                        times_json = json.dumps(valid_times, ensure_ascii=False)
                        if config_mgr.set_config(
                            CONFIG_KEY_SCHEDULED_NOTIFICATIONS_TIMES,
                            times_json,
                            description='定时通知时间配置（格式：[{"hour":8,"minute":0},...]）'
                        ):
                            messages.append(f'通知时间已更新为 {len(valid_times)} 个时间点')
                            logger.info(f"超级管理员 {self.get_current_user().get('username', 'unknown')} 更新了定时通知时间配置")
                        else:
                            success = False
                            messages.append('保存通知时间失败')
                    else:
                        success = False
                        messages.append('无效的时间格式')
            
            if success:
                self.send_json_response({
                    'success': True,
                    'message': '；'.join(messages) if messages else '配置已保存'
                })
            else:
                self.send_json_response({
                    'success': False,
                    'error': '；'.join(messages) if messages else '保存配置失败'
                })
        except Exception as e:
            logger.error(f"设置定时通知配置失败: {e}", exc_info=True)
            self.send_json_response({
                'success': False,
                'error': f'设置配置失败: {str(e)}'
            })
    
    def handle_batch_update_user_status(self):
        """处理批量更新用户状态"""
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                self.send_json_response({'success': False, 'error': '请求体不能为空'})
                return
            
            post_data = self.rfile.read(content_length).decode('utf-8')
            data = json.loads(post_data)
            
            user_ids = data.get('user_ids', [])
            status = data.get('status', '')
            
            if not user_ids or not isinstance(user_ids, list):
                self.send_json_response({'success': False, 'error': '请提供有效的用户ID列表'})
                return
            
            if status not in [STATUS_ACTIVE, 'inactive', 'disabled', STATUS_PENDING, STATUS_REJECTED]:
                self.send_json_response({'success': False, 'error': '无效的状态值'})
                return
            
            if status == 'inactive':
                status = 'disabled'
            
            handoff_batch = data.get('library_handoff_user_id')
            
            success_count = 0
            failed_count = 0
            failed_users = []
            
            for user_id in user_ids:
                try:
                    # 检查用户是否存在
                    user = self.user_manager.get_user_by_id(user_id)
                    if not user:
                        failed_count += 1
                        failed_users.append(user_id)
                        continue
                    
                    # 更新用户状态
                    payload = {'status': status}
                    if handoff_batch is not None:
                        payload['library_handoff_user_id'] = handoff_batch
                    success, message = self.user_manager.update_user(user_id, payload)
                    if success:
                        success_count += 1
                    else:
                        failed_count += 1
                        failed_users.append(user_id)
                except Exception as e:
                    logger.error(f"批量更新用户状态失败 (user_id={user_id}): {e}")
                    failed_count += 1
                    failed_users.append(user_id)
            
            if success_count > 0:
                # 文件已更新，清除用户缓存（确保内存与文件一致）
                self._invalidate_user_cache()
                message = f'成功更新 {success_count} 个用户状态'
                if failed_count > 0:
                    message += f'，{failed_count} 个用户更新失败'
                logger.info(f"批量更新用户状态: {message}")
                self.send_json_response({
                    'success': True,
                    'message': message,
                    'success_count': success_count,
                    'failed_count': failed_count,
                    'failed_users': failed_users
                })
            else:
                self.send_json_response({
                    'success': False,
                    'error': f'所有用户更新失败（共 {failed_count} 个）',
                    'failed_users': failed_users
                })
        except json.JSONDecodeError as e:
            logger.error(f"批量更新用户状态请求：无效的JSON数据: {e}")
            self.send_json_response({'success': False, 'error': '无效的JSON数据'})
        except Exception as e:
            logger.error(f"批量更新用户状态异常: {e}", exc_info=True)
            self.send_json_response({'success': False, 'error': f'服务器错误: {str(e)}'})
    
    
    def _send_attachment_api_error(self, code: int, message: str):
        """附件 API 统一返回 JSON，避免前端 XHR 收到 HTML 误判。"""
        self.send_json_response({'success': False, 'error': message}, status=code)

    def handle_download_attachment(self, announcement_id, filename, version_number=None):
        """处理附件下载
        
        Args:
            announcement_id: 公告ID
            filename: 文件名
            version_number: 版本号（可选，用于下载历史版本的附件）
        """
        user = self.get_current_user()
        if not user:
            self._send_attachment_api_error(401, '请先登录')
            return
        
        # 优先使用正式目录中已发布版本（避免 temp 待审副本导致无权或附件路径不一致）
        announcement = self.announcement_mgr.get_announcement_for_download(announcement_id)
        if not announcement:
            announcement = self.announcement_mgr.get_announcement(announcement_id)
        if not announcement:
            logger.warning(f"附件下载失败: 公告不存在, id={announcement_id}")
            self._send_attachment_api_error(404, '公告不存在')
            return
        
        # 检查权限：已发布公告全员可下；否则仅作者或管理员
        if (announcement.get('status') != 'approved' and 
            announcement.get('author') != user.get('name') and
            not any(self._has_role(user, role) for role in ANNOUNCEMENT_MANAGE_ROLES) and
            not self._is_super_admin(user)):
            self._send_attachment_api_error(403, '无权下载此附件')
            return
        
        # 获取附件路径（如果指定了版本号，从历史版本获取）
        if version_number:
            # 从历史版本获取附件
            version = self.announcement_mgr.get_version(announcement_id, version_number)
            if version:
                version_path = self.announcement_mgr._get_announcement_path(
                    announcement.get('board_id'), 
                    announcement_id, 
                    False
                )
                version_attachments_dir = os.path.join(version_path, 'versions', version_number, 'attachments')
                attachment_path = os.path.join(version_attachments_dir, filename)
                if not os.path.exists(attachment_path):
                    attachment_path = None
            else:
                attachment_path = None
        else:
            attachment_path = self.announcement_mgr.get_attachment(
                announcement_id, filename, metadata=announcement
            )
        
        if not attachment_path or not os.path.exists(attachment_path):
            logger.warning(
                f"附件下载失败: 未找到文件, id={announcement_id}, filename={filename}"
            )
            self._send_attachment_api_error(404, f'附件不存在: {filename}')
            return
        
        try:
            # 获取文件大小
            file_size = os.path.getsize(attachment_path)
            
            # 使用实际文件系统中的文件名（而不是请求中的文件名，避免乱码）
            actual_filename = os.path.basename(attachment_path)
            print(f"附件下载: 使用实际文件名: {actual_filename} (请求文件名: {filename})")
            
            # 获取文件MIME类型（使用实际文件名）
            import mimetypes
            mime_type, _ = mimetypes.guess_type(actual_filename)
            if not mime_type:
                mime_type = 'application/octet-stream'
            
            self.send_response(200)
            self.send_header('Content-Type', mime_type)
            # 使用实际文件名的URL编码，确保中文文件名正确显示
            encoded_filename = urllib.parse.quote(actual_filename.encode('utf-8'))
            self.send_header('Content-Disposition', f'attachment; filename="{encoded_filename}"; filename*=UTF-8\'\'{encoded_filename}')
            self.send_header('Content-Length', str(file_size))
            self.end_headers()
            
            # 使用大缓冲区分块读取和传输，充分利用1000MB带宽
            with open(attachment_path, 'rb') as f:
                while True:
                    chunk = f.read(CHUNK_SIZE)  # 使用1MB块大小
                    if not chunk:
                        break
                    self.wfile.write(chunk)
                    self.wfile.flush()  # 确保数据及时发送，充分利用带宽
            
            logger.info(f"附件下载成功: {filename}, 大小={file_size} bytes")
        except Exception as e:
            logger.error(f"附件下载异常: {e}", exc_info=True)
            self.send_error(500, f"下载失败: {str(e)}")
    
    
    def redirect_to_login(self):
        """重定向到登录页面，保留原始URL以便登录后跳转"""
        # 获取当前请求的完整路径（包括查询参数）
        current_path = self.path
        
        # 如果当前路径已经是登录页面，直接重定向到登录页面，不添加redirect参数（避免循环）
        if current_path == '/login' or current_path.startswith('/login?'):
            self.send_response(302)
            self.send_header('Location', '/login')
            self.end_headers()
            return
        
        if self.path.startswith('/api/'):
            # 如果是API请求，尝试从Referer获取原始页面
            referer = self.headers.get('Referer', '')
            if referer:
                try:
                    from urllib.parse import urlparse
                    parsed = urlparse(referer)
                    current_path = parsed.path
                    if parsed.query:
                        current_path += '?' + parsed.query
                    # 如果Referer也是登录页面，不添加redirect参数
                    if current_path == '/login' or current_path.startswith('/login?'):
                        self.send_response(302)
                        self.send_header('Location', '/login')
                        self.end_headers()
                        return
                except:
                    pass
        
        # 构建登录URL，带上redirect参数
        login_url = '/login'
        if current_path and current_path != '/login' and not current_path.startswith('/api/') and not current_path.startswith('/login'):
            # 对路径进行URL编码
            redirect_param = urllib.parse.quote(current_path, safe='')
            login_url = f'/login?redirect={redirect_param}'
        
        self.send_response(302)
        self.send_header('Location', login_url)
        self.end_headers()
    
    def redirect_to_home(self):
        """重定向到主页"""
        self.send_response(302)
        self.send_header('Location', '/')
        self.end_headers()
    
    def redirect_to_announcement(self, query=''):
        """重定向到公告栏"""
        location = '/announcement'
        if query:
            location = f'{location}?{query}'
        self.send_response(302)
        self.send_header('Location', location)
        self.end_headers()
    

class ThreadingHTTPServer(socketserver.ThreadingMixIn, http.server.HTTPServer):
    """多线程HTTP服务器（高并发优化版），兼容所有Python版本，支持2000+并发"""
    daemon_threads = True
    allow_reuse_address = True
    timeout = REQUEST_TIMEOUT  # 使用配置的超时时间
    
    # 线程池管理
    _active_threads = 0
    _thread_lock = threading.Lock()
    _max_threads = MAX_WORKERS  # 最大线程数限制
    _connection_queue = []  # 连接队列
    _queue_lock = threading.Lock()
    
    def server_bind(self):
        """绑定服务器并设置socket选项（优化网络性能）"""
        self.socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        self.socket.setsockopt(socket.SOL_SOCKET, socket.SO_KEEPALIVE, 1)
        # 设置TCP_NODELAY，禁用Nagle算法，减少延迟（适合高并发场景）
        try:
            self.socket.setsockopt(socket.IPPROTO_TCP, socket.TCP_NODELAY, 1)
        except AttributeError:
            # 某些系统可能不支持TCP_NODELAY
            pass
        # 设置接收和发送缓冲区大小（充分利用1000MB带宽）
        try:
            # 增大缓冲区以充分利用带宽（1MB缓冲区）
            self.socket.setsockopt(socket.SOL_SOCKET, socket.SO_RCVBUF, RECV_BUFFER_SIZE)
            self.socket.setsockopt(socket.SOL_SOCKET, socket.SO_SNDBUF, SEND_BUFFER_SIZE)
            # 设置TCP窗口缩放，支持更大的传输窗口
            try:
                self.socket.setsockopt(socket.IPPROTO_TCP, socket.TCP_WINDOW_CLAMP, SEND_BUFFER_SIZE)
            except AttributeError:
                pass  # 某些系统可能不支持
        except Exception as e:
            logger.warning(f"设置socket缓冲区失败: {e}")
        super().server_bind()
    
    def server_activate(self):
        """激活服务器（设置监听队列大小）"""
        # 设置监听队列大小（backlog，等待队列大小，不是最大连接数）
        # backlog 通常建议设置为 128-512，这里使用 CONNECTION_QUEUE_SIZE 和 512 的较小值
        try:
            backlog = min(CONNECTION_QUEUE_SIZE, 512)  # backlog 不能太大，通常限制在 128-512
            self.socket.listen(backlog)
            logger.info(f"监听队列大小设置为: {backlog}")
        except Exception as e:
            logger.warning(f"设置监听队列大小失败: {e}，使用默认值")
            # 使用默认值（通常是 5）
            try:
                self.socket.listen()
            except:
                pass
        super().server_activate()
    
    def handle_timeout(self):
        """处理超时"""
        pass  # 超时处理，避免服务器崩溃
    
    def process_request(self, request, client_address):
        """处理请求（重写以添加线程数限制和监控，优化高并发性能）"""
        # 使用更高效的线程数检查
        with self._thread_lock:
            current_threads = self._active_threads
            if current_threads >= self._max_threads:
                # 线程数达到上限，尝试加入队列等待
                with self._queue_lock:
                    if len(self._connection_queue) < CONNECTION_QUEUE_SIZE:
                        self._connection_queue.append((request, client_address))
                        return
                
                # 队列也满了，拒绝连接
                try:
                    request.close()
                except:
                    pass
                logger.warning(f"达到最大线程数限制 ({self._max_threads})，连接队列已满，拒绝新连接")
                return
        
        # 增加活跃线程计数
        with self._thread_lock:
            self._active_threads += 1
        
        try:
            # 调用父类方法处理请求
            super().process_request(request, client_address)
        finally:
            # 减少活跃线程计数
            with self._thread_lock:
                self._active_threads -= 1
                # 检查队列中是否有等待的连接
                if self._connection_queue:
                    try:
                        queued_request, queued_address = self._connection_queue.pop(0)
                        # 延迟处理，避免递归过深
                        threading.Timer(0.01, self.process_request, args=(queued_request, queued_address)).start()
                    except (IndexError, Exception):
                        pass
    
    def get_active_threads(self):
        """获取当前活跃线程数"""
        with self._thread_lock:
            return self._active_threads
    
    def get_queue_size(self):
        """获取连接队列大小"""
        with self._queue_lock:
            return len(self._connection_queue)

def sync_dingtalk_data():
    """自动同步钉钉数据（部门列表和用户列表）。Returns: {ok, message, users_synced, users_with_dingtalk_id, error}"""
    try:
        logger.info("=" * 60)
        logger.info("🔄 开始自动同步钉钉数据...")
        logger.info("=" * 60)

        from server.user_manager import UserManager
        UserManager()._ensure_mysql_user_schema()
        
        # 检查配置
        client_id = DINGTALK_CONFIG.get('client_id', '')
        client_secret = DINGTALK_CONFIG.get('client_secret', '')
        corp_id = DINGTALK_CONFIG.get('corp_id', '')
        
        if not client_id or not client_secret or not corp_id:
            msg = "钉钉配置不完整（client_id / client_secret / corp_id）"
            logger.warning(f"⚠️  {msg}")
            return {'ok': False, 'error': msg}
        
        # 步骤1: 获取 access_token（带重试机制，复用现有逻辑）
        logger.info("📝 步骤1: 获取钉钉 AccessToken...")
        import urllib.request
        import urllib.error
        import ssl
        
        url = f"https://api.dingtalk.com/v1.0/oauth2/{corp_id}/token"
        payload = {
            'client_id': client_id,
            'client_secret': client_secret,
            'grant_type': 'client_credentials'
        }
        
        ssl_context = ssl.create_default_context()
        ssl_context.check_hostname = False
        ssl_context.verify_mode = ssl.CERT_NONE
        
        # 重试机制：最多尝试5次，每次间隔2秒
        max_retries = 5
        retry_interval = 2  # 秒
        access_token = None
        
        headers = {
            'Content-Type': 'application/json',
            'Host': 'api.dingtalk.com',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        for attempt in range(1, max_retries + 1):
            try:
                logger.info(f"   尝试第 {attempt} 次获取 AccessToken...")
                request_data_bytes = json.dumps(payload).encode('utf-8')
                req = urllib.request.Request(url, data=request_data_bytes, 
                                            headers=headers, 
                                            method='POST')
                
                with urllib.request.urlopen(req, timeout=30, context=ssl_context) as response:
                    response_text = response.read().decode('utf-8')
                    response_code = response.getcode()
                    
                    # 检查响应是否是HTML（可能是重定向或错误页面）
                    if response_text.strip().startswith('<!DOCTYPE') or response_text.strip().startswith('<html'):
                        logger.warning(f"   第 {attempt} 次尝试返回HTML页面而不是JSON")
                        if attempt < max_retries:
                            logger.info(f"   等待 {retry_interval} 秒后重试...")
                            time.sleep(retry_interval)
                            continue
                        else:
                            logger.error(f"   所有尝试都返回HTML页面")
                            break
                    
                    # 尝试解析JSON
                    try:
                        result = json.loads(response_text)
                        
                        if 'access_token' in result:
                            access_token = result['access_token']
                            expires_in = result.get('expires_in', 7200)
                            logger.info(f"✅ 成功获取 AccessToken (有效期: {expires_in}秒)")
                            break
                        else:
                            error_msg = result.get('error_description', result.get('error', '未知错误'))
                            logger.warning(f"   第 {attempt} 次尝试失败: {error_msg}")
                            if attempt < max_retries:
                                logger.info(f"   等待 {retry_interval} 秒后重试...")
                                time.sleep(retry_interval)
                    except json.JSONDecodeError as json_err:
                        logger.warning(f"   第 {attempt} 次尝试失败: JSON解析错误 - {json_err}")
                        logger.warning(f"   响应内容预览: {response_text[:200]}")
                        if attempt < max_retries:
                            logger.info(f"   等待 {retry_interval} 秒后重试...")
                            time.sleep(retry_interval)
            except urllib.error.HTTPError as e:
                logger.warning(f"   第 {attempt} 次尝试失败: HTTP错误 {e.code}")
                if attempt < max_retries:
                    logger.info(f"   等待 {retry_interval} 秒后重试...")
                    time.sleep(retry_interval)
            except urllib.error.URLError as e:
                logger.warning(f"   第 {attempt} 次尝试失败: 网络错误 {e}")
                if attempt < max_retries:
                    logger.info(f"   等待 {retry_interval} 秒后重试...")
                    time.sleep(retry_interval)
            except Exception as e:
                logger.warning(f"   第 {attempt} 次尝试失败: {e}")
                if attempt < max_retries:
                    logger.info(f"   等待 {retry_interval} 秒后重试...")
                    time.sleep(retry_interval)
        
        if not access_token:
            msg = f"获取钉钉 AccessToken 失败（已重试 {max_retries} 次），请检查 DINGTALK_CLIENT_SECRET 与网络"
            logger.error(f"❌ {msg}")
            return {'ok': False, 'error': msg}
        
        # 步骤2: 获取部门列表并更新（复用现有的Handler方法）
        logger.info("📝 步骤2: 获取并更新部门列表...")
        
        # 创建一个临时的Handler实例来复用现有方法
        # 注意：这里我们直接调用底层方法，因为Handler需要request对象
        from server.department_manager import DepartmentManager
        dept_mgr = DepartmentManager()
        
        # 复用现有的_get_department_list方法和递归逻辑
        # 创建一个模拟的Handler实例来访问方法
        class TempHandler:
            def __init__(self):
                self.department_mgr = dept_mgr
                # 导入必要的模块
                import urllib.request
                import urllib.error
                import ssl
                self.urllib = urllib.request
                self.urllib_error = urllib.error
                self.ssl = ssl
                self.ssl_context = ssl.create_default_context()
                self.ssl_context.check_hostname = False
                self.ssl_context.verify_mode = ssl.CERT_NONE
            
            def _get_department_list(self, access_token, dept_id):
                """复用现有的获取部门列表方法"""
                url = "https://oapi.dingtalk.com/topapi/v2/department/listsub"
                params = {
                    'language': 'zh_CN',
                    'dept_id': dept_id
                }
                url_with_params = f"{url}?access_token={urllib.parse.quote(access_token)}"
                request_data = json.dumps(params).encode('utf-8')
                
                req = self.urllib.Request(
                    url_with_params,
                    data=request_data,
                    headers={'Content-Type': 'application/json'},
                    method='POST'
                )
                
                with self.urllib.urlopen(req, timeout=30, context=self.ssl_context) as response:
                    response_text = response.read().decode('utf-8')
                    result = json.loads(response_text)
                    
                    if result.get('errcode') == 0:
                        return result.get('result', [])
                    else:
                        raise Exception(f"获取部门列表失败: {result.get('errmsg', '未知错误')}")
        
        temp_handler = TempHandler()
        
        # 递归获取所有部门（复用现有逻辑）
        all_departments = []
        visited_dept_ids = set()
        
        def get_departments_recursive(dept_id, parent_id=None, depth=0):
            """递归获取部门列表（严格按照用户提供的流程）
            
            Args:
                dept_id: 当前要获取的部门ID
                parent_id: 父部门ID
                depth: 递归深度（用于日志缩进）
            """
            indent = "  " * depth
            if dept_id in visited_dept_ids:
                logger.info(f"{indent}⚠️  dept_id={dept_id} 已访问过，跳过")
                return
            visited_dept_ids.add(dept_id)
            
            try:
                logger.info(f"{indent}📋 正在获取 dept_id={dept_id} 的子部门...")
                departments = temp_handler._get_department_list(access_token, dept_id)
                logger.info(f"{indent}✅ 获取到 dept_id={dept_id} 下的 {len(departments)} 个子部门")
                
                for dept in departments:
                    dept_info = {
                        'parent_id': dept.get('parent_id', parent_id),
                        'name': dept.get('name', ''),
                        'dept_id': dept.get('dept_id'),
                        'create_dept_group': dept.get('create_dept_group', False)
                    }
                    
                    if dept_info['dept_id']:
                        all_departments.append(dept_info)
                        logger.info(f"{indent}  - 部门: {dept_info['name']} (dept_id={dept_info['dept_id']}, create_dept_group={dept_info['create_dept_group']})")
                        
                        # 如果 create_dept_group 为 true，递归获取子部门
                        if dept_info['create_dept_group']:
                            logger.info(f"{indent}  → 部门 {dept_info['name']} (dept_id={dept_info['dept_id']}) 有子部门，继续递归获取...")
                            get_departments_recursive(dept_info['dept_id'], dept_info['dept_id'], depth + 1)
                        else:
                            logger.info(f"{indent}  → 部门 {dept_info['name']} (dept_id={dept_info['dept_id']}) 无子部门，跳过")
            except Exception as e:
                logger.error(f"{indent}❌ 获取 dept_id={dept_id} 的部门列表失败: {e}")
        
        # 从根部门开始
        logger.info("   开始从根部门 (dept_id=1) 获取部门列表...")
        get_departments_recursive(1, 1)
        
        if all_departments:
            dept_mgr.save_departments(all_departments, append=False)
            logger.info(f"✅ 成功更新 {len(all_departments)} 个部门")
        else:
            logger.warning("⚠️  未获取到任何部门")
            return
        
        # 步骤3: 获取配置的同步部门及其子部门的 ID 列表
        logger.info("📝 步骤3: 分析部门层次关系，确定需要拉取用户的部门...")
        sync_dept_ids = dept_mgr.get_sync_department_ids()
        
        if not sync_dept_ids:
            logger.warning("⚠️  未找到任何同步部门，跳过用户拉取")
            logger.info("=" * 60)
            logger.info(f"✅ 自动同步完成！")
            logger.info(f"   - 部门数量: {len(all_departments)}")
            logger.info(f"   - 用户数量: 0（未找到同步部门）")
            logger.info("=" * 60)
            return
        
        logger.info(f"   找到同步部门及其子部门共 {len(sync_dept_ids)} 个，将只拉取这些部门的用户")
        
        # 步骤4: 拉取配置范围内部门的用户
        logger.info("📝 步骤4: 获取同步部门及其子部门的用户列表...")
        
        from server.user_manager import UserManager
        user_mgr = UserManager()
        
        # 扩展TempHandler以包含获取用户列表的方法
        class TempHandlerWithUsers(TempHandler):
            def _get_department_user_list(self, access_token, dept_id, cursor=0, size=100):
                """复用现有的获取部门用户列表方法"""
                url = "https://oapi.dingtalk.com/topapi/v2/user/list"
                params = {
                    'cursor': cursor,
                    'contain_access_limit': False,
                    'size': size,
                    'order_field': 'modify_desc',
                    'language': 'zh_CN',
                    'dept_id': dept_id
                }
                url_with_params = f"{url}?access_token={urllib.parse.quote(access_token)}"
                request_data = json.dumps(params).encode('utf-8')
                
                req = self.urllib.Request(
                    url_with_params,
                    data=request_data,
                    headers={'Content-Type': 'application/json'},
                    method='POST'
                )
                
                with self.urllib.urlopen(req, timeout=30, context=self.ssl_context) as response:
                    response_text = response.read().decode('utf-8')
                    result = json.loads(response_text)
                    
                    if result.get('errcode') == 0:
                        result_data = result.get('result', {})
                        return {
                            'list': result_data.get('list', []),
                            'has_more': result_data.get('has_more', False),
                            'next_cursor': result_data.get('next_cursor', '0')
                        }
                    else:
                        raise Exception(f"获取部门用户列表失败: {result.get('errmsg', '未知错误')}")
        
        temp_handler = TempHandlerWithUsers()
        
        # 获取所有部门信息，用于查找部门名称
        all_depts = dept_mgr.get_departments()
        dept_id_to_name = {dept.get('dept_id'): dept.get('name', '未知部门') for dept in all_depts}
        
        total_users = 0
        for dept_id in sync_dept_ids:
            dept_name = dept_id_to_name.get(dept_id, f'部门{dept_id}')
            
            try:
                logger.info(f"   正在获取部门 '{dept_name}' (dept_id={dept_id}) 的用户...")
                
                # 获取该部门的所有用户（处理分页，复用现有逻辑）
                all_users = []
                cursor = 0
                has_more = True
                page_count = 0
                
                while has_more:
                    try:
                        page_count += 1
                        logger.info(f"     正在获取第 {page_count} 页用户数据（cursor={cursor}）...")
                        result = temp_handler._get_department_user_list(access_token, dept_id, cursor, 100)
                        users = result.get('list', [])
                        if isinstance(users, dict):
                            users = [users]
                        
                        logger.info(f"     第 {page_count} 页获取到 {len(users)} 个用户")
                        all_users.extend(users)
                        
                        has_more = result.get('has_more', False)
                        if has_more:
                            next_cursor = result.get('next_cursor', '0')
                            try:
                                cursor = int(next_cursor) if isinstance(next_cursor, str) else next_cursor
                            except (ValueError, TypeError):
                                cursor = int(next_cursor) if next_cursor else 0
                        else:
                            logger.info(f"     已获取所有用户，共 {len(all_users)} 个（共 {page_count} 页）")
                            break
                    except Exception as e:
                        logger.error(f"     获取部门 '{dept_name}' 用户列表失败（第 {page_count} 页）: {e}")
                        break
                
                if all_users:
                    # 保存用户（智能合并，复用现有逻辑）
                    # 传递 source_dept_ids 参数，确保即使部门文件损坏也能正确激活用户
                    success = user_mgr.save_dingtalk_users(all_users, append=False, source_dept_ids=sync_dept_ids)
                    if success:
                        total_users += len(all_users)
                        logger.info(f"   ✅ 部门 '{dept_name}': 获取并保存 {len(all_users)} 个用户")
                        # 文件已更新，清除用户缓存（确保内存与文件一致）
                        preloader = get_data_preloader()
                        if preloader:
                            preloader.invalidate_cache('users')
                    else:
                        logger.warning(f"   ⚠️  部门 '{dept_name}': 获取 {len(all_users)} 个用户，但保存失败")
                else:
                    logger.info(f"   ℹ️  部门 '{dept_name}': 无用户")
            except Exception as e:
                logger.error(f"   处理部门 '{dept_name}' 时出错: {e}")
                continue
        
        logger.info("=" * 60)
        logger.info(f"✅ 自动同步完成！")
        logger.info(f"   - 部门数量: {len(all_departments)}")
        logger.info(f"   - 同步部门及其子部门数量: {len(sync_dept_ids)}")
        logger.info(f"   - 用户数量: {total_users}（仅同步配置范围内的部门）")
        logger.info("=" * 60)

        users_with_id = 0
        try:
            from server.db_adapter import get_connection_pool
            pool = get_connection_pool()
            if pool:
                conn = pool.get_connection()
                try:
                    cur = conn.cursor()
                    cur.execute(
                        "SELECT COUNT(*) AS c FROM users WHERE status='active' "
                        "AND dingtalk_userid IS NOT NULL AND dingtalk_userid != ''"
                    )
                    row = cur.fetchone()
                    users_with_id = int(row.get('c', 0) if isinstance(row, dict) else row[0])
                finally:
                    conn.close()
        except Exception as count_err:
            logger.warning(f"统计 dingtalk_userid 数量失败: {count_err}")
        
        preloader = get_data_preloader()
        if preloader:
            preloader.invalidate_cache('users')
            preloader.invalidate_cache('departments')
            logger.info("已清除用户和部门缓存")
        
        return {
            'ok': True,
            'message': f'同步完成：写入 {total_users} 人，库内有效钉钉ID {users_with_id} 人',
            'users_synced': total_users,
            'users_with_dingtalk_id': users_with_id,
            'departments': len(all_departments),
        }
        
    except Exception as e:
        logger.error(f"❌ 自动同步钉钉数据失败: {e}", exc_info=True)
        logger.error(f"   错误详情: {traceback.format_exc()}")
        logger.warning("   服务器将继续启动，但数据可能不是最新的")
        return {'ok': False, 'error': str(e)}


def run_server():
    """启动服务器（高并发优化版）"""
    # 设置socket默认超时
    socket.setdefaulttimeout(REQUEST_TIMEOUT)
    
    # 启动时自动同步钉钉数据
    logger.info("=" * 60)
    logger.info("🔄 正在启动自动同步钉钉数据任务...")
    logger.info("=" * 60)
    try:
        sync_thread = threading.Thread(target=sync_dingtalk_data, daemon=True, name="DingTalkSync")
        sync_thread.start()
        logger.info("✅ 自动同步任务已在后台启动（线程名称: DingTalkSync）")
        # 等待一小段时间，让同步任务开始执行
        time.sleep(0.5)
    except Exception as e:
        logger.error(f"❌ 启动自动同步任务失败: {e}", exc_info=True)
        logger.warning("   服务器将继续启动，但不会自动同步钉钉数据")
    
    # 启动时预加载常用数据到内存（用户、公告、部门、待办）
    logger.info("正在预加载常用数据到内存...")
    try:
        from server.data_preloader import get_data_preloader
        # 使用单例模式获取预加载器实例
        preloader = get_data_preloader()
        # 获取管理器实例（使用类级别的共享实例）
        user_mgr = HardwareRDBHandler._user_manager if HardwareRDBHandler._user_manager else UserManager()
        announcement_mgr = HardwareRDBHandler._announcement_mgr if HardwareRDBHandler._announcement_mgr else AnnouncementManager(base_dir=BASE_DIR)
        department_mgr = HardwareRDBHandler._department_mgr if HardwareRDBHandler._department_mgr else DepartmentManager()
        todo_mgr = HardwareRDBHandler._todo_mgr if HardwareRDBHandler._todo_mgr else TodoManager()
        
        # 设置管理器实例
        preloader.set_managers(user_mgr, announcement_mgr, department_mgr, todo_mgr)
        
        # 在后台线程中预加载，不阻塞服务器启动
        preload_thread = threading.Thread(
            target=preloader.preload_all,
            args=(user_mgr, announcement_mgr, department_mgr, todo_mgr),
            daemon=True,
            name="DataPreloader"
        )
        preload_thread.start()
        logger.info("✅ 数据预加载任务已在后台启动")
    except Exception as e:
        logger.error(f"❌ 启动数据预加载失败: {e}", exc_info=True)
        logger.warning("   服务器将继续启动，但数据预加载功能不可用")
    
    # 启动定时通知任务（每天 8:00、12:30、17:30 发送未读公告通知）
    logger.info("正在启动定时通知任务...")
    try:
        from server.scheduled_notifications import start_scheduled_notifications
        start_scheduled_notifications()
        logger.info("✅ 定时通知任务已启动（每天 8:00、12:30、17:30 发送未读公告通知）")
    except Exception as e:
        logger.error(f"❌ 启动定时通知任务失败: {e}", exc_info=True)
        logger.warning("   服务器将继续启动，但定时通知功能不可用")
    
    # 使用ThreadingHTTPServer支持多线程并发（兼容所有Python版本）
    with ThreadingHTTPServer((HOST, PORT), HardwareRDBHandler) as httpd:
        logger.info(f"硬件研发部管理系统运行在 http://{HOST}:{PORT}")
        logger.info("=" * 60)
        logger.info("🚀 服务器已启用高性能多线程模式，支持超大规模并发访问")
        logger.info(f"📊 最大并发线程数: {MAX_WORKERS}")
        logger.info(f"🔗 最大并发连接数: {MAX_CONNECTIONS}")
        logger.info(f"📋 连接队列大小: {CONNECTION_QUEUE_SIZE}")
        logger.info(f"💾 缓存大小限制: {CACHE_SIZE} 条")
        logger.info(f"⏱️  连接超时时间: {REQUEST_TIMEOUT}秒")
        logger.info(f"🗜️  GZIP压缩: {'启用' if ENABLE_GZIP else '禁用'} (级别: {GZIP_COMPRESSION_LEVEL})")
        logger.info(f"💾 HTTP缓存: {'启用' if ENABLE_CACHE else '禁用'}")
        logger.info(f"📈 当前缓存条目: {api_cache.size()}")
        logger.info(f"💾 内存预加载: 用户={PRELOAD_USERS}, 公告={PRELOAD_ANNOUNCEMENTS}, 部门={PRELOAD_DEPARTMENTS}, 待办={PRELOAD_TODOS}")
        logger.info(f"📦 文件读取缓冲区: {FILE_READ_BUFFER_SIZE/1024:.0f}KB")
        logger.info(f"🚀 数据传输块大小: {CHUNK_SIZE/1024/1024:.1f}MB（充分利用1000MB带宽）")
        logger.info("=" * 60)
        logger.info("按 Ctrl+C 停止服务器")
        logger.info("")
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            logger.info("\n正在关闭服务器...")
            
            # 停止定时通知任务
            try:
                from server.scheduled_notifications import stop_scheduled_notifications
                stop_scheduled_notifications()
                logger.info("定时通知任务已停止")
            except Exception as e:
                logger.error(f"停止定时通知任务失败: {e}", exc_info=True)
            
            # 保存所有待办数据到文件（确保数据不丢失）
            try:
                from server.data_preloader import get_data_preloader
                preloader = get_data_preloader()
                if preloader:
                    logger.info("正在保存待办数据到文件...")
                    preloader.force_save_dirty_todos()
                    preloader.stop_todo_auto_save()
                    logger.info("待办数据已保存完成")
            except Exception as e:
                logger.error(f"保存待办数据失败: {e}", exc_info=True)
            
            active_threads = httpd.get_active_threads()
            queue_size = httpd.get_queue_size()
            if active_threads > 0:
                logger.info(f"等待 {active_threads} 个活跃线程完成...")
            if queue_size > 0:
                logger.info(f"队列中还有 {queue_size} 个等待的连接...")
            logger.info("服务器已停止")

if __name__ == '__main__':
    run_server()
