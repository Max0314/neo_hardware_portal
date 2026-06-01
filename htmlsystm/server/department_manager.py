import os
import threading
from typing import List, Dict, Any, Optional, Tuple
import openpyxl
from openpyxl import Workbook

from server.config import DATA_DIR, DINGTALK_SYNC_DEPARTMENT_IDS
from server.logger import logger

# 部门表头
DEPARTMENT_HEADERS = ["parent_id", "name", "dept_id", "create_dept_group"]


class DepartmentManager:
    """部门管理器，负责部门数据的Excel文件管理"""
    _instance = None
    _lock = threading.Lock()
    _file_checked = False  # 类级别标志，确保每个进程只检查一次
    
    def __new__(cls):
        """单例模式"""
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
                    cls._instance._initialized = False
        return cls._instance
    
    def __init__(self):
        if self._initialized:
            return
        
        # 确保 department 目录存在
        self.department_dir = os.path.join(DATA_DIR, 'department')
        os.makedirs(self.department_dir, exist_ok=True)
        
        # 部门Excel文件路径
        self.departments_file = os.path.join(self.department_dir, 'departments.xlsx')
        
        # 只在第一次初始化时检查文件（每个进程一次）
        if not DepartmentManager._file_checked:
            self._ensure_departments_file()
            DepartmentManager._file_checked = True
        else:
            # 文件已检查过，只确保文件存在
            if not os.path.exists(self.departments_file):
                self._ensure_departments_file()
        
        self._initialized = True
    
    def _ensure_departments_file(self):
        """确保部门文件存在并具备所需列"""
        if not os.path.exists(self.departments_file):
            logger.info(f"创建新的部门文件: {self.departments_file}")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "部门数据"
            
            # 写入表头
            for col, header in enumerate(DEPARTMENT_HEADERS, 1):
                ws.cell(row=1, column=col, value=header)
            
            wb.save(self.departments_file)
            logger.info("部门文件创建成功")
            return
        
        # 检查现有文件的完整性
        try:
            wb = openpyxl.load_workbook(self.departments_file)
            ws = wb.active
            self._ensure_schema(ws)
            wb.save(self.departments_file)
            logger.info("部门文件格式检查完成")
        except Exception as e:
            logger.error(f"检查部门文件时出错: {e}")
            # 如果文件损坏，备份旧文件并创建新文件
            import shutil
            backup_file = self.departments_file + '.backup'
            try:
                if os.path.exists(self.departments_file):
                    shutil.copy2(self.departments_file, backup_file)
                    logger.info(f"已备份损坏的部门文件到: {backup_file}")
            except:
                pass
            
            logger.info("尝试重新创建部门文件...")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "部门数据"
            for col, header in enumerate(DEPARTMENT_HEADERS, 1):
                ws.cell(row=1, column=col, value=header)
            wb.save(self.departments_file)
            logger.info("部门文件已重新创建（将在下次同步时重新拉取部门数据）")
    
    def _ensure_schema(self, ws):
        """确保表头正确"""
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(header)
        
        # 检查缺失的列
        missing_headers = []
        for expected_header in DEPARTMENT_HEADERS:
            if expected_header not in headers:
                missing_headers.append(expected_header)
        
        if missing_headers:
            logger.warning(f"发现缺失的列: {missing_headers}")
            # 添加缺失的列
            for header in missing_headers:
                col = len(headers) + 1
                ws.cell(row=1, column=col, value=header)
                headers.append(header)
        
        # 确保列的顺序正确
        if headers != DEPARTMENT_HEADERS:
            logger.info("调整列顺序...")
            # 创建新工作表并复制数据
            new_ws = ws.parent.create_sheet("部门数据_新")
            for col, header in enumerate(DEPARTMENT_HEADERS, 1):
                new_ws.cell(row=1, column=col, value=header)
            
            # 复制数据
            header_map = {}
            for idx, header in enumerate(headers, 1):
                if header in DEPARTMENT_HEADERS:
                    header_map[header] = idx
            
            for row in range(2, ws.max_row + 1):
                new_row_data = {}
                for header in DEPARTMENT_HEADERS:
                    if header in header_map:
                        old_col = header_map[header]
                        new_row_data[header] = ws.cell(row=row, column=old_col).value
                    else:
                        new_row_data[header] = None
                
                for col, header in enumerate(DEPARTMENT_HEADERS, 1):
                    new_ws.cell(row=row, column=col, value=new_row_data.get(header))
            
            # 删除旧工作表并重命名新工作表
            ws.parent.remove(ws)
            new_ws.title = "部门数据"
            ws = new_ws
    
    def save_departments(self, departments: List[Dict[str, Any]], append: bool = False):
        """保存部门列表到Excel文件
        
        Args:
            departments: 部门列表，每个部门包含 parent_id, name, dept_id, create_dept_group
            append: 是否追加到现有数据（默认False，覆盖）
        """
        try:
            if append and os.path.exists(self.departments_file):
                wb = openpyxl.load_workbook(self.departments_file)
                ws = wb.active
                start_row = ws.max_row + 1
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "部门数据"
                # 写入表头
                for col, header in enumerate(DEPARTMENT_HEADERS, 1):
                    ws.cell(row=1, column=col, value=header)
                start_row = 2
            
            # 写入数据
            for idx, dept in enumerate(departments, start=start_row):
                ws.cell(row=idx, column=1, value=dept.get('parent_id'))
                ws.cell(row=idx, column=2, value=dept.get('name'))
                ws.cell(row=idx, column=3, value=dept.get('dept_id'))
                ws.cell(row=idx, column=4, value=dept.get('create_dept_group', False))
            
            wb.save(self.departments_file)
            logger.info(f"成功保存 {len(departments)} 个部门到文件")
            return True
        except Exception as e:
            logger.error(f"保存部门数据失败: {e}")
            return False
    
    def get_departments(self) -> List[Dict[str, Any]]:
        """从Excel文件读取所有部门"""
        if not os.path.exists(self.departments_file):
            return []
        
        try:
            wb = openpyxl.load_workbook(self.departments_file)
            ws = wb.active
            
            departments = []
            # 从第2行开始读取（第1行是表头）
            for row in range(2, ws.max_row + 1):
                parent_id = ws.cell(row=row, column=1).value
                name = ws.cell(row=row, column=2).value
                dept_id = ws.cell(row=row, column=3).value
                create_dept_group = ws.cell(row=row, column=4).value
                
                if dept_id is not None:  # 跳过空行
                    departments.append({
                        'parent_id': parent_id,
                        'name': name,
                        'dept_id': dept_id,
                        'create_dept_group': bool(create_dept_group) if create_dept_group is not None else False
                    })
            
            return departments
        except Exception as e:
            logger.error(f"读取部门数据失败: {e}")
            return []
    
    def clear_departments(self):
        """清空部门数据（保留表头）"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "部门数据"
            # 写入表头
            for col, header in enumerate(DEPARTMENT_HEADERS, 1):
                ws.cell(row=1, column=col, value=header)
            wb.save(self.departments_file)
            logger.info("部门数据已清空")
            return True
        except Exception as e:
            logger.error(f"清空部门数据失败: {e}")
            return False
    
    def check_file_integrity(self) -> Tuple[bool, str]:
        """检查文件完整性
        
        Returns:
            (is_valid, error_message)
        """
        if not os.path.exists(self.departments_file):
            return True, ""  # 文件不存在是正常的，会在需要时创建
        
        try:
            wb = openpyxl.load_workbook(self.departments_file)
            ws = wb.active
            
            # 检查表头
            headers = []
            for col in range(1, min(ws.max_column + 1, len(DEPARTMENT_HEADERS) + 1)):
                header = ws.cell(row=1, column=col).value
                headers.append(header)
            
            # 检查必需的列
            for required_header in DEPARTMENT_HEADERS:
                if required_header not in headers:
                    return False, f"缺少必需的列: {required_header}"
            
            # 检查文件是否可以正常读取
            test_count = 0
            for row in range(2, min(ws.max_row + 1, 100)):  # 只检查前100行
                dept_id = ws.cell(row=row, column=3).value
                if dept_id is not None:
                    test_count += 1
            
            wb.close()
            return True, ""
        except Exception as e:
            return False, f"文件损坏或格式错误: {str(e)}"
    
    def build_department_tree(self) -> Dict[int, Dict[str, Any]]:
        """建立部门层次关系树
        
        Returns:
            字典，key为dept_id，value为部门信息（包含children列表）
        """
        departments = self.get_departments()
        dept_tree = {}
        
        # 第一遍：创建所有部门的节点
        for dept in departments:
            dept_id = dept.get('dept_id')
            if dept_id:
                dept_tree[dept_id] = {
                    'dept_id': dept_id,
                    'name': dept.get('name', ''),
                    'parent_id': dept.get('parent_id'),
                    'create_dept_group': dept.get('create_dept_group', False),
                    'children': []  # 子部门列表
                }
        
        # 第二遍：建立父子关系
        for dept_id, dept_info in dept_tree.items():
            parent_id = dept_info.get('parent_id')
            if parent_id and parent_id in dept_tree:
                dept_tree[parent_id]['children'].append(dept_id)
        
        return dept_tree
    
    def _collect_department_subtree_ids(self, dept_tree: Dict[int, Dict[str, Any]], root_dept_id: int) -> List[int]:
        """收集指定根部门及其所有子部门的 ID 列表（含根节点）。"""
        result_ids = [root_dept_id]

        def get_all_children(dept_id: int, depth: int = 0):
            indent = "  " * depth
            if dept_id in dept_tree:
                children = dept_tree[dept_id].get('children', [])
                if children:
                    logger.info(f"{indent}部门 {dept_id} 有 {len(children)} 个子部门")
                for child_id in children:
                    child_name = dept_tree.get(child_id, {}).get('name', '')
                    logger.info(f"{indent}  - 子部门: {child_name} (dept_id={child_id})")
                    result_ids.append(child_id)
                    get_all_children(child_id, depth + 1)

        get_all_children(root_dept_id)
        return result_ids

    def get_sync_department_ids(self, root_dept_ids: Optional[List[int]] = None) -> List[int]:
        """获取钉钉用户同步范围的部门 ID 列表（含各根部门及其子部门）。

        优先级：
        1. 参数 root_dept_ids
        2. 环境变量 DINGTALK_SYNC_DEPARTMENT_IDS
        3. 回退：按名称查找「硬件研发部」
        """
        if root_dept_ids is None:
            root_dept_ids = list(DINGTALK_SYNC_DEPARTMENT_IDS)

        if not root_dept_ids:
            logger.info("未配置 DINGTALK_SYNC_DEPARTMENT_IDS，回退按名称查找「硬件研发部」")
            return self.get_hardware_rd_department_ids()

        dept_tree = self.build_department_tree()
        result_set = set()

        for root_id in root_dept_ids:
            try:
                root_id_int = int(root_id)
            except (ValueError, TypeError):
                logger.warning(f"⚠️  无效的部门 ID: {root_id}，已跳过")
                continue

            if root_id_int not in dept_tree:
                logger.warning(f"⚠️  配置的部门 ID {root_id_int} 在部门树中不存在，已跳过")
                continue

            root_name = dept_tree[root_id_int].get('name', '未知部门')
            logger.info(f"✅ 同步根部门: {root_name} (dept_id={root_id_int})")
            logger.info(f"开始递归查找部门 (dept_id={root_id_int}) 的所有子部门...")
            subtree_ids = self._collect_department_subtree_ids(dept_tree, root_id_int)
            result_set.update(subtree_ids)

        result_ids = sorted(result_set)
        if result_ids:
            names = [dept_tree.get(did, {}).get('name', f'部门{did}') for did in result_ids[:10]]
            logger.info(
                f"✅ 配置的同步部门共 {len(result_ids)} 个"
                f"（根部门 ID: {root_dept_ids}）"
            )
            if len(result_ids) <= 10:
                logger.info(f"   部门列表: {names}")
            else:
                logger.info(f"   部门列表（前10个）: {names} ...")
        else:
            logger.warning("⚠️  未找到任何有效的同步部门（请检查 DINGTALK_SYNC_DEPARTMENT_IDS）")

        return result_ids

    def get_hardware_rd_department_ids(self) -> List[int]:
        """获取硬件研发部及其所有子部门的ID列表
        
        严格按照用户提供的流程：
        1. 从Excel中查找"硬件研发部"（精确匹配名称）
        2. 通过部门层次关系树，递归找到所有子部门
        
        Returns:
            硬件研发部及其所有子部门的ID列表
        """
        departments = self.get_departments()
        dept_tree = self.build_department_tree()
        
        # 精确查找硬件研发部的ID（优先精确匹配"硬件研发部"）
        hardware_rd_dept_id = None
        hardware_rd_name = None
        
        # 第一遍：精确匹配"硬件研发部"
        for dept in departments:
            dept_name = str(dept.get('name', '')).strip()
            if dept_name == '硬件研发部':
                hardware_rd_dept_id = dept.get('dept_id')
                hardware_rd_name = dept_name
                logger.info(f"✅ 找到硬件研发部: {dept_name} (dept_id={hardware_rd_dept_id})")
                break
        
        # 如果精确匹配失败，尝试模糊匹配
        if not hardware_rd_dept_id:
            for dept in departments:
                dept_name = str(dept.get('name', '')).strip()
                if '硬件研发' in dept_name and '部' in dept_name:
                    hardware_rd_dept_id = dept.get('dept_id')
                    hardware_rd_name = dept_name
                    logger.info(f"✅ 找到硬件研发部（模糊匹配）: {dept_name} (dept_id={hardware_rd_dept_id})")
                    break
        
        if not hardware_rd_dept_id:
            logger.warning("⚠️  未找到硬件研发部")
            logger.warning(f"   当前部门列表（前10个）: {[d.get('name') for d in departments[:10]]}")
            return []

        logger.info(f"开始递归查找硬件研发部 (dept_id={hardware_rd_dept_id}) 的所有子部门...")
        result_ids = self._collect_department_subtree_ids(dept_tree, hardware_rd_dept_id)

        logger.info(f"✅ 硬件研发部及其子部门共 {len(result_ids)} 个: {result_ids}")
        return result_ids

