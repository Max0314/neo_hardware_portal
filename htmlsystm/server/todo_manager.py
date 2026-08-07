import os
import threading
import zipfile
import openpyxl
from datetime import datetime
from typing import List, Dict, Any, Optional
from server.config import DATA_DIR, USE_MYSQL, USE_TODO_MYSQL
from server.logger import logger


def _normalize_userid_cell(value) -> str:
    """将 Excel 单元格中的 userid 规范为字符串（避免长数字科学计数法/精度丢失）。"""
    if value is None:
        return ''
    if isinstance(value, bool):
        return ''
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        try:
            return format(value, '.0f')
        except Exception:
            return str(value).strip()
    s = str(value).strip()
    if not s:
        return ''
    lower = s.lower()
    if 'e+' in lower or 'e-' in lower:
        try:
            return str(int(float(s)))
        except (ValueError, OverflowError):
            pass
    if s.endswith('.0') and s[:-2].isdigit():
        return s[:-2]
    return s


_TODOS_MIRRORS: dict = {}
_TODOS_MIRRORS_LOCK = threading.Lock()


def _todos_mirror(todos_dir: str):
    """todos 目录 → 对象存储 的写通镜像（进程内单例）。

    待办 Excel 是业务数据，此前不在任何镜像范围内——卷丢失即丢失。
    STORAGE_BACKEND=local 时 store 为 None，镜像空转。
    """
    from server.object_store import build_store_from_env
    from server.tree_mirror import TreeMirror
    with _TODOS_MIRRORS_LOCK:
        mirror = _TODOS_MIRRORS.get(todos_dir)
        if mirror is None:
            mirror = TreeMirror(todos_dir, build_store_from_env('todos'))
            _TODOS_MIRRORS[todos_dir] = mirror
            if mirror.store is not None:
                restored = mirror.restore_all()
                if restored:
                    logger.info('待办目录已从对象存储恢复 %d 个文件', restored)
        return mirror


class TodoManager:
    """待办任务管理器，支持MySQL和Excel两种存储方式"""

    def __init__(self):
        self.todos_dir = os.path.join(DATA_DIR, 'todos')
        os.makedirs(self.todos_dir, exist_ok=True)
        self.use_mysql = USE_MYSQL and USE_TODO_MYSQL
        self._mirror = _todos_mirror(self.todos_dir)

    def _save_workbook(self, wb, file_path: str) -> None:
        """原子保存并写通镜像。

        2026-06-25 曾有两个待办文件因进程在 wb.save() 直写目标途中被打断，
        留下只含 docProps 的半截 zip，此后每次读取都抛
        "There is no item named '[Content_Types].xml'"。先写临时文件再
        os.replace，保证磁盘上要么是旧的完整文件、要么是新的完整文件。
        """
        tmp_path = f'{file_path}.tmp'
        wb.save(tmp_path)
        os.replace(tmp_path, file_path)
        # 整目录差量同步（绝不抛出）：todos 共几十个小文件，哈希开销可忽略，
        # 且统一经状态清单处理新增/覆盖/删除。
        self._mirror.sync_subtree('')

    def _load_workbook_safe(self, file_path: str):
        """加载待办工作簿；损坏文件隔离后按不存在处理。

        返回 None 表示文件不存在或已损坏。损坏文件重命名为 .corrupt-<时间>.bak
        留档（不再参与镜像），避免同一残骸每次轮询都刷一遍 ERROR 堆栈。
        """
        if not os.path.exists(file_path):
            return None
        try:
            return openpyxl.load_workbook(file_path)
        except (zipfile.BadZipFile, KeyError, OSError) as e:
            quarantined = f"{file_path}.corrupt-{datetime.now().strftime('%Y%m%d%H%M%S')}.bak"
            try:
                os.replace(file_path, quarantined)
                self._mirror.sync_subtree('')
            except OSError:
                quarantined = '(隔离失败，文件保留原位)'
            logger.warning('待办文件损坏，已隔离: %s -> %s（%s）',
                           os.path.basename(file_path), os.path.basename(quarantined), e)
            return None
    
    def get_todo_file_path(self, announcement_id: str) -> str:
        """获取公告对应的待办Excel文件路径"""
        return os.path.join(self.todos_dir, f"announcement_{announcement_id}_todos.xlsx")
    
    def create_todo_file(self, announcement_id: str, source_id: str, title: str, userids: List[str], unionids: List[str], task_ids: Dict[str, str] = None, user_source_ids: Dict[str, str] = None, user_names: List[str] = None, user_usernames: List[str] = None) -> bool:
        """创建待办Excel文件并初始化数据
        
        Args:
            announcement_id: 公告ID
            source_id: 基础sourceId（用于兼容旧数据）
            title: 公告标题
            userids: 待办人员userid列表
            unionids: 待办人员unionid列表（与userids一一对应）
            task_ids: 用户ID到TaskId的映射字典（可选）
            user_source_ids: 用户ID到独立sourceId的映射字典（可选，如果提供则使用，否则使用基础source_id）
            user_names: 用户姓名列表（可选，与userids一一对应）
            user_usernames: 用户名列表（可选，与userids一一对应）
        
        Returns:
            是否创建成功
        """
        try:
            file_path = self.get_todo_file_path(announcement_id)
            
            # 如果文件已存在，先删除
            if os.path.exists(file_path):
                os.remove(file_path)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "待办任务"
            
            # 设置表头（添加TaskId列）
            headers = ["公告ID", "公告标题", "SourceId", "用户ID", "UnionId", "TaskId", "用户名", "姓名", "完成状态", "完成时间", "创建时间"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            for idx, (userid, unionid) in enumerate(zip(userids, unionids)):
                row = idx + 2
                # 确保userid是字符串格式，避免类型不匹配
                userid_str = str(userid).strip()
                task_id = task_ids.get(userid, '') if task_ids else ''
                # 如果提供了用户独立的sourceId，使用它；否则使用基础source_id
                user_source_id = user_source_ids.get(userid, source_id) if user_source_ids else source_id
                
                # 获取用户名和姓名（如果提供了列表）
                username = user_usernames[idx] if user_usernames and idx < len(user_usernames) else ''
                name = user_names[idx] if user_names and idx < len(user_names) else ''
                
                ws.cell(row=row, column=1, value=announcement_id)
                ws.cell(row=row, column=2, value=title)
                ws.cell(row=row, column=3, value=user_source_id)  # 使用用户独立的sourceId
                userid_cell = ws.cell(row=row, column=4, value=userid_str)
                userid_cell.number_format = '@'  # 文本格式，防止长 userid 被 Excel 转成数值
                unionid_cell = ws.cell(row=row, column=5, value=str(unionid).strip() if unionid else '')
                unionid_cell.number_format = '@'
                ws.cell(row=row, column=6, value=task_id)  # TaskId
                ws.cell(row=row, column=7, value=username)  # 用户名，从参数获取
                ws.cell(row=row, column=8, value=name)  # 姓名，从参数获取
                ws.cell(row=row, column=9, value='未完成')  # 完成状态
                ws.cell(row=row, column=10, value='')  # 完成时间
                ws.cell(row=row, column=11, value=now)  # 创建时间
            
            self._save_workbook(wb, file_path)
            logger.info(f"成功创建待办Excel文件: {file_path}, 共 {len(userids)} 条记录")
            return True
            
        except Exception as e:
            logger.error(f"创建待办Excel文件失败: {e}", exc_info=True)
            return False
    
    def get_user_todo_status(self, announcement_id: str, userid: str) -> Optional[Dict[str, Any]]:
        """获取用户在该公告的待办状态
        
        重要：优先从内存缓存读取，如果缓存未命中才读取Excel文件
        这样可以避免在高并发场景下的性能问题
        
        Args:
            announcement_id: 公告ID
            userid: 用户ID
        
        Returns:
            待办状态信息，如果未找到返回None
        """
        # 优先尝试从内存缓存获取
        try:
            from server.data_preloader import get_data_preloader
            from server.config import PRELOAD_TODOS
            
            preloader = get_data_preloader()
            if preloader and PRELOAD_TODOS:
                todo_status = preloader.get_user_todo_status_cached(announcement_id, userid)
                if todo_status:
                    logger.debug(f"从内存缓存获取待办状态: announcement_id={announcement_id}, userid={userid}")
                    return todo_status
        except Exception as e:
            logger.debug(f"从内存缓存获取待办状态失败，回退到文件读取: {e}")
        
        # 缓存未命中，从Excel文件读取（作为后备方案）
        try:
            file_path = self.get_todo_file_path(announcement_id)
            if not os.path.exists(file_path):
                logger.debug(f"待办Excel文件不存在: {file_path}, announcement_id={announcement_id}")
                return None
            
            wb = self._load_workbook_safe(file_path)
            if wb is None:
                return None
            ws = wb.active
            
            # 查找用户ID列（第4列）
            userid_str = str(userid).strip()
            logger.debug(f"在Excel中查找userid: {userid_str}, 文件: {file_path}, 总行数: {ws.max_row}")
            
            for row in range(2, ws.max_row + 1):
                cell_userid_str = _normalize_userid_cell(ws.cell(row=row, column=4).value)
                
                if cell_userid_str == userid_str:
                    # 找到用户记录
                    source_id = ws.cell(row=row, column=3).value or ''
                    task_id = ws.cell(row=row, column=6).value or ''  # TaskId在第6列
                    status = ws.cell(row=row, column=9).value or '未完成'  # 完成状态在第9列
                    complete_time = ws.cell(row=row, column=10).value or ''  # 完成时间在第10列
                    unionid = ws.cell(row=row, column=5).value or ''
                    username = ws.cell(row=row, column=7).value or ''  # 用户名在第7列
                    name = ws.cell(row=row, column=8).value or ''  # 姓名在第8列
                    title = ws.cell(row=row, column=2).value or ''  # 公告标题在第2列
                    
                    todo_status = {
                        'announcement_id': announcement_id,
                        'title': title,
                        'source_id': source_id,
                        'task_id': task_id,
                        'userid': userid,
                        'unionid': unionid,
                        'username': username,
                        'name': name,
                        'status': status,
                        'done': status == '已完成',
                        'complete_time': complete_time
                    }
                    
                    # 如果从文件读取到数据，尝试更新缓存（确保缓存是最新的）
                    try:
                        from server.data_preloader import get_data_preloader
                        from server.config import PRELOAD_TODOS
                        import time
                        
                        preloader = get_data_preloader()
                        if preloader and PRELOAD_TODOS:
                            with preloader.cache_lock:
                                if announcement_id in preloader.todos_cache:
                                    todos, _ = preloader.todos_cache[announcement_id]
                                    # 查找并更新对应的待办
                                    updated = False
                                    for todo in todos:
                                        if str(todo.get('userid', '')).strip() == userid_str:
                                            # 更新缓存中的待办数据
                                            todo.update(todo_status)
                                            updated = True
                                            break
                                    
                                    # 如果找到了，更新时间戳；如果没找到，添加到缓存
                                    if updated:
                                        preloader.todos_cache[announcement_id] = (todos, time.time())
                                    else:
                                        # 如果缓存中没有，添加到缓存
                                        todos.append(todo_status)
                                        preloader.todos_cache[announcement_id] = (todos, time.time())
                    except Exception as e:
                        logger.debug(f"更新缓存失败（不影响主流程）: {e}")
                    
                    logger.debug(f"找到用户待办记录: userid={userid_str}, status={status}, done={status == '已完成'}")
                    
                    return todo_status
            
            logger.debug(f"在Excel中未找到匹配的userid: {userid_str}, 文件: {file_path}")
            return None
            
        except Exception as e:
            logger.error(f"获取用户待办状态失败: {e}", exc_info=True)
            return None
    
    def update_todo_status(self, announcement_id: str, userid: str, done: bool = True, update_cache: bool = True, immediate_save: bool = False) -> bool:
        """更新用户的待办状态
        
        Args:
            announcement_id: 公告ID
            userid: 用户ID
            done: 是否完成
            update_cache: 是否同时更新内存缓存（默认True）
            immediate_save: 是否立即保存到文件（默认False，由自动保存线程定期保存）
        
        Returns:
            是否更新成功
        """
        try:
            # 先更新内存缓存（如果启用）
            if update_cache:
                try:
                    from server.data_preloader import get_data_preloader
                    from server.config import PRELOAD_TODOS
                    
                    preloader = get_data_preloader()
                    if preloader and PRELOAD_TODOS:
                        # 更新内存缓存
                        with preloader.cache_lock:
                            if announcement_id in preloader.todos_cache:
                                todos, old_timestamp = preloader.todos_cache[announcement_id]
                                # 查找并更新对应的待办
                                userid_str = str(userid).strip()
                                updated = False
                                for todo in todos:
                                    if str(todo.get('userid', '')).strip() == userid_str:
                                        todo['status'] = '已完成' if done else '未完成'
                                        todo['done'] = done
                                        if done:
                                            todo['complete_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                        else:
                                            todo['complete_time'] = ''
                                        updated = True
                                        break
                                
                                # 如果找到了并更新了，更新时间戳并标记为dirty
                                if updated:
                                    import time
                                    # 更新缓存时间戳，确保后续读取使用最新数据
                                    preloader.todos_cache[announcement_id] = (todos, time.time())
                                    # 标记为dirty，需要保存（由自动保存线程定期保存）
                                    preloader.mark_todos_dirty(announcement_id)
                                    logger.debug(f"已更新内存缓存中的待办状态: announcement_id={announcement_id}, userid={userid}, done={done}")
                                else:
                                    logger.warning(f"在缓存中未找到待办记录: announcement_id={announcement_id}, userid={userid}")
                except Exception as e:
                    logger.warning(f"更新内存缓存失败，继续更新文件: {e}")
            
            # 如果要求立即保存，或者自动保存未启用，则立即更新Excel文件
            from server.config import TODO_AUTO_SAVE_ENABLED
            if immediate_save or not TODO_AUTO_SAVE_ENABLED:
                file_path = self.get_todo_file_path(announcement_id)
                if not os.path.exists(file_path):
                    logger.warning(f"待办Excel文件不存在: {file_path}")
                    return False
                
                # 使用文件锁确保并发安全
                from server.user_manager import get_file_lock
                file_lock = get_file_lock(file_path)
                
                try:
                    with file_lock:
                        wb = self._load_workbook_safe(file_path)
                        if wb is None:
                            logger.warning("待办文件缺失或已隔离，无法更新")
                            return False
                        ws = wb.active
                        
                        # 查找用户ID列（第4列）
                        found = False
                        for row in range(2, ws.max_row + 1):
                            cell_userid_str = _normalize_userid_cell(ws.cell(row=row, column=4).value)
                            if cell_userid_str and cell_userid_str == str(userid).strip():
                                # 更新状态（完成状态在第9列，完成时间在第10列）
                                ws.cell(row=row, column=9, value='已完成' if done else '未完成')
                                if done:
                                    ws.cell(row=row, column=10, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                                else:
                                    ws.cell(row=row, column=10, value='')
                                found = True
                                break
                        
                        if found:
                            self._save_workbook(wb, file_path)
                            # 确保文件写入完成（刷新文件系统缓冲区）
                            try:
                                # 打开文件并同步到磁盘
                                with open(file_path, 'rb+') as f:
                                    os.fsync(f.fileno())
                            except Exception as e:
                                logger.debug(f"文件同步失败（不影响主流程）: {e}")
                            logger.info(f"成功更新待办状态（立即保存）: announcement_id={announcement_id}, userid={userid}, done={done}")
                            
                            # 文件保存后，再次确保缓存已更新（因为可能有其他进程读取文件）
                            if update_cache:
                                try:
                                    from server.data_preloader import get_data_preloader
                                    from server.config import PRELOAD_TODOS
                                    import time
                                    
                                    preloader = get_data_preloader()
                                    if preloader and PRELOAD_TODOS:
                                        with preloader.cache_lock:
                                            if announcement_id in preloader.todos_cache:
                                                todos, _ = preloader.todos_cache[announcement_id]
                                                userid_str = str(userid).strip()
                                                for todo in todos:
                                                    if str(todo.get('userid', '')).strip() == userid_str:
                                                        # 确保缓存状态与文件一致
                                                        todo['status'] = '已完成' if done else '未完成'
                                                        todo['done'] = done
                                                        if done:
                                                            todo['complete_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                                        else:
                                                            todo['complete_time'] = ''
                                                        preloader.todos_cache[announcement_id] = (todos, time.time())
                                                        logger.debug(f"文件保存后，已同步更新缓存: announcement_id={announcement_id}, userid={userid}, done={done}")
                                                        break
                                except Exception as e:
                                    logger.warning(f"文件保存后更新缓存失败: {e}")
                            
                            return True
                        else:
                            logger.warning(f"未找到用户的待办记录: announcement_id={announcement_id}, userid={userid}")
                            return False
                except Exception as e:
                    logger.error(f"保存待办状态到文件失败: {e}", exc_info=True)
                    return False
            else:
                # 只更新内存缓存，由自动保存线程定期保存到文件
                logger.debug(f"待办状态已更新到内存缓存，将在下次自动保存时写入文件: announcement_id={announcement_id}, userid={userid}, done={done}")
                return True
            
        except Exception as e:
            logger.error(f"更新待办状态失败: {e}", exc_info=True)
            return False
    
    def save_todos_to_file(self, announcement_id: str, todos: List[Dict[str, Any]]) -> bool:
        """将内存中的待办数据保存到Excel文件
        
        Args:
            announcement_id: 公告ID
            todos: 待办数据列表
        
        Returns:
            是否保存成功
        """
        try:
            file_path = self.get_todo_file_path(announcement_id)
            if not todos:
                logger.warning(f"待办数据为空，跳过保存: announcement_id={announcement_id}")
                return False
            
            # 如果文件不存在，需要先创建（使用第一个待办的公告信息）
            if not os.path.exists(file_path):
                logger.warning(f"待办Excel文件不存在，无法保存: {file_path}")
                return False
            
            wb = self._load_workbook_safe(file_path)
            if wb is None:
                return False
            ws = wb.active
            
            # 构建userid到待办数据的映射，便于快速查找
            todos_by_userid = {}
            for todo in todos:
                userid = str(todo.get('userid', '')).strip()
                if userid:
                    todos_by_userid[userid] = todo
            
            # 更新Excel文件中的数据
            updated_count = 0
            for row in range(2, ws.max_row + 1):
                cell_userid_str = _normalize_userid_cell(ws.cell(row=row, column=4).value)
                
                if cell_userid_str in todos_by_userid:
                    todo = todos_by_userid[cell_userid_str]
                    # 更新状态（完成状态在第9列，完成时间在第10列）
                    ws.cell(row=row, column=9, value=todo.get('status', '未完成'))
                    ws.cell(row=row, column=10, value=todo.get('complete_time', ''))
                    updated_count += 1
            
            if updated_count > 0:
                self._save_workbook(wb, file_path)
                logger.debug(f"成功保存待办数据到文件: announcement_id={announcement_id}, 更新了{updated_count}条记录")
                return True
            else:
                logger.warning(f"没有找到需要更新的待办记录: announcement_id={announcement_id}")
                return False
            
        except Exception as e:
            logger.error(f"保存待办数据到文件失败: {e}", exc_info=True)
            return False
    
    def get_all_todos(self, announcement_id: str) -> List[Dict[str, Any]]:
        """获取公告的所有待办记录
        
        Args:
            announcement_id: 公告ID
        
        Returns:
            待办记录列表
        """
        try:
            file_path = self.get_todo_file_path(announcement_id)
            if not os.path.exists(file_path):
                logger.debug(f"待办文件不存在: {file_path}")
                return []
            
            wb = self._load_workbook_safe(file_path)
            if wb is None:
                return []
            ws = wb.active
            
            todos = []
            # 确保至少检查到第2行（即使max_row为1，range(2, 2)也是空范围，这是正确的）
            # 但如果max_row为1，说明只有表头，没有数据行
            max_row = ws.max_row if ws.max_row > 1 else 1
            if max_row < 2:
                logger.debug(f"待办文件 {file_path} 只有表头，无数据行 (max_row={max_row})")
                return []
            
            logger.debug(f"读取待办文件: {file_path}, max_row={max_row}, 将从第2行开始读取")
            for row in range(2, max_row + 1):
                # 读取所有列的数据
                announcement_id_val = ws.cell(row=row, column=1).value
                title_val = ws.cell(row=row, column=2).value
                source_id_val = ws.cell(row=row, column=3).value
                userid_val = ws.cell(row=row, column=4).value
                userid_str = _normalize_userid_cell(userid_val)
                unionid_val = ws.cell(row=row, column=5).value
                task_id_val = ws.cell(row=row, column=6).value
                username_val = ws.cell(row=row, column=7).value
                name_val = ws.cell(row=row, column=8).value
                status_val = ws.cell(row=row, column=9).value
                complete_time_val = ws.cell(row=row, column=10).value
                create_time_val = ws.cell(row=row, column=11).value
                
                # 如果announcement_id和userid都为空，跳过这一行（可能是空行）
                if not announcement_id_val and not userid_val:
                    continue
                
                # 即使userid为空，也记录这一行（可能后续会补充）
                # 但至少要有一个标识（announcement_id或userid）
                if not announcement_id_val and not userid_val:
                    continue
                
                todos.append({
                    'announcement_id': str(announcement_id_val).strip() if announcement_id_val else '',
                    'title': str(title_val).strip() if title_val else '',
                    'source_id': str(source_id_val).strip() if source_id_val else '',
                    'userid': userid_str,
                    'unionid': str(unionid_val).strip() if unionid_val else '',
                    'task_id': str(task_id_val).strip() if task_id_val else '',
                    'username': str(username_val).strip() if username_val else '',
                    'name': str(name_val).strip() if name_val else '',
                    'status': str(status_val).strip() if status_val else '未完成',
                    'done': str(status_val).strip() in ['已完成', 'done'] if status_val else False,
                    'complete_time': str(complete_time_val).strip() if complete_time_val else '',
                    'create_time': str(create_time_val).strip() if create_time_val else ''
                })
            
            return todos
            
        except Exception as e:
            logger.error(f"获取待办列表失败: {e}", exc_info=True)
            return []

